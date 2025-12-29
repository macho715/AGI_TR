#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# Py3.11.8 / pandas 2.3.3 / openpyxl 3.1.5

"""
populate_template.py

Fill 07_Stage_Calc input columns from engine stage CSV (e.g., stage_results.csv).

Default mapping (CSV -> 07_Stage_Calc)
- Stage / Stage_ID           -> A (Stage_ID)
- x_stage_m                  -> F (x_stage_m)
- W_stage_t                  -> G (W_stage_t)
- Disp_t                     -> H (Displacement_t)
- Tmean_m                    -> I (Tmean_m)
- Dfwd_m                     -> K (Draft_FWD_m)
- Daft_m                     -> L (Draft_AFT_m)
- GM(m)                      -> U (GM_m)

Notes
- Keeps existing formulas in calc columns.
- If the stage row exists (Stage_ID match), overwrites/FillBlank based on --mode.
- If not found, writes into the first empty row; if no empty rows, extends rows.
- Writes LOG sheet + text log.
"""

from __future__ import annotations

import argparse
import re
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# Headers SSOT support (optional)
try:
    from ssot.headers_registry import load_registry, get_header_row, apply_schema
    from ssot.headers_writer import HeadersWriter

    HEADERS_SSOT_AVAILABLE = True
except ImportError:
    HEADERS_SSOT_AVAILABLE = False


# -----------------------------
# Helpers
# -----------------------------
def _norm_text(x) -> str:
    if x is None:
        return ""
    return str(x).strip()


def _norm_col(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(name).lower())


def pick_column(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
    """
    Find best matching column from df given candidate names.
    - exact match first
    - then normalized match (case/space/symbol insensitive)
    """
    cols = list(df.columns)
    for cand in candidates:
        if cand in cols:
            return cand

    norm_map = {_norm_col(c): c for c in cols}
    for cand in candidates:
        key = _norm_col(cand)
        if key in norm_map:
            return norm_map[key]
    return None


def safe_float(x) -> Optional[float]:
    if x is None:
        return None


def _norm_header(val) -> str:
    return _norm_col(val)


def _last_header_col(ws: Worksheet, header_row: int) -> int:
    last_col = 1
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if v is not None and str(v).strip() != "":
            last_col = c
    return last_col


def find_col_by_header(
    ws: Worksheet, header_row: int, candidates: List[str]
) -> Optional[int]:
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if v is not None and str(v).strip() != "":
            headers[str(v).strip()] = c
    for cand in candidates:
        if cand in headers:
            return headers[cand]
    norm_headers = {_norm_header(k): v for k, v in headers.items()}
    for cand in candidates:
        key = _norm_header(cand)
        if key in norm_headers:
            return norm_headers[key]
    return None


def ensure_column(ws: Worksheet, header_row: int, name: str) -> int:
    existing = find_col_by_header(ws, header_row, [name])
    if existing:
        return existing
    insert_at = _last_header_col(ws, header_row) + 1
    ws.insert_cols(insert_at)
    ws.cell(header_row, insert_at).value = name
    return insert_at
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip()
    if s == "" or s.lower() in ("nan", "none", "null"):
        return None
    try:
        return float(s)
    except Exception:
        try:
            return float(s.replace(",", ""))
        except Exception:
            return None


def get_or_create_sheet(wb, name: str):
    if name in wb.sheetnames:
        return wb[name]
    return wb.create_sheet(name)


def append_log(ws: Worksheet, msg: str) -> None:
    """
    Append log line to LOG sheet (A:Time, B:Message)
    """
    if ws.max_row == 1 and ws["A1"].value is None:
        ws["A1"] = "Time"
        ws["B1"] = "Message"
    r = ws.max_row + 1
    ws.cell(r, 1, datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    ws.cell(r, 2, msg)


def find_header_row(
    ws: Worksheet, header_text: str = "Stage_ID", search_rows: int = 40
) -> Tuple[int, int]:
    """
    Find header row where column A equals header_text. Return (header_row, data_start_row).
    Fallback to (20, 21).
    """
    for r in range(1, search_rows + 1):
        if _norm_text(ws.cell(r, 1).value) == header_text:
            return r, r + 1
    return 20, 21


def find_or_create_stage_row(
    ws: Worksheet,
    stage_id: str,
    data_start: int,
    soft_last_row: int = 200,
) -> int:
    """
    Find existing row where A == stage_id.
    Else first empty row in A.
    Else extend after soft_last_row (or ws.max_row).
    """
    last_scan = max(soft_last_row, ws.max_row)
    for r in range(data_start, last_scan + 1):
        if _norm_text(ws.cell(r, 1).value) == stage_id:
            return r

    for r in range(data_start, last_scan + 1):
        if _norm_text(ws.cell(r, 1).value) == "":
            ws.cell(r, 1).value = stage_id
            return r

    r = last_scan + 1
    ws.cell(r, 1).value = stage_id
    return r


def ensure_stage_formulas(ws: Worksheet, r: int) -> None:
    """
    For rows that didn't have formulas (e.g., template rows beyond initial 9),
    inject core formulas if blank.
    - Only sets if cell is empty (None/"").
    """

    def set_if_blank(col: int, formula: str):
        cell = ws.cell(r, col)
        if cell.value is None or str(cell.value).strip() == "":
            cell.value = formula

    # Parameter cells in 07_Stage_Calc (as defined in create script)
    GATE_AFT_MIN = "$B$4"
    GATE_FWD_MAX = "$B$5"
    TRIM_LIMIT = "$B$6"
    DEPTHREF = "$B$7"
    UKC_MIN = "$B$8"
    SQUAT = "$B$9"
    SAFETY = "$B$10"

    # Columns (A=1..AG=33)
    COL_HOLD = 2
    COL_CRIT = 3
    COL_REMARK = 4
    COL_PROG = 5
    COL_TRIM = 10
    COL_FBMIN = 13
    COL_TIDE_REQ = 14
    COL_TIDE_MARG = 16
    COL_UKC_MIN = 17
    COL_UKC_FWD = 18
    COL_UKC_AFT = 19
    COL_GATE_FWD = 22
    COL_GATE_AFT = 23
    COL_GATE_TRIM = 24
    COL_GATE_UKC = 25
    COL_GATE_RAMP = 26
    COL_HP_BAND = 27
    COL_GONOGO = 28

    set_if_blank(
        COL_HOLD,
        f'=IF($A{r}="","",IFERROR(INDEX(\'06_Stage_Plan\'!$E:$E,MATCH($A{r},\'06_Stage_Plan\'!$A:$A,0)),""))',
    )
    set_if_blank(
        COL_CRIT,
        f'=IF($A{r}="","",IFERROR(INDEX(\'06_Stage_Plan\'!$F:$F,MATCH($A{r},\'06_Stage_Plan\'!$A:$A,0)),""))',
    )
    set_if_blank(
        COL_REMARK,
        f'=IF($A{r}="","",IFERROR(INDEX(\'06_Stage_Plan\'!$H:$H,MATCH($A{r},\'06_Stage_Plan\'!$A:$A,0)),""))',
    )
    set_if_blank(
        COL_PROG,
        f'=IF($A{r}="","",IFERROR(INDEX(\'06_Stage_Plan\'!$B:$B,MATCH($A{r},\'06_Stage_Plan\'!$A:$A,0)),""))',
    )

    set_if_blank(COL_TRIM, f'=IF(OR($K{r}="",$L{r}=""),"",($L{r}-$K{r})*100)')

    set_if_blank(
        COL_FBMIN,
        f'=IF(OR($K{r}="",$L{r}=""),"",MIN('
        f"INDEX('01_SSOT_Master'!$B:$B,MATCH(\"Molded Depth (D)\",'01_SSOT_Master'!$A:$A,0))-$K{r},"
        f"INDEX('01_SSOT_Master'!$B:$B,MATCH(\"Molded Depth (D)\",'01_SSOT_Master'!$A:$A,0))-$L{r}))",
    )

    set_if_blank(COL_UKC_MIN, f'=IF($A{r}="","",{UKC_MIN})')

    set_if_blank(
        COL_TIDE_REQ,
        f'=IF(OR($K{r}="",$L{r}=""),"",MAX(0,(MAX($K{r},$L{r})+{SQUAT}+{SAFETY}+{UKC_MIN})-{DEPTHREF}))',
    )
    set_if_blank(COL_TIDE_MARG, f'=IF(OR($O{r}="",$N{r}=""),"",$O{r}-$N{r})')

    set_if_blank(
        COL_UKC_FWD,
        f'=IF(OR($O{r}="",$K{r}=""),"",({DEPTHREF}+$O{r})-($K{r}+{SQUAT}+{SAFETY}))',
    )
    set_if_blank(
        COL_UKC_AFT,
        f'=IF(OR($O{r}="",$L{r}=""),"",({DEPTHREF}+$O{r})-($L{r}+{SQUAT}+{SAFETY}))',
    )

    set_if_blank(
        COL_GATE_FWD,
        f'=IF($A{r}="","",IF($C{r}<>"Y","N/A",IF($K{r}<={GATE_FWD_MAX},"PASS","LIMIT")))',
    )
    set_if_blank(
        COL_GATE_AFT,
        f'=IF($A{r}="","",IF($L{r}>={GATE_AFT_MIN},"PASS","LIMIT"))',
    )
    set_if_blank(
        COL_GATE_TRIM,
        f'=IF($A{r}="","",IF(ABS($J{r})<={TRIM_LIMIT},"PASS","LIMIT"))',
    )
    set_if_blank(
        COL_GATE_UKC,
        f'=IF($A{r}="","",IF(MIN($R{r},$S{r})>={UKC_MIN},"PASS","LIMIT"))',
    )
    set_if_blank(
        COL_GATE_RAMP,
        f'=IF($A{r}="","",IF($T{r}="","VERIFY",IF($T{r}<=6,"PASS","LIMIT")))',
    )

    set_if_blank(
        COL_HP_BAND,
        f'=IF($B{r}<>"Y","-",IF(OR($V{r}="LIMIT",$W{r}="LIMIT",$X{r}="LIMIT",$Y{r}="LIMIT"),"STOP","GO"))',
    )
    set_if_blank(
        COL_GONOGO,
        f'=IF($A{r}="","",IF($AA{r}="STOP","ABORT",IF($AA{r}="RECALC","NO-GO","GO")))',
    )


def write_value(ws: Worksheet, r: int, col: int, value, mode: str) -> bool:
    """
    mode:
      - overwrite: always write if value is not None
      - fillblank: write only if cell is blank
    returns True if written
    """
    if value is None:
        return False
    cell = ws.cell(r, col)
    if mode == "fillblank":
        if cell.value is not None and str(cell.value).strip() != "":
            return False
    cell.value = value
    return True


# -----------------------------
# Main
# -----------------------------
def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--template", required=True, help="Input template xlsx path")
    ap.add_argument("--stage-csv", required=True, help="stage_results.csv path")
    ap.add_argument(
        "--qa-csv",
        default="",
        help="Optional pipeline_stage_QA.csv to populate Tide/UKC columns",
    )
    ap.add_argument("--out", required=True, help="Output populated xlsx path")
    ap.add_argument(
        "--mode",
        choices=["overwrite", "fillblank"],
        default="overwrite",
        help="Write policy",
    )
    ap.add_argument("--sheet", default="07_Stage_Calc", help="Target sheet name")
    ap.add_argument(
        "--soft-last-row",
        type=int,
        default=200,
        help="Prefer writing within this last row",
    )
    ap.add_argument(
        "--headers-registry",
        default="",
        help="Path to headers_registry.json for SSOT header management (optional).",
    )
    args = ap.parse_args()

    template_path = Path(args.template).resolve()
    stage_csv_path = Path(args.stage_csv).resolve()
    qa_csv_path = Path(args.qa_csv).resolve() if args.qa_csv else None
    out_path = Path(args.out).resolve()

    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")
    if not stage_csv_path.exists():
        raise FileNotFoundError(f"Stage CSV not found: {stage_csv_path}")
    if qa_csv_path and not qa_csv_path.exists():
        raise FileNotFoundError(f"QA CSV not found: {qa_csv_path}")

    log_txt = out_path.with_suffix(".populate.log.txt")

    df = pd.read_csv(stage_csv_path)
    if df.empty:
        raise ValueError("Stage CSV is empty.")

    col_stage = pick_column(
        df, ["Stage_ID", "Stage", "stage", "StageName", "Stage Name"]
    )
    col_x = pick_column(df, ["x_stage_m", "x_m", "x", "xstage"])
    col_w = pick_column(df, ["W_stage_t", "W_t", "W", "Weight_t"])
    col_disp = pick_column(df, ["Disp_t", "Displacement_t", "Displacement", "Disp"])
    col_tmean = pick_column(df, ["Tmean_m", "Tmean", "T_mean_m", "Tmean(m)"])
    col_dfwd = pick_column(
        df, ["Dfwd_m", "Draft_FWD_m", "Fwd Draft(m)", "FWD Draft(m)", "Dfwd"]
    )
    col_daft = pick_column(
        df, ["Daft_m", "Draft_AFT_m", "Aft Draft(m)", "AFT Draft(m)", "Daft"]
    )
    col_gm = pick_column(df, ["GM(m)", "GM_m", "GM", "GM (m)"])

    if not col_stage:
        raise ValueError(
            f"Cannot find Stage column in CSV. Columns: {list(df.columns)}"
        )

    wb = load_workbook(template_path)
    if args.sheet not in wb.sheetnames:
        raise KeyError(f"Sheet not found: {args.sheet}")
    ws = wb[args.sheet]
    ws_log = get_or_create_sheet(wb, "LOG")

    # Headers SSOT support
    writer = None
    registry_header_row = None
    if args.headers_registry and HEADERS_SSOT_AVAILABLE:
        registry_path = Path(args.headers_registry)
        if registry_path.exists():
            try:
                writer = HeadersWriter(registry_path)
                if writer.registry and args.sheet == "07_Stage_Calc":
                    registry_header_row = get_header_row(
                        writer.registry, "BRYAN_STAGE_CALC"
                    )
                    if registry_header_row:
                        append_log(
                            ws_log,
                            f"[Headers SSOT] Using header_row={registry_header_row} from registry",
                        )
            except Exception as e:
                append_log(ws_log, f"[WARN] Headers registry load failed: {e}")

    # Find header row (use registry if available, otherwise auto-detect)
    if registry_header_row:
        header_row = registry_header_row
        data_start = header_row + 1
    else:
        header_row, data_start = find_header_row(ws, "Stage_ID", search_rows=60)

    def _stage_key(val: str) -> str:
        return re.sub(r"\\s+", " ", str(val).strip()).lower()

    qa_df = None
    qa_stage_col = None
    qa_cols_map: Dict[str, Optional[str]] = {}
    qa_idx = None
    if qa_csv_path:
        qa_df = pd.read_csv(qa_csv_path, encoding="utf-8-sig")
        qa_stage_col = pick_column(
            qa_df, ["Stage", "Stage_ID", "Stage Name", "StageName"]
        )
        if not qa_stage_col:
            raise ValueError(
                f"QA CSV missing Stage column. Columns: {list(qa_df.columns)}"
            )

        # Use headers registry if available for better column mapping
        if writer and writer.registry and args.sheet == "07_Stage_Calc":
            try:
                # Apply schema to QA DataFrame
                qa_std = apply_schema(
                    qa_df,
                    writer.registry,
                    "BRYAN_STAGE_CALC",
                    keep_extra=False,
                    fill_missing=False,
                )
                # Map standardized columns to Excel column names
                deliverable = writer.registry.deliverables["BRYAN_STAGE_CALC"]
                for field_key in deliverable.columns:
                    if field_key in qa_std.columns:
                        excel_header = (
                            deliverable.output_headers.get(field_key)
                            or writer.registry.fields[field_key].default_header
                        )
                        qa_cols_map[excel_header] = field_key
                # Use standardized DataFrame
                qa_df = qa_std
                qa_stage_col = (
                    "Stage_ID" if "Stage_ID" in qa_df.columns else qa_stage_col
                )
                append_log(ws_log, "[Headers SSOT] Applied schema to QA CSV")
            except Exception as e:
                append_log(
                    ws_log,
                    f"[WARN] Headers SSOT schema application failed: {e}, using fallback",
                )
                # Fallback to original logic
                qa_cols_map = {
                    "Forecast_tide_m": pick_column(
                        qa_df, ["Forecast_tide_m", "Forecast_Tide_m"]
                    ),
                    "DepthRef_m": pick_column(qa_df, ["DepthRef_m", "Depth_Ref_m"]),
                    "UKC_min_m": pick_column(
                        qa_df, ["UKC_min_m", "UKC_Min_m", "UKC_Min_Required_m"]
                    ),
                    "UKC_fwd_m": pick_column(qa_df, ["UKC_fwd_m", "UKC_FWD_m"]),
                    "UKC_aft_m": pick_column(qa_df, ["UKC_aft_m", "UKC_AFT_m"]),
                    "Tide_required_m": pick_column(
                        qa_df, ["Tide_required_m", "Required_WL_for_UKC_m"]
                    ),
                    "Tide_margin_m": pick_column(qa_df, ["Tide_margin_m"]),
                    "Tide_verdict": pick_column(
                        qa_df, ["Tide_verdict", "Tide_verification"]
                    ),
                }
        else:
            # Original fallback logic
            qa_cols_map = {
                "Forecast_tide_m": pick_column(
                    qa_df, ["Forecast_tide_m", "Forecast_Tide_m"]
                ),
                "DepthRef_m": pick_column(qa_df, ["DepthRef_m", "Depth_Ref_m"]),
                "UKC_min_m": pick_column(
                    qa_df, ["UKC_min_m", "UKC_Min_m", "UKC_Min_Required_m"]
                ),
                "UKC_fwd_m": pick_column(qa_df, ["UKC_fwd_m", "UKC_FWD_m"]),
                "UKC_aft_m": pick_column(qa_df, ["UKC_aft_m", "UKC_AFT_m"]),
                "Tide_required_m": pick_column(
                    qa_df, ["Tide_required_m", "Required_WL_for_UKC_m"]
                ),
                "Tide_margin_m": pick_column(qa_df, ["Tide_margin_m"]),
                "Tide_verdict": pick_column(
                    qa_df, ["Tide_verdict", "Tide_verification"]
                ),
            }

        qa_df["_stage_norm"] = qa_df[qa_stage_col].apply(_stage_key)
        qa_idx = qa_df.set_index("_stage_norm")

        for excel_col, src_col in qa_cols_map.items():
            if src_col:
                ensure_column(ws, header_row, excel_col)

    append_log(
        ws_log,
        f"[populate_template] template={template_path.name}, stage_csv={stage_csv_path.name}, mode={args.mode}",
    )
    append_log(
        ws_log,
        "Detected CSV cols: "
        f"Stage={col_stage}, x={col_x}, W={col_w}, Disp={col_disp}, "
        f"Tmean={col_tmean}, Dfwd={col_dfwd}, Daft={col_daft}, GM={col_gm}",
    )
    if qa_csv_path:
        append_log(
            ws_log,
            "Detected QA cols: "
            + ", ".join([f"{k}={v}" for k, v in qa_cols_map.items() if v]),
        )

    written_count = 0
    stage_count = 0

    for _, row in df.iterrows():
        stage_id = _norm_text(row[col_stage])
        if stage_id == "":
            continue

        stage_count += 1
        r_excel = find_or_create_stage_row(
            ws, stage_id, data_start, soft_last_row=args.soft_last_row
        )

        ensure_stage_formulas(ws, r_excel)

        v_x = safe_float(row[col_x]) if col_x else None
        v_w = safe_float(row[col_w]) if col_w else None
        v_d = safe_float(row[col_disp]) if col_disp else None
        v_t = safe_float(row[col_tmean]) if col_tmean else None
        v_k = safe_float(row[col_dfwd]) if col_dfwd else None
        v_l = safe_float(row[col_daft]) if col_daft else None
        v_g = safe_float(row[col_gm]) if col_gm else None

        wrote_any = False
        wrote_any |= write_value(ws, r_excel, 6, v_x, args.mode)  # x_stage_m
        wrote_any |= write_value(ws, r_excel, 7, v_w, args.mode)  # W_stage_t
        wrote_any |= write_value(ws, r_excel, 8, v_d, args.mode)  # Displacement_t
        wrote_any |= write_value(ws, r_excel, 9, v_t, args.mode)  # Tmean_m
        wrote_any |= write_value(ws, r_excel, 11, v_k, args.mode)  # Draft_FWD_m
        wrote_any |= write_value(ws, r_excel, 12, v_l, args.mode)  # Draft_AFT_m
        wrote_any |= write_value(ws, r_excel, 21, v_g, args.mode)  # GM_m

        if qa_idx is not None:
            key = _stage_key(stage_id)
            if key in qa_idx.index:
                qa_row = qa_idx.loc[key]
                if not isinstance(qa_row, pd.Series):
                    qa_row = qa_row.iloc[0]
                for excel_col, src_col in qa_cols_map.items():
                    if not src_col:
                        continue
                    # Handle both direct column name and field key
                    if src_col in qa_row.index:
                        v = qa_row.get(src_col)
                    else:
                        # Try as field key (if using standardized DataFrame)
                        v = qa_row.get(src_col) if src_col in qa_row.index else None
                    if pd.isna(v) if v is not None else True:
                        continue
                    col_idx = find_col_by_header(ws, header_row, [excel_col])
                    if col_idx:
                        wrote_any |= write_value(ws, r_excel, col_idx, v, args.mode)

        if wrote_any:
            written_count += 1

    append_log(ws_log, f"Stages processed={stage_count}, rows written={written_count}")
    append_log(ws_log, f"Output will be saved: {out_path.name}")

    try:
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.calcOnSave = True
    except Exception:
        pass

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    log_lines = [
        f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] populate_template.py finished",
        f"template: {template_path}",
        f"stage_csv: {stage_csv_path}",
        f"out: {out_path}",
        f"mode: {args.mode}",
        f"detected: Stage={col_stage}, x={col_x}, W={col_w}, Disp={col_disp}, "
        f"Tmean={col_tmean}, Dfwd={col_dfwd}, Daft={col_daft}, GM={col_gm}",
        f"stages processed: {stage_count}, rows written: {written_count}",
    ]
    log_txt.write_text("\n".join(log_lines), encoding="utf-8")

    print(f"[OK] Populated workbook saved: {out_path}")
    print(f"[OK] Log saved: {log_txt}")


if __name__ == "__main__":
    main()
