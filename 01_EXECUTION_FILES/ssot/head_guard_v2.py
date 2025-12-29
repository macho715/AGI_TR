#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Head Guard v2 - JSON Registry Validator
Validates pipeline outputs against headers_registry.json
"""

import sys
import json
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook

from ssot.headers_registry import load_registry, validate_df, HeaderRegistry


class HeadGuardV2:
    """Header validation against JSON registry."""

    def __init__(self, registry_path: Path):
        """Load JSON registry."""
        self.registry: HeaderRegistry = load_registry(registry_path)
        self.manifest = {
            "version": self.registry.version,
            "timestamp": datetime.now().isoformat(),
            "files_validated": [],
            "errors": [],
            "warnings": [],
            "summary": {"total_files": 0, "passed": 0, "failed": 0, "skipped": 0},
        }

    def validate_csv_file(self, file_path: Path, deliverable_id: str) -> Tuple[bool, List[str], str]:
        """Validate CSV file against deliverable schema."""
        if deliverable_id not in self.registry.deliverables:
            return True, [], "SKIPPED"

        try:
            df = pd.read_csv(file_path, encoding="utf-8-sig")
            # Use alias matching (treat_out_headers_as_true=False)
            result = validate_df(df, self.registry, deliverable_id, treat_out_headers_as_true=False)

            errors = []
            if result["missing_keys"]:
                errors.append(f"Missing columns: {result['missing_keys']}")
            if result["unexpected_cols"]:
                # Check if strict mode
                d = self.registry.deliverables[deliverable_id]
                if d.strict:
                    errors.append(f"Unexpected columns: {result['unexpected_cols']}")
                else:
                    self.manifest["warnings"].append(
                        f"{file_path.name}: Unexpected columns (non-strict): {result['unexpected_cols']}"
                    )

            status = "PASS" if len(errors) == 0 else "FAIL"
            return len(errors) == 0, errors, status

        except Exception as e:
            return False, [f"Error reading CSV: {type(e).__name__}: {e}"], "FAIL"

    def validate_excel_file(self, file_path: Path, deliverable_id: str) -> Tuple[bool, List[str], str]:
        """Validate Excel file against deliverable schema."""
        if deliverable_id not in self.registry.deliverables:
            return True, [], "SKIPPED"

        d = self.registry.deliverables[deliverable_id]
        if not d.sheet:
            # No specific sheet, validate all sheets or skip
            return True, [], "SKIPPED"

        errors = []
        try:
            wb = load_workbook(file_path, read_only=True, data_only=False)

            if d.sheet not in wb.sheetnames:
                wb.close()
                return False, [f"Missing required sheet: {d.sheet}"], "FAIL"

            ws = wb[d.sheet]
            header_row = d.header_row or 1

            # Read headers from specified row
            headers = []
            for c in range(1, ws.max_column + 1):
                val = ws.cell(header_row, c).value
                if val:
                    headers.append(str(val).strip())

            # Create DataFrame from headers for validation
            # Note: Excel headers are already in output format, but we need alias matching
            # Convert headers to DataFrame and validate with alias matching
            df = pd.DataFrame(columns=headers)
            result = validate_df(df, self.registry, deliverable_id, treat_out_headers_as_true=False)

            if result["missing_keys"]:
                errors.append(f"Sheet '{d.sheet}' (Row {header_row}) missing columns: {result['missing_keys']}")
            if result["unexpected_cols"] and d.strict:
                errors.append(f"Sheet '{d.sheet}' unexpected columns: {result['unexpected_cols']}")

            wb.close()
            status = "PASS" if len(errors) == 0 else "FAIL"
            return len(errors) == 0, errors, status

        except Exception as e:
            return False, [f"Error reading Excel: {type(e).__name__}: {e}"], "FAIL"

    def validate_directory(self, dir_path: Path) -> bool:
        """Validate all files in directory against registry."""
        all_ok = True

        # Find all deliverables and match files
        for deliverable_id, deliverable in self.registry.deliverables.items():
            file_pattern = deliverable.file_pattern

            # Find matching files
            if "*" in file_pattern:
                pattern = file_pattern.replace("*", "")
                matching_files = list(dir_path.glob(f"*{pattern}*"))
            else:
                matching_files = [dir_path / file_pattern]

            for file_path in matching_files:
                if not file_path.exists():
                    continue

                self.manifest["summary"]["total_files"] += 1

                # Determine file type and validate
                if file_path.suffix.lower() == ".csv":
                    ok, errors, status = self.validate_csv_file(file_path, deliverable_id)
                elif file_path.suffix.lower() in [".xlsx", ".xls"]:
                    ok, errors, status = self.validate_excel_file(file_path, deliverable_id)
                else:
                    ok, errors, status = True, [], "SKIPPED"

                if status == "SKIPPED":
                    self.manifest["summary"]["skipped"] += 1
                    continue

                if not ok:
                    all_ok = False
                    self.manifest["summary"]["failed"] += 1
                    for err in errors:
                        self.manifest["errors"].append(f"{file_path.name}: {err}")
                else:
                    self.manifest["summary"]["passed"] += 1

                self.manifest["files_validated"].append({
                    "file": str(file_path.name),
                    "deliverable_id": deliverable_id,
                    "status": status,
                    "errors": errors,
                })

        return all_ok

    def save_manifest(self, output_path: Path):
        """Save validation manifest."""
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(self.manifest, f, indent=2, ensure_ascii=False)

    def print_summary(self):
        """Print validation summary."""
        print("=" * 80)
        print("Head Guard v2 - JSON Registry Validation Summary")
        print("=" * 80)
        print(f"Registry Version: {self.manifest['version']}")
        print(f"Timestamp: {self.manifest['timestamp']}")
        print()
        print("Summary:")
        print(f"  Total files: {self.manifest['summary']['total_files']}")
        print(f"  Passed: {self.manifest['summary']['passed']}")
        print(f"  Failed: {self.manifest['summary']['failed']}")
        print(f"  Skipped: {self.manifest['summary']['skipped']}")
        print()
        print(f"Files validated: {len(self.manifest['files_validated'])}")
        print(f"Errors: {len(self.manifest['errors'])}")
        print(f"Warnings: {len(self.manifest['warnings'])}")

        if self.manifest["errors"]:
            print("\nErrors:")
            for err in self.manifest["errors"][:20]:
                print(f"  - {err}")
            if len(self.manifest["errors"]) > 20:
                print(f"  ... and {len(self.manifest['errors']) - 20} more")

        if self.manifest["warnings"]:
            print("\nWarnings:")
            for warn in self.manifest["warnings"][:10]:
                print(f"  - {warn}")
            if len(self.manifest["warnings"]) > 10:
                print(f"  ... and {len(self.manifest['warnings']) - 10} more")

        # Show file status breakdown
        if self.manifest["files_validated"]:
            print("\nFile Status Breakdown:")
            status_counts = {}
            for f in self.manifest["files_validated"]:
                status = f.get("status", "UNKNOWN")
                status_counts[status] = status_counts.get(status, 0) + 1
            for status, count in sorted(status_counts.items()):
                print(f"  {status}: {count}")

            print("\nDetailed Results:")
            for f in self.manifest["files_validated"]:
                print(f"  {f['file']} ({f['deliverable_id']}): {f['status']}")
                if f['errors']:
                    for err in f['errors']:
                        print(f"    - {err}")


def main():
    import argparse

    parser = argparse.ArgumentParser(
        description="Validate pipeline outputs against JSON Head Registry"
    )
    parser.add_argument(
        "--registry",
        type=str,
        default="01_EXECUTION_FILES/headers_registry.json",
        help="Path to headers_registry.json",
    )
    parser.add_argument(
        "--final-dir",
        type=str,
        required=True,
        help="Path to output directory to validate",
    )
    parser.add_argument(
        "--manifest",
        type=str,
        default="HEAD_MANIFEST.json",
        help="Output manifest file name",
    )
    parser.add_argument("--fail-fast", action="store_true", help="Exit on first error")

    args = parser.parse_args()

    registry_path = Path(args.registry)
    final_dir = Path(args.final_dir)
    manifest_path = final_dir / args.manifest

    if not registry_path.exists():
        print(f"[ERROR] Registry not found: {registry_path}")
        return 1

    if not final_dir.exists():
        print(f"[ERROR] Directory not found: {final_dir}")
        return 1

    guard = HeadGuardV2(registry_path)
    ok = guard.validate_directory(final_dir)
    guard.save_manifest(manifest_path)
    guard.print_summary()

    return 0 if ok else 1


if __name__ == "__main__":
    sys.exit(main())

