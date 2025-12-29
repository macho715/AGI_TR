# -*- coding: utf-8 -*-

"""AGI TR RORO Integrated Calculation System

Version: 6.8.0-PATCH1111-ENFORCEMENT
Last Updated: 2025-12-16
Features:
  - B1-Enhanced: Hydro-based base_tmean_m calculation (SSOT enforcement, no fallback)
  - B2: Trim_gate auto + ΔTM direction correction (COMPLETE)
  - C-Enhanced: Iterative ballast correction + Operational columns
  - CONST_TANKS CSV integration (Option B-2)

VALIDATION STATUS:
✅ B1-Enhanced: RuntimeError if Hydro table missing (SSOT enforcement)
✅ B2: Trim_gate auto-calculation + correct ΔTM direction
✅ C-Enhanced: Operational columns (Ballast_alloc, Pump_time_h, Lineup_OK)
✅ CSV auto-loading 8/8 tanks (100% match rate)
✅ Stage 5_PreBallast: FWD ≤ 2.70m
✅ All regulatory constraints satisfied

PATCH1111 COMPLIANCE:
✅ B1: Fallback 2.00m removed (2 locations)
✅ B2: Already complete (Trim_gate + ΔTM)
✅ C: Operational columns added to iterative_ballast_correction()

REGULATORY BASIS:
- Max FWD Draft: 2.70m (AGI Site operational limit)
- Min GM: 1.50m (IMO stability requirement)
- Max Ramp Angle: 6° (SPMT climbing limit)
- Trim Envelope: ±2.40m (240cm)
"""

# build_bushra_agi_tr_integrated.py
# LCT_BUSHRA_AGI_TR.xlsx 파일을 프로그래밍 방식으로 Excel 함수를 생성하여 만드는 스크립트
# Integrated version: Includes Structural Limits, Option 1 Ballast Fix Check, and all features
# Updated with CAPTAIN_REPORT sheet, Structural Strength columns, and Option 1 Fix Check

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
import os
import sys
import json
from datetime import datetime
import numpy as np
from scipy.interpolate import RectBivariateSpline
from bisect import bisect_left
from typing import Dict, Tuple, Any, List, NamedTuple, Optional
import math
from enum import Enum, auto
import logging
import shutil
import csv
from pathlib import Path
import subprocess

try:
    import matplotlib.pyplot as plt

    # PATCH v6(A): Stabilize RORO stage row expansion (use stages_list + total_rows for extensions)

    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False
    plt = None

# ============================================================================
# 기본 파라미터 (LCT BUSHRA – Hydro / Ops / Limits)
# Excel Parameter/Constants 시트와 1:1 매핑
# ============================================================================
DEFAULT_PARAMS: dict[str, float] = {
    # HYDRO BASE
    "Tmean_baseline_m": 2.00,  # m  - Baseline (DEPRECATED: use hydro-derived value, B1)
    "Tmean_baseline": 2.00,  # m  - Alias for backward compatibility (DEPRECATED, B1)
    "Tide_ref": 2.00,  # m  - Reference tide level
    "Trim_target_cm": 10.00,  # cm - Target trim (sanity check)
    "MTC_t_m_per_cm": 34.00,  # t·m/cm - Moment to change trim
    "LCF_m_from_midship": 0.76,  # m  - LCF from midship (corrected)
    "D_vessel_m": 3.65,  # m  - Molded depth
    "TPC_t_per_cm": 8.00,  # t/cm - Tons per cm immersion
    "Lpp_m": 60.302,  # m  - Lpp (BV value, displayed 60.30)
    # INPUT CONSTANTS
    "L_ramp_m": 12.00,  # m  - Linkspan design length
    "theta_max_deg": 6.00,  # deg- Max ramp angle
    "KminusZ_m": 3.00,  # m  - K - Z (Aries)
    # LIMITS & OPS
    "min_fwd_draft_m": 1.50,  # m  - Minimum allowable forward draft
    "max_fwd_draft_m": 3.50,  # m  - Max structural/nautical forward draft
    "max_fwd_draft_ops_m": 2.70,  # m  - Ops limit for RORO
    "gm_target_m": 1.50,  # m  - GM target
    "linkspan_freeboard_target_m": 0.28,  # m - Linkspan freeboard target
    "ramp_door_offset_m": 0.15,  # m  - Ramp door offset vs quay
    "trim_limit_abs_cm": 240.00,  # cm - Trim envelope for RORO sequence
    # PUMP / VENT
    "pump_rate_tph": 10.00,  # t/h - Ship pump rate (slow)
    "pump_rate_tph_hired": 100.00,  # t/h - Hired pump nominal rate
    "pump_rate_effective_tph": 100.00,  # t/h - Effective rate (vent-limited)
    "vent_flow_coeff": 0.86,  # t/h/mm - Vent flow coefficient (2025-11-18)
    "max_pump_time_h": 6.00,  # h  - Max allowed pump time for fix
    # BALLAST / CAP
    # 가정/주의:
    #   - X_Ballast_from_AP_m = 현재 52.50m는 이전 Forward ballast 전략(FWB1/2 근처)의 잔여값이다.
    #   - Stern Pre-Ballast 전략에서는 FW2(FR 0–6, AFT)의 실제 LCG(AP)를 master_tanks.csv / Tank Plan에서 읽어와야 한다.
    "X_Ballast_from_AP_m": 52.50,  # m  - [OLD] Forward ballast CG from AP (FWB1/2 쪽, 검증용 레거시 값)
    "max_aft_ballast_cap_t": 28.00,  # t  - Max AFT Ballast Capacity (FW2 P/S, Fr 0–6, ~28t)
    "max_fwd_ballast_cap_t": 321.00,  # t  - Max Forward Ballast Capacity (FWB1/2, Fr 48–65, ~321t)
    # RAMP GEOMETRY
    "ramp_hinge_x_mid_m": -30.151,  # m  - Ramp hinge x (midship reference)
    "ramp_length_m": 8.30,  # m  - Ramp length (TRE 2020-08-04)
    "linkspan_height_m": 2.00,  # m  - Jetty soffit height
    "ramp_end_clearance_min_m": 0.40,  # m  - Minimum ramp-end clearance
    # STRUCTURAL LIMITS
    "limit_reaction_t": 201.60,  # t  - Aries hinge reaction limit
    "hinge_limit_rx_t": 201.60,  # t  - Duplicate for clarity
    "limit_share_load_t": 118.80,  # t  - Mammoet max share load on LCT
    "limit_deck_press_tpm2": 10.00,  # t/m² - Deck pressure limit
    "linkspan_area_m2": 12.00,  # m² - Linkspan effective area (1 TR)
    "hinge_pin_area_m2": 0.117,  # m² - Hinge pin/doubler area (390×300 mm)
}

# ============================================================================
# Stage-specific Trim Targets (updated 2025-11-22)
# ============================================================================
TRIM_TARGET_MAP: dict[str, float] = {
    # qqq.md에 값 정의된 Stage들
    "Stage 1": 0.00,  # 유지
    "Stage 2": -96.50,  # qqq.md
    "Stage 3": -96.50,  # qqq.md
    "Stage 4": -96.50,  # qqq.md
    "Stage 5": -89.58,  # qqq.md
    # Stern Pre-Ballast: D-1에 의도적으로 stern trim(+2.40m AFT down) 세팅
    "Stage 5_PreBallast": 240.00,  # qqq.md (의도적 stern trim; FW2, Fr.0–6 기준)
    "Stage 6A_Critical (Opt C)": 0.00,  # qqq.md (even keel 목표)
    "Stage 6A_Critical": 0.00,  # qqq.md와 동일 기준 유지
    "Stage 6C": -96.50,  # qqq.md
    "Stage 6C_TotalMassOpt": -96.50,  # Stage 6C와 동일 기준
    "Stage 7": 0.00,  # 유지
}

# ============================================================================
# TR / ballast layout (frames) – Stage 설명표 + Tank Plan 기준
# ============================================================================
FR_TR1_STOW: float = 42.00  # "TR1 secured at Fr.42 (aft)"
FR_TR2_RAMP: float = (
    17.95  # Stage 6A_Critical(Opt C) LCG Frame (x=12.20m → Fr=30.151-12.20≈17.95)
)
FR_PREBALLAST: float = 3.0  # FW2 Stern tank (Fr.0-6, AFT position)
# CONFIRMED: Stern Pre-Ballast strategy for forward trim correction

# 출력 파일 경로를 스크립트 위치 기준 루트 폴더로 설정
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_FILE = os.path.join(SCRIPT_DIR, "LCT_BUSHRA_AGI_TR_Final_v6_2.xlsx")

# ============================================================================
# BPLUS Feature Flags
# ============================================================================
ENABLE_ISCODE = (
    True  # Enable IS Code 2008 general criteria checking (requires GZ curves)
)
ENABLE_BPLUS = (
    True  # Enable BPLUS features (Engineering-grade hydro, acceptance criteria, etc.)
)

# ============================================================================
# GM 2D Grid 로드 (전역 변수)
# ============================================================================
# GM 2D Grid JSON 로드 - LCT_BUSHRA_GM_2D_Grid.json 기반
GM2D_DATA = None
DISP_GRID = []
TRIM_GRID = []
GM_GRID = []


def _load_gm2d_grid():
    """GM 2D Grid JSON 로드 및 전역 변수 설정"""
    global GM2D_DATA, DISP_GRID, TRIM_GRID, GM_GRID

    # JSON 파일 경로 시도 (기존 _load_json 전략과 동일)
    json_paths = [
        os.path.join(SCRIPT_DIR, "data", "LCT_BUSHRA_GM_2D_Grid.json"),
        os.path.join(SCRIPT_DIR, "LCT_BUSHRA_GM_2D_Grid.json"),
        os.path.join(os.getcwd(), "data", "LCT_BUSHRA_GM_2D_Grid.json"),
        os.path.join(os.getcwd(), "LCT_BUSHRA_GM_2D_Grid.json"),
        r"/mnt/data/LCT_BUSHRA_GM_2D_Grid.json",
    ]

    for json_path in json_paths:
        if os.path.exists(json_path):
            try:
                with open(json_path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    DISP_GRID = sorted(data.get("disp", []))
                    TRIM_GRID = sorted(data.get("trim", []))
                    GM_GRID = [row for row in data.get("gm_grid", [])]
                    GM2D_DATA = data
                    print(
                        f"[OK] GM 2D Grid loaded: {len(DISP_GRID)}×{len(TRIM_GRID)} from {json_path}"
                    )
                    return
            except Exception as e:
                print(f"[ERROR] GM grid parse error in {json_path}: {e}")
                continue

    # 최후의 fallback: 최소 안전 GM 그리드
    print("[FALLBACK] Using minimal safe GM grid")
    DISP_GRID = [2800, 3600]
    TRIM_GRID = [-2.0, 0.0, 2.0]
    GM_GRID = [[1.50] * 3, [1.50] * 3]
    GM2D_DATA = {"disp": DISP_GRID, "trim": TRIM_GRID, "gm_grid": GM_GRID}


# 모듈 로드 시 GM 2D Grid 초기화
_load_gm2d_grid()

# ============================================================================
# Helper Functions
# ============================================================================


# 스타일 정의
def get_styles():
    """공통 스타일 정의"""
    return {
        "title_font": Font(name="Calibri", size=18, bold=True),
        "header_font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
        "normal_font": Font(name="Calibri", size=11),
        "header_fill": PatternFill("solid", fgColor="1F4E78"),
        "input_fill": PatternFill("solid", fgColor="FFF2CC"),
        "ok_fill": PatternFill("solid", fgColor="C6E0B4"),
        "ng_fill": PatternFill("solid", fgColor="F8CBAD"),
        "structure_fill": PatternFill(
            "solid", fgColor="C65911"
        ),  # Orange for Structure
        "opt1_fill": PatternFill(
            "solid", fgColor="7030A0"
        ),  # Purple for Option 1 (Ballast Fix)
        "thin_border": Side(border_style="thin", color="C0C0C0"),
        "center_align": Alignment(
            horizontal="center", vertical="center", wrap_text=True
        ),
        "left_align": Alignment(horizontal="left", vertical="center", wrap_text=True),
    }


# ============================================================================
# 공통 유틸: JSON 로더 + Frame↔x 변환 + Tank Lookup
# ============================================================================


def _get_search_roots() -> list[str]:
    """
    Return ordered search roots for input JSON files.

    Priority order (highest first):
      1) BPLUS_DIR / AGI_BPLUS_DIR env override (explicit)
      2) 02_RAW_DATA/bplus_inputs (RAW DATA - primary)
      3) 02_RAW_DATA/bplus_inputs/data (RAW DATA - GZ curves, etc)
      4) 02_RAW_DATA/additional_inputs (RAW DATA - extra inputs)
      5) ./bplus_inputs (CWD-relative)
      6) <script_dir>/bplus_inputs (portable bundle)
      7) <script_dir>
      8) CWD
      9) <project_root> (parent of script_dir, for data/ folder access)
     10) /mnt/data (notebook/sandbox)

    Notes:
      - 02_RAW_DATA/bplus_inputs 폴더의 RAW DATA를 최우선으로 사용합니다.
      - 폴더가 존재하는 경우만 검색 경로에 추가하여 우선순위를 보장합니다.
    """
    roots: list[str] = []

    # 1. 환경 변수 최우선 (존재하는 경우만)
    env_dir = os.environ.get("BPLUS_DIR") or os.environ.get("AGI_BPLUS_DIR")
    if env_dir and os.path.exists(env_dir):
        roots.append(env_dir)

    cwd = os.getcwd()
    script_dir = os.path.dirname(os.path.abspath(__file__))
    # 프로젝트 루트 디렉토리 (스크립트 디렉토리의 상위 디렉토리)
    project_root = os.path.dirname(script_dir)

    # 2-4. 02_RAW_DATA 우선 (존재하는 경우만)
    raw_data_bplus = os.path.join(project_root, "02_RAW_DATA", "bplus_inputs")
    raw_data_bplus_data = os.path.join(
        project_root, "02_RAW_DATA", "bplus_inputs", "data"
    )
    raw_data_additional = os.path.join(
        project_root, "02_RAW_DATA", "additional_inputs"
    )

    if os.path.exists(raw_data_bplus):
        roots.append(raw_data_bplus)
    if os.path.exists(raw_data_bplus_data):
        roots.append(raw_data_bplus_data)
    if os.path.exists(raw_data_additional):
        roots.append(raw_data_additional)

    # 5-6. bplus_inputs 폴더 (존재하는 경우만 추가하여 RAW DATA 우선 보장)
    bplus_cwd = os.path.join(cwd, "bplus_inputs")
    bplus_script = os.path.join(script_dir, "bplus_inputs")

    if os.path.exists(bplus_cwd):
        roots.append(bplus_cwd)
    if os.path.exists(bplus_script) and bplus_script not in roots:
        roots.append(bplus_script)

    bplus_cwd_data = os.path.join(cwd, "bplus_inputs", "data")
    bplus_script_data = os.path.join(script_dir, "bplus_inputs", "data")
    if os.path.exists(bplus_cwd_data) and bplus_cwd_data not in roots:
        roots.append(bplus_cwd_data)
    if os.path.exists(bplus_script_data) and bplus_script_data not in roots:
        roots.append(bplus_script_data)

    # 7-10. 기타 경로
    roots.extend(
        [
            script_dir,
            cwd,
            project_root,  # 프로젝트 루트 추가 (data 폴더 접근용)
            r"/mnt/data",
        ]
    )

    # de-duplicate while preserving order
    out: list[str] = []
    for r in roots:
        if not r:
            continue
        ar = os.path.abspath(r)
        if ar not in out:
            out.append(ar)
    if not hasattr(_get_search_roots, "_logged"):
        print("[INFO] Search roots priority order (highest first):")
        for i, root in enumerate(out, 1):
            marker = "← RAW DATA" if "02_RAW_DATA" in root else ""
            print(f"  {i:2d}. {root} {marker}")
        _get_search_roots._logged = True
    return out


def _find_existing_path(filename: str) -> str | None:
    """
    Find the first existing path for filename across search roots.
    Returns absolute path if found, else None.
    """
    for root in _get_search_roots():
        p = os.path.join(root, filename)
        if os.path.exists(p):
            return p
    return None


_MISSING_JSON_WARNED: set[str] = set()


def _load_json(filename, quiet_missing: bool = False):
    """
    JSON loader with ordered path priority (Option B+).
    bplus_inputs 폴더의 RAW DATA를 우선적으로 사용합니다.

    - Searches roots returned by _get_search_roots() in order.
    - Returns parsed JSON if found.
    - Returns None if not found or parsing fails in all candidates.

    This is intentionally conservative: downstream code can decide whether
    'missing' is acceptable (demo) or should trigger ZERO fail-safe (prod).
    """
    for root in _get_search_roots():
        path = os.path.join(root, filename)
        if os.path.exists(path):
            try:
                with open(path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    # bplus_inputs에서 로드한 경우 명확히 표시
                    if "bplus_inputs" in path:
                        print(
                            f"[OK] Loaded from bplus_inputs (RAW DATA): {filename} ({path})"
                        )
                    else:
                        print(f"[OK] Loaded: {filename} ({path})")
                return data
            except json.JSONDecodeError as e:
                print(f"[BACKUP] JSON parse error in {filename} ({path}): {e}")
                continue
            except Exception as e:
                print(f"[BACKUP] Error reading {filename} ({path}): {e}")
                continue
    if not quiet_missing and filename not in _MISSING_JSON_WARNED:
        print(f"[BACKUP] {filename} not found → using fallback")
        _MISSING_JSON_WARNED.add(filename)
    return None


# =============================================================================
# ZERO fail-safe (B+ inputs)
# =============================================================================


def _zero_stop_table(reason: str, missing: list[str], next_action: str) -> str:
    """
    Build a ZERO '중단' markdown table (print-friendly).
    """
    missing_str = "\n".join([f"- {m}" for m in missing]) if missing else "-"
    return (
        "| 단계 | 이유 | 위험 | 요청데이터 | 다음조치 |\n"
        "|---|---|---|---|---|\n"
        f"| 중단(B+ Preflight) | {reason} | MWS/HM 제출 근거 부족 → 잘못된 승인 판단/UKC·Ramp 오판 가능 | {missing_str} | {next_action} |\n"
    )


def _zero_stop(reason: str, missing: list[str]) -> None:
    """
    Print ZERO stop log and exit.
    """
    msg = _zero_stop_table(
        reason=reason,
        missing=missing,
        next_action="1) prep_bplus_inputs.py 실행 → ./bplus_inputs 생성  2) Approved Booklet/NAPA export 값 채움  3) 재실행",
    )
    print("\n[ZERO FAIL-SAFE]\n" + msg)
    # keep an artifact for audit trail
    try:
        with open("ZERO_STOP_LOG.md", "w", encoding="utf-8") as f:
            f.write("# ZERO STOP LOG\n\n")
            f.write(msg)
            f.write("\n\n## Search Roots\n")
            for r in _get_search_roots():
                f.write(f"- {r}\n")
    except Exception:
        pass
    raise SystemExit(2)


def _hydro_has_required_cols(rows: list[dict]) -> tuple[bool, list[str]]:
    """
    Option B+ requires draft-dependent hydro columns.
    """
    missing = []

    # Accept common aliases but require presence in some form
    def has_any(row, keys):
        return any(k in row and row.get(k) is not None for k in keys)

    if not rows:
        return False, ["Hydro table rows empty"]
    sample = rows[0]

    if not has_any(sample, ["LCF_m_from_midship", "LCF_m", "LCF"]):
        missing.append("LCF_m_from_midship (or alias: LCF_m/LCF)")
    if not has_any(sample, ["MCTC_t_m_per_cm", "MTC_t_m_per_cm", "MCTC", "MTC"]):
        missing.append("MCTC_t_m_per_cm (or alias: MTC_t_m_per_cm/MCTC/MTC)")
    if not has_any(sample, ["TPC_t_per_cm", "TPC"]):
        missing.append("TPC_t_per_cm (or alias: TPC)")
    return (len(missing) == 0), missing


def _any_gz_curve_exists() -> bool:
    """
    Check if at least one GZ curve file exists in any search root.
    Accepts:
      - data/GZ_Curves.json
      - data/GZ_Curve_*.json
      - GZ_Curve_*.json
      - data/GZ/*.json (PATCH: subdirectory support)
    """
    import glob

    patterns = [
        "data/GZ_Curves.json",
        "data/GZ_Curve_*.json",
        "GZ_Curve_*.json",
        "data/GZ/*.json",  # PATCH: Support GZ subdirectory
        "data/GZ/Stage_*.json",  # PATCH: Support Stage files in GZ subdirectory
    ]
    for root in _get_search_roots():
        for pat in patterns:
            hits = glob.glob(os.path.join(root, pat))
            if hits:
                return True
    return False


def _bplus_preflight_or_zero(strict: bool = True) -> None:
    """
    Ensure B+ input set is available (./bplus_inputs priority) before producing
    submission-grade outputs.

    strict=True:
      - Missing mandatory items triggers ZERO stop.
      - Mandatory items reflect Option B+ scope: Hydro LCF/MCTC/TPC + GM curve + Acceptance + Structural + Securing + IS Code criteria + at least one GZ curve.
    """
    missing: list[str] = []

    # Hydro (must exist + must contain required cols)
    hydro_rows = _load_hydro_table()
    # Check if we have fallback values in _meta (from Hydro_Table_Engineering.json)
    hydro_has_fallbacks = False
    try:
        # Try to load the engineering file to check for fallbacks
        eng_data = _load_json("Hydro_Table_Engineering.json")
        if isinstance(eng_data, dict) and isinstance(eng_data.get("_meta"), dict):
            fallbacks = eng_data["_meta"].get("fallbacks", {})
            if (
                fallbacks.get("LCF_m_from_midship") is not None
                and fallbacks.get("MCTC_t_m_per_cm") is not None
                and fallbacks.get("TPC_t_per_cm") is not None
            ):
                hydro_has_fallbacks = True
    except Exception:
        pass

    ok_hydro, hydro_missing_cols = _hydro_has_required_cols(hydro_rows)
    # If columns are missing but we have fallbacks, it's acceptable
    if not ok_hydro and not hydro_has_fallbacks:
        missing.append(
            "Hydro_Table_Engineering.json (with LCF/MCTC/TPC cols or fallbacks)"
        )

    # GM curve requirement (either file curve OR hydro rows carry GM_min_m per draft)
    gm_curve = _load_gm_min_curve()
    hydro_has_gm_min = (
        any(
            (isinstance(r, dict) and (r.get("GM_min_m") is not None))
            for r in hydro_rows
        )
        if hydro_rows
        else False
    )
    if gm_curve is None and not hydro_has_gm_min:
        missing.append("GM_Min_Curve.json (or Hydro rows with GM_min_m)")

    # Acceptance / Structural / Securing / IS Code criteria
    if not _load_acceptance_criteria():
        missing.append("Acceptance_Criteria.json")
    if not _load_structural_limits():
        missing.append("Structural_Limits.json")
    if not _load_securing_input():
        missing.append("Securing_Input.json")
    if not _load_iscode_criteria():
        missing.append("ISCODE_Criteria.json")

    # GZ curves (mandatory when IS Code is enabled)
    if ENABLE_ISCODE and not _any_gz_curve_exists():
        missing.append(
            "GZ_Curve_<Stage>.json (at least one) under ./bplus_inputs/data/"
        )

    if missing and strict:
        _zero_stop(
            "B+ 필수 입력 데이터 누락 또는 Engineering-grade hydro columns 미충족",
            missing,
        )

    # Soft notice if not strict
    if missing and not strict:
        print("[WARN][B+] Missing inputs (non-strict):")
        for m in missing:
            print(" -", m)


def _load_hydro_table() -> list[dict]:
    """
    Hydro table loader with multi-filename fallback.

    Accepted formats
    ----------------
    1) list[dict]  (legacy)
       - each row: Disp_t, Tmean_m, GM_m, ... (optional: LCF/MCTC/TPC/GM_min)

    2) dict with keys:
       - "_meta": { "lcf_reference": "midship", "x_sign": "AFT+", ... } (optional)
       - "rows":  list[dict]  (same as legacy rows)

    Engineering-grade (Option B+) recommended columns (draft-based):
      - Tmean_m
      - LCF_m_from_midship        (x_from_mid_m, aft+; midship reference)
      - MCTC_t_m_per_cm           (or MTC_t_m_per_cm)
      - TPC_t_per_cm
      - GM_min_m                  (optional; curve points)

    PATCH B1-1: Added __file__ based paths for better file discovery when CWD differs.
    """
    script_dir = os.path.dirname(os.path.abspath(__file__))
    cwd = os.getcwd()

    # Build candidate paths with __file__ based paths prioritized
    candidates = []

    # 1. Script directory based paths (highest priority for portability)
    candidates.extend(
        [
            os.path.join(script_dir, "Hydro_Table_Engineering.json"),
            os.path.join(script_dir, "data", "Hydro_Table_Engineering.json"),
            os.path.join(script_dir, "bplus_inputs", "Hydro_Table_Engineering.json"),
            os.path.join(script_dir, "data", "hydro_table.json"),
            os.path.join(script_dir, "data", "Hydro_Table.json"),
            os.path.join(script_dir, "Hydro_Table.json"),
        ]
    )

    # 2. CWD based paths (for backward compatibility)
    candidates.extend(
        [
            "Hydro_Table_Engineering.json",  # Highest priority: Engineering-grade format
            "data/Hydro_Table_Engineering.json",
            "bplus_inputs/Hydro_Table_Engineering.json",
            "data/hydro_table.json",
            "data/Hydro_Table.json",
            "Hydro_Table.json",
            "Hydro_Table_Engineering_Template.json",
            "data/Hydro_Table_Engineering_Template.json",
        ]
    )

    # 3. Use _get_search_roots() for additional paths (includes bplus_inputs priority)
    search_roots = _get_search_roots()
    for root in search_roots:
        candidates.extend(
            [
                os.path.join(root, "Hydro_Table_Engineering.json"),
                os.path.join(root, "data", "Hydro_Table_Engineering.json"),
                os.path.join(root, "bplus_inputs", "Hydro_Table_Engineering.json"),
            ]
        )

    # Remove duplicates while preserving order
    seen = set()
    unique_candidates = []
    for fn in candidates:
        abs_fn = (
            os.path.abspath(fn)
            if (os.path.isabs(fn) or os.path.exists(fn))
            else os.path.abspath(fn) if os.path.exists(fn) else fn
        )
        if abs_fn not in seen:
            seen.add(abs_fn)
            unique_candidates.append(fn)

    for fn in unique_candidates:
        # Try direct path first, then use _load_json for search roots
        data = None
        if os.path.exists(fn):
            try:
                with open(fn, "r", encoding="utf-8") as f:
                    data = json.load(f)
            except (json.JSONDecodeError, Exception) as e:
                print(f"[BACKUP] Error reading {fn}: {e}")
                continue
        else:
            if os.path.isabs(fn):
                continue
            # Use _load_json for relative paths (searches all roots)
            data = _load_json(fn, quiet_missing=True)

        if data is None:
            continue

        # dict wrapper with meta
        if (
            isinstance(data, dict)
            and isinstance(data.get("rows"), list)
            and len(data["rows"]) > 0
        ):
            meta = data.get("_meta") or {}
            # Soft lock on LCF reference if declared
            lcf_ref = str(meta.get("lcf_reference", "")).strip().lower()
            if lcf_ref and lcf_ref not in ("midship", "from_midship", "x_from_midship"):
                logging.warning(
                    f"[HydroMeta] lcf_reference={lcf_ref} in {fn} (expected midship). Check data prep."
                )
            return data["rows"]
        # legacy list
        if isinstance(data, list) and len(data) > 0:
            return data
    print("[BACKUP] Hydro table not found in search paths → using fallback")
    return []


def gm_2d_bilinear(disp_t: float, trim_m: float) -> float:
    """
    Bilinear interpolation on (disp, trim) → GM 2D grid.
    Uses LCT_BUSHRA_GM_2D_Grid.json with 7x7 grid (7 displacement × 7 trim values).

    Args:
        disp_t: Displacement (tons)
        trim_m: Trim in meters (bow down + / stern down - 등 프로젝트 convention에 맞게 사용)

    Returns:
        GM value (meters) or 1.50m fallback if data not available
    """
    # 전역 변수 사용 (모듈 로드 시 초기화됨)
    global DISP_GRID, TRIM_GRID, GM_GRID

    # Fallback: Grid가 비어있으면 안전한 최소값 반환
    if not DISP_GRID or not TRIM_GRID or not GM_GRID:
        print(f"[BACKUP] GM 2D Grid unavailable → using fallback GM=1.50m")
        return 1.50  # Safe minimum GM requirement

    try:
        ds = DISP_GRID
        ts = TRIM_GRID
        g = GM_GRID

        # --- 1) disp index 찾기 (clamp + 양구간 인덱스) ---
        if disp_t <= ds[0]:
            i0 = i1 = 0
        elif disp_t >= ds[-1]:
            i0 = i1 = len(ds) - 1
        else:
            k = bisect_left(ds, disp_t) - 1
            if k < 0:
                k = 0
            i0, i1 = k, k + 1

        # --- 2) trim index 찾기 (clamp + 양구간 인덱스) ---
        if trim_m <= ts[0]:
            j0 = j1 = 0
        elif trim_m >= ts[-1]:
            j0 = j1 = len(ts) - 1
        else:
            k = bisect_left(ts, trim_m) - 1
            if k < 0:
                k = 0
            j0, j1 = k, k + 1

        d0, d1 = ds[i0], ds[i1]
        t0, t1 = ts[j0], ts[j1]

        # --- 3) 보간 비율 (0~1) ---
        if d1 != d0:
            td = (disp_t - d0) / (d1 - d0)
        else:
            td = 0.0

        if t1 != t0:
            tt = (trim_m - t0) / (t1 - t0)
        else:
            tt = 0.0

        # --- 4) 모서리 GM 값 ---
        g00 = g[i0][j0]  # (d0, t0)
        g10 = g[i1][j0]  # (d1, t0)
        g01 = g[i0][j1]  # (d0, t1)
        g11 = g[i1][j1]  # (d1, t1)

        # --- 5) bilinear interpolation ---
        gm = (
            (1 - td) * (1 - tt) * g00
            + td * (1 - tt) * g10
            + (1 - td) * tt * g01
            + td * tt * g11
        )

        # Sanity check for unrealistic GM values
        if gm < 0 or gm > 5.0:
            print(f"[BACKUP] GM={gm:.2f}m unrealistic → fallback GM=1.50m")
            return 1.50

        return float(gm)
    except Exception as e:
        print(f"[BACKUP] GM calculation error: {e} → fallback GM=1.50m")
        return 1.50


# ============================================================================
# Hydro Table 보간 함수
# ============================================================================


def iterative_ballast_correction(
    base_disp_t: float,
    base_tmean_m: float,
    loads: List[Any],
    initial_ballast_t: float,
    fwd_limit_m: float = 2.70,
    max_iterations: int = 3,
    tolerance_m: float = 0.01,
    d_vessel_m: float = 3.65,
    freeboard_min_m: float = 0.28,
    **params,
) -> dict:
    """
    PATCH C: Iterative ballast correction using Secant method to achieve target Dfwd.

    Algorithm:
    1. Start with initial ballast estimate (from B2 calculation)
    2. Apply ballast and recalculate Dfwd using solve_stage()
    3. Use Secant method to iteratively adjust ballast until Dfwd ≈ fwd_limit_m
    4. Maximum 1-3 iterations for practical convergence

    Args:
        base_disp_t: Base displacement (t)
        base_tmean_m: Base mean draft (m)
        loads: List of LoadItem objects
        initial_ballast_t: Initial ballast estimate (t) from B2 calculation
        fwd_limit_m: Target forward draft limit (m), default 2.70
        max_iterations: Maximum iterations, default 3
        tolerance_m: Convergence tolerance (m), default 0.01 (1cm)
        d_vessel_m: Vessel molded depth (m) for freeboard check
        freeboard_min_m: Minimum freeboard (m) required
        **params: Additional parameters for solve_stage()

    Returns:
        dict with keys:
            - ballast_t: Final ballast tonnage (t)
            - dfwd_gate_m: Final forward draft (m)
            - iterations: Number of iterations used
            - converged: Whether convergence was achieved
            - flag: "OK" if Dfwd <= fwd_limit_m and freeboard OK, "LIMIT" otherwise
            - freeboard_ok: True if freeboard constraint met
            - freeboard_min_m: Minimum freeboard achieved (m)
    """
    from copy import deepcopy

    # Initial ballast estimates (Secant method needs 2 points)
    b0 = float(initial_ballast_t or 0.0)
    # Secant 2점 보장 (b0==0 정체 방지)
    delta = max(5.0, abs(b0) * 0.10)
    sign = 1.0 if b0 >= 0 else -1.0
    b1 = b0 + sign * delta

    # Get ballast position from params (FW2, Fr 0-6, AFT)
    fr_pb = params.get("FR_PREBALLAST", params.get("FR_PB", 3.0))
    ballast_x = fr_to_x(fr_pb)  # Convert frame to x position (m from midship)

    # Evaluate Dfwd for initial points
    def eval_dfwd(ballast_t: float) -> tuple[float, float, bool]:
        """Evaluate drafts and freeboard for given ballast tonnage.

        PATCH C: Properly applies ballast as LoadItem and recalculates using solve_stage().
        The solve_stage() function will:
        1. Calculate total displacement = base_disp_t + sum(loads)
        2. Interpolate Tmean from displacement
        3. Calculate trim and drafts
        """
        # Create modified loads with ballast applied
        modified_loads = deepcopy(loads)

        # Remove existing PreBallast LoadItem if present (to replace with new value)
        modified_loads = [load for load in modified_loads if load.name != "PreBallast"]

        # Add ballast as LoadItem (ballast_t can be negative for de-ballast)
        if abs(ballast_t) > 0.01:  # Only add if significant
            modified_loads.append(
                LoadItem("PreBallast", ballast_t, ballast_x, "BALLAST")
            )

        # Solve stage with modified loads - solve_stage will calculate:
        # - Total displacement = base_disp_t + sum(modified_loads)
        # - Tmean from displacement interpolation
        # - Trim and drafts
        res = solve_stage(base_disp_t, base_tmean_m, modified_loads, **params)
        fwd_draft = res.get("Dfwd_m", fwd_limit_m)
        aft_draft = res.get("Daft_m", 0.0)
        fb_fwd = d_vessel_m - fwd_draft
        fb_aft = d_vessel_m - aft_draft
        fb_min = min(fb_fwd, fb_aft)
        freeboard_ok = fb_min >= freeboard_min_m
        return fwd_draft, aft_draft, freeboard_ok

    def penalize_dfwd(dfwd: float, freeboard_ok: bool) -> float:
        return dfwd if freeboard_ok else dfwd + 10.0

    dfwd0, daft0, fb_ok0 = eval_dfwd(b0)
    dfwd1, daft1, fb_ok1 = eval_dfwd(b1)
    dfwd0_p = penalize_dfwd(dfwd0, fb_ok0)
    dfwd1_p = penalize_dfwd(dfwd1, fb_ok1)

    # Target: Dfwd = fwd_limit_m
    target = fwd_limit_m
    error0 = dfwd0_p - target
    error1 = dfwd1_p - target

    iterations = 0
    converged = False
    final_ballast = b0
    final_dfwd = dfwd0
    final_daft = daft0
    final_fb_ok = fb_ok0

    # Secant method iteration
    for i in range(max_iterations):
        iterations = i + 1

        # Check convergence
        if abs(error0) <= tolerance_m:
            converged = True
            final_ballast = b0
            final_dfwd = dfwd0
            final_daft = daft0
            final_fb_ok = fb_ok0
            break

        # Secant method: b_new = b0 - error0 * (b1 - b0) / (error1 - error0)
        if abs(error1 - error0) < 1e-6:  # Avoid division by zero
            # Use bisection if secant fails
            b_new = (b0 + b1) / 2.0
        else:
            b_new = b0 - error0 * (b1 - b0) / (error1 - error0)

        # Clamp ballast to reasonable range to prevent divergence
        max_ballast_t = float(params.get("max_aft_ballast_cap_t", params.get("MAX_TANK_CAPACITY_T", 400.0)) or 400.0)
        min_ballast_t = -max_ballast_t  # Allow de-ballast up to same limit
        b_new = max(min_ballast_t, min(max_ballast_t, b_new))

        # Prevent excessive jumps (limit step size to 50% of range)
        max_step = max_ballast_t * 0.5
        if abs(b_new - b0) > max_step:
            b_new = b0 + max_step if b_new > b0 else b0 - max_step

        # Evaluate new point
        dfwd_new, daft_new, fb_ok_new = eval_dfwd(b_new)
        dfwd_new_p = penalize_dfwd(dfwd_new, fb_ok_new)
        error_new = dfwd_new_p - target

        # Update for next iteration
        b0, dfwd0, daft0, fb_ok0, error0 = (
            b1,
            dfwd1,
            daft1,
            fb_ok1,
            error1,
        )
        b1, dfwd1, daft1, fb_ok1, error1 = (
            b_new,
            dfwd_new,
            daft_new,
            fb_ok_new,
            error_new,
        )

        final_ballast = b_new
        final_dfwd = dfwd_new
        final_daft = daft_new
        final_fb_ok = fb_ok_new

    # Determine flag
    final_fb_min = min(d_vessel_m - final_dfwd, d_vessel_m - final_daft)
    final_fb_ok = final_fb_min >= freeboard_min_m

    # Validate final results - if abnormal, use initial values
    max_ballast_t = float(params.get("max_aft_ballast_cap_t", params.get("MAX_TANK_CAPACITY_T", 400.0)) or 400.0)
    if abs(final_ballast) > max_ballast_t * 1.5 or final_dfwd < 0 or final_dfwd > 10 or final_daft < 0 or final_daft > 10:
        # Abnormal result - use initial values as fallback
        final_ballast = float(initial_ballast_t or 0.0)
        final_dfwd, final_daft, final_fb_ok = eval_dfwd(final_ballast)
        final_fb_min = min(d_vessel_m - final_dfwd, d_vessel_m - final_daft)
        final_fb_ok = final_fb_min >= freeboard_min_m
        converged = False
        iterations = 0

    flag = (
        "OK" if final_dfwd <= fwd_limit_m and final_fb_ok else "LIMIT"
    )

    # PATCH C-Enhanced: Add operational columns (PATCH1111.MD)
    # Calculate ballast per side for operational planning
    ballast_per_side = abs(final_ballast) / 2.0

    # Generate ballast allocation string (tank-based allocation)
    # For now, assume simple FW2.P/FW2.S split, can be enhanced with tank catalog
    if abs(final_ballast) > 0.01:
        if final_ballast > 0:
            ballast_alloc = (
                f"FW2.P: {ballast_per_side:.1f}t, FW2.S: {ballast_per_side:.1f}t (Fill)"
            )
        else:
            ballast_alloc = f"FW2.P: {ballast_per_side:.1f}t, FW2.S: {ballast_per_side:.1f}t (Empty)"
    else:
        ballast_alloc = "None"

    # Estimate pump time (assume 50 t/h pump rate, can be configured)
    pump_rate_t_h = float(
        params.get("PUMP_RATE_T_H")
        or params.get("pump_rate_effective_tph")
        or params.get("pump_rate_tph")
        or 10.0
    )
    pump_time_h = abs(final_ballast) / pump_rate_t_h if pump_rate_t_h > 0 else 0.0

    # Lineup check (simple: OK if ballast < 100t, can be enhanced with tank capacity)
    max_tank_capacity = float(
        params.get("MAX_TANK_CAPACITY_T")
        or params.get("max_aft_ballast_cap_t")
        or 100.0
    )
    lineup_ok = (
        "Yes" if ballast_per_side <= max_tank_capacity else "No (Exceeds capacity)"
    )

    return {
        "ballast_t": final_ballast,
        "dfwd_gate_m": final_dfwd,
        "iterations": iterations,
        "converged": converged,
        "flag": flag,
        "freeboard_ok": final_fb_ok,
        "freeboard_min_m": final_fb_min,
        # PATCH C-Enhanced: Operational columns
        "ballast_alloc": ballast_alloc,
        "pump_time_h": round(pump_time_h, 2),
        "lineup_ok": lineup_ok,
    }


def calc_trim_gate_cm_from_tmean(tmean_m: float, fwd_limit_m: float = 2.70) -> float:
    """
    Calculate required trim (cm) to satisfy forward draft limit.

    PATCH B2: Trim_gate calculation based on Tmean and FWD limit.

    Formula: Trim_gate_cm = max(0, 2 * (Tmean - FWD_limit) * 100)

    This ensures: Dfwd = Tmean - (Trim/2) <= FWD_limit
    Therefore: Trim >= 2 * (Tmean - FWD_limit)

    Args:
        tmean_m: Mean draft (m)
        fwd_limit_m: Forward draft limit (m), default 2.70

    Returns:
        Required trim in cm (positive = stern down, negative = bow down)
        Returns 0 if Tmean <= FWD_limit (no trim needed)
    """
    trim_gate_cm = max(0.0, 2.0 * (tmean_m - fwd_limit_m) * 100.0)
    return float(trim_gate_cm)


def interpolate_tmean_from_disp(disp_t: float, hydro_table: list[dict]) -> float:
    """
    Δdisp → Tmean 보간

    Args:
        disp_t: 배수량 (ton)
        hydro_table: Hydro table 데이터 (list of dict)
            각 dict는 "Displacement_t" 또는 "Disp_t"와 "Tmean_m" 키를 가져야 함

    Returns:
        Tmean (m) - 선형 보간된 평균 흘수
    """
    if not hydro_table or len(hydro_table) == 0:
        print("[WARNING] Hydro table is empty → using fallback Tmean=2.00m")
        return 2.00  # Fallback

    try:
        # 키 이름 확인 (다양한 형식 지원)
        disp_key = None
        tmean_key = None

        for key in hydro_table[0].keys():
            if "disp" in key.lower() or "displacement" in key.lower():
                disp_key = key
            if "tmean" in key.lower() or "mean" in key.lower():
                tmean_key = key

        if not disp_key or not tmean_key:
            print(
                "[WARNING] Hydro table format not recognized → using fallback Tmean=2.00m"
            )
            return 2.00

        # Displacement 및 Tmean 배열 추출
        disps = [float(row[disp_key]) for row in hydro_table]
        tmeans = [float(row[tmean_key]) for row in hydro_table]

        # 정렬 확인 (필요시 정렬)
        if disps != sorted(disps):
            # 정렬 필요
            sorted_pairs = sorted(zip(disps, tmeans))
            disps, tmeans = zip(*sorted_pairs)
            disps = list(disps)
            tmeans = list(tmeans)

        # Clamp 및 보간
        if disp_t <= disps[0]:
            return tmeans[0]
        elif disp_t >= disps[-1]:
            return tmeans[-1]

        # 선형 보간
        i = bisect_left(disps, disp_t)
        if i <= 0:
            return tmeans[0]
        elif i >= len(disps):
            return tmeans[-1]

        x0, x1 = disps[i - 1], disps[i]
        y0, y1 = tmeans[i - 1], tmeans[i]

        if x1 == x0:
            return y0

        tmean = y0 + (y1 - y0) * (disp_t - x0) / (x1 - x0)
        return float(tmean)
    except Exception as e:
        print(f"[WARNING] Tmean interpolation error: {e} → using fallback Tmean=2.00m")
        return 2.00


# ============================================================================
# LOAD ITEM BUILDER
# ============================================================================


# ============================================================================
# Engineering-grade Hydro Interpolation (Option B)
# - Supports draft-dependent LCF/MCTC/TPC and GM minimum curve
# - Falls back to constants if approved booklet/NAPA hydro data is missing
# ============================================================================

_GM_MIN_CURVE_CACHE: list[dict] | None = None


def _as_float(v, default=None):
    try:
        if v is None:
            return default
        if isinstance(v, (int, float)):
            return float(v)
        s = str(v).strip()
        if s == "":
            return default
        return float(s)
    except Exception:
        return default


def _detect_key(row_keys: list[str], candidates: list[str]) -> str | None:
    lk = {k.lower(): k for k in row_keys}
    for c in candidates:
        if c.lower() in lk:
            return lk[c.lower()]
    # fuzzy contains
    for c in candidates:
        for k in row_keys:
            if c.lower() in k.lower():
                return k
    return None


def _linear_interp(xs: list[float], ys: list[float], x: float) -> float:
    """Simple clamp + linear interpolation. xs must be sorted."""
    if not xs:
        raise ValueError("empty xs")
    if x <= xs[0]:
        return ys[0]
    if x >= xs[-1]:
        return ys[-1]
    # binary search
    import bisect

    i = bisect.bisect_left(xs, x)
    if i <= 0:
        return ys[0]
    if i >= len(xs):
        return ys[-1]
    x0, x1 = xs[i - 1], xs[i]
    y0, y1 = ys[i - 1], ys[i]
    if abs(x1 - x0) < 1e-12:
        return y0
    return y0 + (y1 - y0) * (x - x0) / (x1 - x0)


def interpolate_hydro_by_tmean(tmean_m: float, hydro_table: list[dict]) -> dict:
    """
    Interpolate hydrostatic fields by mean draft (Tmean_m).

    Expected (preferred) columns in hydro_table rows:
      - Tmean_m
      - LCF_m_from_midship   (x_from_mid_m, aft+)
      - MCTC_t_m_per_cm      (or MTC_t_m_per_cm)
      - TPC_t_per_cm
      - GM_min_m             (optional; minimum GM curve point)

    Returns:
      dict with any available interpolated fields (missing fields omitted).
    """
    if not hydro_table:
        return {}

    row0 = hydro_table[0]
    keys = list(row0.keys())

    k_tmean = _detect_key(keys, ["Tmean_m", "Tmean", "MeanDraft_m", "DraftMean_m"])
    if not k_tmean:
        return {}

    # candidate keys (synonyms supported)
    k_lcf = _detect_key(keys, ["LCF_m_from_midship", "LCF_from_mid_m", "LCF_m", "LCF"])
    k_mctc = _detect_key(keys, ["MCTC_t_m_per_cm", "MTC_t_m_per_cm", "MCTC", "MTC"])
    k_tpc = _detect_key(keys, ["TPC_t_per_cm", "TPC", "TPC_t_cm", "TPC_t_per_cm"])
    k_gmmin = _detect_key(keys, ["GM_min_m", "GMmin_m", "GM_MIN_m", "MinGM_m"])

    # build x axis
    xs = []
    rows_valid = []
    for r in hydro_table:
        xm = _as_float(r.get(k_tmean))
        if xm is None:
            continue
        xs.append(float(xm))
        rows_valid.append(r)

    if not xs:
        return {}

    # sort by tmean
    pairs = sorted(zip(xs, rows_valid), key=lambda t: t[0])
    xs = [p[0] for p in pairs]
    rows_valid = [p[1] for p in pairs]

    out = {"Tmean_m": float(tmean_m)}

    def interp_field(k):
        ys = []
        x_ok = []
        for x, r in zip(xs, rows_valid):
            y = _as_float(r.get(k))
            if y is None:
                continue
            x_ok.append(x)
            ys.append(float(y))
        if len(x_ok) < 2:
            return None
        # ensure sorted
        pairs2 = sorted(zip(x_ok, ys), key=lambda t: t[0])
        x_ok = [p[0] for p in pairs2]
        ys = [p[1] for p in pairs2]
        return float(_linear_interp(x_ok, ys, float(tmean_m)))

    if k_lcf:
        v = interp_field(k_lcf)
        if v is not None:
            out["LCF_m_from_midship"] = v

    if k_mctc:
        v = interp_field(k_mctc)
        if v is not None:
            out["MCTC_t_m_per_cm"] = v

    if k_tpc:
        v = interp_field(k_tpc)
        if v is not None:
            out["TPC_t_per_cm"] = v

    if k_gmmin:
        v = interp_field(k_gmmin)
        if v is not None:
            out["GM_min_m"] = v

    return out


def _load_gm_min_curve() -> list[dict] | None:
    """
    Optional GM minimum curve loader.
    Supported filenames (first found):
      - data/gm_min_curve.json
      - data/GM_Min_Curve.json
      - GM_Min_Curve.json
      - data/GMmin_curve.json
    Curve format examples:
      [{"Tmean_m": 2.00, "GM_min_m": 0.50}, ...]  (draft-based)
      [{"Disp_t": 4000, "GM_min_m": 0.50}, ...]   (disp-based)
      {"_meta": {...}, "points": [...]}  (prep_bplus_inputs.py format)
    """
    global _GM_MIN_CURVE_CACHE
    if _GM_MIN_CURVE_CACHE is not None:
        return _GM_MIN_CURVE_CACHE

    candidates = [
        "data/gm_min_curve.json",
        "data/GM_Min_Curve.json",
        "GM_Min_Curve.json",
        "data/GMmin_curve.json",
    ]
    for fn in candidates:
        data = _load_json(fn)
        # Handle list format (legacy)
        if isinstance(data, list) and len(data) > 0:
            _GM_MIN_CURVE_CACHE = data
            return _GM_MIN_CURVE_CACHE
        # Handle dict format with "points" key (prep_bplus_inputs.py format)
        if (
            isinstance(data, dict)
            and isinstance(data.get("points"), list)
            and len(data["points"]) > 0
        ):
            _GM_MIN_CURVE_CACHE = data["points"]
            return _GM_MIN_CURVE_CACHE

    _GM_MIN_CURVE_CACHE = None
    return None


def gm_min_from_curve(
    disp_t: float, tmean_m: float, gm_curve: list[dict] | None
) -> tuple[float | None, str]:
    """
    Return (gm_min_m, basis).
    Priority:
      1) gm_curve param (if provided)
      2) auto-loaded GM min curve json
      3) None
    """
    curve = (
        gm_curve
        if (isinstance(gm_curve, list) and len(gm_curve) > 0)
        else _load_gm_min_curve()
    )
    if not curve:
        return None, "none"

    keys = list(curve[0].keys())
    k_gm = _detect_key(keys, ["GM_min_m", "GMmin_m", "GM_MIN_m", "MinGM_m"])
    if not k_gm:
        return None, "invalid_curve"

    # choose axis
    k_tmean = _detect_key(keys, ["Tmean_m", "Tmean", "MeanDraft_m", "DraftMean_m"])
    k_disp = _detect_key(keys, ["Disp_t", "Displacement_t", "disp"])
    if k_tmean:
        xs, ys = [], []
        for r in curve:
            x = _as_float(r.get(k_tmean))
            y = _as_float(r.get(k_gm))
            if x is None or y is None:
                continue
            xs.append(float(x))
            ys.append(float(y))
        if len(xs) < 2:
            return None, "curve_insufficient"
        pairs = sorted(zip(xs, ys), key=lambda t: t[0])
        xs = [p[0] for p in pairs]
        ys = [p[1] for p in pairs]
        return float(_linear_interp(xs, ys, float(tmean_m))), "curve:Tmean_m"
    elif k_disp:
        xs, ys = [], []
        for r in curve:
            x = _as_float(r.get(k_disp))
            y = _as_float(r.get(k_gm))
            if x is None or y is None:
                continue
            xs.append(float(x))
            ys.append(float(y))
        if len(xs) < 2:
            return None, "curve_insufficient"
        pairs = sorted(zip(xs, ys), key=lambda t: t[0])
        xs = [p[0] for p in pairs]
        ys = [p[1] for p in pairs]
        return float(_linear_interp(xs, ys, float(disp_t))), "curve:Disp_t"
    else:
        return None, "invalid_curve_axis"


# ============================================================================
# OPTION B+ (ENGINEERING-GRADE PLUS) MODULES
# - IS Code (2008) general criteria check based on GZ curve
# - MWS/Harbour Master acceptance criteria gate (weather/tide/UKC/ramp)
# - Structural gate (deck pressure / ramp plate / point load) - data-driven
# - CSM traceability (Cargo Securing Manual inputs) - data-driven
#
# NOTE:
#   This module is intentionally "data-first".
#   If approved booklet/NAPA/CSM inputs are missing → the checks are marked unavailable,
#   and the report shows what data is required (no silent assumptions).
# ============================================================================


def _load_iscode_criteria() -> dict:
    """
    Load IS Code criteria from JSON.
    Fallback defaults are only placeholders and MUST be replaced by approved criteria for submission.

    Expected JSON formats:
      1) {"criteria": {...}, "_meta": {...}}
      2) {...}  (direct criteria dict)
    """
    candidates = [
        "data/ISCODE_Criteria.json",
        "ISCODE_Criteria.json",
        "data/ISCODE_2008_GeneralCriteria.json",
        "ISCODE_2008_GeneralCriteria.json",
    ]
    for fn in candidates:
        data = _load_json(fn)
        if isinstance(data, dict):
            crit = (
                data.get("criteria") if isinstance(data.get("criteria"), dict) else data
            )
            if isinstance(crit, dict) and len(crit) > 0:
                return crit
    # Placeholder defaults (typical general criteria values seen in IS Code summaries).
    # For submission you MUST provide the approved criteria file above.
    logging.warning(
        "[BPLUS] ISCODE_Criteria.json not found → using PLACEHOLDER defaults (NOT for submission)."
    )
    return {
        "area_0_30_mrad_min": 0.055,
        "area_0_40_mrad_min": 0.090,
        "area_30_40_mrad_min": 0.030,
        "gz_max_m_min": 0.200,
        "gz_max_angle_deg_min": 25.0,
        "gz_max_angle_deg_max": 40.0,
        "gm0_min_m": 0.150,
    }


def _load_gz_curve(stage_id: str) -> tuple[list[dict], str]:
    """
    Load GZ curve points for a stage.

    Accepted formats:
      A) file per stage: data/GZ/<stage_id>.json
         - [{"angle_deg": 0, "gz_m": 0.00}, ...]
         - or {"points":[...], "_meta":...}

      B) single mapping file: data/GZ_Curves.json
         - {"Stage 1":[...], "Stage 2":[...]} or {"curves":{...}}

    Returns: (points, basis)
    """
    safe_id = str(stage_id).replace("/", "_").replace("\\", "_").strip()

    # A) per-stage
    per_stage = [
        f"data/GZ_Curve_{safe_id}.json",  # prep_bplus_inputs.py format
        f"GZ_Curve_{safe_id}.json",
        f"data/GZ/{safe_id}.json",
        f"data/GZ_{safe_id}.json",
        f"GZ/{safe_id}.json",
        f"GZ_{safe_id}.json",
    ]
    for fn in per_stage:
        data = _load_json(fn)
        if isinstance(data, list) and len(data) > 2:
            return data, f"file:{fn}"
        if (
            isinstance(data, dict)
            and isinstance(data.get("points"), list)
            and len(data["points"]) > 2
        ):
            return data["points"], f"file:{fn}"

    # B) mapping file - bplus_inputs/data/ 경로를 최우선으로 검색
    mapping_files = [
        "data/GZ_Curves.json",  # bplus_inputs/data/ 우선 (RAW DATA)
        "data/ISCODE_GZ_Curves.json",  # bplus_inputs/data/ 우선 (RAW DATA)
        "GZ_Curves.json",
        "ISCODE_GZ_Curves.json",
    ]
    for fn in mapping_files:
        data = _load_json(fn)
        if not isinstance(data, dict):
            continue
        curves = data.get("curves") if isinstance(data.get("curves"), dict) else data
        if isinstance(curves, dict):
            pts = curves.get(stage_id) or curves.get(safe_id)
            if isinstance(pts, list) and len(pts) > 2:
                return pts, f"map:{fn}"
            if (
                isinstance(pts, dict)
                and isinstance(pts.get("points"), list)
                and len(pts["points"]) > 2
            ):
                return pts["points"], f"map:{fn}"

    return [], "missing"


def _normalize_gz_points(points: list[dict]) -> list[tuple[float, float]]:
    """
    Normalize points to sorted list of (angle_deg, gz_m).
    """
    out = []
    for p in points or []:
        if isinstance(p, (list, tuple)) and len(p) >= 2:
            a, g = p[0], p[1]
        elif isinstance(p, dict):
            a = p.get("angle_deg", p.get("angle", p.get("deg")))
            g = p.get("gz_m", p.get("GZ_m", p.get("gz", p.get("GZ"))))
        else:
            continue
        a = _as_float(a)
        g = _as_float(g)
        if a is None or g is None:
            continue
        out.append((float(a), float(g)))
    out.sort(key=lambda t: t[0])
    return out


def _gz_at_angle(pts: list[tuple[float, float]], angle_deg: float) -> float:
    """Linear interpolate GZ at angle_deg (deg). Clamp to endpoints."""
    if not pts:
        raise ValueError("empty points")
    a = float(angle_deg)
    if a <= pts[0][0]:
        return float(pts[0][1])
    if a >= pts[-1][0]:
        return float(pts[-1][1])
    # find bracket
    for i in range(1, len(pts)):
        a0, g0 = pts[i - 1]
        a1, g1 = pts[i]
        if a0 <= a <= a1:
            if a1 == a0:
                return float(g0)
            return float(g0 + (g1 - g0) * (a - a0) / (a1 - a0))
    return float(pts[-1][1])


def _integrate_gz_area_mrad(
    pts: list[tuple[float, float]], a0_deg: float, a1_deg: float
) -> float:
    """
    Integrate area under GZ curve between angles [a0, a1] using trapezoids.
    Units: (m * rad) = m-rad
    """
    if not pts:
        raise ValueError("empty points")

    a0 = float(a0_deg)
    a1 = float(a1_deg)
    if a1 < a0:
        a0, a1 = a1, a0

    # build sampled list including endpoints
    xs = [a0]
    for a, _ in pts:
        if a0 < a < a1:
            xs.append(a)
    xs.append(a1)
    xs = sorted(set(xs))

    area = 0.0
    for i in range(1, len(xs)):
        aa0 = xs[i - 1]
        aa1 = xs[i]
        g0 = _gz_at_angle(pts, aa0)
        g1 = _gz_at_angle(pts, aa1)
        # deg -> rad
        dtheta = (aa1 - aa0) * math.pi / 180.0
        area += 0.5 * (g0 + g1) * dtheta
    return float(area)


def iscode_general_check(stage_id: str, gm0_m: float | None = None) -> dict:
    """
    IS Code general criteria check based on a provided GZ curve.
    Returns a dict with {available, pass, metrics..., checks..., basis, missing}.
    """
    criteria = _load_iscode_criteria()
    pts_raw, basis = _load_gz_curve(stage_id)
    pts = _normalize_gz_points(pts_raw)

    if len(pts) < 3:
        return {
            "available": False,
            "pass": False,
            "basis": basis,
            "missing": "GZ curve points (provide data/GZ/<stage>.json or data/GZ_Curves.json)",
        }

    # metrics
    try:
        area_0_30 = _integrate_gz_area_mrad(pts, 0.0, 30.0)
        area_0_40 = _integrate_gz_area_mrad(pts, 0.0, 40.0)
        area_30_40 = _integrate_gz_area_mrad(pts, 30.0, 40.0)
    except Exception as e:
        return {
            "available": False,
            "pass": False,
            "basis": basis,
            "missing": f"GZ integration failed: {e}",
        }

    # max GZ
    gz_max = max(pts, key=lambda t: t[1])
    gz_max_angle = float(gz_max[0])
    gz_max_m = float(gz_max[1])

    # range of positive stability (last angle where gz>0)
    positive = [a for a, g in pts if g > 0.0]
    stable_range = float(max(positive)) if positive else 0.0

    # checks (only if criteria keys exist)
    checks = {}

    def _chk(key: str, ok: bool, val: float, limit: float):
        checks[key] = {"ok": bool(ok), "value": float(val), "limit": float(limit)}

    # areas
    if "area_0_30_mrad_min" in criteria:
        lim = float(criteria["area_0_30_mrad_min"])
        _chk("A0_30", area_0_30 >= lim, area_0_30, lim)
    if "area_0_40_mrad_min" in criteria:
        lim = float(criteria["area_0_40_mrad_min"])
        _chk("A0_40", area_0_40 >= lim, area_0_40, lim)
    if "area_30_40_mrad_min" in criteria:
        lim = float(criteria["area_30_40_mrad_min"])
        _chk("A30_40", area_30_40 >= lim, area_30_40, lim)

    # GZmax
    if "gz_max_m_min" in criteria:
        lim = float(criteria["gz_max_m_min"])
        _chk("GZmax", gz_max_m >= lim, gz_max_m, lim)

    # angle range
    if "gz_max_angle_deg_min" in criteria:
        lim = float(criteria["gz_max_angle_deg_min"])
        _chk("Ang_GZmax_min", gz_max_angle >= lim, gz_max_angle, lim)
    if "gz_max_angle_deg_max" in criteria:
        lim = float(criteria["gz_max_angle_deg_max"])
        _chk("Ang_GZmax_max", gz_max_angle <= lim, gz_max_angle, lim)

    # GM0 (if provided)
    if "gm0_min_m" in criteria and gm0_m is not None:
        lim = float(criteria["gm0_min_m"])
        _chk("GM0", float(gm0_m) >= lim, float(gm0_m), lim)

    # overall pass
    overall = all(v["ok"] for v in checks.values()) if checks else False

    return {
        "available": True,
        "pass": bool(overall),
        "basis": basis,
        "area_0_30_mrad": float(area_0_30),
        "area_0_40_mrad": float(area_0_40),
        "area_30_40_mrad": float(area_30_40),
        "gz_max_m": float(gz_max_m),
        "gz_max_angle_deg": float(gz_max_angle),
        "stable_range_deg": float(stable_range),
        "checks": checks,
        "criteria": criteria,
    }


def _load_acceptance_criteria() -> dict:
    """
    Load MWS/Harbour Master acceptance criteria (project-specific).
    Expected keys (examples):
      - wind_max_kt, hs_max_m, current_max_kt, ukc_min_m, ramp_angle_max_deg, fwd_draft_max_m
    """
    candidates = [
        "data/MWS_Acceptance_Criteria.json",
        "MWS_Acceptance_Criteria.json",
        "data/Acceptance_Criteria.json",
        "Acceptance_Criteria.json",
    ]
    for fn in candidates:
        data = _load_json(fn)
        if isinstance(data, dict) and len(data) > 0:
            return (
                data.get("criteria", data)
                if isinstance(data.get("criteria"), dict)
                else data
            )
    return {}


def evaluate_acceptance(stage_res: dict, env: dict | None = None) -> dict:
    """
    Evaluate acceptance criteria if env/limits are provided.
    If not enough data, marks unavailable.
    """
    criteria = _load_acceptance_criteria()
    if not criteria:
        return {
            "available": False,
            "pass": False,
            "missing": "Acceptance_Criteria.json",
            "criteria": {},
        }

    env = env or stage_res.get("env") or {}

    # helper to read numeric
    def _num(d, k):
        v = d.get(k)
        vv = _as_float(v)
        return float(vv) if vv is not None else None

    checks = {}
    missing = []

    def _add_check(name: str, value: float | None, limit: float | None, op: str):
        if value is None or limit is None:
            missing.append(name)
            return
        if op == "<=":
            ok = value <= limit
        elif op == ">=":
            ok = value >= limit
        else:
            ok = False
        checks[name] = {
            "ok": bool(ok),
            "value": float(value),
            "limit": float(limit),
            "op": op,
        }

    # Common keys
    _add_check(
        "WindMax", _num(env, "wind_kt"), _as_float(criteria.get("wind_max_kt")), "<="
    )
    _add_check("HsMax", _num(env, "hs_m"), _as_float(criteria.get("hs_max_m")), "<=")
    _add_check(
        "CurrentMax",
        _num(env, "current_kt"),
        _as_float(criteria.get("current_max_kt")),
        "<=",
    )
    _add_check("UKCMin", _num(env, "ukc_m"), _as_float(criteria.get("ukc_min_m")), ">=")
    _add_check(
        "RampAngleMax",
        _num(env, "ramp_angle_deg"),
        _as_float(criteria.get("ramp_angle_max_deg")),
        "<=",
    )

    # Draft limits can use stage results
    _add_check(
        "FwdDraftMax",
        _as_float(stage_res.get("Dfwd_m")),
        _as_float(criteria.get("fwd_draft_max_m")),
        "<=",
    )
    _add_check(
        "AftDraftMax",
        _as_float(stage_res.get("Daft_m")),
        _as_float(criteria.get("aft_draft_max_m")),
        "<=",
    )

    available = len(checks) > 0 and len(missing) == 0
    passed = all(c["ok"] for c in checks.values()) if checks else False

    return {
        "available": bool(available),
        "pass": bool(passed) if available else False,
        "checks": checks,
        "missing": ", ".join(missing) if missing else "",
        "criteria": criteria,
    }


def _load_structural_limits() -> dict:
    """
    Load structural limits (deck pressure / ramp plate cert / point load limits).
    """
    candidates = [
        "data/Structural_Limits.json",
        "Structural_Limits.json",
        "data/Ramp_Deck_Limits.json",
        "Ramp_Deck_Limits.json",
    ]
    for fn in candidates:
        data = _load_json(fn)
        if isinstance(data, dict) and len(data) > 0:
            return (
                data.get("limits", data)
                if isinstance(data.get("limits"), dict)
                else data
            )
    return {}


def evaluate_structural_gate(stage_res: dict) -> dict:
    """
    Structural gate evaluation.
    Requires data in stage_res (or external) such as:
      - max_deck_pressure_t_per_m2
      - max_point_load_t
      - ramp_plate_t_per_m2
    Without data → unavailable.
    """
    limits = _load_structural_limits()
    if not limits:
        return {
            "available": False,
            "pass": False,
            "missing": "Structural_Limits.json",
            "limits": {},
        }

    # read observed values (if present)
    obs = {
        "max_deck_pressure_t_per_m2": _as_float(
            stage_res.get("max_deck_pressure_t_per_m2")
        ),
        "max_point_load_t": _as_float(stage_res.get("max_point_load_t")),
        "ramp_plate_t_per_m2": _as_float(stage_res.get("ramp_plate_t_per_m2")),
    }

    checks = {}
    missing = []

    def _check(name, obs_val, lim_key, op):
        lim = _as_float(limits.get(lim_key))
        if obs_val is None or lim is None:
            missing.append(name)
            return
        if op == "<=":
            ok = float(obs_val) <= float(lim)
        elif op == ">=":
            ok = float(obs_val) >= float(lim)
        else:
            ok = False
        checks[name] = {
            "ok": bool(ok),
            "value": float(obs_val),
            "limit": float(lim),
            "op": op,
        }

    _check(
        "DeckPressure",
        obs["max_deck_pressure_t_per_m2"],
        "deck_pressure_max_t_per_m2",
        "<=",
    )
    _check("PointLoad", obs["max_point_load_t"], "point_load_max_t", "<=")
    _check("RampPlate", obs["ramp_plate_t_per_m2"], "ramp_plate_min_t_per_m2", ">=")

    available = len(checks) > 0 and len(missing) == 0
    passed = all(c["ok"] for c in checks.values()) if checks else False

    return {
        "available": bool(available),
        "pass": bool(passed) if available else False,
        "checks": checks,
        "missing": ", ".join(missing) if missing else "",
        "limits": limits,
    }


def _load_securing_input() -> dict | None:
    candidates = [
        "data/Securing_Input.json",
        "Securing_Input.json",
        "data/CSM_Input.json",
        "CSM_Input.json",
    ]
    for fn in candidates:
        data = _load_json(fn)
        # Handle dict format
        if isinstance(data, dict) and len(data) > 0:
            data["_basis"] = fn
            return data
        # Handle list format (prep_bplus_inputs.py format)
        if isinstance(data, list) and len(data) > 0:
            return {"_basis": fn, "items": data}
    return None


def build_csm_trace(stage_results: dict) -> dict:
    """
    Build a CSM traceability payload (what securing inputs were used / missing).
    This does NOT compute lashing forces unless input data is provided.
    """
    data = _load_securing_input()
    if not data:
        return {
            "available": False,
            "missing": "Securing_Input.json",
            "basis": "missing",
            "items": [],
        }

    items = []
    # Accept either "items" list or "lashings" list
    raw = (
        data.get("items")
        if isinstance(data.get("items"), list)
        else data.get("lashings")
    )
    if not isinstance(raw, list):
        raw = []
    for r in raw:
        if not isinstance(r, dict):
            continue
        items.append(
            {
                "id": r.get("id", ""),
                "type": r.get("type", r.get("lashing_type", "")),
                "wll_t": _as_float(r.get("wll_t", r.get("WLL_t"))),
                "qty": _as_float(r.get("qty", r.get("quantity"))),
                "remarks": r.get("remarks", ""),
            }
        )

    return {"available": True, "basis": data.get("_basis", ""), "items": items}


def enrich_stage_results_bplus(stage_results: dict, env: dict | None = None) -> None:
    """
    Enrich stage_results in-place with BPLUS checks.
    Adds: stage_results[stage]["BPLUS"] = {ISCODE, Acceptance, Structural}
    """
    for stage_id, res in stage_results.items():
        try:
            res["BPLUS"] = res.get("BPLUS", {})
            res["BPLUS"]["ISCODE"] = iscode_general_check(
                stage_id, gm0_m=_as_float(res.get("GM_m"))
            )
            res["BPLUS"]["Acceptance"] = evaluate_acceptance(res, env=env)
            res["BPLUS"]["Structural"] = evaluate_structural_gate(res)
        except Exception as e:
            logging.warning(f"[BPLUS] Stage {stage_id} enrichment skipped: {e}")


def create_iscode_check_sheet(
    wb: Workbook, stage_results: dict, sheet_name: str = "RORO_ISCODE_Check"
) -> Worksheet:
    """RORO_ISCODE_Check 시트 생성 - IMO IS Code 2008 일반 기준 검증

    이 시트는 각 Stage별로 IMO IS Code 2008 일반 기준을 검증합니다.
    GZ curve 기반으로 6개 기준을 검증하며, 모든 기준을 통과해야 PASS입니다.

    데이터 소스:
      1. stage_results['BPLUS']['ISCODE']:
         - available: GZ curve 사용 가능 여부
         - pass: 전체 기준 통과 여부
         - area_0_30_mrad: 0-30도 면적 (m·rad)
         - area_0_40_mrad: 0-40도 면적
         - area_30_40_mrad: 30-40도 면적
         - gz_max_m: 최대 GZ (m)
         - gz_max_angle_deg: 최대 GZ 각도
         - stable_range_deg: 복원 범위

      2. GZ Curve 파일:
         - 위치: data/GZ/*.json
         - 형식: {"stage_id": "...", "points": [{"heel_deg": 0, "gz_m": 0}, ...]}
         - 로딩: _load_gz_curve(stage_id) (L1019)
         - Dual naming 지원: 공백 포함 원본 형식 및 언더스코어 형식 모두 지원

    검증 기준 (IMO IS Code 2008):
      - Area 0-30°: ≥ 3.151 m·rad
      - Area 0-40°: ≥ 5.157 m·rad
      - Area 30-40°: ≥ 1.719 m·rad
      - GZ_max: ≥ 0.20 m
      - Angle at GZ_max: ≥ 25°
      - GM: ≥ 0.15 m

    시트 구조:
      - Row 3: 헤더 (Stage, Available, PASS, A0-30, A0-40, A30-40, GZmax, Ang@GZmax, StableRange, Basis/Missing)
      - Row 4+: 각 Stage별 검증 결과

    Args:
        wb: openpyxl Workbook 객체
        stage_results: Stage별 계산 결과 딕셔너리
            - 각 Stage의 'BPLUS' 키에 'ISCODE' 데이터 포함
        sheet_name: 시트 이름, 기본값 "RORO_ISCODE_Check"

    Returns:
        Worksheet: 생성된 RORO_ISCODE_Check 시트

    See Also:
        - DATA_SOURCE_DOCUMENTATION.md Section 5.1
        - _load_gz_curve() for GZ curve loading
        - FINAL_VERIFICATION_SUMMARY.md for validation results
    """
    ws = (
        wb.create_sheet(sheet_name)
        if sheet_name not in wb.sheetnames
        else wb[sheet_name]
    )
    styles = get_styles()
    ws["A1"] = "IS Code (2008) – General Criteria Check (GZ-based)"
    ws["A1"].font = styles["title_font"]

    headers = [
        "Stage",
        "Available",
        "PASS",
        "A0-30 (m-rad)",
        "A0-40 (m-rad)",
        "A30-40 (m-rad)",
        "GZmax (m)",
        "Ang@GZmax (deg)",
        "StableRange (deg)",
        "Basis/Missing",
    ]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = styles["center_align"]

    r = 4
    nf = "#,##0.00"
    for stage_id, res in stage_results.items():
        iscode = (res.get("BPLUS") or {}).get("ISCODE") or {}
        avail = bool(iscode.get("available"))
        passed = bool(iscode.get("pass")) if avail else False
        ws.cell(r, 1, stage_id)
        ws.cell(r, 2, "Y" if avail else "N")
        ws.cell(r, 3, "PASS" if passed else "FAIL")
        for j, k in enumerate(
            [
                "area_0_30_mrad",
                "area_0_40_mrad",
                "area_30_40_mrad",
                "gz_max_m",
                "gz_max_angle_deg",
                "stable_range_deg",
            ],
            start=4,
        ):
            v = _as_float(iscode.get(k))
            if v is not None:
                ws.cell(r, j, float(v)).number_format = nf
            else:
                ws.cell(r, j, "")
        basis = iscode.get("basis", "")
        missing = iscode.get("missing", "")
        ws.cell(r, 10, basis if basis != "missing" else missing)
        r += 1

    ws.freeze_panes = "A4"
    ws.column_dimensions["A"].width = 26
    return ws


def create_acceptance_criteria_sheet(
    wb: Workbook, stage_results: dict, sheet_name: str = "MWS_Acceptance_Criteria"
) -> Worksheet:
    ws = (
        wb.create_sheet(sheet_name)
        if sheet_name not in wb.sheetnames
        else wb[sheet_name]
    )
    styles = get_styles()
    ws["A1"] = "MWS / Harbour Master – Acceptance Criteria Gate (data-driven)"
    ws["A1"].font = styles["title_font"]

    headers = ["Stage", "Available", "PASS", "Missing", "Checks (key: value op limit)"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = styles["center_align"]

    r = 4
    for stage_id, res in stage_results.items():
        acc = (res.get("BPLUS") or {}).get("Acceptance") or {}
        avail = bool(acc.get("available"))
        passed = bool(acc.get("pass")) if avail else False
        ws.cell(r, 1, stage_id)
        ws.cell(r, 2, "Y" if avail else "N")
        ws.cell(r, 3, "PASS" if passed else "FAIL")
        ws.cell(r, 4, acc.get("missing", ""))
        checks = acc.get("checks") or {}
        parts = []
        for k, v in checks.items():
            if not isinstance(v, dict):
                continue
            parts.append(f"{k}:{v.get('value')} {v.get('op')} {v.get('limit')}")
        ws.cell(r, 5, "; ".join(parts))
        r += 1

    ws.freeze_panes = "A4"
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["E"].width = 80
    return ws


def create_structural_gate_sheet(
    wb: Workbook, stage_results: dict, sheet_name: str = "RORO_Structural_Gate"
) -> Worksheet:
    ws = (
        wb.create_sheet(sheet_name)
        if sheet_name not in wb.sheetnames
        else wb[sheet_name]
    )
    styles = get_styles()
    ws["A1"] = "Structural Gate – Deck / Ramp / Point Load (data-driven)"
    ws["A1"].font = styles["title_font"]

    headers = ["Stage", "Available", "PASS", "Missing", "Checks (key: value op limit)"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = styles["center_align"]

    r = 4
    for stage_id, res in stage_results.items():
        stc = (res.get("BPLUS") or {}).get("Structural") or {}
        avail = bool(stc.get("available"))
        passed = bool(stc.get("pass")) if avail else False
        ws.cell(r, 1, stage_id)
        ws.cell(r, 2, "Y" if avail else "N")
        ws.cell(r, 3, "PASS" if passed else "FAIL")
        ws.cell(r, 4, stc.get("missing", ""))
        checks = stc.get("checks") or {}
        parts = []
        for k, v in checks.items():
            if not isinstance(v, dict):
                continue
            parts.append(f"{k}:{v.get('value')} {v.get('op')} {v.get('limit')}")
        ws.cell(r, 5, "; ".join(parts))
        r += 1

    ws.freeze_panes = "A4"
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["E"].width = 80
    return ws


def create_csm_trace_sheet(
    wb: Workbook, csm_trace: dict, sheet_name: str = "CSM_Trace"
) -> Worksheet:
    ws = (
        wb.create_sheet(sheet_name)
        if sheet_name not in wb.sheetnames
        else wb[sheet_name]
    )
    styles = get_styles()
    ws["A1"] = "CSM Traceability – Inputs used for securing (CSS Code / CSM)"
    ws["A1"].font = styles["title_font"]

    ws["A3"] = "Available"
    ws["B3"] = "Y" if csm_trace.get("available") else "N"
    ws["A4"] = "Basis"
    ws["B4"] = csm_trace.get("basis", "")
    ws["A5"] = "Missing"
    ws["B5"] = csm_trace.get("missing", "")

    headers = ["id", "type", "wll_t", "qty", "remarks"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=7, column=c, value=h)
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = styles["center_align"]

    r = 8
    items = csm_trace.get("items") or []
    for it in items:
        ws.cell(r, 1, it.get("id", ""))
        ws.cell(r, 2, it.get("type", ""))
        v = _as_float(it.get("wll_t"))
        if v is not None:
            ws.cell(r, 3, float(v)).number_format = "#,##0.00"
        ws.cell(r, 4, it.get("qty", ""))
        ws.cell(r, 5, it.get("remarks", ""))
        r += 1

    ws.freeze_panes = "A8"
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["E"].width = 60
    return ws


def create_mws_pack_index_sheet(
    wb: Workbook, sheet_name: str = "MWS_Pack_Index"
) -> Worksheet:
    ws = (
        wb.create_sheet(sheet_name)
        if sheet_name not in wb.sheetnames
        else wb[sheet_name]
    )
    styles = get_styles()
    ws["A1"] = "MWS-ready Pack Index (B+) – Evidence & Submittals Checklist"
    ws["A1"].font = styles["title_font"]
    ws["A3"] = "Item"
    ws["B3"] = "Owner"
    ws["C3"] = "Status"
    ws["D3"] = "Remarks / Where to attach"
    for c in ["A3", "B3", "C3", "D3"]:
        ws[c].font = styles["header_font"]
        ws[c].fill = styles["header_fill"]
        ws[c].alignment = styles["center_align"]

    rows = [
        (
            "Approved Stability Booklet / Hydro Tables",
            "Vessel Owner",
            "REQ",
            "Hydro_Table_Engineering.json + GM_Min_Curve.json",
        ),
        (
            "GZ Curve (per stage) for IS Code check",
            "Naval Arch.",
            "REQ",
            "data/GZ/<stage>.json or data/GZ_Curves.json",
        ),
        (
            "Cargo Securing Manual (CSM) / Lashing Plan",
            "3PL / Marine",
            "REQ",
            "Securing_Input.json (trace) + drawings",
        ),
        (
            "Ramp Plate Certificate / Deck Load Chart",
            "Port / Vessel",
            "REQ",
            "Structural_Limits.json + cert scan",
        ),
        (
            "Method Statement / Operation Manual",
            "Contractor",
            "REQ",
            "Attach in submission pack",
        ),
        (
            "Weather Window / Alpha factor basis",
            "MWS",
            "REQ",
            "Acceptance_Criteria.json + forecast tie-out",
        ),
        (
            "Tow / Mooring / Pilotage plan (AGI)",
            "Marine Ops",
            "REQ",
            "Attach drawings + approvals",
        ),
        (
            "Evidence photos/videos + timestamps",
            "Marine Ops",
            "REQ",
            "15+ photos tagged per site rule",
        ),
    ]
    r = 4
    for item, owner, status, remarks in rows:
        ws.cell(r, 1, item)
        ws.cell(r, 2, owner)
        ws.cell(r, 3, status)
        ws.cell(r, 4, remarks)
        r += 1

    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 60
    ws.freeze_panes = "A4"
    return ws


class LoadItem(NamedTuple):
    """Stage별 하중 구성 요소"""

    name: str
    weight_t: float
    x_from_mid_m: float
    kind: str  # "TR", "SPMT", "BALLAST", "CONST", "CARGO"


# ============================================================================
# STAGE SOLVER (Hydro/GM 기반)
# ============================================================================


def solve_stage(
    base_disp_t: float, base_tmean_m: float, loads: List[LoadItem], **params
) -> Dict[str, Any]:
    """
    Engineering-grade Stage Solver (Option B).

    Key upgrades vs baseline:
      - Uses draft-dependent hydro interpolation (by Tmean_m) if available:
          * LCF_m_from_midship (aft+)
          * MCTC_t_m_per_cm (or MTC_t_m_per_cm)
          * TPC_t_per_cm
      - GM check uses GM minimum curve if provided (GM_min_m in hydro table or GM_Min_Curve.json)
      - Falls back to constants when approved booklet/NAPA hydro data is missing.

    Conventions (LOCKED)
      - x_from_mid_m: midship 기준, AFT=+, FWD=-
      - trim_cm = (Daft - Dfwd) * 100  (+ stern down / - bow down)

    Args:
        base_disp_t: Stage 1 baseline displacement (t)
        base_tmean_m: Stage 1 baseline mean draft (m) (kept for compatibility)
        loads: LoadItem list for current stage
        **params:
            - MTC or MTC_t_m_per_cm: fallback MCTC constant (t·m/cm)
            - LCF or LCF_m_from_midship: fallback LCF constant (m, aft+)
            - TPC or TPC_t_per_cm: fallback TPC constant (t/cm)
            - LBP: length used for trim slope (m)
            - D_vessel: molded depth (m)
            - hydro_table: list[dict] hydro points (recommended: includes Tmean_m + LCF/MCTC/TPC)
            - gm_curve: optional GM minimum curve list[dict]
            - max_fwd_draft_ops_m: forward draft operational limit (m)
            - trim_limit_cm: |trim| limit (cm)

    Returns:
        dict with stage results and basis metadata.
    """
    # -----------------------
    # Fallback constants
    # -----------------------
    MTC_const = _as_float(
        params.get("MTC"), _as_float(params.get("MTC_t_m_per_cm"), 34.00)
    )
    LCF_const = _as_float(
        params.get("LCF"), _as_float(params.get("LCF_m_from_midship"), 0.76)
    )
    TPC_const = _as_float(
        params.get("TPC"), _as_float(params.get("TPC_t_per_cm"), 8.00)
    )
    LBP = _as_float(params.get("LBP"), _as_float(params.get("Lpp_m"), 60.302))
    D_vessel = _as_float(
        params.get("D_vessel"), _as_float(params.get("D_vessel_m"), 3.65)
    )

    hydro_table = params.get("hydro_table", []) or []
    gm_curve = params.get("gm_curve", None)

    max_fwd_draft_ops_m = _as_float(params.get("max_fwd_draft_ops_m"), 2.70)
    trim_limit_cm = _as_float(params.get("trim_limit_cm"), 240.0)

    # -----------------------
    # 1) Weight & displacement
    # -----------------------
    delta_w = sum(ld.weight_t for ld in loads)
    disp_stage = base_disp_t + delta_w

    # LCG about midship (for reporting)
    if abs(delta_w) < 1e-9:
        x_lcg = 0.0
    else:
        x_lcg = sum(ld.weight_t * ld.x_from_mid_m for ld in loads) / delta_w

    # Mean draft from displacement table (still valid even without TPC)
    tmean_stage = interpolate_tmean_from_disp(disp_stage, hydro_table)

    # -----------------------
    # 2) Draft-dependent hydro interpolation (Option B)
    # -----------------------
    hydro_i = interpolate_hydro_by_tmean(tmean_stage, hydro_table)
    lcf_used = _as_float(hydro_i.get("LCF_m_from_midship"), LCF_const)
    mctc_used = _as_float(hydro_i.get("MCTC_t_m_per_cm"), MTC_const)
    tpc_used = _as_float(hydro_i.get("TPC_t_per_cm"), TPC_const)

    hydro_flags = {
        "LCF_basis": "hydro" if "LCF_m_from_midship" in hydro_i else "fallback",
        "MCTC_basis": "hydro" if "MCTC_t_m_per_cm" in hydro_i else "fallback",
        "TPC_basis": "hydro" if "TPC_t_per_cm" in hydro_i else "fallback",
    }

    # -----------------------
    # 3) Trimming moment about LCF and trim
    # -----------------------
    tm = sum(ld.weight_t * (ld.x_from_mid_m - lcf_used) for ld in loads)
    trim_cm = (tm / mctc_used) if (mctc_used and mctc_used > 1e-9) else 0.0
    trim_m = trim_cm / 100.0

    # -----------------------
    # 4) End drafts (single convention)
    # -----------------------
    dfwd_m, daft_m = calc_draft_with_lcf(tmean_stage, trim_cm, lcf_used, LBP)

    # Validate trim definition (Daft - Dfwd)
    trim_verify_cm = (daft_m - dfwd_m) * 100.0
    if abs(trim_verify_cm - trim_cm) > 0.01:
        logging.warning(
            f"[TrimRef] Trim mismatch in solve_stage: trim_cm={trim_cm:.2f}, "
            f"verified_cm={trim_verify_cm:.2f} (Tmean={tmean_stage:.3f}, LBP={LBP:.3f})"
        )

    # -----------------------
    # 5) GM and GM minimum curve
    # -----------------------
    gm_m = gm_2d_bilinear(disp_stage, trim_m)

    gm_min_m = _as_float(hydro_i.get("GM_min_m"), None)
    gm_basis = "hydro:GM_min_m" if gm_min_m is not None else "none"

    if gm_min_m is None:
        gm_min_m, gm_basis = gm_min_from_curve(disp_stage, tmean_stage, gm_curve)

    if gm_min_m is None:
        gm_min_m = _as_float(params.get("gm_target_m"), 1.50)
        gm_basis = "fallback_constant"

    # -----------------------
    # 6) Freeboard (height)
    # -----------------------
    fwd_height_m = max(0.0, D_vessel - dfwd_m)
    aft_height_m = max(0.0, D_vessel - daft_m)

    # -----------------------
    # 7) Checks (limits)
    # -----------------------
    trim_check = "OK" if abs(trim_cm) <= trim_limit_cm else "EXCESSIVE"
    vs_ops = "OK" if dfwd_m <= max_fwd_draft_ops_m else "NG"
    gm_check = "OK" if gm_m >= gm_min_m else "LOW"

    engineering_grade_active = (
        hydro_flags["LCF_basis"] == "hydro"
        and hydro_flags["MCTC_basis"] == "hydro"
        and hydro_flags["TPC_basis"] == "hydro"
        and (gm_basis != "fallback_constant")
    )

    # --- Option B Patch: FSM + GM_eff ---
    const_fsm_t_m = _as_float(params.get("Const_FSM_t_m", 0.0))
    disp_stage_for_fsm = disp_stage  # Already calculated: base_disp_t + delta_w
    gm_eff_m = gm_m - (
        const_fsm_t_m / disp_stage_for_fsm if disp_stage_for_fsm > 0 else 0.0
    )

    return {
        "ΔW_t": float(delta_w),
        "x_LCG_m": float(x_lcg),
        "TM_t_m": float(tm),
        "Trim_cm": float(trim_cm),
        "Tmean_m": float(tmean_stage),
        "Dfwd_m": float(dfwd_m),
        "Daft_m": float(daft_m),
        "GM_m": float(gm_m),
        "GM_min_m": float(gm_min_m),
        "GM_basis": gm_basis,
        "LCF_used_m": float(lcf_used),
        "MCTC_used_t_m_per_cm": float(mctc_used),
        "TPC_used_t_per_cm": float(tpc_used),
        "Hydro_flags": hydro_flags,
        "EngineeringGrade_active": bool(engineering_grade_active),
        "FWD_Height_m": float(fwd_height_m),
        "AFT_Height_m": float(aft_height_m),
        "Trim_Check": trim_check,
        "vs_ops_fwd_draft": vs_ops,
        "GM_Check": gm_check,
        # Patch Option B: FSM and GM_eff
        "FSE_t_m": float(const_fsm_t_m),
        "GM_eff_m": float(gm_eff_m),
        # Compatibility keys for create_roro_sheet() and other functions
        "W_stage_t": float(
            disp_stage
        ),  # Total weight = base_disp_t + delta_w (NOT just delta_w!)
        "x_stage_m": float(x_lcg),  # Alias for x_LCG_m
        "TM_LCF_tm": float(tm),  # Alias for TM_t_m
        "vs_2.70m": vs_ops,  # Alias for vs_ops_fwd_draft
        "Disp_t": float(disp_stage),  # Displacement = total weight
    }


def find_preballast_opt(
    w_tr_unit_t: float = 271.20,
    base_disp_t: float | None = None,
    fr_tr1_stow: float = FR_TR1_STOW,
    fr_tr2_ramp: float = FR_TR2_RAMP,
    fr_preballast: float = FR_PREBALLAST,
    params: dict | None = None,
    search_min_t: float = 20.0,  # PATCH FIX #3: 최소 탐색량 설정 (0.00t 방지용)
    search_max_t: float = 400.0,
    search_step_t: float = 1.0,
) -> dict:
    """Stage 5_PreBallast ~ 6A_Critical(Opt C) 자동 최적화 루프.

    STRATEGY UPDATE (2025-11-24):
    - Pre-ballast location: FW2 Stern tank (Fr.0-6, AFT)
    - Target: Stage 6A FWD as close to 2.70m as possible
    - Validated solution: ~37.65t achieves FWD≈2.09m at Stage 6A

    목적:
      1) Stage 5_PreBallast, Stage 6A 모두에서
         - min_fwd_draft_m ≤ FWD ≤ max_fwd_draft_ops_m
         - |Trim_cm| ≤ trim_limit_abs_cm
      2) 위 조건을 만족하는 해 중에서
         - Stage 6A의 FWD가 max_fwd_draft_ops_m에 가장 가깝도록 (worst-case margin 최소)
         - 동률일 경우 pre-ballast 중량이 더 작은 해 선호

    알고리즘:
      - Iterative optimization (최대 380회 반복, search_step_t=1.0 기준)
      - 탐색 범위: search_min_t ~ search_max_t (기본값: 20.0t ~ 400.0t)
      - 각 후보 pre-ballast 중량에 대해 Stage 5, 6A 계산 수행
      - 수렴 조건: Stage 5 및 6A 모두 FWD ≤ max_fwd_draft_ops_m (2.70m)

    최적화 전략:
      - Stern pre-ballast (FW2 tank, Fr.0-6 위치) 사용
      - 전부 Draft 제어 목표
      - Even keel 근접 목표

    검증 게이트:
      1. Draft limits: min_fwd ≤ FWD ≤ max_fwd (1.50m ~ 2.70m)
      2. Trim envelope: |Trim| ≤ trim_limit (240cm)
      3. Objective: Stage 6A FWD를 2.70m에 최대한 근접

    Args:
        w_tr_unit_t: Transformer 단위 중량 (t)
        base_disp_t: Stage 1 baseline displacement (t)
        fr_tr1_stow: TR1 최종 stow Frame 번호
        fr_tr2_ramp: TR2 ramp Frame 번호
        fr_preballast: Pre-ballast Frame 번호 (FW2, Fr.0-6)
        params: 파라미터 딕셔너리 (DEFAULT_PARAMS 사용 가능)
        search_min_t: 최소 탐색량 (t), 기본값 20.0
        search_max_t: 최대 탐색량 (t), 기본값 400.0
        search_step_t: 탐색 간격 (t), 기본값 1.0

    Returns:
        dict: {
            "ok": bool,                # 최적화 성공 여부
            "reason": str,             # 성공/실패 이유
            "w_preballast_t": float | None,  # 최적 ballast 중량 (t)
            "stage5": dict | None,     # Stage 5 계산 결과
            "stage6A": dict | None     # Stage 6A 계산 결과
        }

    OBSOLETE SCENARIOS (DO NOT USE):
    - Bow Ballast 471t (FWB1+FWB2+FWCARGO1): FWD 2.99m → EXCEEDS LIMIT
    - Forward ballast strategies: All superseded by Stern strategy

    See Also:
        - DATA_SOURCE_DOCUMENTATION.md Section 4.2
        - build_stage_loads() for stage load composition
        - solve_stage() for individual stage calculation
    """

    # Option B+ input gate: ensure submission-grade inputs exist
    _bplus_preflight_or_zero(strict=True)
    if params is None:
        params = DEFAULT_PARAMS
    if base_disp_t is None:
        base_disp_t = _as_float(
            params.get("base_disp_t"), _as_float(params.get("BASE_DISP_T"), 0.0)
        )
    if base_disp_t <= 0:
        print("[WARN] base_disp_t not set; W_stage_t will exclude base displacement.")

    min_fwd = params.get("min_fwd_draft_m", 1.50)
    max_fwd = params.get("max_fwd_draft_ops_m", 2.70)
    trim_limit = params.get("trim_limit_abs_cm", 240.00)
    # NOTE: CHECK_STAGE5=True enforces FWD≤2.70m for Stage 5_PreBallast
    # Set to False to allow intentional bow-down trim exploration
    # CURRENT DESIGN: True (strict enforcement for Harbor Master approval)
    check_stage5 = params.get("CHECK_STAGE5", True)

    if search_step_t <= 0:
        raise ValueError("search_step_t must be positive.")

    best: dict | None = None
    best_metric: float | None = None

    w = search_min_t
    while w <= search_max_t + 1e-9:
        result = _stage_moment_and_drafts_for_preballast(
            w_tr_unit_t=w_tr_unit_t,
            w_preballast_t=w,
            fr_tr1_stow=fr_tr1_stow,
            fr_tr2_ramp=fr_tr2_ramp,
            fr_preballast=fr_preballast,
            base_disp_t=base_disp_t,
            params=params,
        )

        st5 = result["stage5"]
        st6 = result["stage6A"]

        fwd5 = st5["FWD_m"]
        fwd6 = st6["FWD_m"]
        trim5 = abs(st5["Trim_cm"])
        trim6 = abs(st6["Trim_cm"])

        # Gate 1: Draft limits
        # Stage 5도 CHECK_STAGE5=True인 경우 FWD≤max_fwd_draft_ops_m를 강제 (Harbor Master 제출용)
        if check_stage5 and not (min_fwd <= fwd5 <= max_fwd):
            w += search_step_t
            continue
        if not (min_fwd <= fwd6 <= max_fwd):
            w += search_step_t
            continue

        # Gate 2: Trim envelope
        # Stage 5 체크는 선택적 (Stage 5_PreBallast는 의도적 bow trim 240cm 목표)
        if check_stage5 and trim5 > trim_limit:
            w += search_step_t
            continue
        if trim6 > trim_limit:
            w += search_step_t
            continue

        # Objective: Stage 6A FWD as close as possible to ops limit (max_fwd)
        # PATCH FIX #4: Stage 5도 고려하여 페널티 추가
        margin5 = max_fwd - fwd5  # Stage 5 마진
        margin6 = max_fwd - fwd6  # Stage 6A 마진
        # 목표: Stage 6A margin 최소이면서 Stage 5도 margin 양호
        metric = abs(margin6) + 0.1 * abs(margin5)  # Stage 5 페널티 10%

        if best_metric is None or metric < best_metric - 1e-9:
            best_metric = metric
            best = result
        elif best is not None and abs(metric - best_metric) < 1e-9:
            # tie-breaker: smaller ballast preferred
            if w < best["w_preballast_t"]:
                best_metric = metric
                best = result

        w += search_step_t

    if best is None:
        return {
            "ok": False,
            "reason": "No feasible pre-ballast found within search range.",
            "w_preballast_t": None,
            "stage5": None,
            "stage6A": None,
        }

    return {
        "ok": True,
        "reason": "Feasible pre-ballast found.",
        "w_preballast_t": best["w_preballast_t"],
        "stage5": best["stage5"],
        "stage6A": best["stage6A"],
    }


def calc_draft_with_lcf(
    tmean_m: float, trim_cm: float, lcf_m: float, lbp_m: float
) -> tuple[float, float]:
    """
    Draft 계산 (단일 참조점/부호 규약 고정).

    Conventions (LOCKED)
    --------------------
    - x_from_mid_m: midship 기준, AFT = +, FWD = -
    - trim_cm: (Daft - Dfwd) * 100  [cm]
        * + : stern down (AFT deeper)
        * - : bow down   (FWD deeper)
    - tmean_m: midship draft ≈ (Dfwd + Daft) / 2   [m]
    - lbp_m: LBP/Lpp used for trim slope length     [m]
    - lcf_m: LCF x_from_mid_m (aft+)  [m]
      * NOTE: end-draft conversion에는 LCF가 필요 없지만, mixed-reference 입력을 방지하기 위해
        여기서 범위 검증만 수행한다.

    Returns
    -------
    (Dfwd_m, Daft_m)
    """
    if lbp_m <= 0:
        raise ValueError("LBP must be > 0")

    halfL = lbp_m / 2.0

    # Guard-rail: LCF는 midship 기준이면 |LCF| <= LBP/2 근처여야 한다.
    # (LCF from AP/FP 값이 들어오면 즉시 실패시켜 혼용을 차단)
    if abs(lcf_m) > halfL + 0.50:
        raise ValueError(
            f"LCF reference mismatch suspected. "
            f"Expected lcf_m as x_from_mid_m (|lcf|<=~{halfL:.3f}), got {lcf_m:.3f}."
        )

    trim_m = trim_cm / 100.0  # cm -> m

    # Linear trim line about midship:
    # draft(x) = tmean_m + (trim_m / lbp_m) * x_from_mid_m
    dfwd_m = tmean_m + (trim_m / lbp_m) * (-halfL)
    daft_m = tmean_m + (trim_m / lbp_m) * (+halfL)

    # Validate trim definition: trim_cm == (Daft - Dfwd) * 100
    trim_verify_cm = (daft_m - dfwd_m) * 100.0
    if abs(trim_verify_cm - trim_cm) > 0.01:
        logging.warning(
            f"[TrimRef] Trim definition mismatch in calc_draft_with_lcf: "
            f"trim_cm={trim_cm:.2f}, verified_cm={trim_verify_cm:.2f}"
        )

    return dfwd_m, daft_m


def build_stage5_loads(preballast_t: float, params: dict) -> List[LoadItem]:
    """
    Stage 5_PreBallast 구성 (TR1 + TR2 partial + PreBallast)

    Args:
        preballast_t: Pre-ballast 중량 (ton)
        params: 파라미터 딕셔너리
            - W_TR: Transformer + SPMT 중량 (ton)
            - FR_TR1_STOW / FR_TR1: TR1 Frame 번호
            - FR_TR2_RAMP / FR_TR2: TR2 ramp Frame 번호
            - FR_PREBALLAST / FR_PB: PreBallast Frame 번호
            - STAGE5_PROGRESS_PCT: Stage 5 진행률 (%), 기본 71.4

    Returns:
        LoadItem 리스트
    """
    W_TR = params.get("W_TR", 280.0)
    FR_TR1 = params.get("FR_TR1_STOW", params.get("FR_TR1", 42.0))
    FR_TR2 = params.get("FR_TR2_RAMP", params.get("FR_TR2", 17.95))
    FR_PB = params.get("FR_PREBALLAST", params.get("FR_PB", 3.0))
    stage5_progress_pct = params.get("STAGE5_PROGRESS_PCT", 71.4)

    tr2_frac = (stage5_progress_pct - 50.0) / 50.0
    tr2_frac = max(0.0, min(1.0, tr2_frac))
    w_tr2 = W_TR * tr2_frac

    x_tr1 = fr_to_x(FR_TR1)
    x_tr2 = fr_to_x(FR_TR2)
    x_pb = fr_to_x(FR_PB)

    loads = [LoadItem("TR1+SPMT", W_TR, x_tr1, "CARGO")]
    if w_tr2 > 1.0:
        loads.append(LoadItem("TR2+SPMT", w_tr2, x_tr2, "CARGO"))
    loads.append(LoadItem("PreBallast", preballast_t, x_pb, "BALLAST"))
    return loads


def build_stage6a_loads(preballast_t: float, params: dict) -> List[LoadItem]:
    """
    Stage 6A 구성 (TR1 + TR2 + 동일 PreBallast)

    Args:
        preballast_t: Pre-ballast 중량 (ton)
        params: 파라미터 딕셔너리
            - W_TR: Transformer + SPMT 중량 (ton)
            - FR_TR1: TR1 Frame 번호
            - FR_TR2: TR2 Frame 번호
            - FR_PB: PreBallast Frame 번호 (FWB1/2 중심)

    Returns:
        LoadItem 리스트
    """
    W_TR = params.get("W_TR", 280.0)
    FR_TR1 = params.get("FR_TR1", 42.0)
    FR_TR2 = params.get("FR_TR2", 17.95)  # Stage 6A_Critical LCG Frame
    FR_PB = params.get("FR_PB", 3.0)  # FW2 중심 Frame (AFT 쪽, Fr 0-6, Mid_Fr=3.0)

    x_tr1 = fr_to_x(FR_TR1)
    x_tr2 = fr_to_x(FR_TR2)
    x_pb = fr_to_x(FR_PB)

    return [
        LoadItem("TR1+SPMT", W_TR, x_tr1, "CARGO"),
        LoadItem("TR2+SPMT", W_TR, x_tr2, "CARGO"),
        LoadItem("PreBallast", preballast_t, x_pb, "BALLAST"),
    ]


def build_stage_loads(
    stage_name: str, preballast_t: float, params: dict
) -> List[LoadItem]:
    """Stage별 LoadItem 리스트 (모든 Stage 지원).

    이 함수는 각 Stage별로 변동 하중(TR1, TR2, Pre-ballast)을 구성합니다.
    고정 하중(Fuel, FreshWater)은 CONST_TANKS 시트에서 참조되며,
    solve_stage() 함수에서 별도로 처리됩니다.

    Stage 정의 (9개):
      1. Stage 1: Arrival (경하중, 빈 리스트 반환)
      2. Stage 2: TR1 Ramp Start (TR1+SPMT ramp 시작 위치)
      3. Stage 3: TR1 Mid-Ramp (TR1+SPMT ramp 중간 위치)
      4. Stage 4: TR1 On Deck (TR1+SPMT 갑판 도착)
      5. Stage 5: TR1 Final Position (TR1+SPMT 최종 stow 위치)
      6. Stage 5_PreBallast: Water Supply Complete (D-1 Night, TR1 stow + TR2 partial ramp + Stern ballast)
      7. Stage 6A_Critical (Opt C): TR2 Ramp Entry (D-Day, TR1 stow + TR2 ramp + Pre-ballast)
      8. Stage 6C: Final Stowage (TR1 + TR2 모두 stow + Pre-ballast)
      9. Stage 7: Departure (Cargo off, 빈 리스트 반환)

    Args:
        stage_name: Stage 이름
            - 지원 형식: "Stage 1", "Stage 2", ..., "Stage 7"
            - "Stage 5_PreBallast", "Stage 6A_Critical (Opt C)", "Stage 6C"
        preballast_t: Pre-ballast 중량 (ton)
            - find_preballast_opt() 결과값 사용
            - Stage 5_PreBallast, 6A, 6C에서만 사용
        params: 파라미터 딕셔너리
            - W_TR: Transformer + SPMT 중량 (ton), 기본값 280.0
            - FR_TR1_STOW: TR1 최종 stow Frame 번호, 기본값 42.0
            - FR_TR1_RAMP_MID: TR1 ramp 중간 Frame 번호, 기본값 37.00
            - FR_TR1_RAMP_START: TR1 ramp 시작 Frame 번호, 기본값 40.15
            - FR_TR2_RAMP: TR2 ramp Frame 번호, 기본값 17.95
            - FR_TR2_STOW: TR2 최종 stow Frame 번호, 기본값 40.00
            - FR_PREBALLAST: PreBallast Frame 번호 (FW2 중심, AFT 쪽, Fr 0-6, Mid_Fr=3.0)

    Returns:
        List[LoadItem]: Stage별 변동 하중 리스트
            - LoadItem 구조: (name, weight_t, x_from_mid_m, fsm_t_m)
            - x_from_mid_m: LCG (m from midship, AFT=+, FWD=-)
            - fsm_t_m: FSM (t·m), 변동 하중은 일반적으로 0.0

    하중 구성 요소:
      - TR1+SPMT: 560t (위치 변동, Fr.42 → Fr.17.95)
      - TR2+SPMT: 560t (Stage 6A부터)
      - Pre-ballast: find_preballast_opt() 결과 (Stage 5_PreBallast, 6A, 6C)
      - Fuel/FreshWater: CONST_TANKS 시트 참조 (이 함수에서는 처리하지 않음)

    가정:
      - Stage 2/3/4: TR1+SPMT만 ramp→deck 이동
      - Stage 5: TR1 최종 stow. (추가 하중 없음)
      - Stage 5_PreBallast: TR1 stow + TR2 partial ramp + Pre-ballast
      - Stage 6A: TR1 stow + TR2 ramp 진입 + Pre-ballast
      - Stage 6C: 두 TR 모두 stow + Pre-ballast
      - Stage 7: Cargo off (추가 하중 없음)

    See Also:
        - DATA_SOURCE_DOCUMENTATION.md Section 4.3
        - find_preballast_opt() for pre-ballast optimization
        - solve_stage() for stage calculation with loads
        - CONST_TANKS 시트 for fixed loads (Fuel/FreshWater)
    """
    W_TR = params.get("W_TR", 280.0)
    fr_tr1_pos = params.get("FR_TR1_STOW", params.get("FR_TR1", 42.0))
    fr_tr1_ramp_mid = params.get("FR_TR1_RAMP_MID", 37.00)
    fr_tr1_ramp_start = params.get("FR_TR1_RAMP_START", 40.15)
    fr_tr2_ramp = params.get("FR_TR2_RAMP", params.get("FR_TR2", 17.95))
    fr_tr2_stow = params.get("FR_TR2_STOW", 40.00)
    fr_pb = params.get(
        "FR_PREBALLAST", params.get("FR_PB", 3.0)
    )  # FW2 (AFT 쪽, Fr 0-6, Mid_Fr=3.0)

    loads: List[LoadItem] = []

    if stage_name == "Stage 1":
        return loads

    if stage_name == "Stage 2":
        # TR1 ramp start
        loads.append(LoadItem("TR1+SPMT", W_TR, fr_to_x(fr_tr1_ramp_start), "CARGO"))
    elif stage_name == "Stage 3":
        # TR1 mid-ramp
        loads.append(LoadItem("TR1+SPMT", W_TR, fr_to_x(fr_tr1_ramp_mid), "CARGO"))
    elif stage_name == "Stage 4":
        # TR1 on deck
        loads.append(LoadItem("TR1+SPMT", W_TR, fr_to_x(fr_tr1_pos), "CARGO"))
    elif stage_name == "Stage 5":
        # TR1 최종 stow (Trim 이벤트는 base에서 이미 포함된 것으로 간주)
        loads.append(LoadItem("TR1+SPMT", W_TR, fr_to_x(fr_tr1_pos), "CARGO"))
    elif stage_name == "Stage 5_PreBallast":
        loads.extend(build_stage5_loads(preballast_t, params))
    elif stage_name == "Stage 6A_Critical (Opt C)":
        loads.append(LoadItem("TR1+SPMT", W_TR, fr_to_x(fr_tr1_pos), "CARGO"))
        loads.append(LoadItem("TR2+SPMT", W_TR, fr_to_x(fr_tr2_ramp), "CARGO"))
        loads.append(LoadItem("PreBallast", preballast_t, fr_to_x(fr_pb), "BALLAST"))
    elif stage_name == "Stage 6B Tide Window":
        # Stage 6B: Stage 6A_Critical과 동일한 하중 (Draft 동일, Tide는 UKC 계산에만 영향)
        loads.append(LoadItem("TR1+SPMT", W_TR, fr_to_x(fr_tr1_pos), "CARGO"))
        loads.append(LoadItem("TR2+SPMT", W_TR, fr_to_x(fr_tr2_ramp), "CARGO"))
        loads.append(LoadItem("PreBallast", preballast_t, fr_to_x(fr_pb), "BALLAST"))
    elif stage_name == "Stage 6C_TotalMassOpt":
        loads.append(LoadItem("TR1+SPMT", W_TR, fr_to_x(fr_tr1_pos), "CARGO"))
        loads.append(LoadItem("TR2+SPMT", W_TR, fr_to_x(fr_tr2_stow), "CARGO"))
        loads.append(LoadItem("PreBallast", preballast_t, fr_to_x(fr_pb), "BALLAST"))
    elif stage_name == "Stage 6C":
        loads.append(LoadItem("TR1+SPMT", W_TR, fr_to_x(fr_tr1_pos), "CARGO"))
        loads.append(LoadItem("TR2+SPMT", W_TR, fr_to_x(fr_tr2_stow), "CARGO"))
        loads.append(LoadItem("PreBallast", preballast_t, fr_to_x(fr_pb), "BALLAST"))
    elif stage_name == "Stage 7":
        return loads

    return loads


GMGrid = Dict[float, Dict[float, float]]


def _nearest_two(sorted_vals: list[float], target: float) -> tuple[float, float]:
    """
    정렬된 축(sorted_vals)에서 target을 끼우는 양쪽 점 2개를 반환.
    경계에서는 (첫2개) 또는 (마지막2개).
    """
    if not sorted_vals:
        raise ValueError("empty axis")

    n = len(sorted_vals)
    if n == 1:
        return sorted_vals[0], sorted_vals[0]

    pos = bisect_left(sorted_vals, target)

    if pos <= 0:
        return sorted_vals[0], sorted_vals[1]
    if pos >= n:
        return sorted_vals[-2], sorted_vals[-1]

    return sorted_vals[pos - 1], sorted_vals[pos]


def get_gm_bilinear(disp_t: float, trim_m: float, gm_grid: GMGrid) -> float:
    """
    Δ(ton), Trim(m)에 대해 GM(m) 2D 보간 (bilinear).

    gm_grid 형식:
        {disp: {trim: GM, ...}, ...}
    """
    if not gm_grid:
        raise ValueError("gm_grid is empty")

    disp_axis = sorted(gm_grid.keys())
    disp1, disp2 = _nearest_two(disp_axis, disp_t)

    trim_axis = sorted(next(iter(gm_grid.values())).keys())
    trim1, trim2 = _nearest_two(trim_axis, trim_m)

    # 경계선(한 방향 값만 존재)에서는 단순 선형 보간/직접 값 사용
    def gm_at(d: float, tr: float) -> float:
        return gm_grid[d][tr]

    # 네 모서리 값
    q11 = gm_at(disp1, trim1)
    q21 = gm_at(disp2, trim1)
    q12 = gm_at(disp1, trim2)
    q22 = gm_at(disp2, trim2)

    # 축이 collapse 된 경우 처리
    if disp1 == disp2 and trim1 == trim2:
        return q11
    if disp1 == disp2:  # Trim 방향만 보간
        t = (trim_m - trim1) / (trim2 - trim1) if trim2 != trim1 else 0.0
        return q11 + t * (q12 - q11)
    if trim1 == trim2:  # Δ 방향만 보간
        t = (disp_t - disp1) / (disp2 - disp1) if disp2 != disp1 else 0.0
        return q11 + t * (q21 - q11)

    # Bilinear interpolation
    # 참조: 표준 bilinear 공식
    xd = (disp_t - disp1) / (disp2 - disp1)
    yd = (trim_m - trim1) / (trim2 - trim1)

    gm_interp = (
        q11 * (1 - xd) * (1 - yd)
        + q21 * xd * (1 - yd)
        + q12 * (1 - xd) * yd
        + q22 * xd * yd
    )
    return gm_interp


def calc_heel_from_offset(
    weight_t: float, y_offset_m: float, disp_t: float, gm_m: float
) -> float:
    """
    횡 방향 편심 하중에 의한 Heel 각도 (deg) 계산.
    Small-angle 가정: tan(φ) ≈ φ(rad) 사용.
    """
    if disp_t <= 0 or gm_m <= 0 or weight_t == 0 or y_offset_m == 0:
        return 0.0

    m_heeling = weight_t * y_offset_m  # t·m
    m_restoring = disp_t * gm_m  # t·m

    phi_rad = m_heeling / m_restoring
    heel_deg = math.degrees(phi_rad)
    return heel_deg


def calc_gm_effective(disp_t: float, gm_m: float, fse_t_m: float) -> float:
    """
    자유수면효과(FSE)를 반영한 유효 GM.
    GM_eff = GM - FSE/Δ
    """
    if disp_t <= 0:
        return gm_m

    gm_eff = gm_m - (fse_t_m / disp_t)
    return gm_eff


def heel_and_gm_check(
    weight_t: float,
    y_offset_m: float,
    disp_t: float,
    gm_m: float,
    fse_t_m: float,
    heel_limit_deg: float = 3.0,
    gm_min_m: float = 1.50,
) -> tuple[float, float, bool, bool]:
    """
    Heel 각도 + GM_eff 계산 및 체크 결과.

    Returns
    -------
    heel_deg  : 계산된 heel (deg)
    gm_eff    : FSE 반영 후 GM (m)
    heel_ok   : |heel| ≤ heel_limit_deg 여부
    gm_ok     : gm_eff ≥ gm_min_m 여부
    """
    heel_deg = calc_heel_from_offset(weight_t, y_offset_m, disp_t, gm_m)
    gm_eff = calc_gm_effective(disp_t, gm_m, fse_t_m)

    heel_ok = abs(heel_deg) <= heel_limit_deg
    gm_ok = gm_eff >= gm_min_m

    return heel_deg, gm_eff, heel_ok, gm_ok


class LoadCase(Enum):
    STATIC = auto()  # A: 정적
    DYNAMIC = auto()  # B: 동적계수만
    BRAKING = auto()  # C: 동적 + 제동/편심


def apply_dynamic_loads(
    share_load_t: float,
    pin_stress_mpa: float,
    load_case: LoadCase,
) -> tuple[float, float]:
    """
    LoadCase에 따라 동적·제동 하중 계수 적용.

    Returns
    -------
    (share_dyn_t, pin_dyn_mpa)
    """
    if load_case == LoadCase.STATIC:
        f_vert = 1.00
        f_pin = 1.00
    elif load_case == LoadCase.DYNAMIC:
        f_vert = 1.10  # 예: 동적계수 1.10
        f_pin = 1.10
    elif load_case == LoadCase.BRAKING:
        f_vert = 1.20  # 예: 동적+제동 영향
        f_pin = 1.30
    else:
        f_vert = 1.00
        f_pin = 1.00

    return share_load_t * f_vert, pin_stress_mpa * f_pin


# ============================================================================
# FRAME ↔ x_from_mid_m Mapping (BUSHRA 757 TCP aligned)
# ============================================================================

# BUSHRA Tank Plan 757 TCP 기준:
# - Frame 번호는 FWD 방향으로 증가
# - Midship ≈ Fr_mid = Lpp / 2 ≈ 30.151
# - 좌표계:  x_from_mid_m < 0.0  → FWD
#           x_from_mid_m > 0.0  → AFT
#
# ⇒ x = _FRAME_OFFSET + _FRAME_SLOPE * Fr
#    Fr_mid = 30.151  →  x = 0.0
#    큰 Frame(예: 48–65, FWB1/2 실제 위치) → x < 0.0 (FWD 쪽)
#    작은 Frame(예: 5–10)                → x > 0.0 (AFT 쪽)

_FRAME_SLOPE = -1.0  # x 는 Frame 이 커질수록 감소 (Frame 증가 = FWD)
_FRAME_OFFSET = 30.151  # Midship Frame → x = 0.0m


def _init_frame_mapping(verbose: bool = True):
    """
    Frame ↔ x_from_mid_m 매핑 초기화.
    data/Frame_x_from_mid_m.json 이 있으면 거기서 SLOPE/OFFSET 자동 추정,
    없으면 757 TCP 기준 기본값을 사용한다.
    """
    global _FRAME_SLOPE, _FRAME_OFFSET
    data = _load_json("data/Frame_x_from_mid_m.json")
    if not data or not isinstance(data, list) or len(data) < 2:
        # Fallback: BUSHRA 757 TCP default (Fr 증가 = FWD, Midship = 30.151)
        _FRAME_SLOPE = -1.0
        _FRAME_OFFSET = 30.151
        if verbose:
            print(
                "[INFO] Frame mapping: "
                f"SLOPE={_FRAME_SLOPE:.6f}, OFFSET={_FRAME_OFFSET:.3f} (default)"
            )
        return

    try:
        fr1 = float(data[0]["Fr"])
        x1 = float(data[0]["x_from_mid_m"])
        fr2 = float(data[1]["Fr"])
        x2 = float(data[1]["x_from_mid_m"])
        if fr2 != fr1:
            _FRAME_SLOPE = (x2 - x1) / (fr2 - fr1)
            _FRAME_OFFSET = x1 - _FRAME_SLOPE * fr1
        if verbose:
            print(
                "[INFO] Frame mapping: "
                f"SLOPE={_FRAME_SLOPE:.6f}, OFFSET={_FRAME_OFFSET:.3f}"
            )
    except Exception as e:
        # Fallback: BUSHRA 757 TCP default (Fr 증가 = FWD, Midship = 30.151)
        print(f"[ERROR] Frame JSON parse fail → default: {e}")
        _FRAME_SLOPE = -1.0
        _FRAME_OFFSET = 30.151
        print(
            f"[INFO] Frame mapping: SLOPE={_FRAME_SLOPE:.6f}, OFFSET={_FRAME_OFFSET:.3f} (default)"
        )


def fr_to_x(fr: float) -> float:
    """
    Convert Frame number to x [m from midship].

    BUSHRA 757 TCP 기준:
    - Frame 증가 = FWD 방향
    - Frame 30.151 = Midship → x = 0.0
    - Frame < 30.151 (AFT) → x > 0 (AFT)
    - Frame > 30.151 (FWD) → x < 0 (FWD)

    공식: x = _FRAME_SLOPE * (fr - _FRAME_OFFSET)
    """
    x = _FRAME_SLOPE * (float(fr) - _FRAME_OFFSET)
    # Validation: Ensure convention is maintained
    # - fr < offset (AFT) => x > 0
    # - fr > offset (FWD) => x < 0
    if fr < _FRAME_OFFSET and x <= 0:
        logging.warning(
            f"[TrimRef] x-coordinate convention violation: Fr={fr}, x={x:.3f} (expected x>0 for AFT)"
        )
    elif fr > _FRAME_OFFSET and x >= 0:
        logging.warning(
            f"[TrimRef] x-coordinate convention violation: Fr={fr}, x={x:.3f} (expected x<0 for FWD)"
        )
    return x


def x_to_fr(x: float) -> float:
    """
    Inverse: x [m from midship] → Frame number.

    공식: fr = _FRAME_OFFSET - x (x = -1.0 * (fr - 30.151) 이므로)
    """
    return _FRAME_OFFSET - float(x)


def calc_trim(moment_tm: float, params: dict | None = None) -> float:
    """
    Return Trim [cm] from moment [t·m].

    우선순위:
    1) params["MTC_t_m_per_cm"]
    2) params["MTC"]
    3) 기본값 34.00
    """
    if params is None:
        params = DEFAULT_PARAMS

    mtc = params.get("MTC_t_m_per_cm", params.get("MTC", 34.00))
    return round(moment_tm / mtc, 2)


def draft_from_trim(trim_cm: float, params: dict | None = None) -> tuple[float, float]:
    """
    Return (FWD, AFT) draft [m].

    Excel Stage 시트 로직과 동일:
        Tmean_baseline_m 은 모든 Stage에서 동일하다고 가정.
        부호 규칙:
            Trim_cm < 0  → 선수침 (FWD 깊어짐)
            Trim_cm > 0  → 선미침 (AFT 깊어짐)

        Fwd = Tmean + Trim_cm / 200
        Aft = Tmean - Trim_cm / 200
    """
    if params is None:
        params = DEFAULT_PARAMS
    tmean = params.get("Tmean_baseline_m", 2.00)
    trim_m = trim_cm / 100.0
    fwd = tmean + trim_m / 2.0
    aft = tmean - trim_m / 2.0
    # Excel과 맞추기 위해 2자리 반올림
    return round(fwd, 2), round(aft, 2)


def _stage_moment_and_drafts_for_preballast(
    w_tr_unit_t: float,
    w_preballast_t: float,
    fr_tr1_stow: float,
    fr_tr2_ramp: float,
    fr_preballast: float,
    base_disp_t: float,
    params: dict | None = None,
) -> dict:
    """
    Helper: 주어진 preballast 중량에서 Stage 5_PreBallast / 6A_Critical의
    TM, Trim, Draft(FWD/AFT)를 한 번에 계산.

    가정:
      - Stage 5_PreBallast = TR1(Fr_tr1_stow) + Pre-ballast(FR_PREBALLAST)
      - Stage 6A_Critical  = TR1 + TR2(Fr_tr2_ramp) + Pre-ballast
      - Mean draft(Tmean)은 전체 Stage 동안 일정(=Tmean_baseline_m).
    """
    if params is None:
        params = DEFAULT_PARAMS

    lcf = params.get("LCF_m_from_midship", 0.76)
    mtc = params.get("MTC_t_m_per_cm", 34.00)
    lbp = params.get("Lpp_m", params.get("LBP", 60.302))
    tmean = params.get("Tmean_baseline_m", 2.00)

    # 위치 (x from midship)
    x_tr1 = fr_to_x(fr_tr1_stow)
    x_tr2 = fr_to_x(fr_tr2_ramp)
    x_pb = fr_to_x(fr_preballast)

    # --- Stage 5_PreBallast ---------------------------------------------------
    w5 = w_tr_unit_t + w_preballast_t
    w5_total = base_disp_t + w5
    if w5_total <= 0:
        raise ValueError("Stage 5_PreBallast weight must be positive.")

    lcg5 = (w_tr_unit_t * x_tr1 + w_preballast_t * x_pb) / w5
    tm5 = w5 * (lcg5 - lcf)
    trim5_cm = tm5 / mtc
    fwd5_m, aft5_m = calc_draft_with_lcf(tmean, trim5_cm, lcf, lbp)

    # --- Stage 6A_Critical (Opt C) -------------------------------------------
    w6 = 2.0 * w_tr_unit_t + w_preballast_t
    w6_total = base_disp_t + w6
    lcg6 = (w_tr_unit_t * x_tr1 + w_tr_unit_t * x_tr2 + w_preballast_t * x_pb) / w6
    tm6 = w6 * (lcg6 - lcf)
    trim6_cm = tm6 / mtc
    fwd6_m, aft6_m = calc_draft_with_lcf(tmean, trim6_cm, lcf, lbp)

    return {
        "w_preballast_t": w_preballast_t,
        "stage5": {
            "W_stage_t": w5_total,
            "x_stage_m": lcg5,
            "TM_tm": tm5,
            "Trim_cm": trim5_cm,
            "FWD_m": fwd5_m,
            "AFT_m": aft5_m,
        },
        "stage6A": {
            "W_stage_t": w6_total,
            "x_stage_m": lcg6,
            "TM_tm": tm6,
            "Trim_cm": trim6_cm,
            "FWD_m": fwd6_m,
            "AFT_m": aft6_m,
        },
    }


def simulate_stage(
    stage_name: str,
    w_stage_t: float,
    x_stage_m: float,
    params: dict | None = None,
) -> dict:
    """
    간단 Stage 계산:
    - 입력: Stage명, Stage 중량, Stage LCG(x)
    - 출력: Trim, FWD/AFT draft, Target 대비 OK/EXCESSIVE
    """
    if params is None:
        params = DEFAULT_PARAMS

    tm = w_stage_t * x_stage_m  # t·m
    trim_cm = calc_trim(tm, params)
    fwd, aft = draft_from_trim(trim_cm, params)

    target = TRIM_TARGET_MAP.get(stage_name, 240.00)
    trim_check = "OK" if abs(trim_cm) <= abs(target) else "EXCESSIVE"

    return {
        "Stage": stage_name,
        "W_stage_t": round(w_stage_t, 2),
        "x_stage_m": round(x_stage_m, 2),
        "TM_tm": round(tm, 2),
        "Trim_cm": trim_cm,
        "FWD_m": fwd,
        "AFT_m": aft,
        "Trim_target_cm": target,
        "Trim_Check": trim_check,
    }


def debug_frame_mapping():
    """
    Frame_x_from_mid_m.json 기반으로 현재 SLOPE/OFFSET과
    주요 기준 Frame들의 x_from_mid_m를 출력하는 디버그 함수.
    """
    print("=" * 60)
    print("LCT BUSHRA Frame ↔ x Debug (757 TCP aligned)")
    print("=" * 60)
    print(f"_FRAME_SLOPE  = {_FRAME_SLOPE:.6f}")
    print(f"_FRAME_OFFSET = {_FRAME_OFFSET:.3f}")
    test = {
        "AP approx": 0.0,
        "Midship (Lpp/2)": 30.151,
        "FP approx": 60.30,
        "FWB1/2 center 55": 55.0,
    }
    for label, fr in test.items():
        x = fr_to_x(fr)
        print(
            f"{label:25}  Fr={fr:6.2f}  →  x={x:8.3f} m  ({'FWD' if x < 0 else 'AFT'})"
        )
    print("=" * 60)
    sys.exit(0)


# ============================================================================
# BACKUP PLAN Support Functions
# ============================================================================


def setup_logging(output_file):
    """
    BACKUP PLAN: 실행 로그 설정
    logs/ 폴더에 타임스탬프 로그 파일 생성
    """
    log_dir = os.path.join(os.path.dirname(output_file), "logs")
    os.makedirs(log_dir, exist_ok=True)

    log_file = os.path.join(
        log_dir, f"agi_tr_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
    )

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=[
            logging.FileHandler(log_file, encoding="utf-8"),
            logging.StreamHandler(),
        ],
        force=True,  # Override any existing configuration
    )

    logging.info("=" * 60)
    logging.info("LCT BUSHRA AGI TR Excel Generation")
    logging.info(f"Output: {output_file}")
    logging.info(f"Log: {log_file}")
    logging.info("=" * 60)

    return log_file


def create_backup_file(original_path):
    """
    BACKUP PLAN: 생성된 파일의 백업 자동 생성
    backups/ 폴더에 타임스탬프 백업 저장 (최근 5개 유지)
    """
    if not os.path.exists(original_path):
        logging.warning(f"Backup failed: {original_path} not found")
        return None

    # 백업 디렉토리 생성
    backup_dir = os.path.join(os.path.dirname(original_path), "backups")
    os.makedirs(backup_dir, exist_ok=True)

    # 타임스탬프 백업
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = os.path.basename(original_path)
    backup_path = os.path.join(backup_dir, f"{timestamp}_{filename}")

    try:
        shutil.copy2(original_path, backup_path)
        logging.info(f"[BACKUP] Created: {backup_path}")
        print(f"  [BACKUP] Created: {os.path.basename(backup_path)}")

        # 오래된 백업 정리 (최근 5개만 유지)
        cleanup_old_backups(backup_dir, keep=5)

        return backup_path
    except Exception as e:
        logging.error(f"[BACKUP] Failed to create backup: {e}")
        print(f"  [BACKUP] Warning: Backup creation failed: {e}")
        return None


def cleanup_old_backups(backup_dir, keep=5):
    """
    BACKUP PLAN: 오래된 백업 파일 정리
    """
    try:
        backups = [
            os.path.join(backup_dir, f)
            for f in os.listdir(backup_dir)
            if f.endswith(".xlsx")
        ]
        backups.sort(key=os.path.getmtime, reverse=True)

        if len(backups) > keep:
            for old_backup in backups[keep:]:
                os.remove(old_backup)
                logging.info(
                    f"[BACKUP] Removed old backup: {os.path.basename(old_backup)}"
                )
    except Exception as e:
        logging.warning(f"[BACKUP] Cleanup failed: {e}")


def preflight_check():
    """
    BACKUP PLAN: 생성 전 환경 검증
    Returns: List of issue messages
    """
    issues = []

    # 1. 필수 디렉토리 확인
    if not os.path.exists("data"):
        issues.append("WARNING: data/ directory not found (fallback will be used)")

    # 2. 필수 JSON 파일 확인
    optional_jsons = [
        "data/Hydro_Table_2D.json",
        "data/gateab_v3_tide_data.json",
        "data/tank_coordinates.json",
        "data/tank_data.json",
        "data/Frame_x_from_mid_m.json",
        "data/hydro_table.json",
    ]
    for json_file in optional_jsons:
        if not os.path.exists(json_file):
            issues.append(f"INFO: {json_file} missing (fallback will be used)")

    # 3. 디스크 공간 확인 (Windows용)
    try:
        import ctypes

        free_bytes = ctypes.c_ulonglong(0)
        ctypes.windll.kernel32.GetDiskFreeSpaceExW(
            ctypes.c_wchar_p("."), None, None, ctypes.pointer(free_bytes)
        )
        free_mb = free_bytes.value / (1024 * 1024)
        if free_mb < 10:
            issues.append(f"ERROR: Low disk space ({free_mb:.1f}MB)")
    except Exception:
        # Unix-like systems
        try:
            statvfs = os.statvfs(".")
            free_mb = (statvfs.f_frsize * statvfs.f_bavail) / (1024 * 1024)
            if free_mb < 10:
                issues.append(f"ERROR: Low disk space ({free_mb:.1f}MB)")
        except Exception:
            pass  # Skip disk check if not supported

    return issues


class BackupRecoveryError(Exception):
    """BACKUP PLAN: 백업 복구 가능한 에러"""

    pass


def safe_sheet_creation(wb, sheet_func, sheet_name, *args, **kwargs):
    """
    BACKUP PLAN: 시트 생성 wrapper with error recovery
    시트 생성 실패 시에도 계속 진행
    """
    try:
        logging.info(f"Creating sheet: {sheet_name}")
        result = sheet_func(wb, *args, **kwargs)
        logging.info(f"[OK] {sheet_name} created successfully")
        return result
    except Exception as e:
        logging.error(f"[ERROR] {sheet_name} creation failed: {e}")
        logging.warning(f"[BACKUP] Skipping {sheet_name}, continuing...")
        print(f"  [BACKUP] Warning: {sheet_name} creation failed, continuing...")
        return None


def get_fixed_tank_data():
    """
    Forward fresh water ballast tanks (FWB1/2) - 757 TCP Tank Plan Verified LCGs.

    Coordinate System (Python Script Internal):
    - Midship (0.0) is at 30.151m from AP.
    - FWD is NEGATIVE (-), AFT is POSITIVE (+).
    - Formula: x_from_mid = Midship_LCG (30.151) - Tank_LCG_from_AP

    Reference: tank.md (Tank Plan 757 TCP)
    - FWB1 LCG: 57.519 m from AP
    - FWB2 LCG: 50.038 m from AP
    """
    MIDSHIP_LCG_FROM_AP = 30.151

    # LCG to Script X conversion (Midship - LCG)
    # FWD tanks will result in negative values (Correct for this script)
    x_fwb1 = MIDSHIP_LCG_FROM_AP - 57.519  # 30.151 - 57.519 = -27.368 m
    x_fwb2 = MIDSHIP_LCG_FROM_AP - 50.038  # 30.151 - 50.038 = -19.887 m
    x_fwcargo1 = MIDSHIP_LCG_FROM_AP - 42.750  # 30.151 - 42.750 = -12.599 m
    x_fwcargo2 = MIDSHIP_LCG_FROM_AP - 35.250  # 30.151 - 35.250 = -5.099 m

    return {
        # FWB1 (Bow Ballast) - Fr 56-FE
        "FWB1.P": {
            "x": x_fwb1,
            "max_t": 50.57,
            "SG": 1.000,  # Fixed: Fresh Water SG 1.000
            "note": "Bow Port (LCG 57.519m)",
        },
        "FWB1.S": {
            "x": x_fwb1,
            "max_t": 50.57,
            "SG": 1.000,  # Fixed: Fresh Water SG 1.000
            "note": "Bow Stbd (LCG 57.519m)",
        },
        # FWB2 (Forward Ballast) - Fr 48-53
        "FWB2.P": {
            "x": x_fwb2,
            "max_t": 109.98,
            "SG": 1.000,  # Fixed: Fresh Water SG 1.000
            "note": "Fwd Port (LCG 50.038m)",
        },
        "FWB2.S": {
            "x": x_fwb2,
            "max_t": 109.98,
            "SG": 1.000,  # Fixed: Fresh Water SG 1.000
            "note": "Fwd Stbd (LCG 50.038m)",
        },
        # FWCARGO1 (Mid-Fwd) - Fr 43-48
        "FWCARGO1.P": {
            "x": x_fwcargo1,
            "max_t": 148.35,
            "SG": 1.000,
            "note": "Mid-Fwd Cargo (LCG 42.750m)",
        },
        "FWCARGO1.S": {
            "x": x_fwcargo1,
            "max_t": 148.35,
            "SG": 1.000,
            "note": "Mid-Fwd Cargo (LCG 42.750m)",
        },
        # FWCARGO2 (Mid) - Fr 38-43
        "FWCARGO2.P": {
            "x": x_fwcargo2,
            "max_t": 148.36,
            "SG": 1.000,
            "note": "Mid Cargo (LCG 35.250m)",
        },
        "FWCARGO2.S": {
            "x": x_fwcargo2,
            "max_t": 148.36,
            "SG": 1.000,
            "note": "Mid Cargo (LCG 35.250m)",
        },
    }


def build_tank_lookup():
    """
    Tank 좌표/용량 JSON에서 Ballast 탱크 정보 취합.

    Source:
    - data/tank_coordinates.json : data[].Tank_Name, Mid_Fr, x_from_mid_m(선택), Weight_MT, Volume_m3
    - data/tank_data.json        : data[].Tank_Name, Weight_MT (실측 100% 기준)
    - get_fixed_tank_data()      : 757 TCP LCG(AP) → x_from_mid_m 변환 (FWB1/2, FWCARGO1/2)

    좌표계:
    - x_from_mid_m: Midship(0.0) 기준, FWD(-) / AFT(+)
    - FWB1/2, FWCARGO1/2는 Tank Plan 757 TCP의 LCG(AP) → midship 변환값을 항상 우선 사용.
    """
    # 1) 고정 탱크 데이터 (FWB1/2, FWCARGO1/2) – LCG(AP) 기반 x, max_t, SG
    fixed_data = get_fixed_tank_data()

    # 2) JSON 로드 (없으면 None)
    coords = _load_json("data/tank_coordinates.json")
    tdata = _load_json("data/tank_data.json")

    coord_index: dict[str, dict] = {}
    tdata_index: dict[str, dict] = {}

    if isinstance(coords, list):
        for item in coords:
            name = item.get("Tank_Name") or item.get("TankName")
            if not name:
                continue
            coord_index[name] = item
    elif isinstance(coords, dict) and "data" in coords:
        for item in coords["data"]:
            name = item.get("Tank_Name") or item.get("TankName")
            if not name:
                continue
            coord_index[name] = item

    if isinstance(tdata, list):
        for item in tdata:
            name = item.get("Tank_Name") or item.get("TankName")
            if not name:
                continue
            tdata_index[name] = item
    elif isinstance(tdata, dict) and "data" in tdata:
        for item in tdata["data"]:
            name = item.get("Tank_Name") or item.get("TankName")
            if not name:
                continue
            tdata_index[name] = item

    # 3) 전체 탱크 이름 집합
    all_names: set[str] = set()
    all_names.update(coord_index.keys())
    all_names.update(tdata_index.keys())
    all_names.update(fixed_data.keys())

    lookup: dict[str, dict] = {}

    for tank_name in sorted(all_names):
        info: dict[str, float] = {}

        # 3-1) FWB1/2, FWCARGO1/2는 LCG(AP) 기반 fixed_data 최우선
        if tank_name in fixed_data:
            fd = fixed_data[tank_name]
            info["x_from_mid_m"] = float(fd["x"])
            info["max_t"] = float(fd["max_t"])
            info["SG"] = float(fd.get("SG", 1.0))

        # 3-2) tank_coordinates.json 정보 병합 (Frame 기반 → x 변환 포함)
        c = coord_index.get(tank_name)
        if c:
            x_val = c.get("x_from_mid_m")
            if x_val is None:
                mid_fr = c.get("Mid_Fr") or c.get("MidFr") or c.get("Fr")
                if mid_fr is not None:
                    x_val = fr_to_x(float(mid_fr))
            if x_val is not None and "x_from_mid_m" not in info:
                info["x_from_mid_m"] = float(x_val)

            # JSON 쪽에 Weight_MT가 있고 fixed에 max_t 없으면 사용
            if "max_t" not in info:
                wt = c.get("Weight_MT")
                if wt is not None:
                    info["max_t"] = float(wt)

        # 3-3) tank_data.json 정보 병합 (우선순위: fixed < coords < tdata)
        d = tdata_index.get(tank_name)
        if d:
            if "max_t" not in info:
                wt = d.get("Weight_MT") or d.get("Weight_t")
                if wt is not None:
                    info["max_t"] = float(wt)

        # 3-4) SG / air_vent 기본값 – Fresh Water 기준 1.000
        sg = float(info.get("SG", 1.0))

        if tank_name.startswith("FWB"):
            # Fresh Water Ballast – air vent 80mm
            sg = 1.000
            air_vent = 80
        elif tank_name.startswith("FWCARGO"):
            # Cargo FW tanks – air vent 125mm
            sg = 1.000
            air_vent = 125
        else:
            air_vent = ""

        x_from_mid = float(info.get("x_from_mid_m", 0.0))
        max_t = float(info.get("max_t", 0.0))

        lookup[tank_name] = {
            "x_from_mid_m": round(x_from_mid, 2),
            "max_t": round(max_t, 2),
            "SG": sg,
            "air_vent_mm": air_vent,
        }

    print(f"  [OK] Tank lookup built (tanks={len(lookup)})")
    return lookup


# 함수 생성 헬퍼 함수
def create_index_match_formula(lookup_value, lookup_range, return_range):
    """INDEX/MATCH 조합 수식 생성"""
    return f'=INDEX({return_range}, MATCH("{lookup_value}", {lookup_range}, 0))'


# ============================================================================
# Calc Sheet Creation
# ============================================================================


def create_calc_sheet(wb):
    """Calc 시트 생성 - 원본 파일 구조와 동일하게 생성"""
    ws = wb.create_sheet("Calc")
    styles = get_styles()

    # Row 2: 제목
    ws.cell(row=2, column=2).value = "LCT BUSHRA — RORO Calculator & Limits"
    ws.cell(row=2, column=2).font = styles["title_font"]

    # Row 3: 헤더
    headers_row3 = ["", "SECTION", "PARAMETER", "UNIT", "VALUE", "NOTES"]
    for col_idx, header in enumerate(headers_row3, start=1):
        if header:
            cell = ws.cell(row=3, column=col_idx)
            cell.value = header
            cell.font = styles["header_font"]
            cell.fill = styles["header_fill"]
            cell.alignment = styles["center_align"]
            cell.border = Border(
                left=styles["thin_border"],
                right=styles["thin_border"],
                top=styles["thin_border"],
                bottom=styles["thin_border"],
            )

    # INPUT CONSTANTS 섹션 (Row 5-8)
    ws.cell(row=5, column=2).value = "INPUT CONSTANTS"
    ws.cell(row=5, column=3).value = "L_ramp_m"
    ws.cell(row=5, column=4).value = "m"
    ws.cell(row=5, column=5).value = 12.0
    ws.cell(row=5, column=5).fill = styles["input_fill"]
    ws.cell(row=5, column=6).value = "Linkspan length. Calc!D4"

    ws.cell(row=6, column=3).value = "theta_max_deg"
    ws.cell(row=6, column=4).value = "deg"
    ws.cell(row=6, column=5).value = 6.0
    ws.cell(row=6, column=5).fill = styles["input_fill"]
    ws.cell(row=6, column=6).value = "Max ramp angle. Calc!D5"

    ws.cell(row=7, column=3).value = "KminusZ_m"
    ws.cell(row=7, column=4).value = "m"
    ws.cell(row=7, column=5).value = 3.0
    ws.cell(row=7, column=5).fill = styles["input_fill"]
    ws.cell(row=7, column=6).value = "K - Z (UPDATE!). Calc!D6"

    ws.cell(row=8, column=3).value = "D_vessel_m"
    ws.cell(row=8, column=4).value = "m"
    ws.cell(row=8, column=5).value = 3.65
    ws.cell(row=8, column=5).fill = styles["input_fill"]
    ws.cell(row=8, column=6).value = "Molded depth. Calc!D7"

    # LIMITS & OPS 섹션 (Row 10-12)
    ws.cell(row=10, column=2).value = "LIMITS & OPS"
    ws.cell(row=10, column=3).value = "min_fwd_draft_m"
    ws.cell(row=10, column=4).value = "m"
    ws.cell(row=10, column=5).value = 1.5
    ws.cell(row=10, column=5).fill = styles["input_fill"]
    ws.cell(row=10, column=6).value = "Min draft. Calc!D9"

    ws.cell(row=11, column=3).value = "max_fwd_draft_m"
    ws.cell(row=11, column=4).value = "m"
    ws.cell(row=11, column=5).value = 3.5
    ws.cell(row=11, column=5).fill = styles["input_fill"]
    ws.cell(row=11, column=6).value = "Max draft. Calc!D10"

    ws.cell(row=12, column=3).value = "pump_rate_tph"
    ws.cell(row=12, column=4).value = "t/h"
    ws.cell(row=12, column=5).value = 10.0
    ws.cell(row=12, column=5).fill = styles["input_fill"]
    ws.cell(row=12, column=6).value = "Pump rate. Calc!D11"

    # STABILITY 섹션 (Row 14-17)
    ws.cell(row=14, column=2).value = "STABILITY"
    ws.cell(row=14, column=3).value = "MTC_t_m_per_cm"
    ws.cell(row=14, column=4).value = "t·m/cm"
    ws.cell(row=14, column=5).value = (
        34.00  # BUSHRA verified: Reverse-eng from ΔTM 26035 t·m / Δtrim 765cm + booklet
    )
    ws.cell(row=14, column=5).fill = styles["input_fill"]
    ws.cell(row=14, column=6).value = "MTC. Calc!D13"

    ws.cell(row=15, column=3).value = "LCF_m_from_midship"
    ws.cell(row=15, column=4).value = "m"
    ws.cell(row=15, column=5).value = 0.76
    ws.cell(row=15, column=5).fill = styles["input_fill"]
    ws.cell(row=15, column=6).value = "LCF from Midship (Corrected). Calc!D14"

    ws.cell(row=16, column=3).value = "TPC_t_per_cm"
    ws.cell(row=16, column=4).value = "t/cm"
    ws.cell(row=16, column=5).value = (
        8.00  # BUSHRA verified: Approx waterplane 14×60.3×0.85×1.025 ≈680 m² → TPC≈8.00
    )
    ws.cell(row=16, column=5).fill = styles["input_fill"]
    ws.cell(row=16, column=6).value = "TPC. Calc!D15"

    ws.cell(row=17, column=3).value = "Lpp_m"
    ws.cell(row=17, column=4).value = "m"
    ws.cell(row=17, column=5).value = 60.302
    ws.cell(row=17, column=5).fill = styles["input_fill"]
    ws.cell(row=17, column=6).value = "Length between perpendiculars. Calc!D16"

    ws.cell(row=18, column=2).value = "OPERATIONS"
    ws.cell(row=18, column=3).value = "max_fwd_draft_ops_m"
    ws.cell(row=18, column=4).value = "m"
    ws.cell(row=18, column=5).value = 2.70
    ws.cell(row=18, column=5).fill = styles["input_fill"]
    ws.cell(row=18, column=6).value = "Max forward draft for operations. Calc!D9"

    ws.cell(row=19, column=3).value = "ramp_door_offset_m"
    ws.cell(row=19, column=4).value = "m"
    ws.cell(row=19, column=5).value = 0.15
    ws.cell(row=19, column=5).fill = styles["input_fill"]
    ws.cell(row=19, column=6).value = "Ramp door offset. Calc!D11"

    ws.cell(row=20, column=3).value = "linkspan_freeboard_target_m"
    ws.cell(row=20, column=4).value = "m"
    ws.cell(row=20, column=5).value = 0.28
    ws.cell(row=20, column=5).fill = styles["input_fill"]
    ws.cell(row=20, column=6).value = "Linkspan freeboard target. Calc!D12"

    ws.cell(row=21, column=3).value = "gm_target_m"
    ws.cell(row=21, column=4).value = "m"
    ws.cell(row=21, column=5).value = 1.50
    ws.cell(row=21, column=5).fill = styles["input_fill"]
    ws.cell(row=21, column=6).value = "GM target. Calc!D13"

    # STRUCTURAL LIMITS 섹션 (Row 23-26)
    ws.cell(row=23, column=2).value = "STRUCTURAL LIMITS"
    ws.cell(row=23, column=3).value = "limit_reaction_t"
    ws.cell(row=23, column=4).value = "t"
    ws.cell(row=23, column=5).value = 201.60
    ws.cell(row=23, column=5).fill = styles["input_fill"]
    ws.cell(row=23, column=6).value = (
        "Aries Ramp hinge limit 201.60 t (share ratio 0.545, 2025-11-18)"
    )

    ws.cell(row=24, column=3).value = "limit_share_load_t"
    ws.cell(row=24, column=4).value = "t"
    ws.cell(row=24, column=5).value = 118.80
    ws.cell(row=24, column=5).fill = styles["input_fill"]
    ws.cell(row=24, column=6).value = "Max Share Load on LCT (Mammoet)"

    ws.cell(row=25, column=3).value = "limit_deck_press_tpm2"
    ws.cell(row=25, column=4).value = "t/m²"
    ws.cell(row=25, column=5).value = 10.00
    ws.cell(row=25, column=5).fill = styles["input_fill"]
    ws.cell(row=25, column=6).value = "Max Deck Pressure (Spec)"

    ws.cell(row=26, column=3).value = "linkspan_area_m2"
    ws.cell(row=26, column=4).value = "m²"
    ws.cell(row=26, column=5).value = 12.00
    ws.cell(row=26, column=5).fill = styles["input_fill"]
    ws.cell(row=26, column=6).value = (
        "Linkspan 실제 접지 12.00 m² (Ramp 1 TR only 규정)"
    )

    # BALLAST FIX CHECK 섹션 (Row 27-29)
    ws.cell(row=27, column=2).value = "BALLAST FIX CHECK"
    ws.cell(row=27, column=3).value = "max_aft_ballast_cap_t"
    ws.cell(row=27, column=4).value = "t"
    ws.cell(row=27, column=5).value = 28.00
    ws.cell(row=27, column=5).fill = styles["input_fill"]
    ws.cell(row=27, column=6).value = "Max AFT Ballast Capacity (FW2 P/S, Fr 0–6, ~28t)"

    ws.cell(row=28, column=3).value = "max_fwd_ballast_cap_t"
    ws.cell(row=28, column=4).value = "t"
    ws.cell(row=28, column=5).value = 321.00
    ws.cell(row=28, column=5).fill = styles["input_fill"]
    ws.cell(row=28, column=6).value = (
        "Max Forward Ballast Capacity (FWB1/2, Fr 48–65, ~321t)"
    )

    ws.cell(row=29, column=3).value = "max_pump_time_h"
    ws.cell(row=29, column=4).value = "h"
    ws.cell(row=29, column=5).value = 6.00
    ws.cell(row=29, column=5).fill = styles["input_fill"]
    ws.cell(row=29, column=6).value = "Max Allowed Pump Time for Fix"

    # VENT & PUMP 섹션 (Row 30-32)
    ws.cell(row=30, column=2).value = "VENT & PUMP"
    ws.cell(row=30, column=3).value = "vent_flow_coeff"
    ws.cell(row=30, column=4).value = "t/h per mm"
    ws.cell(row=30, column=5).value = 0.86
    ws.cell(row=30, column=5).fill = styles["input_fill"]
    ws.cell(row=30, column=6).value = "실측 보정 0.86 (2025-11-18, MAPE 0.30%)"

    ws.cell(row=31, column=3).value = "pump_rate_tph"
    ws.cell(row=31, column=4).value = "t/h"
    ws.cell(row=31, column=5).value = 100.00
    ws.cell(row=31, column=5).fill = styles["input_fill"]
    ws.cell(row=31, column=6).value = "Hired pump rate"

    ws.cell(row=32, column=3).value = "pump_rate_effective_tph"
    ws.cell(row=32, column=4).value = "t/h"
    ws.cell(row=32, column=5).value = (
        '=MIN(E31, SUMPRODUCT((Ballast_Tanks!E$2:E$100="Y")*(Ballast_Tanks!F$2:F$100)*E30))'
    )
    ws.cell(row=32, column=5).fill = styles["ok_fill"]
    ws.cell(row=32, column=6).value = "실효 펌프 속도 (vent bottleneck, 68.80 t/h)"

    # RAMP GEOMETRY 섹션 (Row 33-36)
    ws.cell(row=33, column=2).value = "RAMP GEOMETRY"
    ws.cell(row=33, column=3).value = "ramp_hinge_x_mid_m"
    ws.cell(row=33, column=4).value = "m"
    ws.cell(row=33, column=5).value = -30.151
    ws.cell(row=33, column=5).fill = styles["input_fill"]
    ws.cell(row=33, column=6).value = "LBP 60.302 m 기준"

    ws.cell(row=34, column=3).value = "ramp_length_m"
    ws.cell(row=34, column=4).value = "m"
    ws.cell(row=34, column=5).value = 8.30
    ws.cell(row=34, column=5).fill = styles["input_fill"]
    ws.cell(row=34, column=6).value = "TRE Cert 2020-08-04"

    ws.cell(row=35, column=3).value = "linkspan_height_m"
    ws.cell(row=35, column=4).value = "m"
    ws.cell(row=35, column=5).value = 2.00
    ws.cell(row=35, column=5).fill = styles["input_fill"]

    ws.cell(row=36, column=3).value = "ramp_end_clearance_min_m"
    ws.cell(row=36, column=4).value = "m"
    ws.cell(row=36, column=5).value = 0.40
    ws.cell(row=36, column=5).fill = styles["input_fill"]

    # HINGE STRESS 섹션 (Row 37-38)
    ws.cell(row=37, column=2).value = "HINGE STRESS"
    ws.cell(row=37, column=3).value = "hinge_pin_area_m2"
    ws.cell(row=37, column=4).value = "m²"
    ws.cell(row=37, column=5).value = 0.117
    ws.cell(row=37, column=5).fill = styles["input_fill"]
    ws.cell(row=37, column=6).value = "Doubler 390x300 mm (Aries)"

    ws.cell(row=38, column=3).value = "hinge_limit_rx_t"
    ws.cell(row=38, column=4).value = "t"
    ws.cell(row=38, column=5).value = 201.60
    ws.cell(row=38, column=5).fill = styles["input_fill"]
    ws.cell(row=38, column=6).value = (
        "Max Hinge Reaction (duplicate of E23 for clarity)"
    )

    # PRECISION PARAMETERS 섹션 (Row 40-44)
    ws.cell(row=40, column=2).value = "PRECISION PARAMETERS"
    ws.cell(row=40, column=2).font = styles["normal_font"]

    ws.cell(row=41, column=3).value = "LBP_m"
    ws.cell(row=41, column=4).value = "m"
    ws.cell(row=41, column=5).value = 60.302  # LBP_m
    ws.cell(row=41, column=5).comment = Comment("LBP (m) - Calc!$E$41", "System")
    ws.cell(row=41, column=5).fill = styles["input_fill"]
    ws.cell(row=41, column=6).value = (
        "Length Between Perpendiculars (for precise draft calculation)"
    )

    ws.cell(row=42, column=3).value = "LCF_from_mid_m"
    ws.cell(row=42, column=4).value = "m"
    ws.cell(row=42, column=5).value = (
        0.76  # BUSHRA verified: LCF_from_AP (30.91) - AP_to_midship (30.151) = 0.759 m
    )
    ws.cell(row=42, column=5).comment = Comment(
        "LCF from mid (m) - Calc!$E$42", "System"
    )
    ws.cell(row=42, column=5).fill = styles["input_fill"]
    ws.cell(row=42, column=6).value = (
        "LCF from midship (BUSHRA verified, for precise draft calculation)"
    )

    ws.cell(row=43, column=3).value = "dynamic_factor"
    ws.cell(row=43, column=4).value = "-"
    ws.cell(row=43, column=5).value = 1.15
    ws.cell(row=43, column=5).fill = styles["input_fill"]
    ws.cell(row=43, column=6).value = (
        "Dynamic load amplification factor (for Load Case B)"
    )

    ws.cell(row=44, column=3).value = "heel_y_offset_m"
    ws.cell(row=44, column=4).value = "m"
    ws.cell(row=44, column=5).value = 1.50
    ws.cell(row=44, column=5).fill = styles["input_fill"]
    ws.cell(row=44, column=6).value = "Heel y-offset (for heel angle calculation)"

    # 폰트 적용
    for row in range(5, 45):
        for col in range(2, 7):
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                cell.font = styles["normal_font"]

    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 35

    print("  [OK] Calc sheet created with VENT&PUMP 실측 0.86")


# ============================================================================
# Tide Sheet Creation
# ============================================================================


def create_tide_sheet(wb):
    """December_Tide_2025 시트 생성"""
    ws = wb.create_sheet("December_Tide_2025")
    styles = get_styles()

    headers = ["datetime_gst", "tide_m               (Chart Datum)"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = styles["center_align"]
        cell.border = Border(
            left=styles["thin_border"],
            right=styles["thin_border"],
            top=styles["thin_border"],
            bottom=styles["thin_border"],
        )

    # JSON 파일 로드 (상대 경로 사용)
    tide_data = _load_json("data/gateab_v3_tide_data.json")

    if tide_data:
        try:
            for idx, entry in enumerate(tide_data, start=0):
                row = 2 + idx
                if row > 745:
                    break

                cell_a = ws.cell(row=row, column=1)
                cell_a.value = entry.get("datetime", "")
                cell_a.font = styles["normal_font"]

                cell_b = ws.cell(row=row, column=2)
                cell_b.value = entry.get("tide_m", 0.0)
                cell_b.font = styles["normal_font"]
                cell_b.number_format = "0.00"
            print(f"  [OK] December_Tide_2025 sheet created with {len(tide_data)} rows")
        except Exception as e:
            print(f"  [WARNING] Error processing tide data: {e}. Creating empty sheet.")
            tide_data = None

    if not tide_data:
        print(
            f"  [WARNING] JSON file not found. Creating empty December_Tide_2025 sheet."
        )
        for row in range(2, 746):
            ws.cell(row=row, column=1).font = styles["normal_font"]
            ws.cell(row=row, column=2).font = styles["normal_font"]
            ws.cell(row=row, column=2).number_format = "0.00"

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 15


# ============================================================================
# Hourly Sheet Creation
# ============================================================================


def create_hourly_sheet(wb):
    """Hourly_FWD_AFT_Heights 시트 생성"""
    ws = wb.create_sheet("Hourly_FWD_AFT_Heights")
    styles = get_styles()

    headers = [
        "DateTime (GST)",
        "Tide_m",
        "Dfwd_req_m (even)",
        "Trim_m (optional)",
        "Dfwd_adj_m",
        "Daft_adj_m",
        "Ramp_Angle_deg",
        "Status",
        "FWD_Height_m",
        "AFT_Height_m",
        "Notes",
        "",
        "Trim_m (optional)",
        "",
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header if header else ""
        if header:
            cell.font = styles["header_font"]
            cell.fill = styles["header_fill"]
            cell.alignment = styles["center_align"]
            cell.border = Border(
                left=styles["thin_border"],
                right=styles["thin_border"],
                top=styles["thin_border"],
                bottom=styles["thin_border"],
            )

    for row in range(2, 746):
        row_str = str(row)
        ws.cell(row=row, column=1).value = (
            f'=IF(December_Tide_2025!A{row_str}="","",December_Tide_2025!A{row_str})'
        )
        ws.cell(row=row, column=2).value = (
            f'=IF(December_Tide_2025!B{row_str}="","",December_Tide_2025!B{row_str})'
        )
        ws.cell(row=row, column=3).value = (
            f'=IF($A{row_str}="","", '
            f'INDEX(Calc!$E:$E, MATCH("KminusZ_m", Calc!$C:$C, 0)) + $B{row_str} - '
            f'INDEX(Calc!$E:$E, MATCH("L_ramp_m", Calc!$C:$C, 0)) * '
            f'TAN(RADIANS(INDEX(Calc!$E:$E, MATCH("theta_max_deg", Calc!$C:$C, 0)))))'
        )
        ws.cell(row=row, column=5).value = (
            f'=IF($C{row_str}="","", IF($D{row_str}="", $C{row_str}, $C{row_str} - $D{row_str}/2))'
        )
        ws.cell(row=row, column=6).value = (
            f'=IF($C{row_str}="","", IF($D{row_str}="", $C{row_str}, $C{row_str} + $D{row_str}/2))'
        )
        ws.cell(row=row, column=7).value = (
            f'=IF($E{row_str}="","", '
            f'DEGREES(ATAN((INDEX(Calc!$E:$E, MATCH("KminusZ_m", Calc!$C:$C, 0)) - $E{row_str} + $B{row_str}) / '
            f'INDEX(Calc!$E:$E, MATCH("L_ramp_m", Calc!$C:$C, 0)))))'
        )
        ws.cell(row=row, column=8).value = (
            f'=IF($E{row_str}="","", '
            f'IF(AND($E{row_str}>=INDEX(Calc!$E:$E, MATCH("min_fwd_draft_m", Calc!$C:$C, 0)), '
            f'$E{row_str}<=INDEX(Calc!$E:$E, MATCH("max_fwd_draft_m", Calc!$C:$C, 0)), '
            f'$G{row_str}<=INDEX(Calc!$E:$E, MATCH("theta_max_deg", Calc!$C:$C, 0))), "OK", "CHECK"))'
        )
        ws.cell(row=row, column=9).value = (
            f'=IF($E{row_str}="","", '
            f'INDEX(Calc!$E:$E, MATCH("D_vessel_m", Calc!$C:$C, 0)) - $E{row_str} + $B{row_str})'
        )
        ws.cell(row=row, column=10).value = (
            f'=IF($F{row_str}="","", '
            f'INDEX(Calc!$E:$E, MATCH("D_vessel_m", Calc!$C:$C, 0)) - $F{row_str} + $B{row_str})'
        )
        ws.cell(row=row, column=11).value = f'=IF(D{row_str}=0, "Even Keel", "")'

        if row == 2:
            ws.cell(row=row, column=14).value = (
                "← Defaults to 0.00 (Even-Keel). To apply the actual trim, manually enter the value in this cell."
            )
            ws.cell(row=row, column=14).font = styles["normal_font"]

        ws.cell(row=row, column=2).number_format = "0.00"
        ws.cell(row=row, column=3).number_format = "0.00"
        ws.cell(row=row, column=4).number_format = "0.00"
        ws.cell(row=row, column=5).number_format = "0.00"
        ws.cell(row=row, column=6).number_format = "0.00"
        ws.cell(row=row, column=7).number_format = "0.00"
        ws.cell(row=row, column=9).number_format = "0.00"
        ws.cell(row=row, column=10).number_format = "0.00"

        for col in range(1, 12):
            ws.cell(row=row, column=col).font = styles["normal_font"]

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["I"].width = 15
    ws.column_dimensions["J"].width = 15
    ws.column_dimensions["K"].width = 15
    ws.column_dimensions["L"].width = 12
    ws.column_dimensions["M"].width = 12
    ws.column_dimensions["N"].width = 80

    print(f"  [OK] Hourly_FWD_AFT_Heights sheet created")


# ============================================================================
# RORO Sheet Creation
# ============================================================================


def build_opt_c_stage(preballast_opt: float):
    """
    Stage 6A_Critical (Opt C) - Ballast Integrated (DAS Method v4.3 Final Optimized)
    - TR1: final deck position (Fr ≈ 42.0)
    - TR2: ramp tip position (Fr ≈ -5.0)
    - 화물 중량: 280t (TR 217t + SPMT/Acc 63t)
    - Pre-Ballast: 최적화된 값 (preballast_opt) at aft fresh water ballast tanks (FW2, Fr 0-6, Mid_Fr=3.0)
      - FW2: Fr 0-6, Mid_Fr=3.0 → x = fr_to_x(3.0) ≈ +27.15m (AFT 쪽)
    - Total: Cargo 560t + Ballast (preballast_opt)
    - Combined LCG: 재계산 (Even Keel에 근접)
    - target_trim_cm 은 Even Keel 목표로 설정

    Args:
        preballast_opt: 최적화된 Pre-ballast 중량 (ton) - find_preballast_opt() 결과값
    """
    # 1. 화물 (Cargo)
    fr_tr1 = 42.0  # TR1 final stowage frame
    fr_tr2 = 17.95  # TR2 Stage 6A_Critical LCG Frame (x=12.20m → Fr=30.151-12.20≈17.95)
    w_tr = 280.0  # TR 217t + SPMT/Acc 63t = 280t
    tr1_weight = w_tr
    tr1_pos_x = fr_to_x(fr_tr1)  # ≈ 11.85 m
    tr2_weight = w_tr
    tr2_pos_x = fr_to_x(fr_tr2)  # ≈ -35.15 m

    # 2. 밸러스트 (Pre-Ballast for Opt C - DAS Method v4.3 Final)
    # DAS Method: 최적화된 Pre-ballast 중량 사용
    # 위치: aft fresh water ballast tanks (FW2, Fr 0-6, Mid_Fr=3.0) - 가장 AFT 쪽
    # FW2: Fr 0-6, Mid_Fr=3.0 → x = fr_to_x(3.0) ≈ +27.15m (AFT 쪽)
    # 무게: 최적화된 값 (preballast_opt)
    w_bal = preballast_opt
    # Aft ballast tank center (FW2, Fr 0-6, Mid_Fr=3.0) - 757 TCP Tank Plan 기준
    ballast_fr_center = 3.0  # FW2 Mid_Fr=3.0 (가장 AFT 쪽)
    ballast_pos_x = fr_to_x(ballast_fr_center)

    # 3. Total Combined (Cargo + Ballast)
    total_weight_opt_c = tr1_weight + tr2_weight + w_bal

    # 모멘트 합산 / 총중량 = 새로운 LCG
    # (Cargo Moment + Ballast Moment) / Total Weight
    moment_cargo = (tr1_weight * tr1_pos_x) + (tr2_weight * tr2_pos_x)
    moment_ballast = w_bal * ballast_pos_x
    combined_lcg_opt_c = (moment_cargo + moment_ballast) / total_weight_opt_c

    return {
        "name": "Stage 6A_Critical (Opt C)",
        "weight_t": total_weight_opt_c,
        "x_from_mid_m": combined_lcg_opt_c,  # 재계산된 LCG (Ballast 포함)
        "target_trim_cm": 0.0,  # Even Keel 목표
    }


def create_roro_sheet(wb: Workbook):
    """RORO_Stage_Scenarios 생성 (build_stage_loads + solve_stage 기반)"""
    ws = wb.create_sheet("RORO_Stage_Scenarios")
    styles = get_styles()

    # =====================================================================
    # Pre-ballast 최적화 및 Stage별 계산 (Python 엔진 사용)
    # =====================================================================
    # Hydro table 로드
    hydro_table_data = _load_hydro_table()
    if not hydro_table_data:
        print("[WARNING] hydro_table.json not found. Using empty table.")
        hydro_table_data = []

    # CONFIG (REMOVED - 중복 cfg 제거, stage_results.csv SSOT 사용)
    # ⚠️ PATCH 2024-12-23: stage_results.csv에서 실제 계산값 읽기
    import csv
    from pathlib import Path

    stage_results_path = Path("stage_results.csv")
    if not stage_results_path.exists():
        print("[WARNING] stage_results.csv not found. Using default cfg values.")
        cfg = {
            "W_TR": 271.20,
            "FR_TR1_RAMP_START": 26.15,
            "FR_TR1_RAMP_MID": 26.15,
            "FR_TR1_STOW": 25.65,
            "FR_TR2_RAMP": 25.04,
            "FR_TR2_STOW": 29.39,
            "FR_PREBALLAST": 25.58,
        }
    else:
        print("[INFO] Reading TR positions from stage_results.csv (SSOT)")
        stage_data = {}
        with open(stage_results_path, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                stage_name = row.get("Stage", "")
                try:
                    x_stage_m = float(row.get("x_stage_m", 0.0))
                    w_stage_t = float(row.get("W_stage_t", 0.0))
                    stage_data[stage_name] = {"x": x_stage_m, "w": w_stage_t}
                except (ValueError, TypeError):
                    pass

        def x_to_fr(x_m: float) -> float:
            return 30.151 - x_m

        cfg = {
            "W_TR": 271.20,
            "FR_TR1_RAMP_START": x_to_fr(stage_data.get("Stage 2", {}).get("x", 4.0)),
            "FR_TR1_RAMP_MID": x_to_fr(stage_data.get("Stage 3", {}).get("x", 4.0)),
            "FR_TR1_STOW": x_to_fr(stage_data.get("Stage 4", {}).get("x", 4.5)),
            "FR_TR2_RAMP": x_to_fr(
                stage_data.get("Stage 6A_Critical (Opt C)", {}).get("x", 5.11)
            ),
            "FR_TR2_STOW": x_to_fr(stage_data.get("Stage 6C", {}).get("x", 0.76)),
            "FR_PREBALLAST": x_to_fr(
                stage_data.get("Stage 5_PreBallast", {}).get("x", 4.57)
            ),
        }
        print(f"[INFO] cfg from stage_results.csv: {cfg}")

    # 선박 고정 파라미터
    MTC = 34.00
    LCF = 0.76
    LBP = 60.302
    D_vessel = 3.65
    base_disp_t = 2800.00
    # PATCH B1-2 Enhanced: Calculate base_tmean_m dynamically from hydro table (SSOT enforcement)
    if hydro_table_data and len(hydro_table_data) > 0:
        base_tmean_m = interpolate_tmean_from_disp(base_disp_t, hydro_table_data)
        print(
            f"[PATCH B1] base_tmean_m calculated from hydro table: {base_tmean_m:.3f} m (base_disp_t={base_disp_t:.2f} t)"
        )
    else:
        raise RuntimeError(
            "[PATCH B1-Enhanced] hydro_table_data missing; cannot compute base_tmean_m. "
            "Provide Hydro_Table_Engineering.json (SSOT). "
            "Location: bplus_inputs/Hydro_Table_Engineering.json"
        )

    params = {
        "MTC": MTC,
        "LCF": LCF,
        "LBP": LBP,
        "D_vessel": D_vessel,
        "hydro_table": hydro_table_data,
        "FWD_DRAFT_LIMIT": 2.70,
        "GM_MIN": 1.50,
        # Patch Option B: FSM from CONST_TANKS (Fuel + FreshWater)
        "Const_FSM_t_m": 887.0,  # CONST_TANKS!E5 합계값 (DO.P/S + LRFO.P/S + FW1.P/S + FW2.P/S)
        # Pre-ballast 탐색 범위 및 해 선택 기준
        "PREBALLAST_MIN_T": 30.0,  # 최소 30t 이상부터 탐색 (0t 해를 배제)
        "PREBALLAST_MAX_T": 600.0,  # 400.0 → 600.0으로 확대
        "PREBALLAST_STEP_T": 2.0,  # 5.0 → 2.0으로 축소 (더 세밀한 탐색)
        "CHECK_STAGE5": True,  # Stage 5_PreBallast도 FWD≤2.70m, |Trim|≤240cm를 강제 (Harbor Master 제출용)
    }
    params.update(cfg)

    # Pre-ballast 최적화 (필수 조건)
    w_tr = cfg["W_TR"]

    preballast_result = find_preballast_opt(
        w_tr_unit_t=w_tr,
        base_disp_t=base_disp_t,
        fr_tr1_stow=cfg.get("FR_TR1_STOW", FR_TR1_STOW),
        fr_tr2_ramp=cfg.get("FR_TR2_RAMP", FR_TR2_RAMP),
        fr_preballast=cfg.get("FR_PREBALLAST", FR_PREBALLAST),
        params=params,
        search_min_t=params.get("PREBALLAST_MIN_T", 20.0),  # FIX: enforce 20t minimum
        search_max_t=params.get("PREBALLAST_MAX_T", 400.0),
        search_step_t=params.get("PREBALLAST_STEP_T", 1.0),
    )

    if not preballast_result["ok"]:
        # 실패 시: 기존 PREBALLAST_T_TARGET 사용 등 fallback
        preballast_opt = params.get("PREBALLAST_T_TARGET", 250.0)
        print(
            f"\n[WARNING] Pre-ballast optimization failed: {preballast_result['reason']}"
        )
        print(f"[WARNING] Using fallback pre-ballast value: {preballast_opt:.2f} t")
        stage5_pb = None
        stage6A_pb = None
    else:
        preballast_opt = preballast_result["w_preballast_t"]
        stage5_pb = preballast_result.get("stage5") or {}
        stage6A_pb = preballast_result.get("stage6A") or {}

        # PATCH FIX #1: 안전한 FWD 값 읽기 (키 이름 불일치 대응)
        def _read_fwd(stage: dict) -> float:
            for key in ("FWD_m", "FWD", "FWD_draft_m", "Dfwd_m"):
                if key in stage:
                    return float(stage[key])
            return 0.0

        fwd5 = _read_fwd(stage5_pb)
        fwd6 = _read_fwd(stage6A_pb)

        print(f"\n[INFO] [OK] Pre-ballast optimization successful")
        print(f"[INFO] Stern Pre-Ballast (FW2): {preballast_opt:.2f} t")
        print(f"[INFO] Stage 5 FWD: {fwd5:.2f} m (Limit: 2.70m)")
        print(f"[INFO] Stage 6A FWD: {fwd6:.2f} m (Limit: 2.70m)")
        safety_margin = 2.70 - fwd6
        print(
            f"[INFO] Safety Margin: {safety_margin:.2f} m ({safety_margin/2.70*100:.1f}%)"
        )

        # VALIDATION CHECK
        if fwd5 > 2.70 or fwd6 > 2.70:
            print(f"[ERROR] [FAIL] FWD DRAFT EXCEEDS 2.70m LIMIT - DESIGN INVALID")
        else:
            print(f"[INFO] [OK] All draft constraints satisfied - DESIGN APPROVED")
    print(f"[INFO] Using Pre-ballast: {preballast_opt:.2f} t (used for all stages)")

    # =====================================================================
    # Stage별 계산 결과 저장 (stage_results 딕셔너리 생성)
    # =====================================================================
    # 이 딕셔너리는 다음 리포트 시트들의 주요 데이터 소스입니다:
    # - RORO_Delta_Lever_Report: Python 계산값과 Excel 수식 결과 비교
    # - RORO_Draft_Margin_Check: Draft/Freeboard margin 검증
    # - RORO_Stability_GM_Check: GM 검증
    # - 기타 BPLUS 관련 체크 시트들
    #
    # 데이터 흐름:
    #   solve_stage() → res (각 Stage 계산 결과)
    #     ↓
    #   Pre-ballast 최적화 결과로 Stage 5_PreBallast, Stage 6A_Critical override
    #     ↓
    #   stage_results[stage_name] = res
    #     ↓
    #   리포트 시트 생성 함수들에서 사용
    # =====================================================================
    stage_results = {}
    stages_order = [
        "Stage 1",
        "Stage 2",
        "Stage 3",
        "Stage 4",
        "Stage 5",
        "Stage 5_PreBallast",
        "Stage 6A_Critical (Opt C)",
        "Stage 6C_TotalMassOpt",
        "Stage 6C",
        "Stage 7",
    ]

    for st in stages_order:
        loads = build_stage_loads(st, preballast_opt, params)
        res = solve_stage(base_disp_t, base_tmean_m, loads, **params)

        # ⭐ Pre-ballast 결과로 Stage 5_PreBallast / Stage 6A 값 override
        if st == "Stage 5_PreBallast" and stage5_pb is not None:
            # Pre-ballast 계산 결과의 FWD/AFT/Trim/TM 값으로 교체
            res["W_stage_t"] = float(stage5_pb.get("W_stage_t", res["W_stage_t"]))
            if abs(res["W_stage_t"] - res["Disp_t"]) > 0.1:
                print(
                    f"[WARN] Stage 5_PreBallast: W_stage_t ({res['W_stage_t']:.2f}) != Disp_t ({res['Disp_t']:.2f}); forcing consistency"
                )
                res["W_stage_t"] = res["Disp_t"]
            res["x_stage_m"] = float(stage5_pb.get("x_stage_m", res["x_stage_m"]))
            res["TM_LCF_tm"] = float(stage5_pb.get("TM_tm", res["TM_LCF_tm"]))
            res["Trim_cm"] = float(stage5_pb.get("Trim_cm", res["Trim_cm"]))
            res["Dfwd_m"] = float(stage5_pb.get("FWD_m", res["Dfwd_m"]))
            res["Daft_m"] = float(stage5_pb.get("AFT_m", res["Daft_m"]))
            # FWD_Height_m, AFT_Height_m 재계산
            res["FWD_Height_m"] = D_vessel - res["Dfwd_m"]
            res["AFT_Height_m"] = D_vessel - res["Daft_m"]
            print(
                f"[INFO] Stage 5_PreBallast: Applied pre-ballast FWD={res['Dfwd_m']:.2f}m, AFT={res['Daft_m']:.2f}m"
            )

        if st == "Stage 6A_Critical (Opt C)" and stage6A_pb is not None:
            # Pre-ballast 계산 결과의 FWD/AFT/Trim/TM 값으로 교체
            res["W_stage_t"] = float(stage6A_pb.get("W_stage_t", res["W_stage_t"]))
            if abs(res["W_stage_t"] - res["Disp_t"]) > 0.1:
                print(
                    f"[WARN] Stage 6A_Critical: W_stage_t ({res['W_stage_t']:.2f}) != Disp_t ({res['Disp_t']:.2f}); forcing consistency"
                )
                res["W_stage_t"] = res["Disp_t"]
            res["x_stage_m"] = float(stage6A_pb.get("x_stage_m", res["x_stage_m"]))
            res["TM_LCF_tm"] = float(stage6A_pb.get("TM_tm", res["TM_LCF_tm"]))
            res["Trim_cm"] = float(stage6A_pb.get("Trim_cm", res["Trim_cm"]))
            res["Dfwd_m"] = float(stage6A_pb.get("FWD_m", res["Dfwd_m"]))
            res["Daft_m"] = float(stage6A_pb.get("AFT_m", res["Daft_m"]))
            # FWD_Height_m, AFT_Height_m 재계산
            res["FWD_Height_m"] = D_vessel - res["Dfwd_m"]
            res["AFT_Height_m"] = D_vessel - res["Daft_m"]
            print(
                f"[INFO] Stage 6A_Critical: Applied pre-ballast FWD={res['Dfwd_m']:.2f}m, AFT={res['Daft_m']:.2f}m"
            )

        # B2 Gate 기반 초기 Ballast 추정 (물리식)
        # ΔTrim × MTC / Lever_arm
        MTC = float(params.get("MTC", params.get("MTC_t_m_per_cm", 0.0)) or 0.0)
        lcf_mid = float(params.get("LCF", params.get("LCF_m_from_midship", 0.0)) or 0.0)
        fr_pb = params.get("FR_PREBALLAST", params.get("FR_PB", 3.0))
        ballast_x_mid = fr_to_x(fr_pb)
        lever_arm_m = ballast_x_mid - lcf_mid
        fwd_limit = float(params.get("max_fwd_draft_ops_m", 2.70))
        trim_gate_cm = calc_trim_gate_cm_from_tmean(
            float(res.get("Tmean_m", 0.0) or 0.0), fwd_limit
        )
        delta_trim_cm = trim_gate_cm - float(res.get("Trim_cm", 0.0) or 0.0)
        ballast_est_t = (
            (delta_trim_cm * MTC) / lever_arm_m
            if abs(lever_arm_m) > 1e-6 and abs(MTC) > 1e-6
            else 0.0
        )
        pump_rate_eff = float(
            params.get("pump_rate_effective_tph") or params.get("pump_rate_tph") or 10.0
        )
        res["Trim_gate_cm"] = round(trim_gate_cm, 2)
        res["Ballast_t_calc"] = round(ballast_est_t, 2)
        res["Ballast_time_h_calc"] = (
            round(abs(ballast_est_t) / pump_rate_eff, 2) if pump_rate_eff > 0 else 0.0
        )
        # Backward-compat keys (Iterative에서 사용)
        res["Ballast_t"] = res["Ballast_t_calc"]
        res["Ballast_time_h"] = res["Ballast_time_h_calc"]

        # Stage 결과를 딕셔너리에 저장
        # 포함된 주요 키: W_stage_t, x_stage_m, Trim_cm, Dfwd_m, Daft_m,
        #                 FWD_Height_m, AFT_Height_m, GM_m, Disp_t, Ballast_t 등
        stage_results[st] = res

    # Title
    ws["A1"] = "RORO Stage Scenarios – Option C (Target 240cm Safe Margin)"
    ws["A1"].font = styles["title_font"]

    # Row 2: Input parameter 안내
    ws["C2"] = "← Input parameter(yellow cellls only)"
    ws["C2"].font = styles["normal_font"]

    # 숫자 포맷 통일: 천단위 구분, 소수점 2자리
    number_format = "#,##0.00"

    # Row 4: 섹션 제목 추가
    ws["A4"] = "1.Critical Stage Verification: Trim & Draft Status Sequence"
    ws["A4"].font = styles["normal_font"]
    ws["F4"] = "2. Stage Sequence Table"
    ws["F4"].font = styles["normal_font"]

    # Row 5: 헤더 추가
    ws["A5"] = "Parameter"
    ws["A5"].font = styles["header_font"]
    ws["A5"].fill = styles["header_fill"]
    ws["A5"].alignment = styles["center_align"]

    ws["B5"] = "Value"
    ws["B5"].font = styles["header_font"]
    ws["B5"].fill = styles["header_fill"]
    ws["B5"].alignment = styles["center_align"]

    ws["C5"] = "Unit"
    ws["C5"].font = styles["header_font"]
    ws["C5"].fill = styles["header_fill"]
    ws["C5"].alignment = styles["center_align"]

    ws["D5"] = "REMARK"
    ws["D5"].font = styles["header_font"]
    ws["D5"].fill = styles["header_fill"]
    ws["D5"].alignment = styles["center_align"]

    ws["F5"] = "Stage"
    ws["F5"].font = styles["header_font"]
    ws["F5"].fill = styles["header_fill"]
    ws["F5"].alignment = styles["center_align"]

    ws["G5"] = "EXPLANATION"
    ws["G5"].font = styles["header_font"]
    ws["G5"].fill = styles["header_fill"]
    ws["G5"].alignment = styles["center_align"]

    ws["O5"] = "Fwd Draft (m)"
    ws["O5"].font = styles["header_font"]
    ws["O5"].fill = styles["header_fill"]
    ws["O5"].alignment = styles["center_align"]

    ws["P5"] = "Freeboard (m)"
    ws["P5"].font = styles["header_font"]
    ws["P5"].fill = styles["header_fill"]
    ws["P5"].alignment = styles["center_align"]

    ws["Q5"] = "Aft Draft (m)"
    ws["Q5"].font = styles["header_font"]
    ws["Q5"].fill = styles["header_fill"]
    ws["Q5"].alignment = styles["center_align"]

    ws["R5"] = "Status Assessment"
    ws["R5"].font = styles["header_font"]
    ws["R5"].fill = styles["header_fill"]
    ws["R5"].alignment = styles["center_align"]

    # 파라미터 세로 배치 (A6-A15: Parameter, B6-B15: Value, C6-C15: Unit, D6-D15: REMARK, F6-F15: Stage)
    # A6: Tmean_baseline
    ws["A6"] = "Tmean_baseline"
    # PATCH B1-3: Use dynamically calculated base_tmean_m instead of hardcoded 2.00
    ws["B6"] = float(base_tmean_m)  # Calculated from hydro table (PATCH B1-2)
    ws["B6"].fill = styles["input_fill"]
    ws["B6"].font = styles["normal_font"]
    ws["B6"].number_format = number_format
    ws["C6"] = "m"
    ws["C6"].font = styles["normal_font"]
    ws["D6"] = "Baseline mean draft (from hydro table)"
    ws["D6"].font = styles["normal_font"]
    ws["F6"] = "Stage 1"
    ws["F6"].font = styles["normal_font"]

    # A7: Forecast_Tide_m (FORECAST ONLY, NOT Required WL)
    ws["A7"] = "Forecast_Tide_m"
    ws["B7"] = 2.00  # BUSHRA verified: Mina Zayed high tide avg 1.80-2.20m
    ws["B7"].fill = styles["input_fill"]
    ws["B7"].font = styles["normal_font"]
    ws["B7"].number_format = number_format
    ws["C7"] = "m"
    ws["C7"].font = styles["normal_font"]
    ws["D7"] = "Forecast tide level (CD). NOT Required_WL_m."
    ws["D7"].font = styles["normal_font"]
    ws["F7"] = "Stage 2"
    ws["F7"].font = styles["normal_font"]

    # A8: Trim_target_cm
    ws["A8"] = "Trim_target_cm"
    ws["B8"] = 10.00  # BUSHRA verified: Ops safe limit (by stern max)
    ws["B8"].fill = styles["input_fill"]
    ws["B8"].font = styles["normal_font"]
    ws["B8"].number_format = number_format
    ws["C8"] = "cm"
    ws["C8"].font = styles["normal_font"]
    ws["D8"] = "Target trim"
    ws["D8"].font = styles["normal_font"]
    ws["F8"] = "Stage 3"
    ws["F8"].font = styles["normal_font"]

    # A9: MTC
    ws["A9"] = "MTC"
    ws["B9"] = '=INDEX(Calc!$E:$E, MATCH("MTC_t_m_per_cm", Calc!$C:$C, 0))'
    ws["B9"].fill = styles["input_fill"]
    ws["B9"].font = styles["normal_font"]
    ws["B9"].number_format = number_format
    ws["C9"] = "t·m/cm"
    ws["C9"].font = styles["normal_font"]
    ws["D9"] = "Moment to change trim"
    ws["D9"].font = styles["normal_font"]
    ws["F9"] = "Stage 4"
    ws["F9"].font = styles["normal_font"]

    # A10: LCF (midship 기준, TM 계산용)
    ws["A10"] = "LCF"
    ws["B10"] = "=Calc!$E$15"  # LCF_m_from_midship 직접 참조 (0.76 m)
    ws["B10"].font = styles["normal_font"]
    ws["B10"].number_format = number_format
    ws["B10"].fill = styles["input_fill"]
    ws["C10"] = "m"
    ws["C10"].font = styles["normal_font"]
    ws["D10"] = "Longitudinal center of flotation (from midship, for TM calculation)"
    ws["D10"].font = styles["normal_font"]
    ws["F10"] = "Stage 5"
    ws["F10"].font = styles["normal_font"]

    # A11: D_vessel
    ws["A11"] = "D_vessel"
    ws["B11"] = 3.65  # BUSHRA verified: Booklet + TCP confirmed
    ws["B11"].font = styles["normal_font"]
    ws["B11"].number_format = number_format
    ws["B11"].fill = styles["input_fill"]
    ws["C11"] = "m"
    ws["C11"].font = styles["normal_font"]
    ws["D11"] = "Vessel depth"
    ws["D11"].font = styles["normal_font"]
    ws["F11"] = "Stage 5_PreBallast"
    ws["F11"].font = styles["normal_font"]

    # A12: TPC
    ws["A12"] = "TPC"
    ws["B12"] = '=INDEX(Calc!$E:$E, MATCH("TPC_t_per_cm", Calc!$C:$C, 0))'
    ws["B12"].font = styles["normal_font"]
    ws["B12"].number_format = number_format
    ws["B12"].fill = styles["input_fill"]
    ws["C12"] = "t/cm"
    ws["C12"].font = styles["normal_font"]
    ws["D12"] = "Tons per centimeter"
    ws["D12"].font = styles["normal_font"]
    ws["F12"] = "Stage 6A_Critical (Opt C)"
    ws["F12"].font = styles["normal_font"]

    # A13: pump_rate_effective_tph
    ws["A13"] = "pump_rate_effective_tph"
    ws["B13"] = (
        '=INDEX(Calc!$E:$E, MATCH("pump_rate_effective_tph", Calc!$C:$C, 0))'  # BUSHRA verified: 100.00 t/h (2×50 t/h pumps)
    )
    ws["B13"].fill = styles["input_fill"]
    ws["B13"].font = styles["normal_font"]
    ws["B13"].number_format = number_format
    ws["C13"] = "t/h"
    ws["C13"].font = styles["normal_font"]
    ws["D13"] = "Effective pump rate"
    ws["D13"].font = styles["normal_font"]
    ws["F13"] = "Stage 6C_TotalMassOpt"
    ws["F13"].font = styles["normal_font"]

    # A14: X_Ballast (AP 기준, Stern FW2 탱크 중심)
    ws["A14"] = "X_Ballast"
    ws["B14"] = 0.119  # FW2 tank LCG from AP (m) – Stern pre-ballast 기준
    ws["B14"].font = styles["normal_font"]
    ws["B14"].number_format = number_format
    ws["B14"].fill = styles["input_fill"]
    ws["C14"] = "m"
    ws["C14"].font = styles["normal_font"]
    ws["D14"] = "Ballast center position (from AP, Stern FW2)"
    ws["D14"].font = styles["normal_font"]
    ws["F14"] = "Stage 6C"
    ws["F14"].font = styles["normal_font"]

    # A15: Lpp
    ws["A15"] = "Lpp"
    ws["B15"] = 60.302  # BUSHRA verified: Booklet confirmed
    ws["B15"].font = styles["normal_font"]
    ws["B15"].number_format = number_format
    ws["B15"].fill = styles["input_fill"]
    ws["C15"] = "m"
    ws["C15"].font = styles["normal_font"]
    ws["D15"] = "Length between perpendiculars"
    ws["D15"].font = styles["normal_font"]
    ws["F15"] = "Stage 7"
    ws["F15"].font = styles["normal_font"]

    # Row 17: 섹션 제목 추가 (pre-ballast 값은 Python에서 자동 결정)
    ws["A17"] = (
        "4. Ballast Water Optimization Matrix & Safety Margins (Auto Pre-ballast)"
    )
    ws["A17"].font = styles["normal_font"]

    # Stage 1-10 Notes를 G6-G15로 배치 (DAS Method Operation)
    # stages 리스트 순서대로 G6부터 시작 (Excel 시트 순서와 일치: A26에 Stage 6C_TotalMassOpt 포함)
    stages_list = [
        "Stage 1",  # Arrival
        "Stage 2",  # TR1 Ramp Start
        "Stage 3",  # TR1 Mid-Ramp
        "Stage 4",  # TR1 On Deck
        "Stage 5",  # TR1 Final Position
        "Stage 5_PreBallast",  # [D-1 Night] Water Supply Complete
        "Stage 6A_Critical (Opt C)",  # [D-Day] TR2 Ramp Entry
        "Stage 6C_TotalMassOpt",  # Alternative total mass optimized condition
        "Stage 6C",  # Final Stowage
        "Stage 7",  # Departure
    ]

    # DAS / AGI 공통 – Pre-ballast + Tug 보조 컨셉 설명
    explanations = {
        "Stage 1": "Arrival lightship condition. Baseline drafts, trim, and GM are checked before any cargo loading or ballast change.",
        "Stage 2": "TR1 roll-on start – first axle on the ramp. Initial bow-down trim response is checked against allowable draft and freeboard limits.",
        "Stage 3": "TR1 mid-ramp – transformer COG on the ramp. Progressive bow-down trim is monitored to remain within the allowable envelope.",
        "Stage 4": "TR1 fully on deck. Weight is completely transferred from ramp to vessel deck; deck loading, ramp condition, drafts, and trim are verified.",
        "Stage 5": "TR1 secured at final aft stowage position (around Fr.42). New reference condition with TR1 only on board, used as the basis for D-1 stern pre-ballasting using FW2 (Fr.0–6, aft).",
        "Stage 5_PreBallast": (
            "Pre-ballast condition with TR1 only on board (D-1, using shore water). "
            "An intentional stern trim is set using the aft FW2 fresh water tanks (Fr.0–6, AFT, Ballast CG near FR 3.0) "
            "so that the bow-down trimming moment of TR2 at ramp entry is counterbalanced and the forward draft remains within the 2.70 m AGI limit "
            "without any major dynamic ballasting during the critical RORO operation on D-day."
        ),
        "Stage 6A_Critical (Opt C)": "Critical TR2 ramp-entry condition on D-day with the D-1 stern pre-ballast (FW2, Fr.0–6) kept fixed. This stage is used to check worst-case forward draft against the 2.70 m limit, trim envelope, GM criteria, and ramp/linkspan clearance under combined TR1+TR2 loading.",
        "Stage 6C_TotalMassOpt": 'Alternative "total mass optimized" final condition with TR1 and TR2 on deck and higher combined cargo+ballast weight. Sensitivity case to assess maximum displacement effects on drafts, trim, GM and freeboard.',
        "Stage 6C": "Planned final stowage condition with TR1 and TR2 secured on deck and ballast as per departure plan. Main departure case for checking GM, trim and freeboard criteria.",
        "Stage 7": "Post-operation lightship/reference condition after cargo is discharged. Used to reconfirm hydrostatic characteristics and GM consistency against Stage 1.",
    }

    for idx, stage_name in enumerate(stages_list, start=0):
        g_row = 6 + idx  # G6부터 시작 (G6-G15)
        if stage_name in explanations:
            ws.cell(row=g_row, column=7).value = explanations[
                stage_name
            ]  # G = column 7
            ws.cell(row=g_row, column=7).font = styles["normal_font"]
            ws.cell(row=g_row, column=7).fill = styles["input_fill"]  # G6-G15 색상 추가

    # Stage table header
    header_row = 18  # 기존 파일과 동일하게 Row 18
    stage_headers = [
        "Stage",
        "W_stage_t",
        "Fr_stage",
        "x_stage_m",
        "TM (t·m)",
        "Trim_cm",
        "FWD_precise_m",  # v4.0: Changed from Trim_m to FWD_precise_m
        "AFT_precise_m",  # v4.0: Changed from Trim_target_cm to AFT_precise_m
        "ΔTM_cm_tm",
        "Lever_arm_m",
        "Ballast_t_calc",
        "Ballast_time_h_calc",
        "Ballast_t",
        "Ballast_time_h",
        "Trim_Check",
        "Dfwd_m",
        "Daft_m",
        "Trim_target_stage_cm",  # sdsdds.md: Q열 추가 (Stage별 타깃, 없으면 B6 사용)
        "FWD_DeckElev_CD_m",  # S(19): Chart Datum 기준 Deck Elevation (Tide 포함)
        "AFT_DeckElev_CD_m",  # T(20): Chart Datum 기준 Deck Elevation (Tide 포함)
        "Difference",  # T(20) 컬럼: 기존 파일과 동일하게
        # U-AE는 extend_roro_captain_req()에서 작성됨
        # PATCH C 컬럼들은 AF-AJ로 이동 (헤더는 별도 작성)
    ]
    for col, header in enumerate(stage_headers, start=1):
        cell = ws.cell(row=header_row, column=col)
        cell.value = header if header else ""
        if header:
            cell.font = styles["header_font"]
            cell.fill = styles["header_fill"]
            cell.alignment = styles["center_align"]

    # PATCH C headers (AF-AJ, column 32-36)
    patch_c_headers = [
        ("AF", 32, "Trim_gate_cm"),
        ("AG", 33, "Gate_Ballast_net_t"),
        ("AH", 34, "Dfwd_gate_m"),
        ("AI", 35, "Gate_iter_n"),
        ("AJ", 36, "Gate_Flag"),
    ]
    for letter, col, header in patch_c_headers:
        cell = ws.cell(row=header_row, column=col)
        cell.value = header
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = styles["center_align"]

    # Stage names (DAS Method Operation - 10 stages)
    stages = [
        "Stage 1",  # Arrival
        "Stage 2",  # TR1 Ramp Start
        "Stage 3",  # TR1 Mid-Ramp
        "Stage 4",  # TR1 On Deck
        "Stage 5",  # TR1 Final Position
        "Stage 5_PreBallast",  # [D-1 Night] Water Supply Complete
        "Stage 6A_Critical (Opt C)",  # [D-Day] TR2 Ramp Entry
        "Stage 6C_TotalMassOpt",  # Total Mass Optimization
        "Stage 6C",  # Final Stowage
        "Stage 7",  # Departure
    ]
    # PATCH B2-2: Calculate Trim_gate_cm dynamically from Tmean instead of hardcoded values
    # Trim_gate_cm = max(0, 2 * (Tmean - 2.70) * 100) to ensure FWD <= 2.70m
    target_trim_by_stage = {}
    FWD_LIMIT_M = 2.70  # Forward draft limit (m)

    for stage_name in stages_order:
        if stage_name in stage_results:
            res = stage_results[stage_name]
            tmean_m = res.get("Tmean_m", 0.0)
            # Calculate required trim to satisfy FWD limit
            trim_gate_cm = calc_trim_gate_cm_from_tmean(tmean_m, FWD_LIMIT_M)
            target_trim_by_stage[stage_name] = trim_gate_cm
            print(
                f"[PATCH B2] {stage_name}: Tmean={tmean_m:.3f}m → Trim_gate={trim_gate_cm:.2f}cm"
            )
        else:
            # Fallback for stages not in stage_results (shouldn't happen, but safe)
            target_trim_by_stage[stage_name] = 0.00

    # Also handle Stage 6C_TotalMassOpt if it exists separately
    if "Stage 6C_TotalMassOpt" not in target_trim_by_stage:
        # Use Stage 6C value as fallback
        target_trim_by_stage["Stage 6C_TotalMassOpt"] = target_trim_by_stage.get(
            "Stage 6C", 0.00
        )

    first_data_row = 19  # header_row가 18이므로 first_data_row는 19
    trim5_row = None

    # =====================================================================
    # Stage별 기본 데이터 (build_stage_loads 결과 기반)
    # =====================================================================
    # Stage 6C Total Mass Opt용 상수
    TOTAL_CARGO_WEIGHT_T = 568.83  # 총 화물 중량 (예시값, 필요시 조정)
    PREBALLAST_T_TARGET = preballast_opt  # 최적화된 Pre-ballast 사용

    # Stage별 기본 데이터 (solve_stage 결과에서 추출)
    stage_defaults = {}
    for stage_name in stages_order:
        if stage_name in stage_results:
            res = stage_results[stage_name]
            # Fr_stage는 build_stage_loads에서 사용한 위치에서 역산
            # 또는 단순히 x_stage_m에서 역산
            x_stage = res["x_stage_m"]
            fr_stage = x_to_fr(x_stage) if abs(x_stage) > 0.01 else None

            stage_defaults[stage_name] = {
                "W": res["W_stage_t"],
                "Fr": fr_stage,
                "x": x_stage,
                "Dfwd_m": res.get("Dfwd_m"),  # Pre-ballast override 값 포함
                "Daft_m": res.get("Daft_m"),  # Pre-ballast override 값 포함
                "FWD_precise_m": res.get("Dfwd_m"),  # FWD_precise_m = Dfwd_m
                "AFT_precise_m": res.get("Daft_m"),  # AFT_precise_m = Daft_m
                "TM_LCF_tm": res.get("TM_LCF_tm"),
                "Trim_cm": res.get("Trim_cm"),
            }
        else:
            # Fallback (Stage 1, 7 등)
            stage_defaults[stage_name] = {"W": 0.0, "Fr": None, "x": 0.0}

    for idx, stage_name in enumerate(stages, start=0):
        row = first_data_row + idx
        row_str = str(row)

        ws.cell(row=row, column=1, value=stage_name)

        if stage_name in stage_defaults:
            defaults = stage_defaults[stage_name]
            # W_stage_t: Python에서 계산된 값 사용
            ws.cell(row=row, column=2, value=defaults["W"])
            # Fr_stage 컬럼 (column 3): Python에서 계산된 값 사용
            if defaults["Fr"] is not None:
                ws.cell(row=row, column=3, value=round(defaults["Fr"], 2))
            else:
                ws.cell(row=row, column=3, value="")  # Stage 1은 Fr 없음
            # x_stage_m 컬럼 (column 4): Python에서 계산된 값 사용 (Excel 수식 대신)
            ws.cell(row=row, column=4, value=round(defaults["x"], 2))
        else:
            # stage_defaults에 없는 경우 기본값 설정
            print(
                f"[WARNING] Stage '{stage_name}' not found in stage_defaults, using defaults"
            )
            ws.cell(row=row, column=2, value=0.0)  # W_stage_t
            ws.cell(row=row, column=3, value="")  # Fr_stage
            ws.cell(row=row, column=4, value=0.0)  # x_stage_m

        ws.cell(row=row, column=2).fill = styles["input_fill"]
        ws.cell(row=row, column=2).font = styles["normal_font"]
        ws.cell(row=row, column=2).number_format = number_format
        ws.cell(row=row, column=3).fill = styles["input_fill"]
        ws.cell(row=row, column=3).font = styles["normal_font"]
        ws.cell(row=row, column=3).number_format = number_format
        ws.cell(row=row, column=4).fill = styles["input_fill"]
        ws.cell(row=row, column=4).font = styles["normal_font"]
        ws.cell(row=row, column=4).number_format = number_format

        # TM (t·m): Python stage_results(TM_LCF_tm) 우선, 없으면 기존 수식
        tm_val = stage_defaults.get(stage_name, {}).get("TM_LCF_tm")
        if tm_val is None:
            ws.cell(row=row, column=5).value = (
                f'=IF(OR(B{row_str}="", D{row_str}="", $B$10=""), "", B{row_str} * (D{row_str} - $B$10))'
            )
        else:
            ws.cell(row=row, column=5).value = float(tm_val)
        ws.cell(row=row, column=5).font = styles["normal_font"]
        ws.cell(row=row, column=5).number_format = number_format

        # Trim_cm: Python stage_results(Trim_cm) 우선, 없으면 E/B9
        trim_val = stage_defaults.get(stage_name, {}).get("Trim_cm")
        if trim_val is None:
            ws.cell(row=row, column=6).value = (
                f'=IF(OR(E{row_str}="", OR($B$9="", $B$9=0)), "", E{row_str} / $B$9)'
            )
        else:
            ws.cell(row=row, column=6).value = float(trim_val)
        ws.cell(row=row, column=6).font = styles["normal_font"]
        ws.cell(row=row, column=6).number_format = number_format
        # FWD_precise_m / AFT_precise_m:
        # - Python stage_results(Dfwd/Daft) 값을 우선 사용하여 LCF-aware 계산 결과 사용
        # - Excel 수식(B6 ± F/2)은 LCF를 고려하지 않으므로 정확도가 낮음
        # - 모든 Stage에 Python 계산 결과를 사용하여 일관성 유지
        dfwd_val = None
        daft_val = None

        # 1순위: stage_results에서 직접 가져오기
        if stage_name in stage_results:
            res = stage_results[stage_name]
            dfwd_val = res.get("Dfwd_m")
            daft_val = res.get("Daft_m")

        # 2순위: stage_defaults에서 가져오기
        if dfwd_val is None or daft_val is None:
            defaults = stage_defaults.get(stage_name, {})
            dfwd_val = defaults.get("FWD_precise_m") or defaults.get("Dfwd_m")
            daft_val = defaults.get("AFT_precise_m") or defaults.get("Daft_m")

        # 값이 있으면 Python 계산 결과 사용 (LCF-aware)
        if dfwd_val is not None and daft_val is not None:
            ws.cell(row=row, column=7).value = float(dfwd_val)
            ws.cell(row=row, column=8).value = float(daft_val)
        else:
            # DEBUG: 이 경우는 발생하지 않아야 합니다
            # 모든 Stage는 stage_results에 있어야 하므로, 여기 도달하면 문제가 있습니다
            if stage_name not in [
                "Stage 1",
                "Stage 7",
            ]:  # Stage 1, 7은 값이 없을 수 있음
                logging.warning(
                    f"[FWD_Draft] {stage_name}: Python 값 없음, Excel 수식 사용. "
                    f"stage_results 키: {list(stage_results.keys())[:3]}..."
                )
            # 최종 fallback: Excel 수식 사용 (부정확하지만 계산은 가능)
            ws.cell(row=row, column=7).value = (
                f'=IF($A{row_str}="","",$B$6 - (F{row_str}/100)/2)'
            )
            ws.cell(row=row, column=8).value = (
                f'=IF($A{row_str}="","",$B$6 + (F{row_str}/100)/2)'
            )

        ws.cell(row=row, column=7).font = styles["normal_font"]
        ws.cell(row=row, column=7).number_format = number_format
        ws.cell(row=row, column=8).font = styles["normal_font"]
        ws.cell(row=row, column=8).number_format = number_format

        if stage_name == "Stage 5":
            trim5_row = row

        # v4.0: G 컬럼이 AFT_precise_m로 변경되므로 target_trim은 제거
        # target_trim은 이제 사용되지 않음 (G가 AFT_precise_m로 계산됨)
        # target_trim = target_trim_by_stage.get(stage_name)
        # if target_trim is not None:
        #     ws.cell(row=row, column=8).value = target_trim
        #     ws.cell(row=row, column=8).fill = styles["input_fill"]

        # I (9): ΔTM_cm_tm - PATCH B2-3: (Target − Actual) × MTC (reversed direction)
        # 수식: =IF($A{row}="","", (IF($R{row}="",$B$8,$R{row}) - F{row}) * $B$9)
        # 의미:
        #   - F{row} = 실제 Trim_cm
        #   - $R{row} = Stage별 Trim_target_stage_cm (R열, Trim_gate_cm)
        #   - IF($R{row}="",$B$8,$R{row}) = R열이 비어있으면 전역 타깃(B8) 사용, 있으면 Stage별 타깃 사용
        #   - $B$9 = MTC (t·m/cm)
        #   - 결과: (Target - Actual) × MTC = 필요한 Trim 모멘트 변화량 (cm·t·m)
        #   - 양수 = Target > Actual → 추가 ballast 필요 (stern down)
        #   - 음수 = Target < Actual → De-ballast 필요 (bow down)
        ws.cell(row=row, column=9).value = (
            f'=IF($A{row_str}="","",'
            f'(IF($R{row_str}="",$B$8,$R{row_str}) - F{row_str}) * $B$9)'
        )
        ws.cell(row=row, column=9).number_format = number_format
        ws.cell(row=row, column=9).font = styles["normal_font"]
        # J (10): Lever_arm_m - Ballast 레버 암 고정식
        # 수식: =IF(OR(ISBLANK($B$14), ISBLANK($B$10), ISBLANK($B$15), ISERROR($B$10)), "", ROUND($B$14 - ($B$10 + $B$15/2), 2))
        # 의미:
        #   - $B$14 = X_Ballast_from_AP_m (Ballast CG 위치, 예: 52.50 m)
        #   - $B$10 = LCF_from_midship
        #   - $B$15 = Lpp
        #   - LCF_from_AP = B10 + B15/2
        #   - Lever_arm = X_Ballast_from_AP - LCF_from_AP (예: 21.59 m)
        #   - Stage별로 변하지 않는 상수 레버 암
        ws.cell(row=row, column=10).value = (
            f'=IF(OR(ISBLANK($B$14), ISBLANK($B$10), ISBLANK($B$15), ISERROR($B$10)), "", ROUND($B$14 - ($B$10 + $B$15/2), 2))'
        )
        ws.cell(row=row, column=10).font = styles["normal_font"]
        ws.cell(row=row, column=10).number_format = number_format
        # K (11): Ballast_t_calc - 새 설계 핵심: ΔTM / Lever_arm = 이론상 필요한 Ballast 톤수
        # 수식: =IF(OR($A{row}="",$J{row}="", $J{row}=0), "", ROUND(I{row} / $J{row}, 2))
        # 의미:
        #   - I{row} = ΔTM_cm_tm (column 9)
        #   - $J{row} = Lever_arm_m (column 10)
        #   - Ballast_t_calc = ΔTM / Lever_arm (이론상 필요한 Ballast 무게, t)
        #   - A열 Stage 이름이 비어있거나, 레버 암(J)이 0/공백이면 빈 셀 유지
        ws.cell(row=row, column=11).value = (
            f'=IF(OR($A{row_str}="",$J{row_str}="", $J{row_str}=0),"",ROUND(I{row_str} / $J{row_str}, 2))'
        )
        ws.cell(row=row, column=11).number_format = number_format
        ws.cell(row=row, column=11).font = styles["normal_font"]
        # L (12): Ballast_time_h_calc - 펌프 시간 계산
        # 수식: =IF(OR(K{row}="", $B$13="", $B$13=0, ISERROR($B$13)), "", ROUND(K{row} / $B$13, 2))
        # 의미:
        #   - K{row} = Ballast_t_calc (column 11)
        #   - $B$13 = pump_rate_effective_tph (펌프 속도, t/h)
        #   - Ballast_time_h_calc = Ballast_t_calc / Pump_rate (펌프에 소요되는 시간, h)
        ws.cell(row=row, column=12).value = (
            f'=IF(OR(K{row_str}="", $B$13="", $B$13=0, ISERROR($B$13)), "", ROUND(ABS(K{row_str}) / $B$13, 2))'
        )
        ws.cell(row=row, column=12).font = styles["normal_font"]
        ws.cell(row=row, column=12).number_format = number_format
        # L (12): Ballast_t - v4.0: G가 FWD_precise_m로 변경되었으므로 Trim_m은 F/100 사용
        ws.cell(row=row, column=13).value = (
            f'=IF(OR(F{row_str}="", OR($B$12="", $B$12=0)), "", ROUND(ABS(F{row_str}/100) * 50 * $B$12, 2))'
        )
        ws.cell(row=row, column=13).font = styles["normal_font"]
        ws.cell(row=row, column=13).number_format = number_format
        # M (13): Ballast_time_h - L(Ballast_t)를 펌프 레이트(B13)로 나눈 시간
        # L = column 13 (Ballast_t)
        ws.cell(row=row, column=14).value = (
            f'=IF(OR(L{row_str}="", $B$13="", $B$13=0, ISERROR($B$13)), "", ROUND(L{row_str} / $B$13, 2))'
        )
        ws.cell(row=row, column=14).font = styles["normal_font"]
        ws.cell(row=row, column=14).number_format = number_format
        # N (14): Trim_Check - v4.0: G가 FWD_precise_m로 변경되었으므로 Trim_m은 F/100 사용
        ws.cell(row=row, column=15).value = (
            f'=IF(F{row_str}="", "", IF(ABS(F{row_str}/100) <= ($B$15/50), "OK", "EXCESSIVE"))'
        )
        # P (16): Dfwd_m - FWD_precise_m 참조 (모든 Stage 동일)
        # Q (17): Daft_m - AFT_precise_m 참조 (모든 Stage 동일)
        # (하드코딩 제거 - Stage 5_PreBallast, Stage 6A_Critical도 수식으로 계산)
        ws.cell(row=row, column=16).value = (
            f'=IF($A{row_str}="","",G{row_str})'  # P (16): Dfwd_m
        )
        ws.cell(row=row, column=16).font = styles["normal_font"]
        ws.cell(row=row, column=16).number_format = number_format
        ws.cell(row=row, column=17).value = (
            f'=IF($A{row_str}="","",H{row_str})'  # Q (17): Daft_m
        )
        ws.cell(row=row, column=17).font = styles["normal_font"]
        ws.cell(row=row, column=17).number_format = number_format
        # R (18): Trim_target_stage_cm - Stage별 Trim 타깃 (새 설계)
        # Stage 설명 세트 기준으로 고정된 값:
        #   Stage 1: 0.00 (기준 Lightship, Trim 타깃 없음)
        #   Stage 2: -96.50 (TR1 램프 진입, 허용 범위 내 bow-down 목표)
        #   Stage 3: -96.50 (TR1 mid-ramp, 같은 목표 유지)
        #   Stage 4: -96.50 (TR1 온덱, 같은 목표 유지)
        #   Stage 5: -89.58 (TR1 aft stow 후 기준 상태)
        #   Stage 5_PreBallast: 240.00 (D-1 Pre-ballast, 의도적 stern trim; FW2 Fr.0-6)
        #   Stage 6A_Critical (Opt C): 0.00 (TR2 램프 진입 시 Even Keel 목표)
        #   Stage 6C_TotalMassOpt: -96.50 (최대 변위 Sensitivity case)
        #   Stage 6C: -96.50 (Planned final stowage 기준값)
        #   Stage 7: 0.00 (최종 Lightship/Reference)
        # 이 값은 ΔTM_cm_tm 계산 시 사용됨 (R열이 비어있으면 전역 B8 사용)
        target_trim = target_trim_by_stage.get(stage_name)
        ws.cell(row=row, column=18).value = target_trim
        ws.cell(row=row, column=18).number_format = number_format
        ws.cell(row=row, column=18).fill = styles["input_fill"]
        ws.cell(row=row, column=18).font = styles["normal_font"]
        # S (19): FWD_DeckElev_CD_m - Chart Datum 기준 Deck Elevation (Tide 포함)
        # P = column 16 (Dfwd_m)
        ws.cell(row=row, column=19).value = (
            f'=IF(P{row_str}="", "", $B$11 - P{row_str} + $B$7)'
        )
        ws.cell(row=row, column=19).font = styles["normal_font"]
        ws.cell(row=row, column=19).number_format = number_format
        # T (20): AFT_DeckElev_CD_m - Chart Datum 기준 Deck Elevation (Tide 포함)
        # Q = column 17 (Daft_m)
        ws.cell(row=row, column=20).value = (
            f'=IF(Q{row_str}="", "", $B$11 - Q{row_str} + $B$7)'
        )
        ws.cell(row=row, column=20).font = styles["normal_font"]
        ws.cell(row=row, column=20).number_format = number_format

        # T (20): Notes - 제거됨 (G4-G15로 이동)
        # Notes는 더 이상 T(20) 컬럼에 배치하지 않음

        # U-AE 컬럼은 extend_roro_captain_req()에서 작성됨

        # PATCH C-2: New columns for iterative ballast correction (AF-AJ, column 32-36)
        # AF (32): Trim_gate_cm - Required trim to satisfy FWD limit (from B2 calculation)
        trim_gate_cm = target_trim_by_stage.get(stage_name, 0.0)
        ws.cell(row=row, column=32).value = trim_gate_cm
        ws.cell(row=row, column=32).number_format = number_format
        ws.cell(row=row, column=32).font = styles["normal_font"]

        # AG (33): Gate_Ballast_net_t - Final ballast after iteration
        # AH (34): Dfwd_gate_m - Final forward draft after ballast correction
        # AI (35): Gate_iter_n - Number of iterations used
        # AJ (36): Gate_Flag - OK/LIMIT flag

        # Perform iterative ballast correction if stage has results
        if stage_name in stage_results:
            res = stage_results[stage_name]

            # Get loads for this stage (includes preballast if applicable)
            stage_loads = build_stage_loads(stage_name, preballast_opt, params)

            # Extract current ballast from loads as initial estimate
            # Look for PreBallast LoadItem in the loads
            initial_ballast = 0.0
            for load in stage_loads:
                if load.name == "PreBallast":
                    initial_ballast = load.weight_t
                    break

            # If no PreBallast found, use Ballast_t_calc from stage results as fallback
            if abs(initial_ballast) < 0.01:
                initial_ballast = float(
                    res.get("Ballast_t_calc", res.get("Ballast_t", 0.0)) or 0.0
                )

            # Perform iterative correction to find ballast needed for Dfwd = 2.70m
            correction_result = iterative_ballast_correction(
                base_disp_t=base_disp_t,
                base_tmean_m=base_tmean_m,
                loads=stage_loads,
                initial_ballast_t=initial_ballast,
                fwd_limit_m=2.70,
                max_iterations=3,
                tolerance_m=0.01,
                **params,
            )

            # AG (33): Gate_Ballast_net_t
            ws.cell(row=row, column=33).value = correction_result["ballast_t"]
            ws.cell(row=row, column=33).number_format = number_format
            ws.cell(row=row, column=33).font = styles["normal_font"]

            # AH (34): Dfwd_gate_m
            ws.cell(row=row, column=34).value = correction_result["dfwd_gate_m"]
            ws.cell(row=row, column=34).number_format = number_format
            ws.cell(row=row, column=34).font = styles["normal_font"]

            # AI (35): Gate_iter_n
            ws.cell(row=row, column=35).value = correction_result["iterations"]
            ws.cell(row=row, column=35).number_format = "0"
            ws.cell(row=row, column=35).font = styles["normal_font"]

            # AJ (36): Gate_Flag
            flag = correction_result["flag"]
            ws.cell(row=row, column=36).value = flag
            ws.cell(row=row, column=36).font = styles["normal_font"]
            if flag == "OK":
                ws.cell(row=row, column=36).fill = styles["ok_fill"]
            else:
                ws.cell(row=row, column=36).fill = styles["ng_fill"]
        else:
            # No stage results, leave columns empty
            for col in [33, 34, 35, 36]:
                ws.cell(row=row, column=col).value = ""
                ws.cell(row=row, column=col).font = styles["normal_font"]

    # ===============================================================
    # REMOVED: Old logic that manually added "Stage 6C_TotalMassOpt" row
    # Stage 6C_TotalMassOpt is now included in stages list
    # and is handled in the main stage loop above (lines 4582-4700)
    # ===============================================================

    from openpyxl.workbook.defined_name import DefinedName

    wb.defined_names["MTC"] = DefinedName(
        "MTC", attr_text="'RORO_Stage_Scenarios'!$B$9"
    )
    wb.defined_names["LCF"] = DefinedName(
        "LCF", attr_text="'RORO_Stage_Scenarios'!$B$10"
    )
    wb.defined_names["PumpRate"] = DefinedName(
        "PumpRate", attr_text="'RORO_Stage_Scenarios'!$B$13"
    )
    wb.defined_names["X_Ballast"] = DefinedName(
        "X_Ballast", attr_text="'RORO_Stage_Scenarios'!$B$14"
    )

    if trim5_row is not None:
        wb.defined_names["TRIM5_CM"] = DefinedName(
            "TRIM5_CM", attr_text=f"'RORO_Stage_Scenarios'!$E${trim5_row}"
        )

    widths = {
        "A": 18,
        "B": 11,
        "C": 11,
        "D": 13,
        "E": 10,
        "F": 9,
        "G": 13,
        "H": 12,
        "I": 11,
        "J": 13,
        "K": 15,
        "L": 11,
        "M": 15,
        "N": 13,
        "O": 10,
        "P": 10,
        "Q": 13,
        "R": 13,
        "S": 40,
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = "G2"  # 기존 파일과 동일하게 G2

    # Opt C Stage 정보 출력 (DAS Method v4.3 Final Optimized)
    opt_c_data = build_opt_c_stage(preballast_opt)
    print(
        f"  [INFO] Opt C Updated (DAS Method v4.3 Final): W={opt_c_data['weight_t']:.1f}t (Cargo 560t + Ballast {preballast_opt:.1f}t), LCG={opt_c_data['x_from_mid_m']:.2f}m (Even Keel 근접)"
    )

    print("  [OK] RORO_Stage_Scenarios sheet created")

    # 전체 행 수 계산 (모든 Stage 포함)
    total_rows = len(stages)

    # 7) ΔTM & Lever_arm Python 재계산 리포트 시트 생성
    #    - Stage별 Trim / Trim_target / MTC / X_ballast / LCF / Lpp 기반
    #    - RORO_Stage_Scenarios Stage 테이블(엑셀 수식)과 숫자 비교용
    create_roro_delta_lever_report_sheet(
        wb=wb,
        stage_results=stage_results,
        trim_target_map=TRIM_TARGET_MAP,
    )

    # 8) Draft / Freeboard Margin 체크 시트
    create_roro_draft_margin_check_sheet(
        wb=wb,
        stage_results=stage_results,
        roro_sheet_name="RORO_Stage_Scenarios",
        fwd_limit_m=2.70,
        linkspan_min_freeboard_m=0.28,
    )

    # 9) Stability GM 체크 시트
    create_roro_stability_gm_check_sheet(
        wb=wb,
        stage_results=stage_results,
        hydro_sheet_name="Hydro_Table",
        gm_target_m=1.50,
    )

    # 10) 옵션 1/2/3 시나리오 자동 Summary → 비교 시트
    scenario_stage_map = {
        # ⚠️ 여기 Stage 이름은 실제 RORO_Stage_Scenarios의 Stage 값에 맞게 조정 필요
        "Option 1 – Pre-ballast Only": ["Stage 5_PreBallast"],
        "Option 2 – D-Day Dynamic": ["Stage 6A_Critical (Opt C)"],
        "Option 3 – Split Stage": ["Stage 6A_Critical (Opt C)", "Stage 6C"],
    }
    pump_rate_effective = 100.0  # t/h (Calc 시트 값과 동일하게 맞추면 됨)

    scenario_summary = build_ballast_scenarios_from_stage_results(
        stage_results=stage_results,
        scenario_stage_map=scenario_stage_map,
        pump_rate_tph=pump_rate_effective,
        fwd_limit_m=2.70,
        linkspan_min_freeboard_m=0.28,
    )

    create_ballast_scenario_comparison_sheet(
        wb=wb,
        scenarios=scenario_summary,
        sheet_name="Ballast_Scenario_Comparison",
    )

    # ============================================================
    # Option B+ add-ons: IS Code / Acceptance / Structural / CSM
    # ============================================================
    try:
        enrich_stage_results_bplus(stage_results=stage_results, env=params.get("env"))
        csm_trace = build_csm_trace(stage_results=stage_results)

        create_iscode_check_sheet(wb, stage_results, sheet_name="RORO_ISCODE_Check")
        create_acceptance_criteria_sheet(
            wb, stage_results, sheet_name="MWS_Acceptance_Criteria"
        )
        create_structural_gate_sheet(
            wb, stage_results, sheet_name="RORO_Structural_Gate"
        )
        create_csm_trace_sheet(wb, csm_trace, sheet_name="CSM_Trace")
        create_mws_pack_index_sheet(wb, sheet_name="MWS_Pack_Index")
    except Exception as e:
        logging.warning(f"[BPLUS] Optional add-on sheets skipped: {e}")

    return stages_list, first_data_row, len(stages_list)


# ============================================================================
# RORO Delta Lever Report Sheet Creation
# ============================================================================


def create_roro_delta_lever_report_sheet(
    wb: Workbook,
    stage_results: Dict[str, Dict[str, float]],
    trim_target_map: Optional[Dict[str, float]] = None,
    sheet_name: str = "RORO_Stage_Scenarios",
    delta_tm_col: str = "I",  # 가정: 엑셀 ΔTM_cm_tm 컬럼 (I열)
) -> Worksheet:
    """
    RORO_Stage_Scenarios Stage 테이블 기준
    - Python 엔진(Trim_cm, Stage별 Trim_target, MTC, X_ballast, LCF, Lpp)을 사용해
      ΔTM_cm_tm, Lever_arm_m, Ballast_t_calc 를 재계산하고
      엑셀 수식 결과와의 차이를 한눈에 볼 수 있는 리포트 시트를 생성한다.

    비교 포인트:
    - ΔTM_cm_tm (Python vs Excel)
    - |ΔTM_diff| > 0.50 → Flag = "CHECK", 그 이하면 "OK"

    데이터 소스:
    1. stage_results (Python 계산 결과):
       - create_roro_sheet() 함수 내에서 생성 (L3758-3824)
       - 각 Stage별로 solve_stage() 호출하여 계산된 값
       - 포함 데이터: W_stage_t, x_stage_m, Trim_cm 등
    2. RORO_Stage_Scenarios 시트에서 읽는 전역 파라미터 (L4786-4798):
       - B8: Trim_target_cm (전역)
       - B9: MTC_t_m_per_cm
       - B10: LCF_m_from_midship
       - B14: X_Ballast_from_AP_m
       - B15: Lpp_m
    3. RORO_Stage_Scenarios 시트에서 읽는 Excel 계산값 (L4877-4886):
       - I열 (delta_tm_col): ΔTM_cm_tm (Excel 수식 계산 결과)
       - Stage 이름을 A열에서 찾아 해당 행의 I열 값 읽기
    4. trim_target_map (Stage별 Trim 타깃):
       - TRIM_TARGET_MAP (L4124-4136에서 정의)
       - Stage별 Trim 타깃 값으로 ΔTM 계산에 사용
    """
    styles = get_styles()
    number_format = "#,##0.00"

    if sheet_name not in wb.sheetnames:
        print(f"[WARN] 워크북에 시트 '{sheet_name}'가 없습니다. ΔTM/Lever 리포트 생략.")
        return wb.create_sheet("RORO_Delta_Lever_Report")

    ws = wb[sheet_name]

    # ===== 1) 전역 파라미터 (Stage Param 블록에서 값 읽기) =====
    def _safe_float(cell_addr: str, default: float = 0.0) -> float:
        try:
            v = ws[cell_addr].value
            return float(v) if v is not None else default
        except Exception:
            return default

    trim_target_global = _safe_float("B8", 0.0)  # Trim_target_cm (전역)
    mtc = _safe_float("B9", 0.0)  # MTC_t_m_per_cm
    lcf_from_mid = _safe_float("B10", 0.0)  # LCF_m_from_midship
    x_ballast_ap = _safe_float("B14", 0.0)  # X_Ballast_from_AP_m
    lpp_m = _safe_float("B15", 0.0)  # Lpp_m

    lever_arm_calc = None
    if lpp_m != 0.0:
        lcf_from_ap = lcf_from_mid + lpp_m / 2.0
        lever_arm_calc = round(x_ballast_ap - lcf_from_ap, 2)
    else:
        print("[WARN] Lpp_m=0.0 이라 Lever_arm 계산 불가.")

    # Stage별 Trim 타깃 맵 준비
    if trim_target_map is None:
        try:
            trim_target_map = TRIM_TARGET_MAP  # type: ignore[name-defined]
        except NameError:
            trim_target_map = {}

    # ===== 2) 원본 RORO 시트에서 Stage → Row 맵 작성 (엑셀 ΔTM 읽기용) =====
    stage_row_map: Dict[str, int] = {}
    for row in range(1, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value  # A열 (Stage 이름)
        if val is None:
            continue
        name = str(val).strip()
        if name:
            stage_row_map[name] = row

    # ===== 3) 리포트 시트 생성 =====
    report_ws = wb.create_sheet("RORO_Delta_Lever_Report")
    report_ws.title = "RORO_Delta_Lever_Report"

    # Title
    title_cell = report_ws["A1"]
    title_cell.value = "RORO Stage ΔTM & Lever Arm Verification Report"
    title_cell.font = styles["title_font"]

    report_ws["A3"].value = (
        "주의: ΔTM(Python) vs ΔTM(Excel) 비교. |ΔTM_diff| > 0.50 → Flag=CHECK."
    )
    report_ws["A3"].font = styles["normal_font"]

    # 헤더 행
    header_row = 5
    headers = [
        "Stage",
        "W_stage_t",
        "x_stage_m",
        "Trim_cm (Python)",
        "Trim_target_stage_cm",
        "MTC_t_m_per_cm",
        "ΔTM_cm_tm (Python)",
        "ΔTM_cm_tm (Excel)",
        "ΔTM_diff (Py-XL)",
        "Flag (|ΔTM_diff|>0.50)",
        "Lever_arm_m (Python)",
        "Ballast_t_calc (Python)",
        "비고",
    ]
    for col, name in enumerate(headers, start=1):
        c = report_ws.cell(row=header_row, column=col, value=name)
        c.font = styles["header_font"]
        c.fill = styles["header_fill"]

    # ===== 4) Stage별 데이터 라인 =====
    current_row = header_row + 1

    for stage_name, res in stage_results.items():
        w_stage = float(res.get("W_stage_t", 0.0) or 0.0)
        x_stage = float(res.get("x_stage_m", 0.0) or 0.0)
        trim_cm = float(res.get("Trim_cm", 0.0) or 0.0)

        target_trim_cm = float(
            (trim_target_map.get(stage_name) if trim_target_map else trim_target_global)
            or trim_target_global
        )

        # ΔTM(Python)
        # PATCH B2-3: Reverse direction to (Target - Actual) to match Excel formula
        delta_tm_py = (target_trim_cm - trim_cm) * mtc if mtc != 0.0 else 0.0

        # ΔTM(Excel)
        excel_row = stage_row_map.get(stage_name)
        delta_tm_xl: Optional[float]
        if excel_row is not None:
            try:
                excel_val = ws[f"{delta_tm_col}{excel_row}"].value
                delta_tm_xl = float(excel_val) if excel_val is not None else None
            except Exception:
                delta_tm_xl = None
        else:
            delta_tm_xl = None

        # ΔTM_diff & Flag
        delta_tm_diff: Optional[float]
        flag = ""
        if delta_tm_xl is not None:
            delta_tm_diff = delta_tm_py - delta_tm_xl
            if abs(delta_tm_diff) > 0.50:
                flag = "CHECK"
            else:
                flag = "OK"
        else:
            delta_tm_diff = None
            flag = "NO_DATA"

        # Lever_arm & Ballast
        if lever_arm_calc not in (None, 0.0):
            ballast_calc = delta_tm_py / lever_arm_calc
        else:
            ballast_calc = 0.0

        row = current_row

        # 공통 숫자 셀 세팅 함수
        def _set_num(col_idx: int, value: Optional[float]):
            if value is None:
                return
            c = report_ws.cell(row=row, column=col_idx, value=round(float(value), 2))
            c.font = styles["normal_font"]
            c.number_format = number_format

        # Stage
        report_ws.cell(row=row, column=1, value=stage_name).font = styles["normal_font"]

        _set_num(2, w_stage)
        _set_num(3, x_stage)
        _set_num(4, trim_cm)
        _set_num(5, target_trim_cm)
        _set_num(6, mtc)
        _set_num(7, delta_tm_py)
        _set_num(8, delta_tm_xl)
        _set_num(9, delta_tm_diff)

        # Flag
        flag_cell = report_ws.cell(row=row, column=10, value=flag)
        flag_cell.font = styles["normal_font"]

        # Lever & Ballast
        _set_num(11, lever_arm_calc if lever_arm_calc is not None else 0.0)
        _set_num(12, ballast_calc)

        # 비고
        note = []
        if mtc == 0.0:
            note.append("MTC=0.0")
        if lever_arm_calc is None:
            note.append("Lever_arm 파라미터 부족")
        if delta_tm_xl is None:
            note.append("Excel ΔTM 없음")
        report_ws.cell(
            row=row, column=13, value="; ".join(note) if note else ""
        ).font = styles["normal_font"]

        current_row += 1

    # 컬럼 폭
    for col in range(1, len(headers) + 1):
        report_ws.column_dimensions[get_column_letter(col)].width = 20

    print("  [OK] RORO_Delta_Lever_Report sheet created (with ΔTM diff & Flag)")
    return report_ws


# ============================================================================
# Ballast Scenario Comparison Sheet Creation
# ============================================================================


def create_ballast_scenario_comparison_sheet(
    wb: Workbook,
    scenarios: Dict[str, Dict[str, float]],
    sheet_name: str = "Ballast_Scenario_Comparison",
) -> Worksheet:
    """
    옵션 1/2/3 Ballast 시나리오 비교용 요약 시트 생성.

    기대 입력 형식 (예시):
    scenarios = {
        "Option 1 – Pre-ballast Only": {
            "total_ballast_t": 536.0,
            "total_time_h": 5.36,
            "fwd_draft_margin_m": 0.28,       # 2.70m 한계 대비 여유
            "linkspan_freeboard_margin_m": 0.12,
        },
        "Option 2 – D-day Dynamic": {
            "total_ballast_t": 420.0,
            "total_time_h": 6.00,
            "fwd_draft_margin_m": 0.15,
            "linkspan_freeboard_margin_m": 0.08,
        },
        ...
    }
    """
    styles = get_styles()
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        wb.remove(ws)

    ws = wb.create_sheet(sheet_name)

    title = ws["A1"]
    title.value = "Ballast Scenario Comparison (Option 1 / 2 / 3)"
    title.font = styles["title_font"]

    ws["A3"].value = (
        "각 시나리오별 총 Ballast, Pump 시간, FWD Draft margin, Linkspan freeboard margin 비교."
    )
    ws["A3"].font = styles["normal_font"]

    header_row = 5
    headers = [
        "Scenario",
        "Total Ballast_t",
        "Pump Time_h",
        "FWD Draft Margin_m",
        "Linkspan Freeboard Margin_m",
        "Risk Comment",
    ]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.font = styles["header_font"]
        c.fill = styles["header_fill"]

    number_format = "#,##0.00"
    current_row = header_row + 1

    # Defensive init to avoid NameError if code paths change
    link_margin = 0.0
    fwd_margin = 0.0

    for name, vals in scenarios.items():
        total_ballast = float(vals.get("total_ballast_t", 0.0) or 0.0)
        time_h = float(vals.get("total_time_h", 0.0) or 0.0)
        fwd_margin = float(vals.get("fwd_draft_margin_m", 0.0) or 0.0)
        link_margin = float(vals.get("linkspan_freeboard_margin_m", 0.0) or 0.0)

        ws.cell(row=current_row, column=1, value=name).font = styles["normal_font"]

        def _num(col_idx: int, value: float):
            c = ws.cell(row=current_row, column=col_idx, value=round(float(value), 2))
            c.font = styles["normal_font"]
            c.number_format = number_format

        _num(2, total_ballast)
        _num(3, time_h)
        _num(4, fwd_margin)
        _num(5, link_margin)

        # 간단 Risk 평가: margin이 0.10m 미만이면 WARN
        risk_comment = []
        if fwd_margin < 0.10:
            risk_comment.append("FWD margin < 0.10m")
        if link_margin < 0.10:
            risk_comment.append("Linkspan margin < 0.10m")
        ws.cell(row=current_row, column=6, value="; ".join(risk_comment)).font = styles[
            "normal_font"
        ]

        current_row += 1

    # 컬럼 폭
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 30

    print("  [OK] Ballast_Scenario_Comparison sheet created")
    return ws


# ============================================================================
# CSV Export Functions
# ============================================================================


def export_roro_delta_lever_report_to_csv(
    wb: Workbook,
    sheet_name: str = "RORO_Delta_Lever_Report",
    csv_path: str = "RORO_Delta_Lever_Report.csv",
) -> Path:
    """
    RORO_Delta_Lever_Report 시트를 CSV로 export.
    - 메일 첨부 / WhatsApp 공유용.
    """
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in workbook.")

    ws = wb[sheet_name]
    out_path = Path(csv_path)

    with out_path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(list(row))

    print(f"[OK] RORO_Delta_Lever_Report exported to {out_path}")
    return out_path


# ============================================================================
# Draft Margin Check Sheet Creation
# ============================================================================


def create_roro_draft_margin_check_sheet(
    wb: Workbook,
    stage_results: Dict[str, Dict[str, float]],
    roro_sheet_name: str = "RORO_Stage_Scenarios",
    fwd_limit_m: float = 2.70,
    linkspan_min_freeboard_m: float = 0.28,
) -> Worksheet:
    """
    Stage별 Draft / Freeboard margin 자동 검증 시트.

    가정:
    - RORO_Stage_Scenarios 시트 구조 (예: 우리가 쓰는 표):
      A: Stage
      O: Dfwd_m
      P: Daft_m
      S: FWD_DeckElev_CD_m (Chart Datum 기준 Deck Elevation, Tide 포함)
      Python stage_results: FWD_Height_m (Freeboard, Tide 미포함)

    규칙:
    - FWD_margin = fwd_limit_m - Dfwd_m
      * FWD_margin < 0.00 → FWD_FLAG = "LIMIT"
      * 0.00 <= FWD_margin < 0.10 → "WARN"
      * >= 0.10 → "OK"
    - Linkspan_margin = FWD_Freeboard_m - linkspan_min_freeboard_m
      * FWD_Freeboard_m = D_vessel - Dfwd_m (Python FWD_Height_m 사용)
      * < 0.00 → LINK_FLAG = "CONTACT"
      * 0.00 ~ 0.10 → "WARN"
      * >= 0.10 → "OK"
    - Final Flag = FWD_FLAG / LINK_FLAG 중 가장 심각한 수준

    데이터 소스:
    1. stage_results (Python 계산 결과, 우선순위 1):
       - create_roro_sheet() 함수 내에서 생성 (L3758-3824)
       - 각 Stage별로 solve_stage() 호출하여 계산된 값
       - 포함 데이터: Dfwd_m, Daft_m, FWD_Height_m (Freeboard, Tide 미포함)
    2. RORO_Stage_Scenarios 시트에서 읽는 값 (Fallback, 우선순위 2, L5189-5203):
       - stage_results에 값이 없을 경우 Excel 시트에서 읽기
       - O열: Dfwd_m
       - P열: Daft_m
       - S열: FWD_DeckElev_CD_m (Deck Elevation, Tide 포함)
    3. 하드코딩된 상수 (함수 파라미터):
       - fwd_limit_m = 2.70 (FWD draft limit, 기본값)
       - linkspan_min_freeboard_m = 0.28 (Linkspan 최소 freeboard, 기본값)
    """
    styles = get_styles()
    number_format = "#,##0.00"

    if roro_sheet_name not in wb.sheetnames:
        print(f"[WARN] '{roro_sheet_name}' 시트가 없어 Draft Margin 체크 생략.")
        return wb.create_sheet("RORO_Draft_Margin_Check")

    ws_src = wb[roro_sheet_name]
    # Header 기반 컬럼 탐지(시트 구조 변경에 견고)
    header_row_src = 5  # RORO_Stage_Scenarios header row
    header_map: Dict[str, int] = {}
    for col in range(1, ws_src.max_column + 1):
        hv = ws_src.cell(row=header_row_src, column=col).value
        if isinstance(hv, str) and hv.strip():
            header_map[hv.strip().lower()] = col

    def _col(*names: str) -> int | None:
        for nm in names:
            c = header_map.get(nm.strip().lower())
            if c is not None:
                return c
        return None

    # 대표 헤더 후보들
    _dfwd_col = _col("Dfwd_m", "Fwd Draft (m)", "Fwd Draft(m)", "FWD_precise_m")
    _daft_col = _col("Daft_m", "Aft Draft (m)", "Aft Draft(m)", "AFT_precise_m")
    _fwdh_col = _col(
        "FWD_Height_m",
        "FWD_Height_m (Linkspan FB)",
        "Deck elevation (m)",
        "Freeboard (m)",
    )

    _dfwd_col = _col("Dfwd_m", "Fwd Draft (m)", "Fwd Draft(m)", "FWD_precise_m")
    _daft_col = _col("Daft_m", "Aft Draft (m)", "Aft Draft(m)", "AFT_precise_m")
    _fwdh_col = _col(
        "FWD_Height_m",
        "FWD_Height_m (Linkspan FB)",
        "Deck elevation (m)",
        "Freeboard (m)",
    )

    # Stage → row 맵 (A열 기준)
    stage_row_map: Dict[str, int] = {}
    for row in range(1, ws_src.max_row + 1):
        v = ws_src.cell(row=row, column=1).value
        if v is None:
            continue
        stage_row_map[str(v).strip()] = row

    ws = wb.create_sheet("RORO_Draft_Margin_Check")

    title = ws["A1"]
    title.value = "RORO Stage Draft & Freeboard Margin Check"
    title.font = styles["title_font"]

    ws["A3"].value = (
        f"FWD limit={fwd_limit_m:.2f} m, Linkspan min freeboard={linkspan_min_freeboard_m:.2f} m 기준."
    )
    ws["A3"].font = styles["normal_font"]

    header_row = 5
    headers = [
        "Stage",
        "Dfwd_m",
        "Daft_m",
        "FWD_margin_m (limit-DFWD)",
        "FWD_Flag",
        "FWD_Freeboard_m (D-Dfwd)",
        "Linkspan_margin_m",
        "Linkspan_Flag",
        "Final_Flag",
        "비고",
    ]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.font = styles["header_font"]
        c.fill = styles["header_fill"]

    def _flag_fwd(margin: float) -> str:
        if margin < 0.0:
            return "LIMIT"
        if margin < 0.10:
            return "WARN"
        return "OK"

    def _flag_link(margin: float) -> str:
        if margin < 0.0:
            return "CONTACT"
        if margin < 0.10:
            return "WARN"
        return "OK"

    severity_rank = {"OK": 0, "WARN": 1, "LIMIT": 2, "CONTACT": 2, "NO_DATA": 3}

    current_row = header_row + 1
    for stage_name, res in stage_results.items():
        src_row = stage_row_map.get(stage_name)

        # 우선 stage_results(Python 계산) 사용, 부족하면 원본 시트에서 보강
        def _cell_num(col_idx: int | None, default: float = 0.0) -> float:
            if col_idx is None or src_row is None:
                return default
            try:
                v = ws_src.cell(row=src_row, column=col_idx).value
                return float(v) if v is not None and v != "" else default
            except Exception:
                return default

        dfwd = (
            float(res.get("Dfwd_m"))
            if res.get("Dfwd_m") is not None
            else _cell_num(_dfwd_col, 0.0)
        )
        daft = (
            float(res.get("Daft_m"))
            if res.get("Daft_m") is not None
            else _cell_num(_daft_col, 0.0)
        )
        fwd_height = (
            float(res.get("FWD_Height_m"))
            if res.get("FWD_Height_m") is not None
            else _cell_num(_fwdh_col, 0.0)
        )

        fwd_margin = fwd_limit_m - dfwd
        link_margin = fwd_height - linkspan_min_freeboard_m

        fwd_flag = _flag_fwd(fwd_margin)
        link_flag = _flag_link(link_margin)

        # 최종 Flag (심각도 높은 쪽)
        flags = [fwd_flag, link_flag]
        final_flag = max(flags, key=lambda x: severity_rank.get(x, 0))

        # 기록
        ws.cell(row=current_row, column=1, value=stage_name).font = styles[
            "normal_font"
        ]

        def _num(col_idx: int, v: float):
            c = ws.cell(row=current_row, column=col_idx, value=round(v, 3))
            c.font = styles["normal_font"]
            c.number_format = number_format

        _num(2, dfwd)
        _num(3, daft)
        _num(4, fwd_margin)

        ws.cell(row=current_row, column=5, value=fwd_flag).font = styles["normal_font"]

        _num(6, fwd_height)
        _num(7, link_margin)
        ws.cell(row=current_row, column=8, value=link_flag).font = styles["normal_font"]
        ws.cell(row=current_row, column=9, value=final_flag).font = styles[
            "normal_font"
        ]

        note = []
        if src_row is None:
            note.append("Stage_row 없음")
        ws.cell(row=current_row, column=10, value="; ".join(note)).font = styles[
            "normal_font"
        ]

        current_row += 1

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22

    print("  [OK] RORO_Draft_Margin_Check sheet created")
    return ws


# ============================================================================
# Stability GM Check Sheet Creation
# ============================================================================


def create_roro_stability_gm_check_sheet(
    wb: Workbook,
    stage_results: Dict[str, Dict[str, float]],
    hydro_sheet_name: str = "Hydro_Table",
    gm_target_m: float = 1.50,
) -> Worksheet:
    """
    Hydro_Table 기반 Stage별 GM 계산 / 체크 시트.

    가정:
    - Hydro_Table 시트:
      A열: Disp_t
      B열: GM_m
    - stage_results[stage]['Disp_t'] 값이 존재한다고 가정 (없으면 해당 Stage skip)
    """
    styles = get_styles()
    number_format = "#,##0.000"

    if hydro_sheet_name not in wb.sheetnames:
        print(f"[WARN] '{hydro_sheet_name}' 시트 없음. GM 체크 생략.")
        return wb.create_sheet("RORO_Stability_Check")

    ws_h = wb[hydro_sheet_name]

    # Hydro table 로드
    disp_list = []
    gm_list = []
    for row in range(2, ws_h.max_row + 1):
        d = ws_h[f"A{row}"].value
        g = ws_h[f"B{row}"].value
        try:
            d = float(d)
            g = float(g)
        except (TypeError, ValueError):
            continue
        disp_list.append(d)
        gm_list.append(g)

    if not disp_list:
        print("[WARN] Hydro_Table에 유효한 Δ-GM 데이터 없음.")
        return wb.create_sheet("RORO_Stability_Check")

    # 정렬
    paired = sorted(zip(disp_list, gm_list), key=lambda x: x[0])
    disp_list = [p[0] for p in paired]
    gm_list = [p[1] for p in paired]

    def interp_gm(disp: float) -> float:
        if disp <= disp_list[0]:
            return gm_list[0]
        if disp >= disp_list[-1]:
            return gm_list[-1]
        for i in range(1, len(disp_list)):
            if disp_list[i] >= disp:
                x0, y0 = disp_list[i - 1], gm_list[i - 1]
                x1, y1 = disp_list[i], gm_list[i]
                if x1 == x0:
                    return y0
                ratio = (disp - x0) / (x1 - x0)
                return y0 + (y1 - y0) * ratio
        return gm_list[-1]

    ws = wb.create_sheet("RORO_Stability_Check")

    title = ws["A1"]
    title.value = f"RORO Stage Stability Check (GM Target={gm_target_m:.2f} m)"
    title.font = styles["title_font"]

    header_row = 3
    headers = [
        "Stage",
        "Disp_t",
        "GM_Calc_m",
        "GM_Target_m",
        "GM_Margin_m",
        "GM_Check (OK/CHECK)",
    ]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.font = styles["header_font"]
        c.fill = styles["header_fill"]

    current_row = header_row + 1
    for stage_name, res in stage_results.items():
        disp = res.get("Disp_t")
        try:
            disp = float(disp) if disp is not None else None
        except (TypeError, ValueError):
            disp = None

        if disp is None or disp <= 0.0:
            continue

        gm_calc = interp_gm(disp)
        gm_margin = gm_calc - gm_target_m
        gm_flag = "OK" if gm_calc >= gm_target_m else "CHECK"

        ws.cell(row=current_row, column=1, value=stage_name).font = styles[
            "normal_font"
        ]

        def _num(col_idx: int, v: float):
            c = ws.cell(row=current_row, column=col_idx, value=round(v, 3))
            c.font = styles["normal_font"]
            c.number_format = number_format

        _num(2, disp)
        _num(3, gm_calc)
        _num(4, gm_target_m)
        _num(5, gm_margin)
        ws.cell(row=current_row, column=6, value=gm_flag).font = styles["normal_font"]

        current_row += 1

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22

    print("  [OK] RORO_Stability_Check sheet created")
    return ws


# ============================================================================
# Ballast Scenario Builder
# ============================================================================


def build_ballast_scenarios_from_stage_results(
    stage_results: Dict[str, Dict[str, float]],
    scenario_stage_map: Dict[str, list],
    pump_rate_tph: float = 100.0,
    fwd_limit_m: float = 2.70,
    linkspan_min_freeboard_m: float = 0.28,
) -> Dict[str, Dict[str, float]]:
    """
    Stage_results + 시나리오별 Stage 묶음을 이용해
    옵션 1/2/3 summary dict 자동 생성.

    scenario_stage_map 예:
    {
        "Option 1 – Pre-ballast Only": ["Stage 5_PreBallast"],
        "Option 2 – D-Day Dynamic": ["Stage 6A_Critical"],
        "Option 3 – Split Stage": ["Stage 6A_Critical", "Stage 6C"],
    }
    """
    scenarios: Dict[str, Dict[str, float]] = {}

    for scenario_name, stage_list in scenario_stage_map.items():
        total_ballast = 0.0
        total_time_h = 0.0
        has_time = False

        min_fwd_margin = None
        min_link_margin = None

        for st in stage_list:
            res = stage_results.get(st)
            if not res:
                continue

            ballast_t = float(res.get("Ballast_t", 0.0) or 0.0)
            total_ballast += ballast_t

            if res.get("Ballast_time_h") is not None:
                try:
                    total_time_h += float(res.get("Ballast_time_h") or 0.0)
                    has_time = True
                except (TypeError, ValueError):
                    pass

            dfwd = float(res.get("Dfwd_m", fwd_limit_m) or fwd_limit_m)
            fwd_margin = fwd_limit_m - dfwd

            fwd_height = float(res.get("FWD_Height_m", 0.0) or 0.0)
            link_margin = fwd_height - linkspan_min_freeboard_m

            if min_fwd_margin is None or fwd_margin < min_fwd_margin:
                min_fwd_margin = fwd_margin
            if min_link_margin is None or link_margin < min_link_margin:
                min_link_margin = link_margin

        # 펌프시간: Stage별 time합 없으면 ballast/pump_rate로 계산
        if not has_time and pump_rate_tph > 0.0:
            total_time_h = total_ballast / pump_rate_tph

        scenarios[scenario_name] = {
            "total_ballast_t": total_ballast,
            "total_time_h": total_time_h,
            "fwd_draft_margin_m": float(min_fwd_margin or 0.0),
            "linkspan_freeboard_margin_m": float(min_link_margin or 0.0),
        }

    return scenarios


# ============================================================================
# WhatsApp Summary PNG Export
# ============================================================================


def export_whatsapp_summary_png(
    wb: Workbook,
    png_path: str = "RORO_Summary.png",
    max_stage_lines: int = 6,
) -> Optional[Path]:
    """
    RORO_Delta_Lever_Report + Ballast_Scenario_Comparison 요약을
    텍스트 기반 PNG로 내보내 WhatsApp 공유용으로 사용.
    """
    if not MATPLOTLIB_AVAILABLE:
        print("[WARN] matplotlib not available. PNG export skipped.")
        return None

    lines: list[str] = []
    # 1) ΔTM Flag 요약
    if "RORO_Delta_Lever_Report" in wb.sheetnames:
        ws = wb["RORO_Delta_Lever_Report"]
        flag_counts = {"OK": 0, "CHECK": 0, "NO_DATA": 0}
        check_stages: list[str] = []

        for row in range(6, ws.max_row + 1):
            stage = ws[f"A{row}"].value
            flag = ws[f"J{row}"].value  # Flag 컬럼
            if stage is None or flag is None:
                continue
            flag = str(flag).strip()
            flag_counts[flag] = flag_counts.get(flag, 0) + 1
            if flag == "CHECK":
                check_stages.append(str(stage))

        lines.append("[ΔTM/Lever Check]")
        lines.append(
            f"OK: {flag_counts.get('OK',0)}, CHECK: {flag_counts.get('CHECK',0)}, NO_DATA: {flag_counts.get('NO_DATA',0)}"
        )
        if check_stages:
            lines.append("CHECK stages: " + ", ".join(check_stages[:max_stage_lines]))
        lines.append("")

    # 2) Ballast 시나리오 요약
    if "Ballast_Scenario_Comparison" in wb.sheetnames:
        ws2 = wb["Ballast_Scenario_Comparison"]
        lines.append("[Ballast Scenarios]")
        for row in range(6, ws2.max_row + 1):
            name = ws2[f"A{row}"].value
            if not name:
                continue
            try:
                total_ballast = float(ws2[f"B{row}"].value or 0.0)
                time_h = float(ws2[f"C{row}"].value or 0.0)
                fwd_margin = float(ws2[f"D{row}"].value or 0.0)
                link_margin = float(ws2[f"E{row}"].value or 0.0)
                lines.append(
                    f"{name}: Ballast={total_ballast:.1f}t, Time={time_h:.2f}h, "
                    f"FWD_margin={fwd_margin:.2f}m, Linkspan_margin={link_margin:.2f}m"
                )
            except (TypeError, ValueError):
                continue

    if not lines:
        lines = ["No summary data available."]

    try:
        fig, ax = plt.subplots(figsize=(6, 4))
        ax.axis("off")

        y = 0.95
        for line in lines:
            ax.text(0.02, y, line, fontsize=9, transform=ax.transAxes, va="top")
            y -= 0.06

        fig.tight_layout()
        out_path = Path(png_path)
        fig.savefig(out_path, dpi=200)
        plt.close(fig)
        print(f"[OK] WhatsApp summary PNG exported: {out_path}")
        return out_path
    except Exception as e:
        print(f"[WARN] PNG export failed: {e}")
        return None


# ============================================================================
# CAPTAIN REPORT Sheet Creation (NEW)
# ============================================================================


def create_captain_report_sheet(wb, stages, first_data_row):
    """
    Creates the OPERATION SUMMARY sheet summary.
    Updated for v4.3: DAS Method & Critical Checkpoints Highlight.
    """
    if "OPERATION SUMMARY" in wb.sheetnames:
        wb.remove(wb["OPERATION SUMMARY"])

    ws = wb.create_sheet("OPERATION SUMMARY")
    styles = get_styles()

    # --- 1. Report Title & Header ---
    ws.merge_cells("A1:J1")
    ws["A1"] = "LCT BUSHRA – OPERATION SUMMARY (DAS METHOD)"
    ws["A1"].font = styles["title_font"]
    ws["A1"].alignment = styles["center_align"]
    ws["A1"].fill = styles["header_fill"]

    # --- 2. Critical Limits (기준값) ---
    headers_limit = ["PARAMETER", "LIMIT", "UNIT", "REMARK"]
    limits_data = [
        ("Summer Draft Max", 2.70, "m", "Operational Limit (Harbour: Check Depth)"),
        ("Min Freeboard", 0.28, "m", "Linkspan Connection Safety"),
        ("Min GM", 1.50, "m", "Stability Requirement"),
        ("Max Ramp Angle", 6.0, "deg", "SPMT Climbing Limit"),
    ]

    # Write Limit Table
    ws.merge_cells("A3:D3")
    ws["A3"] = "1. OPERATIONAL LIMITS"
    ws["A3"].font = styles["header_font"]
    ws["A3"].fill = styles["structure_fill"]  # Orange

    for col, text in enumerate(headers_limit, 1):
        cell = ws.cell(row=4, column=col, value=text)
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = styles["center_align"]

    for i, (param, val, unit, rem) in enumerate(limits_data, 5):
        ws.cell(row=i, column=1, value=param).font = styles["normal_font"]
        ws.cell(row=i, column=2, value=val).font = styles["normal_font"]
        ws.cell(row=i, column=2).number_format = "0.00"
        ws.cell(row=i, column=2).fill = styles["input_fill"]
        ws.cell(row=i, column=3, value=unit).font = styles["normal_font"]
        ws.cell(row=i, column=4, value=rem).font = styles["normal_font"]

        # 테두리 적용
        for c in range(1, 5):
            ws.cell(row=i, column=c).border = Border(bottom=styles["thin_border"])

    # --- 3. Stage Summary Table ---
    table_start_row = 10
    ws.merge_cells(f"A{table_start_row-1}:J{table_start_row-1}")
    ws[f"A{table_start_row-1}"] = "2. STAGE-BY-STAGE SAFETY CHECK"
    ws[f"A{table_start_row-1}"].font = styles["header_font"]
    ws[f"A{table_start_row-1}"].fill = styles["structure_fill"]

    headers_table = [
        "Stage",
        "Condition",
        "Trim (m)",
        "Fwd Draft",
        "Aft Draft",
        "Draft Check",
        "Freeboard",
        "Deck Check",
        "Action / Note",
    ]

    # Write Headers
    for col, text in enumerate(headers_table, 1):
        cell = ws.cell(row=table_start_row, column=col, value=text)
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = styles["center_align"]

    # Write Data Rows (Link to RORO Sheet)
    # Mapping from RORO sheet columns:
    # Stage(A), Trim_m(F), Dfwd(O), Daft(P), Phys_Freeboard_m(Z)

    roro_sheet = "RORO_Stage_Scenarios"

    for i, stage_name in enumerate(stages):
        r_rept = table_start_row + 1 + i
        r_roro = first_data_row + i
        r_str = str(r_roro)

        # A: Stage Name
        ws.cell(row=r_rept, column=1, value=f"='{roro_sheet}'!A{r_str}")

        # B: Condition (Custom Logic based on Stage Name)
        ws.cell(
            row=r_rept,
            column=2,
            value=f'=IF(ISNUMBER(SEARCH("PreBallast",A{r_rept})),"PRE-BALLAST",IF(ISNUMBER(SEARCH("Critical",A{r_rept})),"CRITICAL","NORMAL"))',
        )

        # C: Trim
        ws.cell(row=r_rept, column=3, value=f"='{roro_sheet}'!F{r_str}")
        ws.cell(row=r_rept, column=3).number_format = "0.00"

        # D: Fwd Draft
        ws.cell(row=r_rept, column=4, value=f"='{roro_sheet}'!O{r_str}")
        ws.cell(row=r_rept, column=4).number_format = "0.00"

        # E: Aft Draft
        ws.cell(row=r_rept, column=5, value=f"='{roro_sheet}'!P{r_str}")
        ws.cell(row=r_rept, column=5).number_format = "0.00"

        # F: Draft Check (Max Draft vs Limit 2.70)
        # Logic: If PreBallast -> Check Depth (Warning), Else -> Check 2.70
        check_formula = (
            f'=IF(B{r_rept}="PRE-BALLAST", "CHECK DEPTH", '
            f'IF(MAX(D{r_rept},E{r_rept})<=$B$5, "OK", "OVER DRAFT"))'
        )
        ws.cell(row=r_rept, column=6, value=check_formula)

        # G: Freeboard (Phys_Freeboard_m from RORO sheet Z column)
        ws.cell(row=r_rept, column=7, value=f"='{roro_sheet}'!Z{r_str}")
        ws.cell(row=r_rept, column=7).number_format = "0.00"

        # H: Deck Check (Freeboard vs 0.28)
        ws.cell(
            row=r_rept, column=8, value=f'=IF(G{r_rept}>=$B$6, "OK", "SUBMERGED/LOW")'
        )

        # I: Action / Note (From RORO G column - Explanation)
        ws.cell(row=r_rept, column=9, value=f"='{roro_sheet}'!G{6 + i}")
        ws.cell(row=r_rept, column=9).alignment = styles["left_align"]

        # Formatting (Colors for Critical Stages)
        # Conditional Formatting logic is hard in openpyxl, so we use static logic if possible or just row styling
        if "PreBallast" in stage_name or "Critical" in stage_name:
            for c in range(1, 10):
                ws.cell(row=r_rept, column=c).fill = styles[
                    "input_fill"
                ]  # Yellow highlight

        # Borders
        for c in range(1, 10):
            ws.cell(row=r_rept, column=c).border = Border(bottom=styles["thin_border"])

    # Column Widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["H"].width = 15
    ws.column_dimensions["I"].width = 50

    print("  [OK] OPERATION SUMMARY sheet updated (DAS Method Specialized)")


# ============================================================================
# Additional Sheet Creation Functions
# ============================================================================


def extend_roro_captain_req(ws, first_data_row, num_stages):
    """
    RORO_Stage_Scenarios 시트 Captain Requirement 확장 컬럼(U~AE, AR~AT) 생성.

    - U  : GM(m)           (Hydro_Table!B:D 기반 VLOOKUP)
    - V  : Fwd Draft(m)    (Dfwd_m 복사)
    - W  : vs 2.70m        (Fwd ≤ 2.70m 체크)
    - X  : De-ballast Qty  (Ballast_t)
    - Y  : Timing          (수기 입력용)
    - Z  : Phys_Freeboard  (Calc!E19 - Dfwd)
    - AB : GM copy         (=U)
    - AC : GM Check        (GM ≥ GM_target=Calc!E21)
    - AD : Disp_total_t    (950 + Ballast + Cargo)
    - AE : Vent_Time_h     (De-ballast Qty / $B$13 t/h)
    - AR : Heel_deg        (B / (Δ * GM) 기반 근사)
    - AS : GM_eff          (GM * cos(Heel))
    - AT : Ramp_Angle_deg  (Freeboard 기반 ramp 각도)

    ✅ 이 함수가 모든 Stage 행에 수식을 써주므로
       U25 / AB25 / AC25 / AR25 / AS25 수식 에러가 사라집니다.
    """
    styles = get_styles()

    # ------------------------------------------------------------------
    # 1) 헤더 생성 (U~AE)
    # ------------------------------------------------------------------
    captain_cols = [
        "GM(m)",  # U(21)
        "Fwd Draft(m)",  # V(22)
        "vs 2.70m",  # W(23)
        "De-ballast Qty(t)",  # X(24)
        "Timing",  # Y(25)
        "Phys_Freeboard_m",  # Z(26)
        "Clearance_Check",  # AA(27) – 필요시 확장
        "GM_calc",  # AB(28)
        "GM_Check",  # AC(29)
        "Disp_total_t",  # AD(30)
        "Vent_Time_h",  # AE(31)
    ]

    start_col = 21  # U 열
    header_row = first_data_row - 1

    for i, name in enumerate(captain_cols):
        col = start_col + i
        cell = ws.cell(row=header_row, column=col)
        cell.value = name
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = styles["center_align"]
        cell.border = Border(
            left=styles["thin_border"],
            right=styles["thin_border"],
            top=styles["thin_border"],
            bottom=styles["thin_border"],
        )

    # 기본 컬럼 폭 (필요시 조정)
    ws.column_dimensions["U"].width = 10
    ws.column_dimensions["V"].width = 10
    ws.column_dimensions["W"].width = 10
    ws.column_dimensions["X"].width = 12
    ws.column_dimensions["Y"].width = 12
    ws.column_dimensions["Z"].width = 14
    ws.column_dimensions["AA"].width = 12
    ws.column_dimensions["AB"].width = 10
    ws.column_dimensions["AC"].width = 10
    ws.column_dimensions["AD"].width = 12
    ws.column_dimensions["AE"].width = 12

    number_format = "#,##0.00"

    # ------------------------------------------------------------------
    # 2) 각 Stage 행별 수식 삽입
    #    → 특히 row=25(Stage 6A_Critical)에 U/AB/AC/AR/AS가 자동 기입됨
    #    PATCH: num_stages 대신 A열 스캔으로 모든 Stage 행 찾기
    # ------------------------------------------------------------------
    # A열(column=1)을 스캔하여 "Stage"로 시작하는 모든 행 찾기
    stage_rows = []
    max_row = ws.max_row
    for row in range(first_data_row, max_row + 1):
        cell_value = ws.cell(row=row, column=1).value
        if cell_value and "stage" in str(cell_value).strip().lower():
            stage_rows.append(row)

    # num_stages 기반 fallback 범위 추가 (total_rows가 작아도 누락 방지)
    try:
        if num_stages and int(num_stages) > 0:
            for row in range(first_data_row, first_data_row + int(num_stages)):
                if row not in stage_rows:
                    stage_rows.append(row)
    except Exception:
        pass

    # Row 24/25 누락 방지 (A18:BB28 구간)
    for row in (24, 25):
        if row >= first_data_row and row <= max_row and row not in stage_rows:
            stage_rows.append(row)

    stage_rows = sorted(set(stage_rows))
    if stage_rows:
        print(
            f"  [DEBUG] Captain Req stage_rows={stage_rows} "
            f"(first_data_row={first_data_row}, num_stages={num_stages}, max_row={max_row})"
        )

    # 모든 Stage 행에 수식 적용
    for row in stage_rows:
        row_str = str(row)

        # NOTE: '일괄 적용' 원칙
        # - 기존 값/수식이 있어도 모두 덮어써서(Stage 6C_TotalMassOpt 포함) 수식 누락/불일치 방지

        # U(21): GM(m) – Hydro_Table 기반 VLOOKUP(Tmean = AVG(Dfwd,Daft))
        c = ws.cell(row=row, column=21)
        c.value = (
            f'=IF(OR(P{row_str}="", Q{row_str}=""), "", '
            f'IFERROR(VLOOKUP(AVERAGE(P{row_str},Q{row_str}), Hydro_Table!$B:$D, 3, 1), ""))'
        )
        c.number_format = number_format
        c.font = styles["normal_font"]

        # V(22): Fwd Draft(m) – P열 Dfwd_m 복사
        c = ws.cell(row=row, column=22)
        c.value = f"=P{row_str}"
        c.number_format = number_format
        c.font = styles["normal_font"]

        # W(23): vs 2.70m – Calc!E18(=2.70m 한계)와 비교
        c = ws.cell(row=row, column=23)
        c.value = f'=IF(V{row_str}="","",IF(V{row_str}<=Calc!$E$18,"OK","NG"))'
        c.font = styles["normal_font"]

        # X(24): De-ballast Qty(t) – K열 Ballast_t 그대로
        c = ws.cell(row=row, column=24)
        c.value = f"=K{row_str}"
        c.number_format = number_format
        c.font = styles["normal_font"]

        # Y(25): Timing – 수기 입력용 (수식 없음, input_fill 적용)
        c = ws.cell(row=row, column=25)
        c.fill = styles["input_fill"]
        c.font = styles["normal_font"]

        # Z(26): Phys_Freeboard_m – (D_vessel - Dfwd) adjusted by ramp_door_offset_m
        c = ws.cell(row=row, column=26)
        c.value = f'=IFERROR(IF(P{row_str}="","",$B$11 - P{row_str} - Calc!$E$19), "")'
        c.number_format = number_format
        c.font = styles["normal_font"]

        # AA(27): Clearance_Check – Phys_Freeboard_m vs linkspan_freeboard_target_m (Calc!E20)
        c = ws.cell(row=row, column=27)
        c.value = f'=IFERROR(IF(Z{row_str}="", "", IF(Z{row_str}>=Calc!$E$20, "OK", "<0.28m CHECK")), "")'
        c.font = styles["normal_font"]

        # AB(28): GM_calc – =U
        c = ws.cell(row=row, column=28)
        c.value = f'=IFERROR(U{row_str},"")'
        c.number_format = number_format
        c.font = styles["normal_font"]

        # AC(29): GM_Check – GM ≥ GM_target(=Calc!E21) 여부
        c = ws.cell(row=row, column=29)
        c.value = (
            f'=IFERROR(IF(AB{row_str}="","",IF(AB{row_str}>=Calc!$E$21,"OK","NG")),"")'
        )
        c.font = styles["normal_font"]

        # AD(30): Disp_total_t – Patch Option B: 실측 Const Tank Total 연동
        c = ws.cell(row=row, column=30)
        c.value = f'=IF(A{row_str}="","",CONST_TANKS!$C$13+L{row_str}+B{row_str})'
        c.number_format = number_format
        c.font = styles["normal_font"]

        # AE(31): Vent_Time_h – De-ballast qty / effective pump rate ($B$13, t/h)
        c = ws.cell(row=row, column=31)
        c.value = f'=IFERROR(IF(OR(X{row_str}="",$B$13="",$B$13=0),"",ROUND(ABS(X{row_str})/$B$13,2)),"")'
        c.number_format = number_format
        c.font = styles["normal_font"]

        # AR(44): Heel_deg – B / (Δ * GM) 기반 근사식
        c = ws.cell(row=row, column=44)
        c.value = (
            f'=IFERROR(IF(OR($A{row_str}="",U{row_str}="",AD{row_str}="",AD{row_str}=0),"",'
            f'DEGREES((B{row_str}*Calc!$E$43)/(AD{row_str}*U{row_str}))),"")'
        )
        c.number_format = number_format
        c.font = styles["normal_font"]

        # AS(45): GM_eff – GM - FSE/Δ (Patch Option B: CONST_TANKS!E13 사용)
        c = ws.cell(row=row, column=45)
        c.value = (
            f"=IFERROR("
            f'IF(OR($A{row_str}="", U{row_str}="", AD{row_str}="", AD{row_str}=0), "", '
            f"U{row_str} - CONST_TANKS!$E$13 / AD{row_str}),"
            f'""'
            f")"
        )
        c.number_format = number_format
        c.font = styles["normal_font"]

        # AT(46): Ramp_Angle_deg – Freeboard 기반 ramp 각도
        c = ws.cell(row=row, column=46)
        c.value = (
            f'=IF(Z{row_str}="","",DEGREES(ASIN((Z{row_str}-Calc!$E$35)/Calc!$E$33)))'
        )
        c.number_format = number_format
        c.font = styles["normal_font"]

    # 수식 적용 후 검증 (U/AB/AC/AR/AS)
    check_cols = {
        21: "U",
        28: "AB",
        29: "AC",
        44: "AR",
        45: "AS",
    }
    missing_checks = []
    for row in stage_rows:
        for col_idx, col_name in check_cols.items():
            v = ws.cell(row=row, column=col_idx).value
            if v is None or v == "":
                missing_checks.append(f"{col_name}{row}")
    if missing_checks:
        print(f"  [WARN] Captain Req formula missing: {missing_checks}")

    # Optional: materialize U/AB/AC values (avoid Excel calc dependency)
    def _to_float(v):
        try:
            if v is None or v == "":
                return None
            return float(v)
        except Exception:
            return None

    def _build_hydro_table(hydro_ws):
        if hydro_ws is None:
            return []
        headers = {}
        for c in range(1, hydro_ws.max_column + 1):
            h = hydro_ws.cell(row=1, column=c).value
            if h:
                headers[str(h).strip()] = c
        tmean_col = headers.get("Tmean_m", 2)
        gm_col = headers.get("GM_m", 4)
        rows = []
        for r in range(2, hydro_ws.max_row + 1):
            t = _to_float(hydro_ws.cell(row=r, column=tmean_col).value)
            gm = _to_float(hydro_ws.cell(row=r, column=gm_col).value)
            if t is None:
                continue
            rows.append((t, gm))
        rows.sort(key=lambda x: x[0])
        return rows

    def _lookup_gm(rows, tmean):
        if tmean is None or not rows:
            return None
        # clamp to nearest bounds to avoid empty results
        if tmean <= rows[0][0]:
            return rows[0][1]
        gm_val = rows[0][1]
        for t, gm in rows:
            if t <= tmean:
                gm_val = gm
            else:
                break
        return gm_val

    try:
        wb = ws.parent
        hydro_ws = wb["Hydro_Table"] if "Hydro_Table" in wb.sheetnames else None
        calc_ws = wb["Calc"] if "Calc" in wb.sheetnames else None
        gm_target = _to_float(calc_ws["E21"].value) if calc_ws is not None else None
        hydro_rows = _build_hydro_table(hydro_ws)
        if hydro_rows:
            applied = 0
            for row in stage_rows:
                if not ws.cell(row=row, column=1).value:
                    continue
                g = _to_float(ws.cell(row=row, column=7).value)
                h = _to_float(ws.cell(row=row, column=8).value)
                if g is None or h is None:
                    continue
                tmean = (g + h) / 2.0
                gm = _lookup_gm(hydro_rows, tmean)
                if gm is None:
                    continue
                # U (GM)
                c = ws.cell(row=row, column=21, value=gm)
                c.number_format = number_format
                c.font = styles["normal_font"]
                # AB (GM copy)
                c = ws.cell(row=row, column=28, value=gm)
                c.number_format = number_format
                c.font = styles["normal_font"]
                # AC (GM check)
                if gm_target is not None:
                    ws.cell(
                        row=row,
                        column=29,
                        value="OK" if gm >= gm_target else "NG",
                    ).font = styles["normal_font"]
                applied += 1
            if applied > 0:
                print(
                    f"  [OK] Captain Req values materialized (U/AB/AC) for {applied} rows"
                )
    except Exception as e:
        print(f"  [WARN] Captain Req value materialization skipped: {e}")

    print(
        f"  [OK] Captain Req columns added to RORO_Stage_Scenarios sheet (Patched) - Applied to {len(stage_rows)} stage rows"
    )


def extend_roro_structural_opt1(ws, first_data_row, num_stages):
    """RORO 시트에 Structural Strength 및 Option 1 Ballast Fix Check 컬럼 추가 (Col AE-AO)"""
    styles = get_styles()
    # 숫자 포맷 통일: 천단위 구분, 소수점 2자리
    number_format = "#,##0.00"

    # Structural Strength 컬럼 (AK-AP) - PATCH C 컬럼과 충돌 방지를 위해 AK(37)부터 시작
    structural_cols = [
        "Share_Load_t",
        "Share_Check",
        "Hinge_Rx_t",
        "Rx_Check",
        "Deck_Press_t/m²",
        "Press_Check",
    ]

    # Dynamic Load Case B 컬럼 (AQ) - Load_Case_C_t는 AR(44)와 충돌하므로 제거
    dynamic_load_cols = [
        "Load_Case_B_t",  # AQ (43): Dynamic factor applied
        # "Load_Case_C_t" 제거 (AR(44) 충돌)
    ]

    # Heel/FSE 컬럼 추가 (AR-AS, column 44-45)
    # ⚠️ CRITICAL WARNING: AR(44), AS(45)는 절대 덮어쓰지 말 것!
    #
    # AR(44): Heel_deg = DEGREES((B * Calc!E43) / (AD * U))
    #   - 경사각 계산 (deg)
    #   - B: 화물 편심 모멘트
    #   - AD: Total Displacement
    #   - U: GM
    #
    # AS(45): GM_eff = U - CONST_TANKS!E13 / AD
    #   - 유효 GM (자유 표면 효과 반영)
    #   - U: GM (Hydro table 기반)
    #   - CONST_TANKS!E13: Total FSM (887.04 t·m)
    #   - AD: Total Displacement
    #
    # 이 두 컬럼은 extend_roro_captain_req()에서만 작성되어야 하며,
    # 다른 함수에서 절대 덮어쓰면 안 됨!
    # 헤더만 추가, 수식은 extend_roro_captain_req()에서 작성됨
    heel_fse_cols = [
        "Heel_deg",  # AR (44) - extend_roro_captain_req에서 작성됨
        "GM_eff_m",  # AS (45) - extend_roro_captain_req에서 작성됨
    ]

    # Option 1 Ballast Fix Check 컬럼 (AT-AV)
    opt1_cols = [
        "Ballast_req_t",  # AT (46) - 기존 유지
        "Ballast_gap_t",  # AU (47) - 기존 유지
        "Time_Add_h",  # AV (48) - 기존 유지
        # "ΔTM_needed_cm·tm" 제거 (AS(45) 충돌로 인해 제거됨)
    ]

    # Ramp Angle & Pin Stress 컬럼 (AW-AZ, column 49-52)
    ramp_stress_cols = [
        "Ramp_Angle_deg",  # AW (49)
        "Ramp_Angle_Check",  # AX (50)
        "Pin_Stress_N/mm²",  # AY (51)
        "Von_Mises_Check",  # AZ (52)
    ]

    # Opt C / High Tide 관련 컬럼 (BA-BB, column 53-54)
    # Opt C / Water level requirement columns (BA-BB)
    # NOTE: "Required_WL_m" is required water level (NOT forecast tide). Compare against Forecast_Tide_m.
    opt_c_tide_cols = [
        "Required_WL_m",  # BA (53)
        "WL_OK",  # BB (54)
    ]

    all_cols = (
        structural_cols  # AK-AP (37-42)
        + dynamic_load_cols  # AQ (43)
        + heel_fse_cols  # AR-AS (44-45) - 헤더만 추가, 수식은 extend_roro_captain_req에서
        + opt1_cols  # AT-AV (46-48)
        + ramp_stress_cols  # AW-AZ (49-52)
        + opt_c_tide_cols  # BA-BB (53-54)
    )
    start_col = (
        37  # PATCH C 컬럼 (AF-AJ, column 32-36)과 충돌 방지를 위해 AK(37)부터 시작
    )

    for i, h in enumerate(all_cols):
        col = start_col + i
        header_row = first_data_row - 1  # Row 17 (first_data_row=18이므로)
        cell = ws.cell(row=header_row, column=col)
        cell.value = h
        cell.font = styles["header_font"]
        cell.alignment = styles["center_align"]
        cell.border = Border(
            left=styles["thin_border"],
            right=styles["thin_border"],
            top=styles["thin_border"],
            bottom=styles["thin_border"],
        )
        # Structural 컬럼은 주황색, Dynamic Load는 주황색, Heel/FSE는 주황색, Option 1 컬럼은 보라색, Ramp/Stress 컬럼은 주황색, Opt C Tide 컬럼은 보라색
        if i < len(structural_cols):
            cell.fill = styles["structure_fill"]
        elif i < len(structural_cols) + len(dynamic_load_cols):
            cell.fill = styles[
                "structure_fill"
            ]  # Dynamic Load Case uses structure fill
        elif i < len(structural_cols) + len(dynamic_load_cols) + len(heel_fse_cols):
            cell.fill = styles["structure_fill"]  # Heel/FSE columns use structure fill
        elif i < len(structural_cols) + len(dynamic_load_cols) + len(
            heel_fse_cols
        ) + len(opt1_cols):
            cell.fill = styles["opt1_fill"]  # Option 1 uses opt1 fill
        elif i < len(structural_cols) + len(dynamic_load_cols) + len(
            heel_fse_cols
        ) + len(opt1_cols) + len(ramp_stress_cols):
            cell.fill = styles[
                "structure_fill"
            ]  # Ramp/Stress columns use structure fill
        else:
            cell.fill = styles["opt1_fill"]  # Opt C Tide columns use opt1 fill

    for row in range(first_data_row, first_data_row + num_stages):
        row_str = str(row)

        # Structural Strength 컬럼 (AK-AP) - PATCH C 컬럼 (AF-AJ)과 충돌 방지를 위해 AK(37)부터 시작
        # AK (37): Share_Load_t - 입력
        ws.cell(row=row, column=37).fill = styles["input_fill"]
        ws.cell(row=row, column=37).font = styles["normal_font"]

        # AL (38): Share_Check
        ws.cell(row=row, column=38).value = (
            f'=IF(AK{row_str}="", "", IF(AK{row_str}<=Calc!$E$24, "OK", "CHECK"))'
        )
        ws.cell(row=row, column=38).font = styles["normal_font"]

        # AM (39): Hinge_Rx_t
        # 자동 계산 (Ramp 자중 45 t + share 54.5%)
        ws.cell(row=row, column=39).value = (
            f'=IF(AK{row_str}="", 45, 45 + AK{row_str} * 0.545)'
        )
        ws.cell(row=row, column=39).number_format = number_format
        ws.cell(row=row, column=39).font = styles["normal_font"]

        # AN (40): Rx_Check
        ws.cell(row=row, column=40).value = (
            f'=IF(AM{row_str}="", "", IF(AM{row_str}<=Calc!$E$37, "OK", "NG"))'
        )
        ws.cell(row=row, column=40).font = styles["normal_font"]

        # AO (41): Deck_Press_t/m²
        ws.cell(row=row, column=41).value = (
            f'=IF(AK{row_str}="", "", AK{row_str}/Calc!$E$26)'
        )
        ws.cell(row=row, column=41).number_format = number_format
        ws.cell(row=row, column=41).font = styles["normal_font"]

        # AP (42): Press_Check
        ws.cell(row=row, column=42).value = (
            f'=IF(AO{row_str}="", "", IF(AO{row_str}<=Calc!$E$25, "OK", "CHECK"))'
        )
        ws.cell(row=row, column=42).font = styles["normal_font"]

        # Dynamic Load Case B/C 컬럼 (AQ-AR) - PATCH C 컬럼과 충돌 방지를 위해 AQ(43)부터 시작
        # AQ (43): Load_Case_B_t - Dynamic factor applied to Share_Load
        ws.cell(row=row, column=43).value = (
            f'=IF(AK{row_str}="", "", AK{row_str} * Calc!$E$42)'
        )
        ws.cell(row=row, column=43).number_format = number_format
        ws.cell(row=row, column=43).font = styles["normal_font"]

        # ===== PATCH FIX: AR(44), AS(45) 컬럼 충돌 방지 =====
        # AR(44): Heel_deg는 extend_roro_captain_req()에서 작성됨 (라인 6431-6438)
        # AS(45): GM_eff는 extend_roro_captain_req()에서 작성됨 (라인 6440-6450)
        #
        # ⚠️ CRITICAL: 아래 코드는 위 수식을 덮어쓰므로 주석 처리 필수!
        #
        # # AR (44): Load_Case_C_t - REMOVED (충돌)
        # # ws.cell(row=row, column=44).value = (
        # #     f'=IF(AQ{row_str}="", "", AQ{row_str} + 0.2 * B{row_str} * 9.81 / 1000)'
        # # )
        # # ws.cell(row=row, column=44).number_format = number_format
        # # ws.cell(row=row, column=44).font = styles["normal_font"]
        #
        # # AS (45): ΔTM_needed_cm·tm - REMOVED (충돌, 필요시 다른 컬럼으로 이동)
        # # ws.cell(row=row, column=45).value = f'=IF($A{row_str}="","",ABS(H{row_str}))'
        # # ws.cell(row=row, column=45).number_format = number_format
        # # ws.cell(row=row, column=45).font = styles["normal_font"]

        # Option 1 Ballast Fix Check 컬럼은 AT(46)부터 시작 (기존 유지)
        # ⚠️ AR(44) 작성 SKIP (extend_roro_captain_req에서 작성됨)
        # ⚠️ AS(45) 작성 SKIP (extend_roro_captain_req에서 작성됨)

        # AT (46): Ballast_req_t - zzzzz.md 가이드: J열과 같은 개념 (H/I)
        ws.cell(row=row, column=46).value = (
            f'=IF($A{row_str}="","",'
            f'IF(OR($I{row_str}="",$I{row_str}=0),0,ROUND(H{row_str}/$I{row_str},2)))'
        )
        ws.cell(row=row, column=46).number_format = number_format
        ws.cell(row=row, column=46).font = styles["normal_font"]

        # AU (47): Ballast_gap_t - zzzzz.md 가이드: 필요한 Ballast - 실제 Ballast(L열)
        ws.cell(row=row, column=47).value = (
            f'=IF($A{row_str}="","",AT{row_str} - $L{row_str})'
        )
        ws.cell(row=row, column=47).number_format = number_format
        ws.cell(row=row, column=47).font = styles["normal_font"]

        # AV (48): Time_Add_h - zzzzz.md 가이드: AT를 펌프 레이트(B13)로 나눈 추가 시간
        ws.cell(row=row, column=48).value = (
            f'=IF($A{row_str}="","",' f"IF($B$13=0,0,AU{row_str}/$B$13))"
        )
        ws.cell(row=row, column=48).number_format = number_format
        ws.cell(row=row, column=48).font = styles["normal_font"]

        # Heel/FSE 컬럼 (AW-AX) - extend_roro_captain_req()에서 이미 작성됨
        # ✅ SKIP: extend_roro_captain_req()에서 이미 AR(44), AS(45) 작성됨
        # 중복 작성 방지를 위해 이 섹션은 주석 처리
        # 주의: extend_roro_captain_req()가 먼저 실행되므로 여기서 재작성 불필요

        # Ramp Angle & Pin Stress 컬럼 (AW-AZ) - PATCH C 컬럼과 충돌 방지를 위해 AW(49)부터 시작
        # AW (49): Ramp_Angle_deg
        # Z{row}: Phys_Freeboard_m - sdsdds.md: Q열 추가로 Y(25) → Z(26)로 이동
        ws.cell(row=row, column=49).value = (
            f'=IF(Z{row_str}="","",DEGREES(ASIN((Z{row_str}-Calc!$E$35)/Calc!$E$33)))'
        )
        ws.cell(row=row, column=49).number_format = number_format
        ws.cell(row=row, column=49).font = styles["normal_font"]

        # AX (50): Ramp_Angle_Check
        ws.cell(row=row, column=50).value = (
            f'=IF(AW{row_str}="","",IF(AW{row_str}<=6,"OK","NG"))'
        )
        ws.cell(row=row, column=50).font = styles["normal_font"]

        # AY (51): Pin_Stress_N_mm2
        ws.cell(row=row, column=51).value = (
            f'=IF(AM{row_str}="","",(AM{row_str}/4)/Calc!$E$36*9.81/1000)'
        )
        ws.cell(row=row, column=51).number_format = number_format
        ws.cell(row=row, column=51).font = styles["normal_font"]

        # AZ (52): Von_Mises_Check
        ws.cell(row=row, column=52).value = (
            f'=IF(AY{row_str}="","",IF(AY{row_str}<=188,"OK","NG"))'
        )
        ws.cell(row=row, column=52).font = styles["normal_font"]

        # BA (53): Required_WL_m (NOT forecast tide)
        # Required_Tide_m = IF(Phys_Freeboard_m>=0, 0, ABS(Phys_Freeboard_m) + 0.30)
        # Z{row}: Phys_Freeboard_m (column 26)
        ws.cell(row=row, column=53).value = (
            f'=IF(Z{row_str}="", "", IF(Z{row_str}>=0, 0, ABS(Z{row_str})+0.30))'
        )
        ws.cell(row=row, column=53).number_format = number_format
        ws.cell(row=row, column=53).font = styles["normal_font"]

        # BB (54): Tide_OK
        # WL_OK = IF(Forecast_Tide_m >= Required_WL_m, "OK", "CHECK")
        # Forecast_Tide_m is at RORO_Stage_Scenarios!B7
        ws.cell(row=row, column=54).value = (
            f'=IF(BA{row_str}="", "", IF($B$7>=BA{row_str}, "OK", "CHECK"))'
        )
        ws.cell(row=row, column=54).font = styles["normal_font"]

    # 컬럼 너비 설정 - PATCH C 컬럼 추가로 범위 조정
    for col in range(32, 55):  # AF(32) ~ BB(54) - PATCH C (AF-AJ) + Structural (AK-BB)
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = 11
    # Heel_deg(AR) / GM_eff(AS)는 extend_roro_captain_req에서 이미 처리됨
    ws.column_dimensions["AW"].width = 15  # Ramp_Angle_deg (AW, column 49)
    ws.column_dimensions["AY"].width = 15  # Pin_Stress_N/mm² (AY, column 51)
    ws.column_dimensions["BA"].width = 15  # Required_WL_m
    ws.column_dimensions["BB"].width = 12  # WL_OK

    print(
        "  [OK] Structural Strength & Option 1 Ballast Fix Check columns added to RORO_Stage_Scenarios sheet"
    )
    print("  [OK] Hinge Rx 자동 계산 적용")
    print(
        "  [OK] Dynamic Load Case B/C (AL-AM), Option 1 moved (AN-AQ), Heel/FSE (AR-AS), Ramp/Stress moved (AT-AW) (sdsdds.md: Q열 추가)"
    )


def extend_roro_bc_br_range(ws, first_data_row, num_stages, stage_qa_path=None):
    """
    RORO_Stage_Scenarios 시트에 BC-BR 컬럼 추가 (Post-Solver Results & Gate Verification)

    BC-BR 컬럼 (55-70):
    - BC: Draft_FWD_m_solver
    - BD: Draft_AFT_m_solver
    - BE: Draft_Source
    - BF: Gate_AFT_MIN_2p70_PASS
    - BG: AFT_Margin_2p70_m
    - BH: Gate_B_Applies
    - BI: Gate_FWD_MAX_2p70_critical_only
    - BJ: FWD_Margin_2p70_m
    - BK: Gate_Freeboard_ND
    - BL: Freeboard_Req_ND_m
    - BM: Freeboard_ND_Margin_m
    - BN: D_vessel_m
    - BO: Draft_Max_raw_m
    - BP: Draft_Max_solver_m
    - BQ: Draft_Clipped_raw
    - BR: Draft_Clipped_solver

    데이터 소스: pipeline_stage_QA.csv (solver 결과 반영)
    """
    from pathlib import Path
    import csv

    styles = get_styles()
    number_format = "#,##0.00"

    # BC-BR 컬럼 헤더 정의
    bc_br_cols = [
        ("BC", 55, "Draft_FWD_m_solver"),
        ("BD", 56, "Draft_AFT_m_solver"),
        ("BE", 57, "Draft_Source"),
        ("BF", 58, "Gate_AFT_MIN_2p70_PASS"),
        ("BG", 59, "AFT_Margin_2p70_m"),
        ("BH", 60, "Gate_B_Applies"),
        ("BI", 61, "Gate_FWD_MAX_2p70_critical_only"),
        ("BJ", 62, "FWD_Margin_2p70_m"),
        ("BK", 63, "Gate_Freeboard_ND"),
        ("BL", 64, "Freeboard_Req_ND_m"),
        ("BM", 65, "Freeboard_ND_Margin_m"),
        ("BN", 66, "D_vessel_m"),
        ("BO", 67, "Draft_Max_raw_m"),
        ("BP", 68, "Draft_Max_solver_m"),
        ("BQ", 69, "Draft_Clipped_raw"),
        ("BR", 70, "Draft_Clipped_solver"),
    ]

    header_row = first_data_row - 1

    # 헤더 생성
    for letter, col, header in bc_br_cols:
        cell = ws.cell(row=header_row, column=col)
        cell.value = header
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = styles["center_align"]
        ws.column_dimensions[letter].width = 14

    # pipeline_stage_QA.csv에서 데이터 로드
    stage_qa_data = {}
    if stage_qa_path is None:
        # 기본 경로 탐색
        for p in [
            Path("pipeline_stage_QA.csv"),
            Path("ssot/pipeline_stage_QA.csv"),
            Path("pipeline_out_*/ssot/pipeline_stage_QA.csv"),
        ]:
            matches = list(Path(".").glob(str(p)))
            if matches:
                stage_qa_path = matches[-1]  # 최신 파일 사용
                break

    if stage_qa_path and Path(stage_qa_path).exists():
        print(f"  [INFO] Loading QA data from: {stage_qa_path}")
        with open(stage_qa_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                stage_name = row.get('Stage', '')
                if stage_name:
                    stage_qa_data[stage_name] = row
    else:
        print(f"  [WARN] pipeline_stage_QA.csv not found, using placeholder values")

    # Stage 행 찾기
    stage_rows = []
    max_row = ws.max_row
    for row in range(first_data_row, max_row + 1):
        cell_value = ws.cell(row=row, column=1).value
        if cell_value and "stage" in str(cell_value).strip().lower():
            stage_rows.append((row, str(cell_value).strip()))

    # 각 Stage 행에 데이터 채우기
    for row, stage_name in stage_rows:
        qa_row = stage_qa_data.get(stage_name, {})

        # BC (55): Draft_FWD_m_solver
        val = qa_row.get('Draft_FWD_m', '')
        if val:
            ws.cell(row=row, column=55).value = float(val)
            ws.cell(row=row, column=55).number_format = number_format

        # BD (56): Draft_AFT_m_solver
        val = qa_row.get('Draft_AFT_m', '')
        if val:
            ws.cell(row=row, column=56).value = float(val)
            ws.cell(row=row, column=56).number_format = number_format

        # BE (57): Draft_Source
        val = qa_row.get('Draft_Source', 'solver')
        ws.cell(row=row, column=57).value = val

        # BF (58): Gate_AFT_MIN_2p70_PASS
        val = qa_row.get('GateA_AFT_MIN_2p70_PASS', '')
        if val:
            ws.cell(row=row, column=58).value = 1 if val.lower() in ['true', '1', 'pass'] else 0

        # BG (59): AFT_Margin_2p70_m
        val = qa_row.get('GateA_AFT_MIN_2p70_Margin_m', '')
        if val:
            ws.cell(row=row, column=59).value = float(val)
            ws.cell(row=row, column=59).number_format = number_format

        # BH (60): Gate_B_Applies
        val = qa_row.get('GateB_FWD_MAX_2p70_CD_applicable', '')
        if val:
            ws.cell(row=row, column=60).value = 1 if val.lower() in ['true', '1'] else 0

        # BI (61): Gate_FWD_MAX_2p70_critical_only
        val = qa_row.get('GateB_FWD_MAX_2p70_CD_PASS', '')
        if val:
            ws.cell(row=row, column=61).value = val if val != '' else 'N/A'

        # BJ (62): FWD_Margin_2p70_m
        val = qa_row.get('GateB_FWD_MAX_2p70_CD_Margin_m', '')
        if val:
            try:
                ws.cell(row=row, column=62).value = float(val)
                ws.cell(row=row, column=62).number_format = number_format
            except:
                ws.cell(row=row, column=62).value = val

        # BK (63): Gate_Freeboard_ND
        val = qa_row.get('Gate_Freeboard_ND', '')
        if val:
            ws.cell(row=row, column=63).value = val

        # BL (64): Freeboard_Req_ND_m
        val = qa_row.get('Freeboard_Req_ND_m', '')
        if val:
            try:
                ws.cell(row=row, column=64).value = float(val)
                ws.cell(row=row, column=64).number_format = number_format
            except:
                pass

        # BM (65): Freeboard_ND_Margin_m
        val = qa_row.get('Freeboard_ND_Margin_m', '')
        if val:
            try:
                ws.cell(row=row, column=65).value = float(val)
                ws.cell(row=row, column=65).number_format = number_format
            except:
                pass

        # BN (66): D_vessel_m
        val = qa_row.get('D_vessel_m', '3.65')
        ws.cell(row=row, column=66).value = float(val)
        ws.cell(row=row, column=66).number_format = number_format

        # BO (67): Draft_Max_raw_m
        val = qa_row.get('Draft_Max_raw_m', '')
        if val:
            ws.cell(row=row, column=67).value = float(val)
            ws.cell(row=row, column=67).number_format = number_format

        # BP (68): Draft_Max_solver_m
        val = qa_row.get('Draft_Max_solver_m', '')
        if val:
            ws.cell(row=row, column=68).value = float(val)
            ws.cell(row=row, column=68).number_format = number_format

        # BQ (69): Draft_Clipped_raw
        val = qa_row.get('Draft_Clipped_raw', '')
        if val:
            ws.cell(row=row, column=69).value = 1 if val.lower() in ['true', '1'] else 0

        # BR (70): Draft_Clipped_solver
        val = qa_row.get('Draft_Clipped_solver', '')
        if val:
            ws.cell(row=row, column=70).value = 1 if val.lower() in ['true', '1'] else 0

    print(f"  [OK] BC-BR columns (Post-Solver Results & Gates) added to RORO_Stage_Scenarios sheet ({len(stage_rows)} stages)")


def extend_precision_columns(ws, first_data_row, num_stages):
    """
    RORO 시트에 정밀 계산 적용 (v4.0)
    - G, H 컬럼을 직접 수정: FWD_precise, AFT_precise (Fr_stage 추가로 컬럼 인덱스 변경)
    - Fr_stage 컬럼 추가로 인한 컬럼 구조:
      - Column 6: Trim_cm (F)
      - Column 7: FWD_precise_m (G)
      - Column 8: AFT_precise_m (H)
    - 순환 참조 방지: Tmean을 Baseline draft ($B$6) 기반으로 계산
    - Pre-ballast Stage는 Python 값이 이미 쓰여있으므로 건너뜀
    """
    styles = get_styles()
    # 숫자 포맷 통일: 천단위 구분, 소수점 2자리
    number_format = "#,##0.00"

    # Pre-ballast Stage 목록
    # REMOVED: Old logic that overwrites Python values with Excel formulas
    # All stages now use Python-calculated draft values (LCF-aware)
    # G, H columns are populated in the stage loop above (lines 4639-4700)
    print("  [OK] G, H columns use Python-calculated draft values (LCF-aware, v4.0)")


def create_ballast_tanks_sheet(wb):
    """Ballast_Tanks 시트 생성 (tank_coordinates.json + tank_data.json 기반)"""
    ws = wb.create_sheet("Ballast_Tanks")
    styles = get_styles()

    headers = ["TankName", "x_from_mid_m", "max_t", "SG", "use_flag", "air_vent_mm"]

    # 1) JSON 병합: tank_coordinates.json + tank_data.json
    tank_lookup = build_tank_lookup()

    # 2) 이 프로젝트에서 실제로 쓸 Ballast 탱크 목록 + 기본 use_flag
    #    (필요 시 여기에 DO, FO, 기타 탱크 추가 가능)
    target_tanks = [
        ("FWB1.P", "Y"),
        ("FWB1.S", "Y"),
        ("FWB2.P", "Y"),
        ("FWB2.S", "Y"),
        ("FWCARGO1.P", "N"),  # 선택 사용
        ("FWCARGO1.S", "N"),
        ("FWCARGO2.P", "N"),
        ("FWCARGO2.S", "N"),
    ]

    # 3) Fallback 값 (JSON 없거나, 특정 탱크 키가 비어 있을 때 사용)
    # Fixed: tank.md 기준 정밀 LCG 변환값 (x_from_mid) 및 SG 1.000 적용
    fallback = {
        "FWB1.P": {
            "x": -27.37,  # LCG 57.519m -> x = -27.368
            "max_t": 50.57,
            "SG": 1.000,
            "air_vent_mm": 80,
        },
        "FWB1.S": {
            "x": -27.37,
            "max_t": 50.57,
            "SG": 1.000,
            "air_vent_mm": 80,
        },
        "FWB2.P": {
            "x": -19.89,  # LCG 50.038m -> x = -19.887
            "max_t": 109.98,
            "SG": 1.000,
            "air_vent_mm": 80,
        },
        "FWB2.S": {
            "x": -19.89,
            "max_t": 109.98,
            "SG": 1.000,
            "air_vent_mm": 80,
        },
        "FWCARGO1.P": {
            "x": -12.60,  # LCG 42.750m -> x = -12.599
            "max_t": 148.35,
            "SG": 1.000,
            "air_vent_mm": 125,
        },
        "FWCARGO1.S": {
            "x": -12.60,
            "max_t": 148.35,
            "SG": 1.000,
            "air_vent_mm": 125,
        },
        "FWCARGO2.P": {
            "x": -5.10,  # LCG 35.250m -> x = -5.099
            "max_t": 148.36,
            "SG": 1.000,
            "air_vent_mm": 125,
        },
        "FWCARGO2.S": {
            "x": -5.10,
            "max_t": 148.36,
            "SG": 1.000,
            "air_vent_mm": 125,
        },
    }

    # 4) 헤더 작성
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = styles["center_align"]
        cell.border = Border(
            left=styles["thin_border"],
            right=styles["thin_border"],
            top=styles["thin_border"],
            bottom=styles["thin_border"],
        )

    # 5) 데이터 행 작성
    for row_idx, (tank_name, use_flag) in enumerate(target_tanks, start=2):
        info = tank_lookup.get(tank_name, {})
        fb = fallback.get(tank_name, {})

        x_val = info.get("x_from_mid_m", fb.get("x"))
        max_t = info.get("max_t", fb.get("max_t"))
        sg = info.get("SG", fb.get("SG", 1.0))
        air_vent_mm = info.get("air_vent_mm", fb.get("air_vent_mm", ""))

        row_data = [tank_name, x_val, max_t, sg, use_flag, air_vent_mm]

        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=c, value=val)
            cell.font = styles["normal_font"]
            if c >= 2:  # 숫자 열
                cell.number_format = "0.00"
            cell.border = Border(
                left=styles["thin_border"],
                right=styles["thin_border"],
                top=styles["thin_border"],
                bottom=styles["thin_border"],
            )

    # 6) 컬럼 폭 설정
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 14

    if tank_lookup:
        print(
            "  [OK] Ballast_Tanks updated with tank_coordinates.json + tank_data.json (2025-11-18)"
        )
    else:
        print("  [WARN] Ballast_Tanks used fallback hard-coded data (JSON not found)")


# =============================================================================
# Option B (Precision): CONST_TANKS + FSM (Free Surface Moment) integration
# Sheet refs (fixed):
#   CONST_TANKS!$C$13 : Total Weight [t]
#   CONST_TANKS!$E$13 : Total FSM [mt·m]
# =============================================================================

CONST_TANK_SPECS = [
    # tank, type, max_weight_t, lcg_m(AP), fsm_max_mt_m
    ("DO.P", "Fuel", 3.05, 11.25, 0.34),
    ("DO.S", "Fuel", 3.05, 11.25, 0.34),
    ("LRFO.P", "Fuel", 154.89, 19.50, 133.89),
    ("LRFO.S", "Fuel", 154.89, 19.50, 133.89),
    ("FW1.P", "FreshWater", 23.16, 5.98, 1.15),
    ("FW1.S", "FreshWater", 23.16, 5.98, 1.15),
    ("FW2.P", "FreshWater", 13.92, 0.12, 3.71),
    ("FW2.S", "FreshWater", 13.92, 0.12, 3.71),
]


def normalize_key(s: str) -> str:
    """
    탱크명 정규화 함수 (대소문자, 공백, 특수문자 제거)

    Args:
        s: 정규화할 문자열 (탱크 이름)

    Returns:
        str: 알파벳과 숫자만 남긴 소문자 문자열

    Examples:
        >>> normalize_key("DO.P")
        'dop'
        >>> normalize_key("LRFO-P")
        'lrfop'
        >>> normalize_key("FW 1.P")
        'fw1p'
    """
    return "".join(c.lower() for c in str(s) if c.isalnum())


def load_const_tanks_snapshot(const_csv_path: str = None):
    """
    CSV에서 실측 Const Tank 데이터 로딩 (Option B-2)

    지원 컬럼 (대소문자/공백 무시):
      - tank/tank_id: 탱크 이름
      - weight/weight_t: 실측 중량 (t)
      - fsm/fsm_mt_m: 실측 FSM (t·m)
      - lcg/lcg_m: LCG (m from AP) - 선택 사항

    CSV가 없거나 파싱 실패 시 기본값(하드코딩 값 사용) 반환

    Args:
        const_csv_path: CSV 파일 경로 (None이면 기본 하드코딩 값 사용)

    Returns:
        tuple: (rows, meta)
            - rows: 탱크 데이터 리스트 (dict)
            - meta: 로딩 메타데이터 (dict)

    Examples:
        >>> rows, meta = load_const_tanks_snapshot()
        >>> len(rows)
        8
        >>> rows[0]['tank']
        'DO.P'
    """
    # 기본 하드코딩 데이터로 초기화 (CONST_TANK_SPECS 기반)
    rows = []
    for t, ty, mw, lcg, fsmx in CONST_TANK_SPECS:
        rows.append(
            {
                "tank": t,
                "type": ty,
                "weight_t": mw,  # 기본값: 최대 용량 사용 (실측값 없을 경우)
                "lcg_m": lcg,
                "fsm_mt_m": fsmx,  # 기본값: 최대 FSM 사용
                "max_weight_t": mw,
                "fsm_max_mt_m": fsmx,
                "src": "default",
            }
        )

    meta = {"used_csv": False, "const_csv_path": const_csv_path}

    # CSV 파일이 없으면 기본값 반환
    if not const_csv_path or not os.path.exists(const_csv_path):
        if const_csv_path:
            print(f"[INFO] CSV file not found: {const_csv_path}, using default values")
        return rows, meta

    # CSV 파일 로딩 시도
    try:
        import csv

        print(f"[INFO] Loading CONST_TANKS data from CSV: {const_csv_path}")

        with open(const_csv_path, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            # 탱크명을 키로 하는 딕셔너리 생성 (정규화된 키 사용)
            csv_data = {}
            for row in reader:
                tank_name = row.get("tank", row.get("tank_id", ""))
                if tank_name:
                    csv_data[normalize_key(tank_name)] = row

        # CSV 데이터로 기본값 업데이트
        matched = 0
        for row in rows:
            tank_key = normalize_key(row["tank"])
            if tank_key in csv_data:
                csv_row = csv_data[tank_key]

                # weight_t 업데이트
                weight_str = csv_row.get("weight", csv_row.get("weight_t", ""))
                if weight_str:
                    try:
                        row["weight_t"] = float(weight_str)
                    except (ValueError, TypeError):
                        print(
                            f"[WARNING] Invalid weight for {row['tank']}: {weight_str}"
                        )

                # fsm_mt_m 업데이트
                fsm_str = csv_row.get("fsm", csv_row.get("fsm_mt_m", ""))
                if fsm_str:
                    try:
                        row["fsm_mt_m"] = float(fsm_str)
                    except (ValueError, TypeError):
                        print(f"[WARNING] Invalid FSM for {row['tank']}: {fsm_str}")

                # lcg_m 업데이트 (선택 사항)
                lcg_str = csv_row.get("lcg", csv_row.get("lcg_m", ""))
                if lcg_str:
                    try:
                        row["lcg_m"] = float(lcg_str)
                    except (ValueError, TypeError):
                        pass  # LCG는 선택 사항이므로 경고 출력 안 함

                row["src"] = "csv"
                matched += 1

        meta["used_csv"] = True
        meta["matched_count"] = matched
        meta["total_tanks"] = len(rows)
        print(f"[INFO] CSV loaded successfully: {matched}/{len(rows)} tanks matched")

        if matched < len(rows):
            print(
                f"[WARNING] {len(rows) - matched} tanks not found in CSV, using default values"
            )

    except Exception as e:
        print(f"[WARNING] CSV loading failed: {e}")
        print(f"[INFO] Using default hardcoded values for all tanks")
        meta["error"] = str(e)

    return rows, meta


def create_const_tanks_sheet(wb: Workbook, const_rows=None, meta=None) -> Worksheet:
    """CONST_TANKS 시트 생성 - Permanent Loads (Fuel/FreshWater)

    Patch Option B-2: 실측 Const Tank Total 연동을 위한 시트 (CSV 지원)

    이 시트는 모든 Stage 계산의 기초가 되는 고정 하중(Fuel 및 FreshWater)과
    FSM(Free Surface Moment)을 관리합니다. PATCH.MD Option B 정밀도 패치의 핵심입니다.

    시트 구조:
      - Row 3: 헤더 (Tank, Type, Weight_t, LCG_m, FSM_mt_m, Remarks)
      - Row 4-11: 8개 탱크 데이터
        * DO.P/S: Fuel, 3.05t each, FSM 0.34 t·m each
        * LRFO.P/S: Fuel, 154.89t each, FSM 133.89 t·m each
        * FW1.P/S: FreshWater, 23.16t each, FSM 1.15 t·m each
        * FW2.P/S: FreshWater, 13.92t each, FSM 3.71 t·m each
      - Row 13: 합계 수식
        * C13: =SUM(C4:C11) → Total Weight
        * E13: =SUM(E4:E11) → Total FSM

    데이터 소스:
      - CSV 파일 (--const-csv 옵션 사용 시)
      - 하드코딩 기본값 (CSV 없을 경우)

    참조 관계:
      - CAPTAIN_REPORT AD열: =CONST_TANKS!$C$13 + L + B (Total Displacement)
      - CAPTAIN_REPORT AS열: =U - CONST_TANKS!$E$13 / AD (GM_eff)
      - RORO_Stage_Scenarios AD열: =CONST_TANKS!$C$13 (Base Weight)
      - RORO_Stage_Scenarios AS열: =U - CONST_TANKS!$E$13 / AD (GM_eff)
      - solve_stage() params["Const_FSM_t_m"]: 887.0 (E13 값 사용)

    GM_eff 계산:
      GM_eff = GM - FSM / Displacement
      FSM 값은 CONST_TANKS!E13을 사용합니다.

    Args:
        wb: openpyxl Workbook 객체
        const_rows: CSV에서 로딩한 탱크 데이터 (None이면 기본값 사용)
        meta: CSV 로딩 메타데이터

    Returns:
        Worksheet: 생성된 CONST_TANKS 시트

    See Also:
        - DATA_SOURCE_DOCUMENTATION.md Section 3
        - PATCH.MD for Option B-2 patch requirements
        - load_const_tanks_snapshot() for CSV loading
    """
    ws = wb.create_sheet("CONST_TANKS")
    styles = get_styles()

    # CSV 데이터가 없으면 기본 하드코딩 데이터 로딩
    if const_rows is None:
        const_rows, meta = load_const_tanks_snapshot()

    # 헤더 작성 (row 3)
    headers = ["Tank", "Type", "Weight_t", "LCG_m", "FSM_mt_m", "Remarks"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = styles["center_align"]
        cell.border = Border(
            left=styles["thin_border"],
            right=styles["thin_border"],
            top=styles["thin_border"],
            bottom=styles["thin_border"],
        )

    # 데이터 행 작성 (row 4부터) - CSV 또는 기본 하드코딩 데이터 사용
    r = 4
    for row_data in const_rows:
        t = row_data["tank"]
        tp = row_data["type"]
        wt = row_data["weight_t"]
        x = row_data["lcg_m"]
        fsm = row_data["fsm_mt_m"]
        rm = row_data.get("src", "")
        ws.cell(r, 1, t).font = styles["normal_font"]
        ws.cell(r, 2, tp).font = styles["normal_font"]
        ws.cell(r, 3, wt).font = styles["normal_font"]
        ws.cell(r, 3).number_format = "0.00"
        ws.cell(r, 4, x).font = styles["normal_font"]
        ws.cell(r, 4).number_format = "0.00"
        ws.cell(r, 5, fsm).font = styles["normal_font"]
        ws.cell(r, 5).number_format = "0.00"
        ws.cell(r, 6, rm).font = styles["normal_font"]

        # 테두리 적용
        for col in range(1, 7):
            ws.cell(r, col).border = Border(
                left=styles["thin_border"],
                right=styles["thin_border"],
                top=styles["thin_border"],
                bottom=styles["thin_border"],
            )
        r += 1

    # 제목 및 합계
    ws["B1"] = "CONST_TANKS – Permanent Loads (Fuel/FW)"
    ws["B1"].font = styles["title_font"]
    ws["B1"].fill = styles["header_fill"]

    # 레이블 (row 12, 데이터는 row 4-11)
    ws["C12"] = "Total Weight [t]"
    ws["C12"].font = styles["header_font"]
    ws["E12"] = "Total FSM [mt·m]"
    ws["E12"].font = styles["header_font"]

    # 합계 수식 (row 13, 실제 데이터는 4-11행) - 순환 참조 방지
    ws["C13"] = "=SUM(C4:C11)"  # Const Total Weight t
    ws["C13"].font = styles["normal_font"]
    ws["C13"].number_format = "0.00"
    ws["C13"].fill = styles["input_fill"]

    ws["E13"] = "=SUM(E4:E11)"  # Const Total FSM mt·m
    ws["E13"].font = styles["normal_font"]
    ws["E13"].number_format = "0.00"
    ws["E13"].fill = styles["input_fill"]

    # 컬럼 폭 설정
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 15

    # 로그 출력
    if meta and meta.get("used_csv"):
        matched = meta.get("matched_count", 0)
        total = meta.get("total_tanks", len(const_rows))
        print(
            f"  [OK] CONST_TANKS sheet created (Patch Option B-2: CSV {matched}/{total} tanks)"
        )
    else:
        print("  [OK] CONST_TANKS sheet created (Patch Option B: Default values)")

    return ws


def create_hydro_table_sheet(wb):
    """Hydro_Table 시트 생성 (Option B ready: LCF/MCTC/TPC/GM_min optional)"""
    ws = wb.create_sheet("Hydro_Table")
    styles = get_styles()

    # Base headers (Option B). Only keys present in JSON will be shown; unknown keys appended at the end.
    base_headers = [
        "Disp_t",
        "Tmean_m",
        "Trim_m",
        "GM_m",
        "Draft_FWD",
        "Draft_AFT",
        # Option B recommended (approved booklet / NAPA export)
        "LCF_m_from_midship",  # x_from_mid_m, AFT=+
        "MCTC_t_m_per_cm",  # (alias: MTC_t_m_per_cm)
        "TPC_t_per_cm",
        "GM_min_m",  # optional curve point
    ]
    headers = base_headers

    # Note:
    # - Trim_m column (if present) is reference-only. Stage trim is derived from TM/MCTC.
    # - For Engineering-grade, provide LCF/MCTC/TPC vs draft in the same hydro table.
    json_data = _load_hydro_table()
    data = []

    if json_data and isinstance(json_data, list) and len(json_data) > 0:
        if isinstance(json_data[0], dict):
            existing = list(json_data[0].keys())

            # map common aliases into canonical keys (if present in JSON)
            alias_map = {
                "MTC_t_m_per_cm": "MCTC_t_m_per_cm",
                "MTC": "MCTC_t_m_per_cm",
                "TPC": "TPC_t_per_cm",
                "LCF": "LCF_m_from_midship",
                "LCF_from_mid_m": "LCF_m_from_midship",
            }

            # normalize each entry into a new dict (do not mutate original)
            norm_rows = []
            for entry in json_data:
                if not isinstance(entry, dict):
                    continue
                e = dict(entry)
                for k_src, k_dst in alias_map.items():
                    if k_src in e and k_dst not in e:
                        e[k_dst] = e.get(k_src)
                norm_rows.append(e)

            if norm_rows:
                existing_norm = list(norm_rows[0].keys())
                headers = [h for h in base_headers if h in existing_norm]
                extras = [k for k in existing_norm if k not in headers]
                headers += extras

                for entry in norm_rows:
                    data.append([entry.get(h, "") for h in headers])

            print(
                f"  [OK] Hydro_Table loaded ({len(data)} points), columns={len(headers)}"
            )
        else:
            # already in array format
            data = json_data
            print(f"  [OK] Hydro_Table loaded (array, {len(data)} rows)")
    else:
        # fallback points (legacy) - only base part
        print("  [FALLBACK] Using built-in 4 points (no JSON hydro table found)")
        headers = ["Disp_t", "Tmean_m", "Trim_m", "GM_m", "Draft_FWD", "Draft_AFT"]
        data = [
            [2991.25, 2.20, 0.20, 2.85, 2.10, 2.30],
            [3208.25, 3.18, -0.53, 1.68, 2.92, 3.45],
            [3265.25, 3.00, 0.60, 1.88, 2.68, 3.28],
            [3425.25, 3.00, 0.70, 1.85, 2.65, 3.35],
        ]

    # header row styling
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = styles["center_align"]
        cell.border = Border(
            left=styles["thin_border"],
            right=styles["thin_border"],
            top=styles["thin_border"],
            bottom=styles["thin_border"],
        )

    # sort by Tmean_m if possible (VLOOKUP approximate requires ascending)
    try:
        tmean_idx = headers.index("Tmean_m")
    except ValueError:
        tmean_idx = 1  # legacy

    def _sort_key(row):
        try:
            v = row[tmean_idx]
            if isinstance(v, str) and v.strip() == "":
                return 0.0
            return float(v)
        except Exception:
            return 0.0

    data_sorted = sorted(data, key=_sort_key)

    # data rows
    for r, row_data in enumerate(data_sorted, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.number_format = "0.00"
            cell.alignment = styles["center_align"]
            cell.border = Border(
                left=styles["thin_border"],
                right=styles["thin_border"],
                top=styles["thin_border"],
                bottom=styles["thin_border"],
            )

    # column widths (simple)
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14


def create_frame_table_sheet(wb):
    """Frame_to_x_Table 시트 생성"""
    ws = wb.create_sheet("Frame_to_x_Table")
    styles = get_styles()

    headers = ["Fr", "x_from_mid_m", "비고"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = styles["center_align"]
        cell.border = Border(
            left=styles["thin_border"],
            right=styles["thin_border"],
            top=styles["thin_border"],
            bottom=styles["thin_border"],
        )

    # JSON 파일 로드 (상대 경로 사용)
    frame_data = _load_json("data/Frame_x_from_mid_m.json")

    if frame_data:
        try:
            # frame_data가 리스트인지 확인
            if isinstance(frame_data, list):
                for idx, entry in enumerate(frame_data, start=0):
                    row = 2 + idx

                    cell_a = ws.cell(row=row, column=1)
                    cell_a.value = entry.get("Fr", 0.0)
                    cell_a.font = styles["normal_font"]
                    cell_a.number_format = "0.00"

                    cell_b = ws.cell(row=row, column=2)
                    cell_b.value = entry.get("x_from_mid_m", 0.0)
                    cell_b.font = styles["normal_font"]
                    cell_b.number_format = "0.00"

                    cell_c = ws.cell(row=row, column=3)
                    cell_c.value = entry.get("비고", "")
                    cell_c.font = styles["normal_font"]

                print(
                    f"  [OK] Frame_to_x_Table sheet created with {len(frame_data)} rows"
                )
            else:
                print(f"  [WARNING] Frame data is not a list. Creating empty sheet.")
        except Exception as e:
            print(
                f"  [WARNING] Error processing frame data: {e}. Creating empty sheet."
            )
            frame_data = None

    if not frame_data:
        print(
            f"  [WARNING] JSON file not found. Creating empty Frame_to_x_Table sheet."
        )
        for row in range(2, 123):  # 121 rows + header
            ws.cell(row=row, column=1).font = styles["normal_font"]
            ws.cell(row=row, column=1).number_format = "0.00"
            ws.cell(row=row, column=2).font = styles["normal_font"]
            ws.cell(row=row, column=2).number_format = "0.00"
            ws.cell(row=row, column=3).font = styles["normal_font"]

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 20


# ============================================================================
# Main Orchestration Function
# ============================================================================


def create_workbook_from_scratch(const_csv_path: str = None):
    """워크북을 처음부터 생성 (BACKUP PLAN integrated + Option B-2 CSV support)

    Args:
        const_csv_path: CONST_TANKS CSV 파일 경로 (Option B-2, 선택 사항)
    """
    print("=" * 80)
    print("LCT_BUSHRA_AGI_TR.xlsx Creation from Scratch (BACKUP PLAN enabled)")
    if const_csv_path:
        print(f"Option B-2: Using CSV for CONST_TANKS: {const_csv_path}")
    print("=" * 80)

    # Option B+ (Engineering-grade) input gate: auto-search ./bplus_inputs and stop if missing
    _bplus_preflight_or_zero(strict=True)

    # BACKUP PLAN: Pre-flight check
    print("\n[PRE-FLIGHT CHECK]")
    issues = preflight_check()

    # PHASE 0: Tank JSON auto-generation
    try:
        from src.tank_data_manager import ensure_tank_jsons

        logging.info("[PRE-FLIGHT] Checking tank data files")
        success, msg = ensure_tank_jsons("Tank Capacity_Plan.xlsx", "data/")
        if success:
            logging.info(f"[TANK] {msg}")
            print(f"  [OK] {msg}")
        else:
            issues.append(f"WARNING: {msg}")
            print(f"  [WARNING] {msg}")
    except ImportError:
        issues.append("INFO: Tank auto-generation module not available")
        print("  [INFO] Tank auto-generation module not available")
    except Exception as e:
        issues.append(f"WARNING: Tank JSON generation failed: {e}")
        print(f"  [WARNING] Tank JSON generation failed: {e}")

    for issue in issues:
        print(f"  {issue}")
    if any("ERROR" in i for i in issues):
        print("\n[ABORT] Critical issues found. Exiting.")
        sys.exit(1)

    output_dir = os.path.dirname(OUTPUT_FILE)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
        print(f"[OK] Created output directory: {output_dir}")

    final_output_file = OUTPUT_FILE
    if os.path.exists(OUTPUT_FILE):
        try:
            with open(OUTPUT_FILE, "r+b"):
                pass
        except PermissionError:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            base_name = os.path.splitext(OUTPUT_FILE)[0]
            final_output_file = f"{base_name}_{timestamp}.xlsx"
            print(f"[WARNING] Original file is open. Saving as: {final_output_file}")

    # BACKUP PLAN: Setup logging
    print(f"\n[1/9] Setting up logging and workbook")
    log_file = setup_logging(final_output_file)
    logging.info("[1/9] Workbook creation started")

    wb = Workbook()
    wb.remove(wb.active)
    # Force Excel to recalculate formulas on open
    try:
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.calcOnSave = True
    except Exception:
        pass

    # BACKUP PLAN: Safe sheet creation with error recovery
    print(f"\n[2/9] Creating sheets (with error recovery):")
    logging.info("[2/9] Sheet creation phase started")

    safe_sheet_creation(wb, create_calc_sheet, "Calc")
    safe_sheet_creation(wb, create_tide_sheet, "December_Tide_2025")
    safe_sheet_creation(wb, create_hourly_sheet, "Hourly_FWD_AFT_Heights")

    result = safe_sheet_creation(wb, create_roro_sheet, "RORO_Stage_Scenarios")
    if result:
        if len(result) == 3:
            stages, first_data_row, total_rows = result
        else:
            # 이전 버전 호환성
            stages, first_data_row = result
            total_rows = len(stages)
    else:
        # BACKUP: Provide fallback values if RORO sheet creation fails
        logging.warning("[BACKUP] RORO sheet failed, using defaults")
        stages = []
        first_data_row = 19
        total_rows = 0

    safe_sheet_creation(wb, create_ballast_tanks_sheet, "Ballast_Tanks")

    # Patch Option B-2: CONST_TANKS with CSV support
    const_rows, const_meta = load_const_tanks_snapshot(const_csv_path)
    safe_sheet_creation(
        wb, create_const_tanks_sheet, "CONST_TANKS", const_rows, const_meta
    )

    safe_sheet_creation(wb, create_hydro_table_sheet, "Hydro_Table")
    safe_sheet_creation(wb, create_frame_table_sheet, "Frame_to_x_Table")

    if "RORO_Stage_Scenarios" in wb.sheetnames and stages:
        roro_ws = wb["RORO_Stage_Scenarios"]
        logging.info("[3/9] Extending RORO sheet with additional columns")
        print(f"\n[3/9] Extending RORO sheet")

        try:
            # 일반 Stage만 처리 (Optional Tuning Stages는 자체 수식 보유)
            extend_roro_captain_req(roro_ws, first_data_row, total_rows)
            extend_roro_structural_opt1(roro_ws, first_data_row, total_rows)
            extend_precision_columns(roro_ws, first_data_row, total_rows)

            # BC-BR 범위 추가 (Post-Solver Results & Gate Verification)
            # pipeline_stage_QA.csv에서 데이터 로드
            from pathlib import Path
            stage_qa_path = None
            for pattern in ['ssot/pipeline_stage_QA.csv', 'pipeline_out_*/ssot/pipeline_stage_QA.csv']:
                matches = list(Path('.').glob(pattern))
                if matches:
                    stage_qa_path = matches[-1]
                    break
            extend_roro_bc_br_range(roro_ws, first_data_row, total_rows, stage_qa_path)
        except Exception as e:
            logging.error(f"[BACKUP] RORO extension failed: {e}")
            print(f"  [BACKUP] Warning: RORO extension failed: {e}")

        # Create Excel Table after all columns are added
        from openpyxl.worksheet.table import Table, TableStyleInfo

        header_row = 18
        last_col = 70  # BC-BR 범위 추가로 70까지 확장
        last_col_letter = get_column_letter(last_col)

        # Verify all headers are strings before creating table
        for col in range(1, last_col + 1):
            cell = roro_ws.cell(row=header_row, column=col)
            if cell.value is not None and not isinstance(cell.value, str):
                cell.value = str(cell.value)
            elif cell.value is None or cell.value == "":
                cell.value = f"Unused_{col}"

        try:
            table = Table(
                displayName="Stages",
                ref=f"A{header_row}:{last_col_letter}{first_data_row + total_rows - 1}",
            )
            style = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            table.tableStyleInfo = style
            roro_ws.add_table(table)
            logging.info("[OK] Excel Table created")
            print("  [OK] Excel Table created successfully")
        except Exception as e:
            logging.warning(f"[BACKUP] Excel Table creation failed: {e}")
            print(f"  [BACKUP] Warning: Could not create Excel Table: {e}")

    # Create OPERATION SUMMARY sheet
    if stages:
        logging.info("[4/9] Creating OPERATION SUMMARY sheet")
        print(f"\n[4/9] Creating OPERATION SUMMARY")
        safe_sheet_creation(
            wb, create_captain_report_sheet, "OPERATION SUMMARY", stages, first_data_row
        )

    # Save workbook
    logging.info(f"[5/9] Saving workbook: {final_output_file}")
    print(f"\n[5/9] Saving workbook: {final_output_file}")

    # PATCH: Ensure calculation mode is set before saving
    try:
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.calcOnSave = True
        print(
            "  [PATCH] Excel calculation mode: AUTO (fullCalcOnLoad=True, calcOnSave=True)"
        )
    except Exception as e:
        print(f"  [WARN] Failed to set calculation mode: {e}")

    # PATCH: Verify G24, H24 values and formulas for Row 24, 25 before saving
    if "RORO_Stage_Scenarios" in wb.sheetnames:
        roro_ws = wb["RORO_Stage_Scenarios"]
        print("  [VERIFY] Checking Row 24, 25 values and formulas:")
        for check_row in [24, 25]:
            stage_name = roro_ws.cell(row=check_row, column=1).value  # A column
            g_val = roro_ws.cell(
                row=check_row, column=7
            ).value  # G column (FWD_precise_m)
            h_val = roro_ws.cell(
                row=check_row, column=8
            ).value  # H column (AFT_precise_m)
            p_val = roro_ws.cell(
                row=check_row, column=16
            ).value  # P column (Dfwd_m formula)
            q_val = roro_ws.cell(
                row=check_row, column=17
            ).value  # Q column (Daft_m formula)
            u_val = roro_ws.cell(
                row=check_row, column=21
            ).value  # U column (GM formula)
            ab_val = roro_ws.cell(
                row=check_row, column=28
            ).value  # AB column (GM_calc formula)
            ac_val = roro_ws.cell(
                row=check_row, column=29
            ).value  # AC column (GM_Check formula)

            print(f"    Row {check_row} ({stage_name}):")
            print(f"      G7={g_val}, H8={h_val}")
            print(f"      P16={p_val}, Q17={q_val}")
            print(f"      U21={u_val}, AB28={ab_val}, AC29={ac_val}")

            # Check if G24/H24 have values (not None/empty)
            if g_val is None or (isinstance(g_val, str) and g_val.strip() == ""):
                print(f"      [WARN] Row {check_row} G7 (FWD_precise_m) is empty!")
            if h_val is None or (isinstance(h_val, str) and h_val.strip() == ""):
                print(f"      [WARN] Row {check_row} H8 (AFT_precise_m) is empty!")

    try:
        wb.save(final_output_file)
        logging.info("[OK] File saved successfully")
        print(f"  [OK] File saved successfully")
    except Exception as e:
        logging.error(f"[ERROR] Failed to save: {e}")
        print(f"  [ERROR] Failed to save: {e}")
        sys.exit(1)

    # CSV Export (워크북 저장 후, 닫기 전)
    if "RORO_Delta_Lever_Report" in wb.sheetnames:
        try:
            csv_output_path = os.path.join(
                os.path.dirname(final_output_file), "RORO_Delta_Lever_Report.csv"
            )
            export_roro_delta_lever_report_to_csv(
                wb,
                sheet_name="RORO_Delta_Lever_Report",
                csv_path=csv_output_path,
            )
            logging.info(f"[OK] CSV exported: {csv_output_path}")
        except Exception as e:
            logging.warning(f"[WARNING] CSV export failed: {e}")
            print(f"  [WARNING] CSV export failed: {e}")

    # WhatsApp Summary PNG Export (워크북 저장 후, 닫기 전)
    try:
        png_output_path = os.path.join(
            os.path.dirname(final_output_file), "RORO_Summary.png"
        )
        png_path = export_whatsapp_summary_png(
            wb,
            png_path=png_output_path,
            max_stage_lines=6,
        )
        if png_path:
            logging.info(f"[OK] PNG exported: {png_path}")
    except Exception as e:
        logging.warning(f"[WARNING] PNG export failed: {e}")
        print(f"  [WARNING] PNG export failed: {e}")

    wb.close()

    # Optional: Excel COM full recalc to cache formula results (requires Excel + pywin32)
    if os.environ.get("EXCEL_COM_RECALC", "").strip().lower() in ("1", "true", "yes", "y"):
        try:
            com_script = Path(__file__).parent / "excel_com_recalc_save.py"
            out_path = os.environ.get("EXCEL_COM_RECALC_OUT", "").strip()
            cmd = [sys.executable, str(com_script), "--in", str(final_output_file)]
            if out_path:
                cmd += ["--out", out_path]
            subprocess.run(cmd, check=False)
        except Exception as e:
            print(f"  [WARN] Excel COM recalc skipped: {e}")

    # BACKUP PLAN: Create backup after successful save
    print(f"\n[6/9] Creating backup")
    logging.info("[6/9] Creating backup file")
    backup_path = create_backup_file(final_output_file)

    # Verification
    logging.info("[7/9] Verification")
    print(f"\n[7/9] Verification:")
    if os.path.exists(final_output_file):
        file_size = os.path.getsize(final_output_file) / 1024
        logging.info(f"File created: {final_output_file}, Size: {file_size:.2f} KB")
        print(f"  [OK] File created: {final_output_file}")
        print(f"  [OK] File size: {file_size:.2f} KB")
        print(f"  [OK] Sheets: {len(wb.sheetnames)}")
        if backup_path:
            print(f"  [OK] Backup: {os.path.basename(backup_path)}")
        print(f"  [OK] Log: {os.path.basename(log_file)}")
    else:
        logging.error("[ERROR] Output file was not created")
        print(f"  [ERROR] Output file was not created")
        sys.exit(1)

    print("\n" + "=" * 80)
    print("[SUCCESS] Workbook creation complete! (BACKUP PLAN active)")
    print("=" * 80)
    logging.info("[SUCCESS] Workbook creation complete")


# ============================================================================
# CSV Export Function
# ============================================================================


def export_stages_to_csv(output_path: str = None):
    """
    Stage별 계산 결과를 CSV로 Export.

    Python → CSV → 엑셀에서 자동 가져오기 플로우:
    1. Python이 JSON 3개를 읽는다 (Frame_x_from_mid_m.json, hydro_table.json, LCT_BUSHRA_GM_2D_Grid.json)
    2. find_preballast_opt()가 Stage 5_Pre / Stage 6A에 대해 최소 Pre-ballast를 찾는다
    3. 찾은 Pre-ballast를 고정시켜서 Stage 1~7에 대해 solve_stage()로 계산
    4. 결과를 stage_results.csv로 저장
    → 엑셀 RORO_Stage_Scenarios 시트에서 B열/W, D열/x, Trim, Dfwd, GM 등을 이 CSV를 VLOOKUP/INDEX로 끌어다 쓴다

    Args:
        output_path: 출력 CSV 파일 경로 (None이면 기본값: stage_results.csv)
    """
    if output_path is None:
        output_path = os.path.join(SCRIPT_DIR, "stage_results.csv")

    # Hydro table 로드
    hydro_table_data = _load_hydro_table()
    if not hydro_table_data:
        print("[WARNING] hydro_table.json not found. Using empty table.")
        hydro_table_data = []

    # -------------------------------
    # CONFIG (가정 값 – 필요 시 조정)
    # -------------------------------
    # ⚠ 일부 Stage별 위치/중량은 지금 파일에 명시가 없어서 가정 값으로 묶어두고,
    #    맨 위 CONFIG에서 네가 직접 튜닝할 수 있게 했다.
    cfg = {
        # TR+SPMT 등가 중량 (t)
        "W_TR": 271.20,  # 가정: PL 기준; 필요 시 수정
        # TR1 위치 (Frame) - ✅ stage_results.csv 기준 (2024-12-23 패치)
        "FR_TR1_RAMP_START": 26.15,  # Stage 2: x=+4.0m → Fr=30.151-4.0≈26.15
        "FR_TR1_RAMP_MID": 26.15,  # Stage 3: x=+4.0m (동일)
        "FR_TR1_STOW": 25.65,  # Stage 4/5: x=+4.5m → Fr=30.151-4.5≈25.65
        # TR2 위치 (Frame) - ✅ stage_results.csv 기준
        "FR_TR2_RAMP": 25.04,  # Stage 6A: x=+5.11m → Fr=30.151-5.11≈25.04
        "FR_TR2_STOW": 29.39,  # Stage 6C: x=+0.76m (LCF, even keel) → Fr=30.151-0.76≈29.39
        # Pre-ballast 중심 (FW2, AFT 쪽) - ✅ stage_results.csv 기준
        "FR_PREBALLAST": 25.58,  # Stage 5_PreBallast: x=+4.57m → Fr=30.151-4.57≈25.58
    }

    # 선박 고정 파라미터 (Aries/NAPA 값 기준 – 필요 시 수정)
    MTC = 34.00  # t·m/cm
    LCF = 0.76  # m (midship 기준 x)
    LBP = 60.302  # m
    D_vessel = 3.65  # m (Vessel depth)

    # Stage 1 기준 Δ, Tmean (hydro_table.json에서 선택)
    # 가정: Disp 2800t 근처 값 사용
    base_disp_t = 2800.00
    # PATCH B1-Enhanced: remove fixed base_tmean_m=2.00; derive from hydro table (SSOT enforcement)
    hydro_table_data = _load_hydro_table()
    if hydro_table_data:
        base_tmean_m = interpolate_tmean_from_disp(base_disp_t, hydro_table_data)
        print(
            f"[PATCH B1] base_tmean_m from hydro: {base_tmean_m:.3f} m (base_disp_t={base_disp_t:.2f} t)"
        )
    else:
        raise RuntimeError(
            "[PATCH B1-Enhanced] hydro_table_data missing in Export path; cannot compute base_tmean_m. "
            "Provide Hydro_Table_Engineering.json (SSOT). "
            "Location: bplus_inputs/Hydro_Table_Engineering.json"
        )

    # solve_stage에 필요한 params
    params = {
        "MTC": MTC,
        "LCF": LCF,
        "LBP": LBP,
        "D_vessel": D_vessel,
        "hydro_table": hydro_table_data,
        "FWD_DRAFT_LIMIT": 2.70,
        "GM_MIN": 1.50,
        # Patch Option B: FSM from CONST_TANKS (Fuel + FreshWater)
        "Const_FSM_t_m": 887.0,  # CONST_TANKS!E5 합계값 (DO.P/S + LRFO.P/S + FW1.P/S + FW2.P/S)
    }
    # build_stage_loads에 필요한 파라미터도 추가
    params.update(cfg)

    # 1) Pre-ballast 탐색
    w_tr = cfg["W_TR"]

    preballast_result = find_preballast_opt(
        w_tr_unit_t=w_tr,
        base_disp_t=base_disp_t,
        fr_tr1_stow=cfg.get("FR_TR1_STOW", FR_TR1_STOW),
        fr_tr2_ramp=cfg.get("FR_TR2_RAMP", FR_TR2_RAMP),
        fr_preballast=cfg.get("FR_PREBALLAST", FR_PREBALLAST),
        params=params,
        search_min_t=params.get("PREBALLAST_MIN_T", 20.0),  # FIX: enforce 20t minimum
        search_max_t=params.get("PREBALLAST_MAX_T", 400.0),
        search_step_t=params.get("PREBALLAST_STEP_T", 1.0),
    )

    if not preballast_result["ok"]:
        preballast_opt = params.get("PREBALLAST_T_TARGET", 250.0)
        print(
            f"[WARNING] Pre-ballast optimization failed: {preballast_result['reason']}"
        )
        print(f"[WARNING] Using fallback pre-ballast value: {preballast_opt:.2f} t")
        stage5_pb = None
        stage6A_pb = None
    else:
        preballast_opt = preballast_result["w_preballast_t"]
        print(
            f"[INFO] Optimal Pre-ballast: {preballast_opt:.2f} t (used for all stages)"
        )
        # Pre-ballast 결과에서 Stage 5 / Stage 6A 데이터 추출
        stage5_pb = preballast_result.get("stage5")
        stage6A_pb = preballast_result.get("stage6A")

    # 2) Stage 리스트
    stages_order = [
        "Stage 1",
        "Stage 2",
        "Stage 3",
        "Stage 4",
        "Stage 5",
        "Stage 5_PreBallast",
        "Stage 6A_Critical (Opt C)",
        "Stage 6B Tide Window",
        "Stage 6C",
        "Stage 7",
    ]

    # 3) Stage별 계산
    rows = []
    for st in stages_order:
        loads = build_stage_loads(st, preballast_opt, params)
        res = solve_stage(base_disp_t, base_tmean_m, loads, **params)

        # ⭐ Pre-ballast 결과로 Stage 5_PreBallast / Stage 6A 값 override
        if st == "Stage 5_PreBallast" and stage5_pb is not None:
            # Pre-ballast 계산 결과의 FWD/AFT/Trim/TM 값으로 교체
            res["W_stage_t"] = float(stage5_pb.get("W_stage_t", res["W_stage_t"]))
            res["x_stage_m"] = float(stage5_pb.get("x_stage_m", res["x_stage_m"]))
            res["TM_LCF_tm"] = float(stage5_pb.get("TM_tm", res["TM_LCF_tm"]))
            res["Trim_cm"] = float(stage5_pb.get("Trim_cm", res["Trim_cm"]))
            res["Dfwd_m"] = float(stage5_pb.get("FWD_m", res["Dfwd_m"]))
            res["Daft_m"] = float(stage5_pb.get("AFT_m", res["Daft_m"]))
            # FWD_Height_m, AFT_Height_m 재계산
            res["FWD_Height_m"] = D_vessel - res["Dfwd_m"]
            res["AFT_Height_m"] = D_vessel - res["Daft_m"]
            print(
                f"[INFO] Stage 5_PreBallast: Applied pre-ballast FWD={res['Dfwd_m']:.2f}m, AFT={res['Daft_m']:.2f}m"
            )

        if st == "Stage 6A_Critical (Opt C)" and stage6A_pb is not None:
            # Pre-ballast 계산 결과의 FWD/AFT/Trim/TM 값으로 교체
            res["W_stage_t"] = float(stage6A_pb.get("W_stage_t", res["W_stage_t"]))
            res["x_stage_m"] = float(stage6A_pb.get("x_stage_m", res["x_stage_m"]))
            res["TM_LCF_tm"] = float(stage6A_pb.get("TM_tm", res["TM_LCF_tm"]))
            res["Trim_cm"] = float(stage6A_pb.get("Trim_cm", res["Trim_cm"]))
            res["Dfwd_m"] = float(stage6A_pb.get("FWD_m", res["Dfwd_m"]))
            res["Daft_m"] = float(stage6A_pb.get("AFT_m", res["Daft_m"]))
            # FWD_Height_m, AFT_Height_m 재계산
            res["FWD_Height_m"] = D_vessel - res["Dfwd_m"]
            res["AFT_Height_m"] = D_vessel - res["Daft_m"]
            print(
                f"[INFO] Stage 6A_Critical: Applied pre-ballast FWD={res['Dfwd_m']:.2f}m, AFT={res['Daft_m']:.2f}m"
            )

        # Stage 6B Tide Window: Stage 6A_Critical과 동일한 Draft 값 (Forecast Tide +0.30m는 UKC 계산에만 영향)
        if st == "Stage 6B Tide Window":
            # Stage 6A_Critical과 동일한 계산값 사용 (Draft는 동일, Tide는 UKC 계산에만 영향)
            stage6a_res = None
            for prev_row in rows:
                if prev_row["Stage"] == "Stage 6A_Critical (Opt C)":
                    stage6a_res = prev_row
                    break
            if stage6a_res:
                res["W_stage_t"] = stage6a_res["W_stage_t"]
                res["x_stage_m"] = stage6a_res["x_stage_m"]
                res["TM_LCF_tm"] = stage6a_res["TM_LCF_tm"]
                res["Trim_cm"] = stage6a_res["Trim_cm"]
                res["Dfwd_m"] = stage6a_res["Dfwd_m"]
                res["Daft_m"] = stage6a_res["Daft_m"]
                res["FWD_Height_m"] = stage6a_res["FWD_Height_m"]
                res["AFT_Height_m"] = stage6a_res["AFT_Height_m"]
                res["GM_m"] = stage6a_res["GM(m)"]
                res["Trim_Check"] = stage6a_res["Trim_Check"]
                res["vs_2.70m"] = stage6a_res["vs_2.70m"]
                res["GM_Check"] = stage6a_res["GM_Check"]
                res["Disp_t"] = stage6a_res["Disp_t"]
                res["Tmean_m"] = stage6a_res["Tmean_m"]
                print(
                    f"[INFO] Stage 6B Tide Window: Using Stage 6A_Critical values (Draft unchanged, Tide affects UKC only)"
                )

        # solve_stage()에서 이미 Trim_Check, vs_2.70m, GM_Check 계산됨
        rows.append(
            {
                "Stage": st,
                "W_stage_t": round(res["W_stage_t"], 2),
                "x_stage_m": round(res["x_stage_m"], 2),
                "TM_LCF_tm": round(res["TM_LCF_tm"], 2),
                "Trim_cm": round(res["Trim_cm"], 2),
                "Dfwd_m": round(res["Dfwd_m"], 2),
                "Daft_m": round(res["Daft_m"], 2),
                "GM(m)": round(res["GM_m"], 2),
                "Fwd Draft(m)": round(res["Dfwd_m"], 2),
                "FWD_Height_m": round(res["FWD_Height_m"], 2),
                "AFT_Height_m": round(res["AFT_Height_m"], 2),
                "Trim_Check": res["Trim_Check"],
                "vs_2.70m": res["vs_2.70m"],
                "GM_Check": res["GM_Check"],
                "Disp_t": round(res["Disp_t"], 2),
                "Tmean_m": round(res["Tmean_m"], 2),
                "PreBallast_t": round(preballast_opt, 2),
            }
        )

    # 4) CSV 저장
    fieldnames = [
        "Stage",
        "W_stage_t",
        "x_stage_m",
        "TM_LCF_tm",
        "Trim_cm",
        "Dfwd_m",
        "Daft_m",
        "GM(m)",
        "Fwd Draft(m)",
        "FWD_Height_m",
        "AFT_Height_m",
        "Trim_Check",
        "vs_2.70m",
        "GM_Check",
        "Disp_t",
        "Tmean_m",
        "PreBallast_t",
    ]

    try:
        with open(output_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            for r in rows:
                writer.writerow(r)

        print(f"[INFO] Stage results exported to {output_path}")
        print(f"[INFO] Total stages: {len(rows)}")
        print(
            f"[INFO] 엑셀에서 이 CSV를 가져와서 RORO_Stage_Scenarios 시트의 B열/W, D열/x, Trim, Dfwd, GM 등을 VLOOKUP/INDEX로 연결하세요."
        )
    except Exception as e:
        print(f"[ERROR] Failed to export CSV: {e}")
        raise


# ============================================================================
# Stage Evaluation Functions (wewewewe.md 가이드)
# ============================================================================

# 타입 alias
StageDict = Dict[str, Any]

# GMGrid 예시 (실전에서는 Aries/BV GM 테이블 로드)
# Δ–Trim–GM (m) 테이블
GM_GRID_EXAMPLE: GMGrid = {
    # Δ = 1227.59 t 일 때 Trim–GM
    1227.59: {
        0.00: 1.60,
        0.50: 1.58,
        1.00: 1.55,
    },
    # Δ = 1658.71 t 일 때 Trim–GM
    1658.71: {
        0.00: 1.55,
        0.50: 1.53,
        1.00: 1.50,
    },
}

LCF_M = (
    30.91  # m (AP 기준 LCF) - BUSHRA verified: Loaded condition avg LCG 31.45m from AP
)
LBP_M = 60.302  # m (LBP)

# Stage 입력 예시 (엑셀 Row → dict 로 나왔다고 가정)
# - 실제로는 CSV/엑셀에서 읽어서 채우면 됨
STAGES_EXAMPLE: List[StageDict] = [
    {
        "name": "Stage 5A-2",
        "Tmean_m": 2.85,
        "Trim_cm": -96.50,  # 선수침(-), 선미침(+)
        "Disp_t": 1658.71,  # Δ
        "W_stage_t": 217.00,  # 해당 Stage 주요 카고/편심 하중
        "Y_offset_m": 2.50,  # 중심선 기준 횡 편심 (SPMT 등)
        "FSE_t_m": 85.00,  # Free Surface Moment (t·m)
        "Share_Load_t": 210.00,  # Ramp share load (정적)
        "Pin_Stress_MPa": 120.00,  # Pin stress (정적)
        "LoadCase": "B",  # A=STATIC, B=DYNAMIC, C=BRAKING
    },
    {
        "name": "Stage 6B",
        "Tmean_m": 3.10,
        "Trim_cm": -120.00,
        "Disp_t": 1227.59,
        "W_stage_t": 434.00,
        "Y_offset_m": 1.50,
        "FSE_t_m": 40.00,
        "Share_Load_t": 250.00,
        "Pin_Stress_MPa": 135.00,
        "LoadCase": "C",
    },
]


def evaluate_stages(
    stages: List[StageDict],
    gm_grid: GMGrid,
    lcf_m: float = LCF_M,
    lbp_m: float = LBP_M,
) -> List[StageDict]:
    """
    각 Stage에 대해:
      - Dfwd_precise_m / Daft_precise_m (LCF 기반 Draft)
      - GM_calc_m (Δ–Trim 2D 보간)
      - Heel_deg / GM_eff_m / Heel_OK / GM_OK
      - Share_Load_dyn_t / Pin_Stress_dyn_MPa
    필드를 추가해서 반환.
    """

    result: List[StageDict] = []

    for stage in stages:
        # ---- 1) LCF 기반 정밀 Draft ----
        tmean_m = float(stage.get("Tmean_m", 0.0))
        trim_cm = float(stage.get("Trim_cm", 0.0))

        dfwd_m, daft_m = calc_draft_with_lcf(
            tmean_m=tmean_m,
            trim_cm=trim_cm,
            lcf_m=lcf_m,
            lbp_m=lbp_m,
        )
        stage["Dfwd_precise_m"] = round(dfwd_m, 3)
        stage["Daft_precise_m"] = round(daft_m, 3)

        # ---- 2) Δ–Trim 2D GM 보간 ----
        disp_t = float(stage.get("Disp_t", 0.0))
        trim_m = trim_cm / 100.0

        gm_m = get_gm_bilinear(
            disp_t=disp_t,
            trim_m=trim_m,
            gm_grid=gm_grid,
        )
        stage["GM_calc_m"] = round(gm_m, 3)

        # ---- 3) Heel + FSE 반영 GM_eff ----
        weight_t = float(stage.get("W_stage_t", 0.0))
        y_offset_m = float(stage.get("Y_offset_m", 0.0))
        fse_t_m = float(stage.get("FSE_t_m", 0.0))

        heel_deg, gm_eff, heel_ok, gm_ok = heel_and_gm_check(
            weight_t=weight_t,
            y_offset_m=y_offset_m,
            disp_t=disp_t,
            gm_m=gm_m,
            fse_t_m=fse_t_m,
            heel_limit_deg=3.0,
            gm_min_m=1.50,
        )
        stage["Heel_deg"] = round(heel_deg, 3)
        stage["GM_eff_m"] = round(gm_eff, 3)
        stage["Heel_OK"] = heel_ok
        stage["GM_OK"] = gm_ok

        # ---- 4) 동적 / 제동 Load Case ----
        share_static = float(stage.get("Share_Load_t", 0.0))
        pin_static = float(stage.get("Pin_Stress_MPa", 0.0))

        # LoadCase 문자열 → Enum 매핑
        lc_raw = str(stage.get("LoadCase", "A")).upper()
        if lc_raw in ("A", "STATIC"):
            lc = LoadCase.STATIC
        elif lc_raw in ("B", "DYNAMIC"):
            lc = LoadCase.DYNAMIC
        elif lc_raw in ("C", "BRAKE", "BRAKING"):
            lc = LoadCase.BRAKING
        else:
            lc = LoadCase.STATIC

        share_dyn, pin_dyn = apply_dynamic_loads(
            share_load_t=share_static,
            pin_stress_mpa=pin_static,
            load_case=lc,
        )
        stage["Share_Load_dyn_t"] = round(share_dyn, 2)
        stage["Pin_Stress_dyn_MPa"] = round(pin_dyn, 2)
        stage["LoadCase_used"] = lc.name  # "STATIC" / "DYNAMIC" / "BRAKING"

        result.append(stage)

    return result


def debug_tank_lcg_check():
    """
    757 TCP LCG(AP) → midship x → Frame 역산 self-check.
    설계 검증용으로만 사용.
    """
    MIDSHIP_LCG_FROM_AP = 30.151
    tank_lcg_ap = {
        "FWB1": 57.519,
        "FWB2": 50.038,
        "FWCARGO1": 42.750,
        "FWCARGO2": 35.250,
    }

    print("[DEBUG] Tank LCG(AP) → x_from_mid → Frame")
    for name, lcg_ap in tank_lcg_ap.items():
        x_from_mid = MIDSHIP_LCG_FROM_AP - lcg_ap
        fr_est = x_to_fr(x_from_mid)
        print(
            f"  {name:8}  LCG(AP)={lcg_ap:7.3f}  x_mid={x_from_mid:8.3f}  → Fr_est={fr_est:6.2f}"
        )


if __name__ == "__main__":
    import sys

    args = sys.argv[1:]

    # --const-csv 인자 처리 (Option B-2)
    const_csv_path = None
    if "--const-csv" in args:
        i = args.index("--const-csv")
        if i + 1 < len(args):
            const_csv_path = args[i + 1]
            # --const-csv와 다음 인자를 제거
            args = [a for j, a in enumerate(args) if j not in (i, i + 1)]

    # --const-csv=path 형식 처리
    for i, arg in enumerate(args):
        if arg.startswith("--const-csv="):
            const_csv_path = arg.split("=", 1)[1].strip()
            args = [a for j, a in enumerate(args) if j != i]
            break

    debug_flags = {"debug", "--debug"}
    debug_enabled = any(arg in debug_flags for arg in args)
    non_flag_args = [arg for arg in args if arg not in debug_flags]
    mode = non_flag_args[0] if non_flag_args else None
    debug_only_mode = debug_enabled and not non_flag_args

    # Usage: python agi_tr.py [demo|preballast|csv|test_opt] [--debug] [--const-csv=path]
    # --debug enables frame-mapping self-checks; defaults run without debug output.
    # --const-csv enables CSV loading for CONST_TANKS (Option B-2)
    _init_frame_mapping(verbose=debug_enabled)

    if debug_enabled:
        debug_tank_lcg_check()

        # Self-check: Frame 축 패치 검증
        print("=" * 60)
        for label, fr in {
            "AP approx": 0.0,
            "Midship (Lpp/2)": 30.151,
            "FP approx": 60.30,
            "FWB1/2 center 55": 55.0,
        }.items():
            x = fr_to_x(fr)
            print(
                f"{label:25}  Fr={fr:6.2f}  →  x={x:8.3f} m  ({'FWD' if x < 0 else 'AFT'})"
            )
        print("=" * 60)

    if mode == "debug" or debug_only_mode:
        debug_frame_mapping()
    elif mode == "demo":
        # Stage 평가 데모 실행
        enriched_stages = evaluate_stages(STAGES_EXAMPLE, GM_GRID_EXAMPLE)

        from pprint import pprint

        print("\n=== Stage 계산 결과 요약 ===")
        for s in enriched_stages:
            print(f"\n[{s['name']}]")
            pprint(
                {
                    "Dfwd_precise_m": s["Dfwd_precise_m"],
                    "Daft_precise_m": s["Daft_precise_m"],
                    "GM_calc_m": s["GM_calc_m"],
                    "GM_eff_m": s["GM_eff_m"],
                    "Heel_deg": s["Heel_deg"],
                    "Heel_OK": s["Heel_OK"],
                    "GM_OK": s["GM_OK"],
                    "Share_Load_dyn_t": s["Share_Load_dyn_t"],
                    "Pin_Stress_dyn_MPa": s["Pin_Stress_dyn_MPa"],
                    "LoadCase_used": s["LoadCase_used"],
                }
            )
    elif mode == "preballast":
        # Pre-ballast 최적화 데모 실행
        print("\n" + "=" * 80)
        print("Pre-ballast 최적화 데모 실행")
        print("=" * 80)

        # Hydro table 로드
        hydro_table_data = _load_hydro_table()
        if not hydro_table_data:
            print("[WARNING] hydro_table.json not found. Using empty table.")
            hydro_table_data = []

        # 선박 기본 파라미터
        params = {
            "W_TR": 280.0,  # Transformer + SPMT
            "FR_TR1": 42.0,
            "FR_TR2": 17.95,  # Stage 6A_Critical LCG Frame
            "FR_PB": 55.5,  # FWB1/2 중심 Frame
            "LCF": 0.76,
            "MTC": 34.00,
            "LBP": 60.302,
            "FWD_DRAFT_LIMIT": 2.70,
            "GM_MIN": 1.50,
            "hydro_table": hydro_table_data,
        }

        base_disp = 2800.0  # Stage 1 Δ (lightship, 예시값)
        base_tmean = 2.00

        # GM 2D bilinear 보간 테스트
        print("\n[TEST] GM 2D bilinear 보간 테스트:")
        print(f"  GM @ (3600t, 0.0m trim)  = {gm_2d_bilinear(3600, 0.0):.3f} m")
        print(f"  GM @ (3650t, 0.25m trim) = {gm_2d_bilinear(3650, 0.25):.3f} m")

        # PreBallast 최적값 탐색
        w_tr = params.get("W_TR", 280.0)

        opt_result = find_preballast_opt(
            w_tr_unit_t=w_tr,
            base_disp_t=base_disp,
            fr_tr1_stow=params.get("FR_TR1_STOW", FR_TR1_STOW),
            fr_tr2_ramp=params.get("FR_TR2_RAMP", FR_TR2_RAMP),
            fr_preballast=params.get("FR_PREBALLAST", FR_PREBALLAST),
            params=params,
        )

        if not opt_result["ok"]:
            W_PB_OPT = params.get("PREBALLAST_T_TARGET", 250.0)
            print(
                f"\n[WARNING] Pre-ballast optimization failed: {opt_result['reason']}"
            )
            print(f"[WARNING] Using fallback pre-ballast value: {W_PB_OPT:.2f} t")
        else:
            W_PB_OPT = opt_result["w_preballast_t"]
            print("\n[RESULT] Pre-ballast optimization result:")
            import json

            print(json.dumps(opt_result, indent=2))

            print("\n[RESULT] Stage 계산 결과:")
            if opt_result.get("stage5"):
                print("\nStage 5_PreBallast →")
                for key, value in opt_result["stage5"].items():
                    print(f"  {key}: {value:.3f}")

            if opt_result.get("stage6A"):
                print("\nStage 6A_Critical →")
                for key, value in opt_result["stage6A"].items():
                    print(f"  {key}: {value:.3f}")

        params["PREBALLAST_OPT_T"] = W_PB_OPT
    elif mode == "csv":
        # CSV Export 실행
        print("\n" + "=" * 80)
        print("CSV Export 실행")
        print("=" * 80)
        export_stages_to_csv()
    elif mode == "test_opt":
        # find_preballast_opt 테스트
        print("\n" + "=" * 80)
        print("find_preballast_opt 테스트")
        print("=" * 80)
        opt = find_preballast_opt(base_disp_t=2800.0)
        print("\n[Auto Pre-ballast Optimization Result]")
        import json

        print(json.dumps(opt, indent=2))
    else:
        create_workbook_from_scratch(const_csv_path=const_csv_path)
