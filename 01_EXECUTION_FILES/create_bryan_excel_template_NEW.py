#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
create_bryan_excel_template.py

Goal
- Generate Bryan Submission Data Pack Template aligned to the analyzed report structure:
  16 sheets in fixed order (same sheet names).
- Provide enough rows/columns so all data can be entered.
- Use NamedRanges for cross-sheet DataValidation lists (Stage_List, Tanks_List).

Sheet order (per analysis doc)
- 00_README
- 01_SSOT_Master
- 02_Vessel_Hydro
- 03_Berth_Tide
- 04_Cargo_SPMT
- 05_Ballast_Tanks
- 06_Stage_Plan
- 07_Stage_Calc
- 08_HoldPoint_Log
- 09_Evidence_Log
- 10_RACI
- 11_Assumptions_Issues
- 12_DataDictionary
- Pack_A_Inputs_SSOT
- Pack_B_Calculation_Formulas
- Pack_D_Tide_Window_RACI
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.workbook.defined_name import DefinedName


# =============================================================================
# Styles
# =============================================================================
@dataclass(frozen=True)
class Styles:
    font_title: Font
    font_header: Font
    font_normal: Font

    fill_title: PatternFill
    fill_section: PatternFill
    fill_input: PatternFill
    fill_calc: PatternFill
    fill_note: PatternFill

    border_thin: Border

    align_center: Alignment
    align_left: Alignment


def build_styles() -> Styles:
    thin = Side(style="thin", color="C0C0C0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    return Styles(
        font_title=Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
        font_header=Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
        font_normal=Font(name="Calibri", size=11, bold=False, color="000000"),
        fill_title=PatternFill("solid", fgColor="1F4E78"),
        fill_section=PatternFill("solid", fgColor="D9D9D9"),
        fill_input=PatternFill(
            "solid", fgColor="D9E1F2"
        ),  # light blue: "Fill blue cells only"
        fill_calc=PatternFill("solid", fgColor="F2F2F2"),
        fill_note=PatternFill("solid", fgColor="FFF2CC"),  # light yellow note
        border_thin=border,
        align_center=Alignment(horizontal="center", vertical="center", wrap_text=True),
        align_left=Alignment(horizontal="left", vertical="center", wrap_text=True),
    )


# =============================================================================
# Utilities
# =============================================================================
def set_col_widths(ws, widths: Dict[str, float]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[col].width = float(w)


def make_header_row(
    ws, row_idx: int, headers: Sequence[str], styles: Styles, start_col: int = 1
) -> None:
    for i, h in enumerate(headers, start=start_col):
        c = ws.cell(row=row_idx, column=i, value=h)
        c.font = styles.font_header
        c.fill = styles.fill_title
        c.alignment = styles.align_center
        c.border = styles.border_thin


def apply_row_style(
    ws,
    row_idx: int,
    col_from: int,
    col_to: int,
    styles: Styles,
    fill: Optional[PatternFill] = None,
    align: Optional[Alignment] = None,
    font: Optional[Font] = None,
) -> None:
    for col in range(col_from, col_to + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.border = styles.border_thin
        if fill is not None:
            cell.fill = fill
        if align is not None:
            cell.alignment = align
        if font is not None:
            cell.font = font


def add_named_range(wb: Workbook, name: str, refers_to: str) -> None:
    """
    refers_to example: "'06_Stage_Plan'!$A$2:$A$200"
    """
    # Remove if exists
    try:
        if name in wb.defined_names:
            del wb.defined_names[name]
    except Exception:
        try:
            # Fallback for older openpyxl versions
            existing = [dn for dn in wb.defined_names.definedName if dn.name == name]
            for dn in existing:
                wb.defined_names.definedName.remove(dn)
        except Exception:
            pass

    # Add new named range (compatible with both old and new openpyxl)
    try:
        # New openpyxl API (dict-like)
        wb.defined_names[name] = refers_to
    except (TypeError, AttributeError):
        # Fallback for older openpyxl versions
        try:
            dn = DefinedName(name=name, attr_text=refers_to)
            if hasattr(wb.defined_names, "definedName"):
                wb.defined_names.definedName.append(dn)
            else:
                wb.defined_names.append(dn)
        except Exception:
            # If all else fails, skip named range (non-critical)
            pass


def safe_add_table(ws, table_name: str, ref: str) -> None:
    """
    Adds an Excel table with a standard style. If table creation fails, just skip.
    """
    try:
        tab = Table(displayName=table_name, ref=ref)
        tab.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(tab)
    except Exception:
        # Table creation is optional
        pass


# =============================================================================
# Sheet builders
# =============================================================================
def build_readme_sheet(wb: Workbook, styles: Styles) -> None:
    ws = wb.create_sheet("00_README")
    ws["A1"] = "Workbook"
    ws["B1"] = "Bryan Submission Data Pack (Template)"
    ws["A2"] = "Version"
    ws["B2"] = "v1.0"
    ws["A3"] = "Created"
    ws["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws["A4"] = "Purpose"
    ws["B4"] = (
        "Provide a single SSOT-style input/calculation template for ballast operations submission."
    )
    ws["A6"] = "How to use"
    ws["B6"] = (
        "1) Fill blue cells only. 2) Keep SSOT_Master as single source. 3) Stage_Calc uses these inputs."
    )
    ws["A8"] = "Key Sheets"
    ws["B8"] = "01_SSOT_Master, 07_Stage_Calc, 08_HoldPoint_Log, 09_Evidence_Log"
    ws["A10"] = "Notes"
    ws["B10"] = "Designed for AGI site operations."

    for r in [1, 2, 3, 4, 6, 8, 10]:
        ws[f"A{r}"].fill = styles.fill_section
        ws[f"A{r}"].font = Font(bold=True)
        ws[f"A{r}"].alignment = styles.align_left
        ws[f"A{r}"].border = styles.border_thin

        ws[f"B{r}"].border = styles.border_thin
        ws[f"B{r}"].alignment = styles.align_left

    ws["B1"].fill = styles.fill_title
    ws["B1"].font = styles.font_title
    ws["B1"].alignment = styles.align_left

    set_col_widths(ws, {"A": 22.0, "B": 115.0})


def build_ssot_master_sheet(wb: Workbook, styles: Styles) -> None:
    ws = wb.create_sheet("01_SSOT_Master")
    make_header_row(
        ws,
        1,
        ["Control / Input", "Value", "Unit", "Notes", "Owner", "Source/Ref"],
        styles,
    )
    ws.row_dimensions[1].height = 22.0

    ssot_rows = [
        (
            "SITE",
            "AGI",
            "",
            "AGI site operations",
            "Logistics",
            "",
        ),
        ("Project / Shipment ID", "", "", "Internal reference", "Logistics", ""),
        ("Vessel / LCT Name", "", "", "", "Marine", ""),
        (
            "Datum for Tide",
            "CD",
            "",
            "Chart Datum unless specified otherwise",
            "Marine",
            "Port tide table",
        ),
        ("Timezone", "GST", "", "Local timezone for logs", "Marine", ""),
        ("LBP", "60.30", "m", "Length between perpendiculars", "Marine", "Vessel spec"),
        ("Beam", "24.00", "m", "", "Marine", "Vessel spec"),
        (
            "Molded Depth (D)",
            "3.65",
            "m",
            "Used for freeboard estimate",
            "Marine",
            "Vessel spec",
        ),
        (
            "Lightship Displacement",
            "2800",
            "t",
            "Baseline displacement (confirm)",
            "Marine",
            "",
        ),
        ("Baseline Tmean", "2.00", "m", "Baseline mean draft (confirm)", "Marine", ""),
        ("Gate-A: AFT Draft MIN", "2.70", "m", "AFT >= this", "Marine", "Ops"),
        (
            "Gate-B: FWD Draft MAX",
            "2.70",
            "m",
            "FWD <= this (critical stages)",
            "Marine",
            "Ops",
        ),
        ("Gate-C: TRIM ABS MAX", "240", "cm", "ABS(Trim_cm) <= this", "Marine", "Ops"),
        ("UKC min", "0.50", "m", "Under keel clearance minimum", "Marine", "Port"),
        (
            "DepthRef (charted/berth)",
            "5.50",
            "m",
            "Depth reference (same datum as tide)",
            "Marine",
            "Port",
        ),
        ("Squat", "0.00", "m", "Allowance", "Marine", ""),
        ("Safety Allowance", "0.00", "m", "Allowance", "Marine", ""),
        ("Pump rate (ship)", "10.0", "t/h", "Default ship pump rate", "Marine", ""),
        (
            "Vent flow coeff",
            "0.86",
            "t/h/mm",
            "Vent-limited effective rate",
            "Marine",
            "",
        ),
        ("Ramp length", "8.30", "m", "Ramp geometry", "Marine", ""),
        ("Linkspan height", "2.00", "m", "Reference height", "Marine", ""),
    ]

    for r, row in enumerate(ssot_rows, start=2):
        for c, v in enumerate(row, start=1):
            ws.cell(r, c, v)

        # styles
        ws.cell(r, 1).fill = styles.fill_section
        ws.cell(r, 1).font = Font(bold=True)
        ws.cell(r, 1).alignment = styles.align_left

        ws.cell(r, 2).fill = styles.fill_input
        ws.cell(r, 2).alignment = styles.align_left

        for c in range(3, 7):
            ws.cell(r, c).fill = styles.fill_calc
            ws.cell(r, c).alignment = styles.align_left

        for c in range(1, 7):
            ws.cell(r, c).border = styles.border_thin

    # DataValidation for SITE
    dv_site = DataValidation(type="list", formula1='"AGI"', allow_blank=False)
    ws.add_data_validation(dv_site)
    dv_site.add("B2")

    set_col_widths(
        ws, {"A": 32.0, "B": 22.0, "C": 10.0, "D": 55.0, "E": 16.0, "F": 18.0}
    )
    ws.freeze_panes = "A2"


def build_vessel_hydro_sheet(wb: Workbook, styles: Styles, n_rows: int = 60) -> None:
    ws = wb.create_sheet("02_Vessel_Hydro")

    ws.merge_cells("A1:H1")
    ws["A1"] = "Hydro Table (input)"
    ws["A1"].fill = styles.fill_title
    ws["A1"].font = styles.font_title
    ws["A1"].alignment = styles.align_center
    ws.row_dimensions[1].height = 20.0

    ws.merge_cells("A2:H2")
    ws["A2"] = (
        "Fill hydrostatic points sorted by Displacement ascending. Minimum 2 points required."
    )
    ws["A2"].fill = styles.fill_note
    ws["A2"].alignment = styles.align_left

    headers = [
        "#",
        "Disp_t",
        "Tmean_m",
        "TPC_t/cm",
        "MTC_t·m/cm",
        "LCF_m",
        "GM_m",
        "Source/Remark",
    ]
    make_header_row(ws, 3, headers, styles)
    ws.row_dimensions[3].height = 22.0

    start = 4
    end = start + n_rows - 1
    for i in range(start, end + 1):
        ws.cell(i, 1, i - (start - 1)).fill = styles.fill_calc
        ws.cell(i, 1).alignment = styles.align_center
        ws.cell(i, 1).border = styles.border_thin
        for c in range(2, 9):
            cell = ws.cell(i, c)
            cell.fill = styles.fill_input
            cell.border = styles.border_thin
            cell.alignment = styles.align_center

    # Excel table (optional)
    safe_add_table(ws, "HydroTable", f"A3:H{end}")

    set_col_widths(
        ws,
        {
            "A": 6.0,
            "B": 12.0,
            "C": 12.0,
            "D": 14.0,
            "E": 16.0,
            "F": 10.0,
            "G": 10.0,
            "H": 55.0,
        },
    )
    ws.freeze_panes = "A4"


def build_berth_tide_sheet(wb: Workbook, styles: Styles, tide_rows: int = 120) -> None:
    ws = wb.create_sheet("03_Berth_Tide")

    ws.merge_cells("A1:E1")
    ws["A1"] = "Berth / Depth Reference (input)"
    ws["A1"].fill = styles.fill_title
    ws["A1"].font = styles.font_title
    ws.row_dimensions[1].height = 20.0

    ws.merge_cells("A2:E2")
    ws["A2"] = "DepthRef must use same datum as tide (e.g., CD)."
    ws["A2"].fill = styles.fill_note

    make_header_row(ws, 3, ["Field", "Value", "Unit", "Notes", "Source"], styles)
    ws.row_dimensions[3].height = 22.0

    berth_data = [
        ("Site/Location", "", "", "Berth/jetty identification", ""),
        ("DepthRef", "5.50", "m", "Use latest sounding if available", ""),
        ("Datum", "CD", "", "e.g., CD", ""),
        ("UKC min", "0.50", "m", "", ""),
        ("Squat", "0.00", "m", "", ""),
        ("Safety Allowance", "0.00", "m", "", ""),
    ]
    for i, (field, value, unit, notes, source) in enumerate(berth_data, start=4):
        ws.cell(i, 1, field).fill = styles.fill_section
        ws.cell(i, 1).font = Font(bold=True)
        ws.cell(i, 2, value).fill = styles.fill_input
        ws.cell(i, 3, unit).fill = styles.fill_calc
        ws.cell(i, 4, notes).fill = styles.fill_calc
        ws.cell(i, 5, source).fill = styles.fill_calc
        for c in range(1, 6):
            ws.cell(i, c).border = styles.border_thin
            ws.cell(i, c).alignment = (
                styles.align_left if c in (1, 4, 5) else styles.align_center
            )

    # Tide forecast section
    ws.merge_cells("A11:E11")
    ws["A11"] = "Tide Forecast (input)"
    ws["A11"].fill = styles.fill_title
    ws["A11"].font = styles.font_title
    ws.row_dimensions[11].height = 20.0

    ws.merge_cells("A12:E12")
    ws["A12"] = "Fill official tide forecast (same datum)."
    ws["A12"].fill = styles.fill_note

    make_header_row(
        ws, 13, ["DateTime", "Tide_m", "Source", "Verified (Y/N)", "Remarks"], styles
    )
    ws.row_dimensions[13].height = 22.0

    first = 14
    last = first + tide_rows - 1
    dv_verified = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    ws.add_data_validation(dv_verified)
    dv_verified.add(f"D{first}:D{last}")

    for r in range(first, last + 1):
        # DateTime
        ws.cell(r, 1).fill = styles.fill_input
        ws.cell(r, 1).number_format = "yyyy-mm-dd hh:mm"
        # Tide
        ws.cell(r, 2).fill = styles.fill_input
        ws.cell(r, 2).number_format = "0.00"
        # Source
        ws.cell(r, 3).fill = styles.fill_input
        # Verified
        ws.cell(r, 4).fill = styles.fill_input
        # Remarks
        ws.cell(r, 5).fill = styles.fill_input

        for c in range(1, 6):
            ws.cell(r, c).border = styles.border_thin
            ws.cell(r, c).alignment = (
                styles.align_center if c in (1, 2, 4) else styles.align_left
            )

    safe_add_table(ws, "TideForecast", f"A13:E{last}")
    set_col_widths(ws, {"A": 20.0, "B": 12.0, "C": 18.0, "D": 16.0, "E": 45.0})
    ws.freeze_panes = "A14"


def build_cargo_spmt_sheet(wb: Workbook, styles: Styles) -> None:
    ws = wb.create_sheet("04_Cargo_SPMT")
    ws.protection.sheet = False
    ws.protection.enable = False
    for row in ws.iter_rows(min_row=1, max_row=100, min_col=1, max_col=10):
        for cell in row:
            cell.protection = Protection(locked=False)

    ws.merge_cells("A1:J1")
    ws["A1"] = "Cargo / SPMT Inputs (TR1/TR2 + SPMT)"
    ws["A1"].fill = styles.fill_title
    ws["A1"].font = styles.font_title
    ws["A1"].protection = Protection(locked=True)
    ws.row_dimensions[1].height = 20.0

    ws.merge_cells("A2:J2")
    ws["A2"] = "Fill cargo weights and CoG. Provide stage position if variable."
    ws["A2"].fill = styles.fill_note
    ws["A2"].protection = Protection(locked=True)

    headers = [
        "Item",
        "Weight_t",
        "CoG_x_m",
        "CoG_y_m",
        "CoG_z_m",
        "Final_Frame",
        "Ramp_Frame",
        "L_m",
        "W_m",
        "Remarks/Source",
    ]
    make_header_row(ws, 3, headers, styles)
    ws.row_dimensions[3].height = 22.0
    for col in range(1, 11):
        ws.cell(3, col).protection = Protection(locked=True)

    items = [
        ("TR1 + SPMT", "", "", "", "", "42", "17.95", "", "", ""),
        ("TR2 + SPMT", "", "", "", "", "40", "17.95", "", "", ""),
        ("SPMT Self-weight (if separated)", "", "", "", "", "", "", "", "", ""),
        ("Other / Dunnage / Lumber", "", "", "", "", "", "", "", "", ""),
    ]
    input_cols = {2, 3, 4, 5, 8, 9, 10}
    for i, row in enumerate(items, start=4):
        for j, v in enumerate(row, start=1):
            cell = ws.cell(i, j, v)
            cell.border = styles.border_thin
            cell.alignment = styles.align_center if j < 10 else styles.align_left
            is_input = j in input_cols
            cell.fill = styles.fill_input if is_input else styles.fill_calc
            if not is_input:
                cell.protection = Protection(locked=True)

    # Total (simple)
    ws.cell(9, 1, "Total cargo on LCT (TR1+TR2, excluding SPMT if separated)").fill = (
        styles.fill_section
    )
    ws.cell(9, 1).font = Font(bold=True)
    ws.cell(9, 1).protection = Protection(locked=True)
    ws.cell(9, 2, '=IFERROR(SUM(B4:B5),"")').fill = styles.fill_calc
    ws.cell(9, 2).border = styles.border_thin
    ws.cell(9, 2).number_format = "0.00"
    ws.cell(9, 2).protection = Protection(locked=True)
    ws.cell(9, 3, "t").fill = styles.fill_calc
    ws.cell(9, 3).border = styles.border_thin
    ws.cell(9, 3).protection = Protection(locked=True)
    for col in range(4, 11):
        ws.cell(9, col).protection = Protection(locked=True)

    set_col_widths(
        ws,
        {
            "A": 22.0,
            "B": 12.0,
            "C": 12.0,
            "D": 12.0,
            "E": 12.0,
            "F": 12.0,
            "G": 12.0,
            "H": 10.0,
            "I": 10.0,
            "J": 45.0,
        },
    )
    ws.freeze_panes = "A4"


def build_ballast_tanks_sheet(
    wb: Workbook, styles: Styles, tank_rows: int = 80, delta_rows: int = 120
) -> None:
    ws = wb.create_sheet("05_Ballast_Tanks")

    ws.merge_cells("A1:K1")
    ws["A1"] = "Ballast Tank Catalog (input)"
    ws["A1"].fill = styles.fill_title
    ws["A1"].font = styles.font_title
    ws.row_dimensions[1].height = 20.0

    ws.merge_cells("A2:K2")
    ws["A2"] = (
        "List all relevant tanks and initial ROB. Use same tank names as in ballast solver."
    )
    ws["A2"].fill = styles.fill_note

    # Constants (L/M)
    ws["L1"] = "VENT_COEFF (t/h/mm)"
    ws["M1"] = 0.86
    ws["L2"] = "PUMP_RATE (t/h) default"
    ws["M2"] = 10.0
    ws["L1"].font = Font(bold=True)
    ws["L2"].font = Font(bold=True)
    ws["M1"].number_format = "0.00"
    ws["M2"].number_format = "0.00"
    ws["M1"].fill = styles.fill_input
    ws["M2"].fill = styles.fill_input

    headers = [
        "Tank",
        "Side",
        "Family",
        "Capacity_t",
        "Initial_ROB_t",
        "Max_allow_t",
        "x_m",
        "Frame",
        "PumpGroup",
        "Vent_dia_mm",
        "Remarks/Source",
    ]
    make_header_row(ws, 3, headers, styles)
    ws.row_dimensions[3].height = 22.0

    # Tank catalog rows
    cat_first = 4
    cat_last = cat_first + tank_rows - 1

    for r in range(cat_first, cat_last + 1):
        for c in range(1, 12):
            cell = ws.cell(r, c)
            cell.border = styles.border_thin
            cell.alignment = styles.align_center if c != 11 else styles.align_left
            # Input columns
            if c in (1, 2, 4, 5, 6, 10, 11):
                cell.fill = styles.fill_input
            else:
                cell.fill = styles.fill_calc

    # Named range for tank list (used in delta section & elsewhere)
    add_named_range(
        wb, "Tanks_List", f"'05_Ballast_Tanks'!$A${cat_first}:$A${cat_last}"
    )

    # Stage-wise delta section
    sec_title_row = cat_last + 2
    sec_note_row = sec_title_row + 1
    header_row = sec_title_row + 2
    data_first = header_row + 1
    data_last = data_first + delta_rows - 1

    ws.merge_cells(f"A{sec_title_row}:K{sec_title_row}")
    ws[f"A{sec_title_row}"] = "Stage-wise Tank Deltas (input per stage)"
    ws[f"A{sec_title_row}"].fill = styles.fill_title
    ws[f"A{sec_title_row}"].font = styles.font_title
    ws.row_dimensions[sec_title_row].height = 20.0

    ws.merge_cells(f"A{sec_note_row}:K{sec_note_row}")
    ws[f"A{sec_note_row}"] = (
        "Fill delta (t) for each tank per stage. + = fill, - = discharge."
    )
    ws[f"A{sec_note_row}"].fill = styles.fill_note

    stage_cols = [
        "Tank",
        "Stage 1",
        "Stage 2",
        "Stage 3",
        "Stage 4",
        "Stage 5",
        "Stage 5_PreBallast",
        "Stage 6A_Critical (Opt C)",
        "Stage 6C",
        "Stage 7",
        "Remarks",
    ]
    make_header_row(ws, header_row, stage_cols, styles)
    ws.row_dimensions[header_row].height = 22.0

    dv_tank = DataValidation(type="list", formula1="=Tanks_List", allow_blank=True)
    ws.add_data_validation(dv_tank)
    dv_tank.add(f"A{data_first}:A{data_last}")

    for r in range(data_first, data_last + 1):
        for c in range(1, 12):
            cell = ws.cell(r, c)
            cell.border = styles.border_thin
            cell.alignment = styles.align_center if c != 11 else styles.align_left
            cell.fill = styles.fill_input if c <= 11 else styles.fill_calc
            if 2 <= c <= 10:
                cell.number_format = "0.00"

    set_col_widths(
        ws,
        {
            "A": 18.0,
            "B": 10.0,
            "C": 10.0,
            "D": 10.0,
            "E": 10.0,
            "F": 10.0,
            "G": 16.0,
            "H": 20.0,
            "I": 12.0,
            "J": 12.0,
            "K": 30.0,
            "L": 20.0,
            "M": 14.0,
        },
    )
    ws.freeze_panes = "A4"


def build_stage_plan_sheet(wb: Workbook, styles: Styles) -> None:
    """
    06_Stage_Plan: stage list (from analysis doc).
    """
    ws = wb.create_sheet("06_Stage_Plan")

    headers = [
        "Stage_ID",
        "Progress_%",
        "Planned_Start",
        "Planned_End",
        "HoldPoint",
        "GateB_Critical(Y/N)",
        "Stage_Status",
        "Remarks",
    ]
    make_header_row(ws, 1, headers, styles)
    ws.row_dimensions[1].height = 22.0

    # Stage IDs per analysis doc:contentReference[oaicite:1]{index=1}
    stages = [
        ("Stage 1", "N/A", "", "", "-", "N", "PLANNED", "Arrival (lightship)"),
        ("Stage 2", "N/A", "", "", "Y", "N", "PLANNED", "TR1 ramp start"),
        ("Stage 3", "N/A", "", "", "-", "N", "PLANNED", "TR1 mid-ramp"),
        ("Stage 4", "N/A", "", "", "-", "N", "PLANNED", "TR1 on deck"),
        ("Stage 5", "N/A", "", "", "-", "N", "PLANNED", "TR1 final position"),
        (
            "Stage 5_PreBallast",
            "N/A",
            "",
            "",
            "-",
            "Y",
            "PLANNED",
            "Water supply complete (TR1 stowed)",
        ),
        (
            "Stage 6A_Critical (Opt C)",
            "N/A",
            "",
            "",
            "-",
            "Y",
            "PLANNED",
            "TR2 ramp entry",
        ),
        ("Stage 6C", "N/A", "", "", "-", "N", "PLANNED", "Final stowage"),
        ("Stage 7", "N/A", "", "", "-", "N", "PLANNED", "Departure (cargo off)"),
    ]

    for i, row in enumerate(stages, start=2):
        for j, v in enumerate(row, start=1):
            cell = ws.cell(i, j, v)
            cell.border = styles.border_thin
            cell.alignment = styles.align_center if j != 8 else styles.align_left
            # allow editing for planned start/end/remarks
            if j in (3, 4, 8):
                cell.fill = styles.fill_input
            else:
                cell.fill = styles.fill_calc

    # Enough empty rows for extension
    for r in range(2 + len(stages), 202):
        for c in range(1, 9):
            cell = ws.cell(r, c)
            cell.border = styles.border_thin
            cell.alignment = styles.align_center if c != 8 else styles.align_left
            cell.fill = styles.fill_input if c in (3, 4, 8) else styles.fill_calc

    # Named range for Stage list (for DataValidation in other sheets)
    add_named_range(wb, "Stage_List", "'06_Stage_Plan'!$A$2:$A$200")

    dv_status = DataValidation(
        type="list",
        formula1='"PLANNED,IN-PROGRESS,COMPLETED,ABORTED"',
        allow_blank=True,
    )
    dv_critical = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    dv_hold = DataValidation(type="list", formula1='"Y,-"', allow_blank=True)

    ws.add_data_validation(dv_status)
    ws.add_data_validation(dv_critical)
    ws.add_data_validation(dv_hold)

    dv_status.add("G2:G200")
    dv_critical.add("F2:F200")
    dv_hold.add("E2:E200")

    set_col_widths(
        ws,
        {
            "A": 22.0,
            "B": 12.0,
            "C": 18.0,
            "D": 18.0,
            "E": 10.0,
            "F": 18.0,
            "G": 14.0,
            "H": 40.0,
        },
    )
    ws.freeze_panes = "A2"

    # Optional table
    safe_add_table(ws, "StagePlan", "A1:H200")


def build_stage_calc_sheet(wb: Workbook, styles: Styles, stage_rows: int = 200) -> None:
    """
    07_Stage_Calc: submission table.
    - Columns A..AG (33 cols)
    - Row 1 title
    - Row 2..10 parameters
    - Row 19 title, Row 20 header, Row 21.. stage rows
    """
    ws = wb.create_sheet("07_Stage_Calc")

    ws.merge_cells("A1:AG1")
    ws["A1"] = "Stage-wise Draft / Trim / Tide / Ramp Gates (submission table)"
    ws["A1"].fill = styles.fill_title
    ws["A1"].font = styles.font_title
    ws["A1"].alignment = styles.align_left
    ws.row_dimensions[1].height = 22.0

    # Parameter mini-table (A2:D10)
    ws["A2"] = "Parameter"
    ws["B2"] = "Value"
    ws["C2"] = "Unit"
    ws["D2"] = "Notes"
    for c in ("A2", "B2", "C2", "D2"):
        ws[c].fill = styles.fill_section
        ws[c].font = Font(bold=True)
        ws[c].border = styles.border_thin
        ws[c].alignment = styles.align_center

    params = [
        ("TOTAL_CARGO_T", 500, "t", "TR1+TR2 total (edit)"),
        ("GATE_AFT_MIN", 2.70, "m", "AFT draft must be >= this"),
        ("GATE_FWD_MAX", 2.70, "m", "FWD draft must be <= this (critical stages)"),
        ("TRIM_LIMIT_CM", 240, "cm", "ABS(trim_cm) <= this"),
        ("DEPTHREF_M", 5.50, "m", "Berth depth reference (same datum as tide)"),
        ("UKC_MIN_M", 0.50, "m", "Under keel clearance minimum"),
        ("SQUAT_M", 0.00, "m", "Allowance"),
        ("SAFETY_M", 0.00, "m", "Allowance"),
    ]
    for i, (p, v, u, note) in enumerate(params, start=3):
        ws.cell(i, 1, p).fill = styles.fill_section
        ws.cell(i, 1).font = Font(bold=True)
        ws.cell(i, 2, v).fill = styles.fill_input
        ws.cell(i, 3, u).fill = styles.fill_calc
        ws.cell(i, 4, note).fill = styles.fill_calc
        for c in range(1, 5):
            ws.cell(i, c).border = styles.border_thin
            ws.cell(i, c).alignment = (
                styles.align_left if c in (1, 4) else styles.align_center
            )

    # Stage Table Title
    ws.merge_cells("A19:AG19")
    ws["A19"] = "Stage-wise Calculation Table"
    ws["A19"].fill = styles.fill_title
    ws["A19"].font = styles.font_title
    ws["A19"].alignment = styles.align_left
    ws.row_dimensions[19].height = 20.0

    # Stage Table Header Row (Row 20)
    headers = [
        "Stage_ID",
        "HoldPoint",
        "GateB_Critical",
        "Stage_Remark",
        "Progress_%",
        "x_stage_m",
        "W_stage_t",
        "Displacement_t",
        "Tmean_m",
        "Trim_cm",
        "Draft_FWD_m",
        "Draft_AFT_m",
        "Freeboard_min_m",
        "Tide_required_m",
        "Forecast_tide_m",
        "Tide_margin_m",
        "UKC_min_m",
        "UKC_fwd_m",
        "UKC_aft_m",
        "Ramp_angle_deg",
        "GM_m",
        "Gate_FWD",
        "Gate_AFT",
        "Gate_TRIM",
        "Gate_UKC",
        "Gate_RAMP",
        "HoldPoint_Band",
        "Go/No-Go",
        "Ballast_Action",
        "Ballast_Time_h",
        "Ballast_Net_t",
        "Ballast_Alloc",
        "Notes",
    ]
    make_header_row(ws, 20, headers, styles)
    ws.row_dimensions[20].height = 26.0

    # Data validation for Stage_ID list
    dv_stage = DataValidation(type="list", formula1="=Stage_List", allow_blank=True)
    ws.add_data_validation(dv_stage)
    dv_stage.add("A21:A200")

    dv_hold = DataValidation(type="list", formula1='"Y,-"', allow_blank=True)
    dv_yesno = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    ws.add_data_validation(dv_hold)
    ws.add_data_validation(dv_yesno)
    dv_hold.add("B21:B200")  # HoldPoint
    dv_yesno.add("C21:C200")  # GateB_Critical

    dv_band = DataValidation(
        type="list", formula1='"GO,RECALC,STOP,-"', allow_blank=True
    )
    ws.add_data_validation(dv_band)
    dv_band.add("AA21:AA200")

    # Pre-fill first 9 stage rows from Stage_Plan (row 21..29)
    stage_ids = [
        "Stage 1",
        "Stage 2",
        "Stage 3",
        "Stage 4",
        "Stage 5",
        "Stage 5_PreBallast",
        "Stage 6A_Critical (Opt C)",
        "Stage 6C",
        "Stage 7",
    ]

    def _set_formula(r: int, col: int, formula: str) -> None:
        cell = ws.cell(r, col)
        cell.value = formula

    # Column indices (A=1 ... AG=33)
    COL_STAGE = 1
    COL_HOLD = 2
    COL_CRIT = 3
    COL_REMARK = 4
    COL_PROG = 5
    COL_X = 6
    COL_W = 7
    COL_DISP = 8
    COL_TMEAN = 9
    COL_TRIM = 10
    COL_DFWD = 11
    COL_DAFT = 12
    COL_FBMIN = 13
    COL_TIDE_REQ = 14
    COL_TIDE_FC = 15
    COL_TIDE_MARG = 16
    COL_UKC_MIN = 17
    COL_UKC_FWD = 18
    COL_UKC_AFT = 19
    COL_RAMP = 20
    COL_GM = 21
    COL_GATE_FWD = 22
    COL_GATE_AFT = 23
    COL_GATE_TRIM = 24
    COL_GATE_UKC = 25
    COL_GATE_RAMP = 26
    COL_HP_BAND = 27
    COL_GONOGO = 28
    COL_BALLAST_ACT = 29
    COL_BALLAST_T = 30
    COL_BALLAST_NET = 31
    COL_BALLAST_ALLOC = 32
    COL_NOTES = 33

    # Parameter cells for formulas
    # Row mapping in param table:
    # B4 = GATE_AFT_MIN, B5 = GATE_FWD_MAX, B6 = TRIM_LIMIT_CM, B7 = DEPTHREF_M, B8=UKC_MIN, B9=SQUAT, B10=SAFETY
    GATE_AFT_MIN = "$B$4"
    GATE_FWD_MAX = "$B$5"
    TRIM_LIMIT = "$B$6"
    DEPTHREF = "$B$7"
    UKC_MIN = "$B$8"
    SQUAT = "$B$9"
    SAFETY = "$B$10"

    for idx, sid in enumerate(stage_ids, start=0):
        r = 21 + idx
        ws.cell(r, COL_STAGE, sid).fill = styles.fill_input
        ws.cell(r, COL_STAGE).border = styles.border_thin
        ws.cell(r, COL_STAGE).alignment = styles.align_left

        # Pull HoldPoint / Critical / Remark / Progress from Stage_Plan (INDEX/MATCH)
        _set_formula(
            r,
            COL_HOLD,
            f'=IF($A{r}="","",IFERROR(INDEX(\'06_Stage_Plan\'!$E:$E,MATCH($A{r},\'06_Stage_Plan\'!$A:$A,0)),""))',
        )
        _set_formula(
            r,
            COL_CRIT,
            f'=IF($A{r}="","",IFERROR(INDEX(\'06_Stage_Plan\'!$F:$F,MATCH($A{r},\'06_Stage_Plan\'!$A:$A,0)),""))',
        )
        _set_formula(
            r,
            COL_REMARK,
            f'=IF($A{r}="","",IFERROR(INDEX(\'06_Stage_Plan\'!$H:$H,MATCH($A{r},\'06_Stage_Plan\'!$A:$A,0)),""))',
        )
        _set_formula(
            r,
            COL_PROG,
            f'=IF($A{r}="","",IFERROR(INDEX(\'06_Stage_Plan\'!$B:$B,MATCH($A{r},\'06_Stage_Plan\'!$A:$A,0)),""))',
        )

        # Trim (cm) from Drafts
        _set_formula(r, COL_TRIM, f'=IF(OR($K{r}="",$L{r}=""),"",($L{r}-$K{r})*100)')
        # Freeboard_min (m) using SSOT molded depth (search SSOT row "Molded Depth (D)")
        _set_formula(
            r,
            COL_FBMIN,
            f'=IF(OR($K{r}="",$L{r}=""),"",MIN('
            f"INDEX('01_SSOT_Master'!$B:$B,MATCH(\"Molded Depth (D)\",'01_SSOT_Master'!$A:$A,0))-$K{r},"
            f"INDEX('01_SSOT_Master'!$B:$B,MATCH(\"Molded Depth (D)\",'01_SSOT_Master'!$A:$A,0))-$L{r}))",
        )

        # UKC min
        _set_formula(r, COL_UKC_MIN, f'=IF($A{r}="","",{UKC_MIN})')

        # Tide required
        _set_formula(
            r,
            COL_TIDE_REQ,
            f'=IF(OR($K{r}="",$L{r}=""),"",MAX(0,(MAX($K{r},$L{r})+{SQUAT}+{SAFETY}+{UKC_MIN})-{DEPTHREF}))',
        )
        # Tide margin
        _set_formula(r, COL_TIDE_MARG, f'=IF(OR($O{r}="",$N{r}=""),"",$O{r}-$N{r})')

        # UKC forward/aft
        _set_formula(
            r,
            COL_UKC_FWD,
            f'=IF(OR($O{r}="",$K{r}=""),"",({DEPTHREF}+$O{r})-($K{r}+{SQUAT}+{SAFETY}))',
        )
        _set_formula(
            r,
            COL_UKC_AFT,
            f'=IF(OR($O{r}="",$L{r}=""),"",({DEPTHREF}+$O{r})-($L{r}+{SQUAT}+{SAFETY}))',
        )

        # Gate checks
        _set_formula(
            r,
            COL_GATE_FWD,
            f'=IF($A{r}="","",IF($C{r}<>"Y","N/A",IF($K{r}<={GATE_FWD_MAX},"PASS","LIMIT")))',
        )
        _set_formula(
            r,
            COL_GATE_AFT,
            f'=IF($A{r}="","",IF($L{r}>={GATE_AFT_MIN},"PASS","LIMIT"))',
        )
        _set_formula(
            r,
            COL_GATE_TRIM,
            f'=IF($A{r}="","",IF(ABS($J{r})<={TRIM_LIMIT},"PASS","LIMIT"))',
        )
        _set_formula(
            r,
            COL_GATE_UKC,
            f'=IF($A{r}="","",IF(MIN($R{r},$S{r})>={UKC_MIN},"PASS","LIMIT"))',
        )
        # Gate_RAMP left as VERIFY if ramp angle blank
        _set_formula(
            r,
            COL_GATE_RAMP,
            f'=IF($A{r}="","",IF($T{r}="","VERIFY",IF($T{r}<=6,"PASS","LIMIT")))',
        )
        # Holdpoint band / Go-NoGo (simple)
        _set_formula(
            r,
            COL_HP_BAND,
            f'=IF($B{r}<>"Y","-",IF(OR($V{r}="LIMIT",$W{r}="LIMIT",$X{r}="LIMIT",$Y{r}="LIMIT"),"STOP","GO"))',
        )
        _set_formula(
            r,
            COL_GONOGO,
            f'=IF($A{r}="","",IF($AA{r}="STOP","ABORT",IF($AA{r}="RECALC","NO-GO","GO")))',
        )

    # Fill styles for stage calc area (rows 21..200)
    for r in range(21, 21 + stage_rows):
        for c in range(1, 34):
            cell = ws.cell(r, c)
            cell.border = styles.border_thin
            # input columns (blue)
            if c in (1, 6, 7, 11, 12, 15, 20, 21, 29, 30, 31, 32, 33):
                cell.fill = styles.fill_input
            else:
                cell.fill = styles.fill_calc
            cell.alignment = (
                styles.align_center if c not in (4, 33) else styles.align_left
            )

    set_col_widths(
        ws,
        {
            "A": 22.0,
            "B": 10.0,
            "C": 14.0,
            "D": 40.0,
            "E": 12.0,
            "F": 12.0,
            "G": 12.0,
            "H": 14.0,
            "I": 12.0,
            "J": 12.0,
            "K": 12.0,
            "L": 12.0,
            "M": 14.0,
            "N": 14.0,
            "O": 14.0,
            "P": 14.0,
            "Q": 12.0,
            "R": 12.0,
            "S": 12.0,
            "T": 12.0,
            "U": 10.0,
            "V": 12.0,
            "W": 12.0,
            "X": 12.0,
            "Y": 12.0,
            "Z": 12.0,
            "AA": 14.0,
            "AB": 12.0,
            "AC": 18.0,
            "AD": 14.0,
            "AE": 14.0,
            "AF": 28.0,
            "AG": 40.0,
        },
    )
    ws.freeze_panes = "A21"
    safe_add_table(ws, "StageCalc", "A20:AG200")


def build_holdpoint_log_sheet(wb: Workbook, styles: Styles, n_rows: int = 200) -> None:
    ws = wb.create_sheet("08_HoldPoint_Log")

    ws.merge_cells("A1:L1")
    ws["A1"] = "Hold Point Measurement Log (Measured vs Predicted)"
    ws["A1"].fill = styles.fill_title
    ws["A1"].font = styles.font_title
    ws.row_dimensions[1].height = 22.0

    # Thresholds (B2/B3 used in formulas)
    ws["A2"] = "HP_GO_CM"
    ws["B2"] = 2
    ws["C2"] = "cm"
    ws["D2"] = "GO if max(|Δ|) <= this"

    ws["A3"] = "HP_RECALC_CM"
    ws["B3"] = 4
    ws["C3"] = "cm"
    ws["D3"] = "RECALC if <= this else STOP"

    for r in (2, 3):
        for c in range(1, 5):
            ws.cell(r, c).border = styles.border_thin
            ws.cell(r, c).alignment = (
                styles.align_left if c in (1, 4) else styles.align_center
            )
            ws.cell(r, c).fill = styles.fill_calc if c != 2 else styles.fill_input

    headers = [
        "Time",
        "Stage_ID",
        "Meas_FWD_m",
        "Meas_AFT_m",
        "Meas_Trim_cm",
        "Pred_FWD_m",
        "Pred_AFT_m",
        "ΔFWD_cm",
        "ΔAFT_cm",
        "ΔTrim_cm",
        "Action",
        "Comments",
    ]
    make_header_row(ws, 5, headers, styles)
    ws.row_dimensions[5].height = 24.0

    first = 6
    last = first + n_rows - 1

    # Stage list validation
    dv_stage = DataValidation(type="list", formula1="=Stage_List", allow_blank=True)
    ws.add_data_validation(dv_stage)
    dv_stage.add(f"B{first}:B{last}")

    for i in range(first, last + 1):
        # Predicted values auto lookup from 07_Stage_Calc
        ws.cell(i, 6).value = (
            f'=IF($B{i}="","",IFERROR(INDEX(\'07_Stage_Calc\'!$K:$K,MATCH($B{i},\'07_Stage_Calc\'!$A:$A,0)),""))'
        )
        ws.cell(i, 7).value = (
            f'=IF($B{i}="","",IFERROR(INDEX(\'07_Stage_Calc\'!$L:$L,MATCH($B{i},\'07_Stage_Calc\'!$A:$A,0)),""))'
        )

        # Meas_Trim_cm
        ws.cell(i, 5).value = f'=IF(OR(C{i}="",D{i}=""),"",(D{i}-C{i})*100)'
        # ΔFWD_cm
        ws.cell(i, 8).value = f'=IF(OR(C{i}="",F{i}=""),"",(C{i}-F{i})*100)'
        # ΔAFT_cm
        ws.cell(i, 9).value = f'=IF(OR(D{i}="",G{i}=""),"",(D{i}-G{i})*100)'
        # ΔTrim_cm
        ws.cell(i, 10).value = (
            f'=IF(OR(E{i}="",F{i}="",G{i}=""),"",E{i}-(G{i}-F{i})*100)'
        )
        # Action
        ws.cell(i, 11).value = (
            f'=IF(OR(H{i}="",I{i}="",J{i}=""),"",'
            f'IF(MAX(ABS(H{i}),ABS(I{i}),ABS(J{i}))<=$B$2,"GO",'
            f'IF(MAX(ABS(H{i}),ABS(I{i}),ABS(J{i}))<=$B$3,"RECALC","STOP")))'
        )

        # Input vs calc styling
        input_cols = (1, 2, 3, 4, 12)  # Time, Stage_ID, Meas, Comments
        calc_cols = (5, 6, 7, 8, 9, 10, 11)

        for c in input_cols:
            ws.cell(i, c).fill = styles.fill_input
        for c in calc_cols:
            ws.cell(i, c).fill = styles.fill_calc

        for c in range(1, 13):
            ws.cell(i, c).border = styles.border_thin
            ws.cell(i, c).alignment = (
                styles.align_center if c != 12 else styles.align_left
            )

        ws.cell(i, 1).number_format = "yyyy-mm-dd hh:mm"
        for c in (3, 4, 6, 7):
            ws.cell(i, c).number_format = "0.00"

    set_col_widths(
        ws,
        {
            "A": 18.0,
            "B": 22.0,
            "C": 14.0,
            "D": 14.0,
            "E": 14.0,
            "F": 14.0,
            "G": 14.0,
            "H": 14.0,
            "I": 14.0,
            "J": 14.0,
            "K": 12.0,
            "L": 45.0,
        },
    )
    ws.freeze_panes = "A6"
    safe_add_table(ws, "HoldPointLog", f"A5:L{last}")


def build_evidence_log_sheet(wb: Workbook, styles: Styles, rows: int = 200) -> None:
    ws = wb.create_sheet("09_Evidence_Log")

    headers = [
        "Site",
        "Category",
        "Document / Evidence Item",
        "Req",
        "Status",
        "Due/Submitted",
        "Portal",
        "Ref No.",
        "File/Link",
        "Remarks",
    ]
    make_header_row(ws, 1, headers, styles)
    ws.row_dimensions[1].height = 22.0

    evidence_items = [
        (
            "COMMON",
            "Booking",
            "Booking Confirmation",
            "Y",
            "PENDING",
            "",
            "ATLP/MTG",
            "",
            "",
            "",
        ),
        (
            "COMMON",
            "Insurance",
            "Insurance Certificate",
            "Y",
            "PENDING",
            "",
            "ATLP/MTG",
            "",
            "",
            "",
        ),
        (
            "COMMON",
            "Manifest",
            "Cargo Manifest",
            "Y",
            "PENDING",
            "",
            "ATLP/MTG",
            "",
            "",
            "",
        ),
        (
            "COMMON",
            "Contacts",
            "Emergency Contact List",
            "Y",
            "PENDING",
            "",
            "ATLP/MTG",
            "",
            "",
            "",
        ),
        (
            "AGI",
            "Ramp",
            "Ramp Angle Calculation",
            "Y",
            "PENDING",
            "",
            "MTG",
            "",
            "",
            "",
        ),
        (
            "AGI",
            "Berth",
            "Berth Load Chart / Depth Confirmation",
            "Y",
            "PENDING",
            "",
            "MTG",
            "",
            "",
            "",
        ),
        (
            "COMMON",
            "Evidence",
            "Hold Point Draft Photo (FWD/AFT)",
            "Y",
            "PENDING",
            "",
            "ATLP/MTG",
            "",
            "",
            "",
        ),
    ]

    # Write defaults
    r0 = 2
    for i, row in enumerate(evidence_items, start=r0):
        for j, v in enumerate(row, start=1):
            cell = ws.cell(i, j, v)
            cell.border = styles.border_thin
            cell.alignment = (
                styles.align_left if j in (3, 9, 10) else styles.align_center
            )
            cell.fill = styles.fill_input if j in (5, 6, 8, 9, 10) else styles.fill_calc

    # Fill remaining blank rows
    end = r0 + rows - 1
    for r in range(r0 + len(evidence_items), end + 1):
        for c in range(1, 11):
            cell = ws.cell(r, c)
            cell.border = styles.border_thin
            cell.alignment = (
                styles.align_left if c in (3, 9, 10) else styles.align_center
            )
            cell.fill = styles.fill_input if c in (5, 6, 8, 9, 10) else styles.fill_calc

    # Data validation
    dv_site = DataValidation(type="list", formula1='"COMMON,AGI"', allow_blank=True)
    dv_status = DataValidation(
        type="list", formula1='"PENDING,SUBMITTED,APPROVED,REJECTED"', allow_blank=True
    )
    ws.add_data_validation(dv_site)
    ws.add_data_validation(dv_status)
    dv_site.add(f"A2:A{end}")
    dv_status.add(f"E2:E{end}")

    set_col_widths(
        ws,
        {
            "A": 10.0,
            "B": 14.0,
            "C": 38.0,
            "D": 8.0,
            "E": 14.0,
            "F": 16.0,
            "G": 12.0,
            "H": 16.0,
            "I": 40.0,
            "J": 40.0,
        },
    )
    ws.freeze_panes = "A2"
    safe_add_table(ws, "EvidenceLog", f"A1:J{end}")


def build_raci_sheet(wb: Workbook, styles: Styles) -> None:
    """
    10_RACI sheet per analysis doc.:contentReference[oaicite:2]{index=2}
    """
    ws = wb.create_sheet("10_RACI")
    headers = [
        "Activity / Deliverable",
        "SCT Logistics",
        "DSV/3PL",
        "Vessel Master",
        "Chief Officer/Ballast",
        "Port Authority",
        "HSE",
        "Notes",
    ]
    make_header_row(ws, 1, headers, styles)
    ws.row_dimensions[1].height = 22.0

    raci_items = [
        (
            "Hydrostatic data pack + sign convention",
            "A",
            "C",
            "R",
            "R",
            "I",
            "I",
            "Provide Hydro table and calculation method",
        ),
        (
            "Stage-wise plan + timeline",
            "A",
            "C",
            "C",
            "R",
            "I",
            "I",
            "Include load transfer %, ballast sequence",
        ),
        (
            "Ballast tank list + initial ROB",
            "C",
            "I",
            "I",
            "R",
            "I",
            "I",
            "Include pump/vent constraints",
        ),
        (
            "Tide forecast + depth confirmation",
            "C",
            "I",
            "R",
            "R",
            "A",
            "I",
            "Use official source; align datum",
        ),
        (
            "HoldPoint measurement & decision",
            "C",
            "I",
            "R",
            "R",
            "I",
            "A",
            "Record drafts/trim; decide GO/RECALC/STOP",
        ),
        (
            "ATLP/MTG submissions",
            "R",
            "C",
            "I",
            "I",
            "A",
            "C",
            "Site-specific lead times",
        ),
        (
            "Photo/video evidence collection",
            "A",
            "C",
            "I",
            "I",
            "I",
            "R",
            "Tag by stage/time/GPS",
        ),
        (
            "Final report & archive",
            "A",
            "C",
            "C",
            "R",
            "I",
            "C",
            "7-day submission requirement",
        ),
    ]

    for i, row in enumerate(raci_items, start=2):
        for j, v in enumerate(row, start=1):
            cell = ws.cell(i, j, v)
            cell.fill = styles.fill_calc
            cell.border = styles.border_thin
            cell.alignment = styles.align_left if j in (1, 8) else styles.align_center

    set_col_widths(
        ws,
        {
            "A": 38.0,
            "B": 14.0,
            "C": 14.0,
            "D": 14.0,
            "E": 18.0,
            "F": 14.0,
            "G": 10.0,
            "H": 40.0,
        },
    )
    ws.freeze_panes = "A2"


def build_assumptions_issues_sheet(
    wb: Workbook, styles: Styles, rows: int = 200
) -> None:
    ws = wb.create_sheet("11_Assumptions_Issues")
    make_header_row(
        ws,
        1,
        [
            "ID",
            "Type",
            "Description",
            "Value",
            "Unit",
            "Impact",
            "Owner",
            "Due",
            "Status",
            "Evidence/Ref",
        ],
        styles,
    )
    ws.row_dimensions[1].height = 22.0

    defaults = [
        (
            1,
            "Assumption",
            "Hydro points provided are accurate and sorted by displacement",
            "",
            "",
            "High",
            "Marine",
            "",
            "OPEN",
            "",
        ),
        (
            2,
            "Assumption",
            "DepthRef uses same datum as tide (CD)",
            "",
            "",
            "High",
            "Marine",
            "",
            "OPEN",
            "",
        ),
        (
            3,
            "Issue",
            "Lightship displacement/draft not confirmed",
            "",
            "",
            "Med",
            "Marine",
            "",
            "OPEN",
            "",
        ),
        (
            4,
            "Gap",
            "No clear explanation of draft mark positions",
            "",
            "",
            "Med",
            "Marine",
            "",
            "OPEN",
            "",
        ),
    ]
    for i, row in enumerate(defaults, start=2):
        for j, v in enumerate(row, start=1):
            cell = ws.cell(i, j, v)
            cell.fill = styles.fill_input if j in (4, 5, 8, 9, 10) else styles.fill_calc
            cell.border = styles.border_thin
            cell.alignment = styles.align_left if j == 3 else styles.align_center

    # Fill blanks
    end = 1 + rows
    for r in range(2 + len(defaults), end + 1):
        for c in range(1, 11):
            cell = ws.cell(r, c)
            cell.fill = styles.fill_input if c in (4, 5, 8, 9, 10) else styles.fill_calc
            cell.border = styles.border_thin
            cell.alignment = styles.align_left if c == 3 else styles.align_center

    dv_type = DataValidation(
        type="list", formula1='"Assumption,Issue,Gap"', allow_blank=True
    )
    dv_impact = DataValidation(type="list", formula1='"Low,Med,High"', allow_blank=True)
    dv_status = DataValidation(
        type="list", formula1='"OPEN,CLOSED,MITIGATED"', allow_blank=True
    )
    ws.add_data_validation(dv_type)
    ws.add_data_validation(dv_impact)
    ws.add_data_validation(dv_status)
    dv_type.add(f"B2:B{end}")
    dv_impact.add(f"F2:F{end}")
    dv_status.add(f"I2:I{end}")

    set_col_widths(
        ws,
        {
            "A": 8.0,
            "B": 14.0,
            "C": 55.0,
            "D": 16.0,
            "E": 10.0,
            "F": 12.0,
            "G": 14.0,
            "H": 16.0,
            "I": 14.0,
            "J": 40.0,
        },
    )
    ws.freeze_panes = "A2"


def build_data_dictionary_sheet(wb: Workbook, styles: Styles) -> None:
    ws = wb.create_sheet("12_DataDictionary")
    make_header_row(
        ws, 1, ["Sheet", "Field", "Unit", "Type", "Definition / Rule"], styles
    )
    ws.row_dimensions[1].height = 22.0

    dict_items = [
        (
            "01_SSOT_Master",
            "GATE_AFT_MIN_M",
            "m",
            "Input",
            "Minimum aft draft allowed (AFT >=).",
        ),
        (
            "01_SSOT_Master",
            "GATE_FWD_MAX_M",
            "m",
            "Input",
            "Maximum forward draft allowed (FWD <=).",
        ),
        (
            "01_SSOT_Master",
            "TRIM_LIMIT_CM",
            "cm",
            "Input",
            "Absolute trim limit. Trim = (Daft - Dfwd)*100.",
        ),
        (
            "03_Berth_Tide",
            "Tide_m",
            "m",
            "Input",
            "Tide height above datum (e.g., CD).",
        ),
        (
            "04_Cargo_SPMT",
            "Weight_t",
            "t",
            "Input",
            "Cargo weight including SPMT if not separated.",
        ),
        (
            "05_Ballast_Tanks",
            "Delta_t",
            "t",
            "Input",
            "Per stage tank delta: + ballast, - discharge.",
        ),
        ("07_Stage_Calc", "Trim_cm", "cm", "Calc", "Trim = (Daft - Dfwd)*100."),
        (
            "07_Stage_Calc",
            "Tide_required_m",
            "m",
            "Calc",
            "Required tide so UKC >= UKC_MIN.",
        ),
        (
            "08_HoldPoint_Log",
            "Action",
            "-",
            "Calc",
            "GO/RECALC/STOP based on thresholds.",
        ),
    ]

    for i, row in enumerate(dict_items, start=2):
        for j, v in enumerate(row, start=1):
            cell = ws.cell(i, j, v)
            cell.fill = styles.fill_calc
            cell.border = styles.border_thin
            cell.alignment = styles.align_left if j in (2, 5) else styles.align_center

    set_col_widths(ws, {"A": 18.0, "B": 22.0, "C": 12.0, "D": 14.0, "E": 60.0})
    ws.freeze_panes = "A2"


def build_pack_a_inputs_ssot_sheet(wb: Workbook, styles: Styles) -> None:
    ws = wb.create_sheet("Pack_A_Inputs_SSOT")

    ws.merge_cells("A1:E1")
    ws["A1"] = "Pack A: Inputs & Assumptions (SSOT)"
    ws["A1"].fill = styles.fill_title
    ws["A1"].font = styles.font_title
    ws["A1"].alignment = styles.align_left

    # Section 1
    ws.merge_cells("A3:E3")
    ws["A3"] = "1. Vessel / Show-stopper inputs"
    ws["A3"].fill = styles.fill_section
    ws["A3"].font = Font(bold=True)

    make_header_row(
        ws, 4, ["Parameter", "Value", "Unit", "Source/Reference", "Note"], styles
    )

    vessel_params = [
        ("LBP (Length Between Perpendiculars)", "", "m", "", ""),
        ("Beam", "", "m", "", ""),
        ("Depth (D, molded)", "3.65", "m", "Vessel spec", ""),
        ("Lightship Displacement", "2800", "t", "Vessel spec", ""),
        ("Gate AFT MIN", "2.70", "m", "Ops", ""),
        ("Gate FWD MAX", "2.70", "m", "Ops", ""),
        ("Trim ABS limit", "240", "cm", "Ops", ""),
        ("UKC min", "0.50", "m", "Port", ""),
        ("DepthRef", "5.50", "m", "Port", ""),
    ]
    for i, row in enumerate(vessel_params, start=5):
        for j, v in enumerate(row, start=1):
            cell = ws.cell(i, j, v)
            cell.border = styles.border_thin
            cell.fill = (
                styles.fill_input
                if j == 2
                else (styles.fill_section if j == 1 else styles.fill_calc)
            )
            cell.alignment = (
                styles.align_left if j in (1, 4, 5) else styles.align_center
            )
        ws.cell(i, 1).font = Font(bold=True)

    # Keep it simple but leave room
    set_col_widths(ws, {"A": 35.0, "B": 18.0, "C": 12.0, "D": 25.0, "E": 35.0})
    ws.freeze_panes = "A5"


def build_pack_b_calculation_formulas_sheet(wb: Workbook, styles: Styles) -> None:
    ws = wb.create_sheet("Pack_B_Calculation_Formulas")

    ws.merge_cells("A1:D1")
    ws["A1"] = "Pack B: Calculation Sheet (Formula + Sign Convention)"
    ws["A1"].fill = styles.fill_title
    ws["A1"].font = styles.font_title
    ws["A1"].alignment = styles.align_left

    # (1) Draft/Trim
    ws.merge_cells("A3:D3")
    ws["A3"] = "(1) Draft / Trim (Hydro 기반)"
    ws["A3"].fill = styles.fill_section
    ws["A3"].font = Font(bold=True)

    make_header_row(
        ws,
        4,
        ["Parameter/Formula", "Formula/Definition", "Unit/Note", "Reference"],
        styles,
    )
    formulas = [
        ("Tmean (m)", "T_mean = (D_fwd + D_aft) / 2", "", ""),
        (
            "Delta Tmean (m)",
            "Delta T_mean = Sum(Delta W) / (TPC * 100)",
            "TPC: t/cm",
            "",
        ),
        (
            "Delta Trim (m)",
            "Delta Trim = Sum(Delta W*(x-LCF)) / MTC",
            "MTC: t·m/cm",
            "",
        ),
        ("Trim (cm)", "Trim_cm = (D_aft - D_fwd) * 100", "", ""),
        ("Sign: Trim", "+ = Aft deeper / - = Bow down", "", ""),
        ("Sign: x", "+Aft / -Fwd (midship ref)", "", ""),
    ]
    for i, row in enumerate(formulas, start=5):
        for j, v in enumerate(row, start=1):
            cell = ws.cell(i, j, v)
            cell.border = styles.border_thin
            cell.alignment = styles.align_left
            cell.fill = styles.fill_calc
        ws.cell(i, 1).fill = styles.fill_section
        ws.cell(i, 1).font = Font(bold=True)

    # (2) UKC / Tide Required
    start2 = 13
    ws.merge_cells(f"A{start2}:D{start2}")
    ws[f"A{start2}"] = "(2) UKC / Tide Required"
    ws[f"A{start2}"].fill = styles.fill_section
    ws[f"A{start2}"].font = Font(bold=True)

    make_header_row(
        ws, start2 + 1, ["Item", "Definition", "Unit/Note", "Reference"], styles
    )
    rows2 = [
        ("Depth_available", "DepthRef + Forecast_Tide", "m", "03_Berth_Tide"),
        ("UKC", "Depth_available - (Draft + Squat + Safety)", "m", ""),
        (
            "Tide_required",
            "max(0, (Draft_max + Squat + Safety + UKC_min) - DepthRef)",
            "m",
            "",
        ),
    ]
    for i, row in enumerate(rows2, start=start2 + 2):
        for j, v in enumerate(row, start=1):
            cell = ws.cell(i, j, v)
            cell.border = styles.border_thin
            cell.alignment = styles.align_left
            cell.fill = styles.fill_calc
        ws.cell(i, 1).fill = styles.fill_section
        ws.cell(i, 1).font = Font(bold=True)

    set_col_widths(ws, {"A": 30.0, "B": 60.0, "C": 22.0, "D": 22.0})
    ws.freeze_panes = "A5"


def build_pack_d_tide_window_raci_sheet(wb: Workbook, styles: Styles) -> None:
    ws = wb.create_sheet("Pack_D_Tide_Window_RACI")

    ws.merge_cells("A1:F1")
    ws["A1"] = "Pack D: Tide Window & RACI"
    ws["A1"].fill = styles.fill_title
    ws["A1"].font = styles.font_title
    ws["A1"].alignment = styles.align_left

    ws.merge_cells("A3:F3")
    ws["A3"] = "Stage-wise Tide Window"
    ws["A3"].fill = styles.fill_section
    ws["A3"].font = Font(bold=True)

    make_header_row(
        ws,
        4,
        [
            "Stage",
            "Required WL (m)",
            "Forecast Tide (m)",
            "Margin (m)",
            "Window Time",
            "Source",
        ],
        styles,
    )

    stages = [
        "Stage 1",
        "Stage 2",
        "Stage 3",
        "Stage 4",
        "Stage 5",
        "Stage 5_PreBallast",
        "Stage 6A_Critical (Opt C)",
        "Stage 6C",
        "Stage 7",
    ]
    for i, stage in enumerate(stages, start=5):
        ws.cell(i, 1, stage).fill = styles.fill_calc
        ws.cell(i, 1).border = styles.border_thin
        ws.cell(i, 1).alignment = styles.align_left
        # link required/forecast to Stage_Calc if desired (leave as input but add margin formula)
        ws.cell(i, 2).fill = styles.fill_input
        ws.cell(i, 3).fill = styles.fill_input
        ws.cell(i, 4).value = f'=IF(OR(B{i}="",C{i}=""),"",C{i}-B{i})'
        ws.cell(i, 4).fill = styles.fill_calc
        ws.cell(i, 5).fill = styles.fill_input
        ws.cell(i, 6).fill = styles.fill_input

        for c in range(2, 7):
            ws.cell(i, c).border = styles.border_thin
            ws.cell(i, c).alignment = (
                styles.align_center if c != 6 else styles.align_left
            )

    # RACI section (simple placeholder)
    ws.merge_cells("A16:F16")
    ws["A16"] = "RACI Matrix (optional)"
    ws["A16"].fill = styles.fill_section
    ws["A16"].font = Font(bold=True)

    set_col_widths(
        ws, {"A": 26.0, "B": 18.0, "C": 18.0, "D": 14.0, "E": 18.0, "F": 22.0}
    )
    ws.freeze_panes = "A5"


# =============================================================================
# Main generator
# =============================================================================
def generate(output_xlsx: Path) -> None:
    styles = build_styles()
    wb = Workbook()
    # remove default sheet
    wb.remove(wb.active)

    # Sheet order per analysis doc:contentReference[oaicite:3]{index=3}
    build_readme_sheet(wb, styles)
    build_ssot_master_sheet(wb, styles)
    build_vessel_hydro_sheet(wb, styles)
    build_berth_tide_sheet(wb, styles)
    build_cargo_spmt_sheet(wb, styles)
    build_ballast_tanks_sheet(wb, styles)
    build_stage_plan_sheet(wb, styles)
    build_stage_calc_sheet(wb, styles)
    build_holdpoint_log_sheet(wb, styles)
    build_evidence_log_sheet(wb, styles)
    build_raci_sheet(wb, styles)
    build_assumptions_issues_sheet(wb, styles)
    build_data_dictionary_sheet(wb, styles)
    build_pack_a_inputs_ssot_sheet(wb, styles)
    build_pack_b_calculation_formulas_sheet(wb, styles)
    build_pack_d_tide_window_raci_sheet(wb, styles)

    # Calculation mode (Excel will recalc on open)
    try:
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.calcOnSave = True
    except Exception:
        pass

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)


def main() -> None:
    import argparse

    ap = argparse.ArgumentParser()
    ap.add_argument(
        "--out",
        default="Bryan_Submission_Data_Pack_Template_GENERATED.xlsx",
        help="Output xlsx path",
    )
    args = ap.parse_args()

    out_path = Path(args.out).resolve()
    generate(out_path)
    print(f"[OK] Template generated: {out_path}")


if __name__ == "__main__":
    main()
