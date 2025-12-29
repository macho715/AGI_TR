#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
bryan_template_unified.py

Unified entrypoint for Bryan template workflows:
- create (generate template)
- populate (fill 07_Stage_Calc from stage_results.csv)
- validate (basic checks on populated file)
- one-click (create -> optional SPMT import -> populate -> optional validate)

This script uses create_bryan_excel_template_NEW.py for template generation
and embeds populate/validate logic for a single-file workflow.
"""

from __future__ import annotations

import argparse
import importlib.util
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


BASE_DIR = Path(__file__).resolve().parent


def _load_module(name: str, path: Path):
    spec = importlib.util.spec_from_file_location(name, path)
    if spec is None or spec.loader is None:
        raise ImportError(f"Cannot load module: {path}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def _resolve_script_path(rel_path: str) -> Path:
    """
    Resolve dependent script path with fallback to parent dir and first filename match.
    """
    candidate = (BASE_DIR / rel_path).resolve()
    if candidate.exists():
        return candidate
    alt = (BASE_DIR.parent / rel_path).resolve()
    if alt.exists():
        return alt
    name = Path(rel_path).name
    matches = sorted(BASE_DIR.parent.rglob(name))
    if matches:
        return matches[0].resolve()
    return candidate


# -------------------------------------------------------------------------
# SPMT cargo import into 04_Cargo_SPMT
# -------------------------------------------------------------------------
def _norm_header(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(text).lower())


def _find_header_row(ws: Worksheet, target_headers: List[str], max_rows: int = 10) -> int:
    target_norm = [_norm_header(h) for h in target_headers]
    for r in range(1, max_rows + 1):
        row_vals = [_norm_header(ws.cell(r, c).value or "") for c in range(1, 11)]
        if all(h in row_vals for h in target_norm[:2]):
            return r
    return 0


def _read_spmt_rows(ws: Worksheet, header_row: int) -> Dict[str, Dict[str, object]]:
    headers = {
        _norm_header(ws.cell(header_row, c).value or ""): c for c in range(1, 11)
    }
    data: Dict[str, Dict[str, object]] = {}
    for r in range(header_row + 1, header_row + 8):
        item = ws.cell(r, headers.get("item", 1)).value
        if item is None or str(item).strip() == "":
            continue
        row = {}
        for key in (
            "item",
            "weightt",
            "cogxm",
            "cogym",
            "cogzm",
            "finalframe",
            "rampframe",
            "lm",
            "wm",
            "remarkssource",
        ):
            col = headers.get(key)
            if col:
                row[key] = ws.cell(r, col).value
        data[str(item).strip()] = row
    return data


def _apply_spmt_rows(target_ws: Worksheet, rows: Dict[str, Dict[str, object]]) -> None:
    header_row = _find_header_row(
        target_ws,
        ["Item", "Weight_t", "CoG_x_m"],
        max_rows=10,
    )
    if header_row == 0:
        raise ValueError("04_Cargo_SPMT header row not found.")
    headers = {
        _norm_header(target_ws.cell(header_row, c).value or ""): c
        for c in range(1, 11)
    }
    for r in range(header_row + 1, header_row + 8):
        item = target_ws.cell(r, headers.get("item", 1)).value
        if item is None or str(item).strip() == "":
            continue
        src = rows.get(str(item).strip())
        if not src:
            continue
        for key, col_key in (
            ("weightt", "weightt"),
            ("cogxm", "cogxm"),
            ("cogym", "cogym"),
            ("cogzm", "cogzm"),
            ("finalframe", "finalframe"),
            ("rampframe", "rampframe"),
            ("lm", "lm"),
            ("wm", "wm"),
            ("remarkssource", "remarkssource"),
        ):
            col = headers.get(col_key)
            if col:
                target_ws.cell(r, col).value = src.get(key)


def import_spmt_cargo(spmt_xlsx: Path, target_xlsx: Path) -> None:
    wb_spmt = openpyxl.load_workbook(spmt_xlsx, data_only=True)
    if "Cargo_SPMT_Inputs" not in wb_spmt.sheetnames:
        raise ValueError("SPMT workbook missing Cargo_SPMT_Inputs sheet.")
    ws_spmt = wb_spmt["Cargo_SPMT_Inputs"]
    header_row = _find_header_row(ws_spmt, ["Item", "Weight_t"], max_rows=10)
    if header_row == 0:
        raise ValueError("SPMT Cargo_SPMT_Inputs header row not found.")
    rows = _read_spmt_rows(ws_spmt, header_row)
    wb_spmt.close()

    wb_target = openpyxl.load_workbook(target_xlsx)
    if "04_Cargo_SPMT" not in wb_target.sheetnames:
        raise ValueError("Target workbook missing 04_Cargo_SPMT sheet.")
    _apply_spmt_rows(wb_target["04_Cargo_SPMT"], rows)
    wb_target.save(target_xlsx)
    wb_target.close()


# -------------------------------------------------------------------------
# Populate 07_Stage_Calc (embedded from populate_template.py)
# -------------------------------------------------------------------------
def _norm_text(x) -> str:
    if x is None:
        return ""
    return str(x).strip()


def _norm_col(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(name).lower())


def pick_column(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
    cols = list(df.columns)
    for cand in candidates:
        if cand in cols:
            return cand
    norm_map = {_norm_col(c): c for c in cols}
    for cand in candidates:
        key = _norm_col(cand)
        if key in norm_map:
            return norm_map[key]
    return None


def safe_float(x) -> Optional[float]:
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip()
    if s == "" or s.lower() in ("nan", "none", "null"):
        return None
    try:
        return float(s)
    except Exception:
        try:
            return float(s.replace(",", ""))
        except Exception:
            return None


def get_or_create_sheet(wb, name: str):
    if name in wb.sheetnames:
        return wb[name]
    return wb.create_sheet(name)


def append_log(ws: Worksheet, msg: str) -> None:
    if ws.max_row == 1 and ws["A1"].value is None:
        ws["A1"] = "Time"
        ws["B1"] = "Message"
    r = ws.max_row + 1
    ws.cell(r, 1, datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    ws.cell(r, 2, msg)


def find_header_row(
    ws: Worksheet, header_text: str = "Stage_ID", search_rows: int = 40
) -> Tuple[int, int]:
    for r in range(1, search_rows + 1):
        if _norm_text(ws.cell(r, 1).value) == header_text:
            return r, r + 1
    return 20, 21


def find_or_create_stage_row(
    ws: Worksheet,
    stage_id: str,
    data_start: int,
    soft_last_row: int = 200,
) -> int:
    last_scan = max(soft_last_row, ws.max_row)
    for r in range(data_start, last_scan + 1):
        if _norm_text(ws.cell(r, 1).value) == stage_id:
            return r
    for r in range(data_start, last_scan + 1):
        if _norm_text(ws.cell(r, 1).value) == "":
            ws.cell(r, 1).value = stage_id
            return r
    r = last_scan + 1
    ws.cell(r, 1).value = stage_id
    return r


def ensure_stage_formulas(ws: Worksheet, r: int) -> None:
    def set_if_blank(col: int, formula: str):
        cell = ws.cell(r, col)
        if cell.value is None or str(cell.value).strip() == "":
            cell.value = formula

    GATE_AFT_MIN = "$B$4"
    GATE_FWD_MAX = "$B$5"
    TRIM_LIMIT = "$B$6"
    DEPTHREF = "$B$7"
    UKC_MIN = "$B$8"
    SQUAT = "$B$9"
    SAFETY = "$B$10"

    COL_HOLD = 2
    COL_CRIT = 3
    COL_REMARK = 4
    COL_PROG = 5
    COL_TRIM = 10
    COL_FBMIN = 13
    COL_TIDE_REQ = 14
    COL_TIDE_MARG = 16
    COL_UKC_MIN = 17
    COL_UKC_FWD = 18
    COL_UKC_AFT = 19
    COL_GATE_FWD = 22
    COL_GATE_AFT = 23
    COL_GATE_TRIM = 24
    COL_GATE_UKC = 25
    COL_GATE_RAMP = 26
    COL_HP_BAND = 27
    COL_GONOGO = 28

    set_if_blank(
        COL_HOLD,
        f'=IF($A{r}="","",IFERROR(INDEX(\'06_Stage_Plan\'!$E:$E,MATCH($A{r},\'06_Stage_Plan\'!$A:$A,0)),""))',
    )
    set_if_blank(
        COL_CRIT,
        f'=IF($A{r}="","",IFERROR(INDEX(\'06_Stage_Plan\'!$F:$F,MATCH($A{r},\'06_Stage_Plan\'!$A:$A,0)),""))',
    )
    set_if_blank(
        COL_REMARK,
        f'=IF($A{r}="","",IFERROR(INDEX(\'06_Stage_Plan\'!$H:$H,MATCH($A{r},\'06_Stage_Plan\'!$A:$A,0)),""))',
    )
    set_if_blank(
        COL_PROG,
        f'=IF($A{r}="","",IFERROR(INDEX(\'06_Stage_Plan\'!$B:$B,MATCH($A{r},\'06_Stage_Plan\'!$A:$A,0)),""))',
    )

    set_if_blank(COL_TRIM, f'=IF(OR($K{r}="",$L{r}=""),"",($L{r}-$K{r})*100)')
    set_if_blank(
        COL_FBMIN,
        f'=IF(OR($K{r}="",$L{r}=""),"",MIN('
        f'INDEX(\'01_SSOT_Master\'!$B:$B,MATCH("Molded Depth (D)",\'01_SSOT_Master\'!$A:$A,0))-$K{r},'
        f'INDEX(\'01_SSOT_Master\'!$B:$B,MATCH("Molded Depth (D)",\'01_SSOT_Master\'!$A:$A,0))-$L{r}))',
    )
    set_if_blank(COL_UKC_MIN, f'=IF($A{r}="","",{UKC_MIN})')
    set_if_blank(
        COL_TIDE_REQ,
        f'=IF(OR($K{r}="",$L{r}=""),"",MAX(0,(MAX($K{r},$L{r})+{SQUAT}+{SAFETY}+{UKC_MIN})-{DEPTHREF}))',
    )
    set_if_blank(COL_TIDE_MARG, f'=IF(OR($O{r}="",$N{r}=""),"",$O{r}-$N{r})')
    set_if_blank(
        COL_UKC_FWD,
        f'=IF(OR($O{r}="",$K{r}=""),"",({DEPTHREF}+$O{r})-($K{r}+{SQUAT}+{SAFETY}))',
    )
    set_if_blank(
        COL_UKC_AFT,
        f'=IF(OR($O{r}="",$L{r}=""),"",({DEPTHREF}+$O{r})-($L{r}+{SQUAT}+{SAFETY}))',
    )
    set_if_blank(
        COL_GATE_FWD,
        f'=IF($A{r}="","",IF($C{r}<>"Y","N/A",IF($K{r}<={GATE_FWD_MAX},"PASS","LIMIT")))',
    )
    set_if_blank(
        COL_GATE_AFT,
        f'=IF($A{r}="","",IF($L{r}>={GATE_AFT_MIN},"PASS","LIMIT"))',
    )
    set_if_blank(
        COL_GATE_TRIM,
        f'=IF($A{r}="","",IF(ABS($J{r})<={TRIM_LIMIT},"PASS","LIMIT"))',
    )
    set_if_blank(
        COL_GATE_UKC,
        f'=IF($A{r}="","",IF(MIN($R{r},$S{r})>={UKC_MIN},"PASS","LIMIT"))',
    )
    set_if_blank(
        COL_GATE_RAMP,
        f'=IF($A{r}="","",IF($T{r}="","VERIFY",IF($T{r}<=6,"PASS","LIMIT")))',
    )
    set_if_blank(
        COL_HP_BAND,
        f'=IF($B{r}<>"Y","-",IF(OR($V{r}="LIMIT",$W{r}="LIMIT",$X{r}="LIMIT",$Y{r}="LIMIT"),"STOP","GO"))',
    )
    set_if_blank(
        COL_GONOGO,
        f'=IF($A{r}="","",IF($AA{r}="STOP","ABORT",IF($AA{r}="RECALC","NO-GO","GO")))',
    )


def write_value(ws: Worksheet, r: int, col: int, value, mode: str) -> bool:
    if value is None:
        return False
    cell = ws.cell(r, col)
    if mode == "fillblank":
        if cell.value is not None and str(cell.value).strip() != "":
            return False
    cell.value = value
    return True


def populate_stage_calc(
    template_path: Path,
    stage_csv_path: Path,
    out_path: Path,
    mode: str,
    sheet_name: str = "07_Stage_Calc",
    soft_last_row: int = 200,
) -> None:
    df = pd.read_csv(stage_csv_path)
    if df.empty:
        raise ValueError("Stage CSV is empty.")

    col_stage = pick_column(
        df, ["Stage_ID", "Stage", "stage", "StageName", "Stage Name"]
    )
    col_x = pick_column(df, ["x_stage_m", "x_m", "x", "xstage"])
    col_w = pick_column(df, ["W_stage_t", "W_t", "W", "Weight_t"])
    col_disp = pick_column(df, ["Disp_t", "Displacement_t", "Displacement", "Disp"])
    col_tmean = pick_column(df, ["Tmean_m", "Tmean", "T_mean_m", "Tmean(m)"])
    col_dfwd = pick_column(
        df, ["Dfwd_m", "Draft_FWD_m", "Fwd Draft(m)", "FWD Draft(m)", "Dfwd"]
    )
    col_daft = pick_column(
        df, ["Daft_m", "Draft_AFT_m", "Aft Draft(m)", "AFT Draft(m)", "Daft"]
    )
    col_gm = pick_column(df, ["GM(m)", "GM_m", "GM", "GM (m)"])

    # Tide/UKC optional columns (if provided by pipeline SSOT)
    col_forecast_tide = pick_column(df, ["Forecast_tide_m", "Forecast_Tide_m", "Tide_forecast_m", "tide_forecast_m"])
    col_tide_required = pick_column(df, ["Tide_required_m", "Required_WL_for_UKC_m", "tide_required_m"])
    col_tide_margin = pick_column(df, ["Tide_margin_m", "tide_margin_m"])
    col_ukc_min = pick_column(df, ["UKC_min_m", "UKC_Min_Required_m", "UKC_Min_m", "ukc_min_m"])
    col_ukc_fwd = pick_column(df, ["UKC_fwd_m", "UKC_FWD_m", "ukc_fwd_m"])
    col_ukc_aft = pick_column(df, ["UKC_aft_m", "UKC_AFT_m", "ukc_aft_m"])

    if not col_stage:
        raise ValueError(f"Cannot find Stage column in CSV. Columns: {list(df.columns)}")

    wb = load_workbook(template_path)
    if sheet_name not in wb.sheetnames:
        raise KeyError(f"Sheet not found: {sheet_name}")
    ws = wb[sheet_name]
    ws_log = get_or_create_sheet(wb, "LOG")

    header_row, data_start = find_header_row(ws, "Stage_ID", search_rows=60)
    # Build header map (normalized) for optional Tide/UKC columns
    header_map = {}
    try:
        for c in range(1, ws.max_column + 1):
            key = _norm_col(ws.cell(header_row, c).value)
            if key:
                header_map[key] = c
    except Exception:
        header_map = {}

    append_log(
        ws_log,
        f"[populate] template={template_path.name}, stage_csv={stage_csv_path.name}, mode={mode}",
    )
    append_log(
        ws_log,
        "Detected CSV cols: "
        f"Stage={col_stage}, x={col_x}, W={col_w}, Disp={col_disp}, "
        f"Tmean={col_tmean}, Dfwd={col_dfwd}, Daft={col_daft}, GM={col_gm}",
    )

    for _, row in df.iterrows():
        stage_id = _norm_text(row[col_stage])
        if stage_id == "":
            continue

        r_excel = find_or_create_stage_row(
            ws, stage_id, data_start, soft_last_row=soft_last_row
        )
        ensure_stage_formulas(ws, r_excel)

        v_x = safe_float(row[col_x]) if col_x else None
        v_w = safe_float(row[col_w]) if col_w else None
        v_d = safe_float(row[col_disp]) if col_disp else None
        v_t = safe_float(row[col_tmean]) if col_tmean else None
        v_k = safe_float(row[col_dfwd]) if col_dfwd else None
        v_l = safe_float(row[col_daft]) if col_daft else None
        v_g = safe_float(row[col_gm]) if col_gm else None

        write_value(ws, r_excel, 6, v_x, mode)
        write_value(ws, r_excel, 7, v_w, mode)
        write_value(ws, r_excel, 8, v_d, mode)
        write_value(ws, r_excel, 9, v_t, mode)
        write_value(ws, r_excel, 11, v_k, mode)
        write_value(ws, r_excel, 12, v_l, mode)
        write_value(ws, r_excel, 21, v_g, mode)

        # Optional Tide/UKC writes if template has matching headers
        try:
            if isinstance(header_map, dict) and header_map:
                def _wopt(h: str, v):
                    cidx = header_map.get(_norm_col(h))
                    if cidx:
                        write_value(ws, r_excel, int(cidx), v, mode)

                if col_forecast_tide:
                    _wopt("Forecast_tide_m", safe_float(row.get(col_forecast_tide)))
                if col_tide_required:
                    _wopt("Tide_required_m", safe_float(row.get(col_tide_required)))
                if col_tide_margin:
                    _wopt("Tide_margin_m", safe_float(row.get(col_tide_margin)))
                if col_ukc_min:
                    _wopt("UKC_min_m", safe_float(row.get(col_ukc_min)))
                if col_ukc_fwd:
                    _wopt("UKC_fwd_m", safe_float(row.get(col_ukc_fwd)))
                if col_ukc_aft:
                    _wopt("UKC_aft_m", safe_float(row.get(col_ukc_aft)))
        except Exception:
            pass


    try:
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.calcOnSave = True
    except Exception:
        pass

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    wb.close()


# -------------------------------------------------------------------------
# Validation (embedded)
# -------------------------------------------------------------------------
def validate_populated_template(excel_path: Path) -> bool:
    wb = openpyxl.load_workbook(excel_path, data_only=False)
    if "07_Stage_Calc" not in wb.sheetnames:
        print("ERROR: '07_Stage_Calc' sheet not found")
        return False

    ws = wb["07_Stage_Calc"]
    data_start_row = 21

    col_map = {
        "Stage_ID": "A",
        "Draft_FWD_m": "K",
        "Draft_AFT_m": "L",
        "GM_m": "U",
        "Gate_FWD": "V",
        "Gate_AFT": "W",
        "Gate_TRIM": "X",
        "Gate_UKC": "Y",
    }

    issues: List[str] = []
    stage_count = 0

    for r in range(data_start_row, data_start_row + 9):
        stage_cell = ws[f"{col_map['Stage_ID']}{r}"]
        if not stage_cell.value or str(stage_cell.value).strip() == "":
            continue

        stage_count += 1
        stage_name = str(stage_cell.value).strip()

        fwd_val = ws[f"{col_map['Draft_FWD_m']}{r}"].value
        aft_val = ws[f"{col_map['Draft_AFT_m']}{r}"].value
        if fwd_val is None or aft_val is None:
            issues.append(f"{stage_name}: Draft values missing")

        gm_val = ws[f"{col_map['GM_m']}{r}"].value
        if gm_val is None:
            issues.append(f"{stage_name}: GM_m missing")

        for col_name, col_letter in col_map.items():
            cell = ws[f"{col_letter}{r}"]
            if cell.data_type == "f":
                formula = str(cell.value)
                if any(token in formula for token in ("#REF!", "#VALUE!", "#NAME?")):
                    issues.append(
                        f"{stage_name}: {col_name} has error in formula: {formula}"
                    )

    print(f"Stages found: {stage_count}/9")
    print(f"Issues found: {len(issues)}")
    if issues:
        for issue in issues:
            print(f"  - {issue}")
        return False
    return True


# -------------------------------------------------------------------------
# Orchestration
# -------------------------------------------------------------------------
def create_template(out_path: Path) -> None:
    mod_path = _resolve_script_path("create_bryan_excel_template_NEW.py")
    # Use subprocess to avoid Python 3.13 dataclass loading issues
    import subprocess
    result = subprocess.run(
        [sys.executable, str(mod_path), "--out", str(out_path)],
        cwd=str(BASE_DIR),
        capture_output=True,
        text=True
    )
    if result.returncode != 0:
        print(f"STDOUT: {result.stdout}")
        print(f"STDERR: {result.stderr}")
        raise RuntimeError(f"Template generation failed: {result.stderr}")


def cmd_create(args: argparse.Namespace) -> int:
    create_template(Path(args.out))
    if args.spmt_xlsx:
        import_spmt_cargo(Path(args.spmt_xlsx), Path(args.out))
    print(f"[OK] Template generated: {args.out}")
    return 0


def cmd_populate(args: argparse.Namespace) -> int:
    populate_stage_calc(
        template_path=Path(args.template),
        stage_csv_path=Path(args.stage_csv),
        out_path=Path(args.out),
        mode=args.mode,
    )
    if args.spmt_xlsx:
        import_spmt_cargo(Path(args.spmt_xlsx), Path(args.out))
    print(f"[OK] Populated workbook saved: {args.out}")
    return 0


def cmd_validate(args: argparse.Namespace) -> int:
    ok = validate_populated_template(Path(args.file))
    return 0 if ok else 1


def cmd_one_click(args: argparse.Namespace) -> int:
    out_template = Path(args.out_template)
    out_populated = Path(args.out_populated)
    create_template(out_template)
    if args.spmt_xlsx:
        import_spmt_cargo(Path(args.spmt_xlsx), out_template)
    populate_stage_calc(
        template_path=out_template,
        stage_csv_path=Path(args.stage_csv),
        out_path=out_populated,
        mode=args.mode,
    )
    if args.validate:
        ok = validate_populated_template(out_populated)
        if not ok:
            return 1
    print(f"[OK] Template: {out_template}")
    print(f"[OK] Populated: {out_populated}")
    return 0


def build_parser() -> argparse.ArgumentParser:
    ap = argparse.ArgumentParser(description="Unified Bryan template workflow.")
    sub = ap.add_subparsers(dest="cmd", required=True)

    p1 = sub.add_parser("create", help="Generate template.")
    p1.add_argument("--out", required=True, help="Output template path")
    p1.add_argument("--spmt-xlsx", default="", help="Optional SPMT inputs workbook")
    p1.set_defaults(func=cmd_create)

    p2 = sub.add_parser("populate", help="Populate 07_Stage_Calc from CSV.")
    p2.add_argument("--template", required=True, help="Template xlsx path")
    p2.add_argument("--stage-csv", required=True, help="stage_results.csv path")
    p2.add_argument("--out", required=True, help="Output populated path")
    p2.add_argument("--mode", choices=["overwrite", "fillblank"], default="overwrite")
    p2.add_argument("--spmt-xlsx", default="", help="Optional SPMT inputs workbook")
    p2.set_defaults(func=cmd_populate)

    p3 = sub.add_parser("validate", help="Validate populated template.")
    p3.add_argument("--file", required=True, help="Workbook to validate")
    p3.set_defaults(func=cmd_validate)

    p4 = sub.add_parser("one-click", help="Create -> populate -> optional validate.")
    p4.add_argument("--stage-csv", required=True, help="stage_results.csv path")
    p4.add_argument("--out-template", required=True, help="Output template path")
    p4.add_argument("--out-populated", required=True, help="Output populated path")
    p4.add_argument("--mode", choices=["overwrite", "fillblank"], default="overwrite")
    p4.add_argument("--spmt-xlsx", default="", help="Optional SPMT inputs workbook")
    p4.add_argument("--validate", action="store_true", help="Run validation after populate")
    p4.set_defaults(func=cmd_one_click)

    return ap


def main() -> int:
    ap = build_parser()
    args = ap.parse_args()
    return int(args.func(args))


if __name__ == "__main__":
    raise SystemExit(main())
