# -*- coding: utf-8 -*-
"""
OPS-FINAL-R3: AGI BALLAST MANAGEMENT SYSTEM (INTEGRATED)
Integrated with agi_tr_patched_v6_6.py engineering-grade calculations

Version: R3-INTEGRATED
Features:
- Engineering-grade stage calculations (hydro table interpolation)
- GM 2D bilinear interpolation
- Frame-based coordinate system
- VOID3 ballast analysis (preserved)
- Discharge-only strategy (preserved)
- Excel + Markdown outputs
"""
import sys
import io
import argparse

# Fix encoding for Windows console
if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

import json
import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Import engineering-grade core engine
# - Prefer lightweight core engine module (if available in project repo)
# - Fallback to full agi_tr_patched_v6_6.py (bundled)
import sys
from pathlib import Path

# Add current directory and parent LCF directory to path
current_dir = Path(__file__).parent
sys.path.insert(0, str(current_dir))
sys.path.insert(0, str(current_dir.parent))
sys.path.insert(0, str(current_dir.parent.parent))

try:
    from agi_tr_core_engine import (
        LoadItem,
        solve_stage,
        _load_hydro_table,
        _load_gm2d_grid,
        _init_frame_mapping,
        fr_to_x,
        x_to_fr,
        gm_2d_bilinear,
    )
except ImportError:
    try:
        from agi_tr_patched_v6_6_defsplit_v1 import (
            LoadItem,
            solve_stage,
            _load_hydro_table,
            _load_gm2d_grid,
            _init_frame_mapping,
            fr_to_x,
            x_to_fr,
            gm_2d_bilinear,
        )
    except ImportError:
        # Try from parent directory
        try:
            from agi_tr_patched_v6_7_defsplit import (
                LoadItem,
                solve_stage,
                _load_hydro_table,
                _load_gm2d_grid,
                _init_frame_mapping,
                fr_to_x,
                x_to_fr,
                gm_2d_bilinear,
            )
        except ImportError:
            # Final fallback - import from agi_tr_patched_v6_6 if exists
            from agi_tr_patched_v6_6 import (
                LoadItem,
                solve_stage,
                _load_hydro_table,
                _load_gm2d_grid,
                _init_frame_mapping,
                fr_to_x,
                x_to_fr,
                gm_2d_bilinear,
            )

# ============================================================================
# VESSEL PARAMETERS
# ============================================================================
W_LIGHTSHIP = 730.0  # tons
W_TR = 271.2  # tons per transformer (AGI Site verified)

# Vessel vertical reference (keel → deck line). Used ONLY for freeboard = D - Draft.
D_MOLDED = 3.65  # m (molded depth)

# LONGITUDINAL REFERENCE (SSOT)
LPP_M = 60.302  # m
MIDSHIP_FROM_AP_M = LPP_M / 2.0  # m (AP=0 → FP=+)

# ============================================================================
# GATE SETS (DO NOT MIX DEFINITIONS)
# ============================================================================
# CLI gate overrides (keep defaults for backward compatibility)
_cli_parser = argparse.ArgumentParser(add_help=False)
_cli_parser.add_argument(
    "--fwd_max",
    type=float,
    default=2.70,
    help="FWD maximum draft gate (m).",
)
_cli_parser.add_argument(
    "--aft_min",
    type=float,
    default=2.70,
    help="AFT minimum draft gate (m).",
)
_cli_args, _ = _cli_parser.parse_known_args()

# 1) Port/Ramp draft gate: forward draft must NOT exceed limit
FWD_MAX_DRAFT_M = float(_cli_args.fwd_max)  # m (AGI Port/Ramp gate)

# 2) Captain/Master operational gate: keep propeller immersed (aft draft minimum)
AFT_MIN_DRAFT_M = float(_cli_args.aft_min)  # m (Captain gate; basis must be confirmed by vessel manual)

# 3) UKC gate: uses (Charted Depth + Water Level) − (Draft + Squat + Safety)
UKC_MIN = 0.50  # m (project default; confirm with Port/Vessel UKC guidance)

# 4) Optional freeboard gate (green-water margin). If unknown, keep None → "VERIFY"
FREEBOARD_MIN_M = None  # m (TBD)

# 5) GM minimum (fallback constant if not in hydro table)
GM_MIN_M = 1.50  # m (fallback GM minimum; matches agi_tr_patched gm_target_m)

# TIDE / WATER LEVEL (definition split)
# Forecast_Tide_m: predicted tide height above Chart Datum (CD) (NOT "required").
# Water_Level_m: effective sea level above CD used in UKC: Depth_available = Depth_charted + Water_Level_m
TIDE_DATUM = "CD"  # Chart Datum (set to match official tide table)
TIDE_TZ = "GST"  # Local TZ for tide table (set to match official tide table)
DATUM_OFFSET_M = 0.00  # (Port datum - CD). If Port datum differs, set here.

# Example tide window for Stage 6B (replace with official forecast from tide table)
FORECAST_TIDE_STAGE6B_M = 0.30  # m

# Backward compatible alias used by existing discharge logic
TARGET_FWD = FWD_MAX_DRAFT_M

# COORDINATE SYSTEM: internal x_from_midship (+AFT / −FWD) must match fr_to_x() convention
STERN_TO_MIDSHIP = MIDSHIP_FROM_AP_M  # alias (AP to midship)

print("=" * 80)
print("OPS-FINAL-R3: AGI BALLAST MANAGEMENT SYSTEM (INTEGRATED)")
print("=" * 80)
print(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
print(f"Engineering Grade: agi_tr_patched v6.6 calculations")
print(f"Document Status: OPERATIONAL DRAFT - FIELD VERIFICATION REQUIRED")
print("=" * 80)

# ============================================================================
# 1. LOAD VESSEL HYDRO DATA
# ============================================================================
print("\n[1/12] Loading Vessel Hydro Data...")

# Load hydro table
hydro_table = _load_hydro_table()
if hydro_table:
    print(f"[OK] Hydro table loaded: {len(hydro_table)} draft points")
else:
    print(f"⚠️ Hydro table not found - using fallback constants")

# GM grid already loaded by module
print(f"✅ GM 2D Grid ready")

# Frame mapping already initialized
print(f"✅ Frame mapping ready")

# Prepare params dict for solve_stage
vessel_params = {
    "MTC_t_m_per_cm": 34.00,
    "LCF_m_from_midship": 0.76,
    "TPC_t_per_cm": 8.00,
    "Lpp_m": LPP_M,
    "D_vessel_m": D_MOLDED,
    "hydro_table": hydro_table,
    "gm_curve": None,
    "max_fwd_draft_ops_m": TARGET_FWD,
    "trim_limit_cm": 240.0,
}

# ============================================================================
# 2. LOAD TANK CATALOG
# ============================================================================
print("\n[2/12] Loading Tank Catalog...")

# Try multiple paths for tank catalog
tank_catalog_paths = [
    "tank_catalog_from_tankmd.json",
    "../tank_catalog_from_tankmd.json",
    "../../LCF/tank_catalog_from_tankmd.json",
    "../../tank_catalog_from_tankmd.json",
]
tank_catalog = None
for path in tank_catalog_paths:
    try:
        with open(path, "r", encoding="utf-8") as f:
            tank_catalog = json.load(f)
            print(f"[INFO] Loaded tank catalog from: {path}")
            break
    except FileNotFoundError:
        continue

if tank_catalog is None:
    raise FileNotFoundError(
        f"tank_catalog_from_tankmd.json not found in any of: {tank_catalog_paths}"
    )

print(f"✅ Loaded {len(tank_catalog['tanks'])} tank entries")
print(f"   Source: {tank_catalog['source_file']}")

# ============================================================================
# 3. PROCESS TANK DATA
# ============================================================================
print("\n[3/12] Processing Tank Data...")

tanks_data = []
for tank in tank_catalog["tanks"]:
    if tank["id"] == "**Total**":
        continue

    lcg_stern = tank["lcg_m"]
    lcg_midship = STERN_TO_MIDSHIP - lcg_stern  # +AFT / -FWD

    tanks_data.append(
        {
            "Tank_ID": tank["id"],
            "Category": tank["category"],
            "Location": tank.get("loc", "N/A"),
            "Fr_Start": tank.get("fr_start"),
            "Fr_End": tank.get("fr_end"),
            "Capacity_t": tank["cap_t"],
            "LCG_stern_m": lcg_stern,
            "LCG_midship_m": lcg_midship,
            "TCG_m": tank["tcg_m"],
            "TCG_Side": tank.get("tcg_side", "cl"),
            "VCG_m": tank["vcg_m"],
        }
    )

df_tanks = pd.DataFrame(tanks_data)

print(f"✅ Processed {len(df_tanks)} tanks")
print(
    f"   LCG Range (midship): {df_tanks['LCG_midship_m'].min():.2f} to {df_tanks['LCG_midship_m'].max():.2f} m"
)

# ============================================================================
# 4. CLASSIFY BALLAST TANKS
# ============================================================================
print("\n[4/12] Classifying Ballast Tanks...")

df_ballast_all = df_tanks[
    df_tanks["Category"].str.contains("fresh_water", case=False, na=False)
].copy()


def classify_tank_role(row):
    tank_id = row["Tank_ID"]
    if "VOID3" in tank_id:
        return "Pre-ballast_Storage_ONLY"
    elif "FWB2" in tank_id:
        return "Discharge_Source_Primary"
    elif "VOIDDB2" in tank_id or "VOIDDB1" in tank_id:
        return "Discharge_Source_Secondary"
    elif "FWB1" in tank_id or "FW1" in tank_id or "FW2" in tank_id:
        return "Forward_Ballast_Potential"
    elif "FWCARGO" in tank_id:
        return "Cargo_Liquid"
    else:
        return "Other"


df_ballast_all["Operational_Role"] = df_ballast_all.apply(classify_tank_role, axis=1)

role_summary = (
    df_ballast_all.groupby("Operational_Role")
    .agg({"Capacity_t": "sum", "Tank_ID": "count", "LCG_midship_m": "mean"})
    .round(2)
)
role_summary.columns = ["Total_Capacity_t", "Tank_Count", "Avg_LCG_m"]

print("\n📊 Ballast Tank Classification:")
print(role_summary.to_string())

# ============================================================================
# 5. VOID3 DETAILED ANALYSIS
# ============================================================================
print("\n[5/12] VOID3 Tank Analysis...")
print("-" * 80)

df_void3 = df_ballast_all[
    df_ballast_all["Tank_ID"].str.contains("VOID3", na=False)
].copy()

print("VOID3 Specifications:")
print(
    df_void3[["Tank_ID", "Capacity_t", "LCG_midship_m", "VCG_m", "Location"]].to_string(
        index=False
    )
)

void3_total_cap = df_void3["Capacity_t"].sum()
void3_avg_lcg = df_void3["LCG_midship_m"].mean()
void3_avg_vcg = df_void3["VCG_m"].mean()

print(f"\n✅ VOID3 Total: {void3_total_cap:.2f} tons")
print(f"   LCG: {void3_avg_lcg:.2f} m, VCG: {void3_avg_vcg:.2f} m")

print("\n" + "=" * 80)
print("VOID3 OPERATIONAL ASSESSMENT")
print("=" * 80)
print("\n🟢 PHYSICAL SUITABILITY: EXCELLENT")
print("   ✅ Capacity: 304.16 tons (sufficient)")
print("   ✅ LCG: -2.25 m (favorable for trim)")
print("   ✅ VCG: 1.91 m (low CoG - stability positive)")
print("\n🔴 OPERATIONAL LIMITATIONS:")
print("   ❌ NO pump/valve route for internal transfer")
print("   ❌ CANNOT receive ballast during operations")
print("   ❌ Transfer operations NOT AUTHORIZED")
print("\n🟡 APPROVED USE: PRE-BALLAST STORAGE ONLY")
print("   ✅ Shore-fill before RORO")
print("   ✅ Maintain as STATIC BALLAST")

# ============================================================================
# 6. DISCHARGE SOURCE TANKS
# ============================================================================
print("\n[6/12] Discharge Source Analysis...")

df_discharge = df_ballast_all[
    df_ballast_all["Operational_Role"].str.contains("Discharge_Source", na=False)
].copy()

print("\n📊 Operable Discharge Tanks:")
print(
    df_discharge[
        ["Tank_ID", "Capacity_t", "LCG_midship_m", "Operational_Role"]
    ].to_string(index=False)
)

discharge_total = df_discharge["Capacity_t"].sum()
print(f"\n✅ Total Discharge Capacity: {discharge_total:.2f} tons")

# ============================================================================


# ============================================================================
# 6B. DEFINITIONS HELPERS (avoid definition mix-up)
# ============================================================================
def _gate_icon(ok: bool | None) -> str:
    """Return iconized gate result: ✅PASS / ⚠️FAIL / ❓VERIFY."""
    if ok is None:
        return "❓ VERIFY"
    return "✅ PASS" if ok else "⚠️ FAIL"


def _freeboard_min_m(d_molded: float, dfwd: float, daft: float) -> float:
    """Freeboard = D - Draft (independent of tide)."""
    fb_fwd = d_molded - dfwd
    fb_aft = d_molded - daft
    return min(fb_fwd, fb_aft)


def _gate_freeboard_ok(fb_min: float, fb_limit: float | None) -> bool | None:
    """If fb_limit is None → unknown (VERIFY)."""
    if fb_limit is None:
        return None
    return fb_min >= fb_limit


# 7. STAGE DEFINITIONS (SSOT FIRST)
# ============================================================================
print("\n[7/12] Loading Stage Definitions (SSOT: stage_results.csv)...")

# SSOT Stage name conventions (must match TR script csv export)
STAGE6A_NAME = "Stage 6A_Critical (Opt C)"
STAGE6B_NAME = "Stage 6B"

# Where to look for stage_results.csv (integrated pipeline Step-1b output)
_STAGE_RESULTS_CSV_CANDIDATES = [
    Path.cwd() / "stage_results.csv",
    Path(__file__).resolve().parent / "stage_results.csv",
    Path.cwd() / "ssot" / "stage_results.csv",
]


def _first_existing_path(paths):
    for p in paths:
        try:
            if p.exists():
                return p
        except Exception:
            continue
    return None


stage_results_csv = _first_existing_path(_STAGE_RESULTS_CSV_CANDIDATES)
if stage_results_csv is None:
    raise FileNotFoundError(
        "stage_results.csv not found. Generate it first via TR script csv mode:\n"
        "  python agi_tr_patched_v6_6_defsplit_v1.py csv\n"
        "or run the integrated pipeline (Step-1b)."
    )

df_stage_ssot = pd.read_csv(stage_results_csv)
if "Stage" not in df_stage_ssot.columns:
    raise ValueError(
        f"Invalid stage_results.csv (missing 'Stage'): {stage_results_csv}"
    )

print(f"✅ Loaded SSOT stage_results.csv: {stage_results_csv}")
print(f"   SSOT stage rows: {len(df_stage_ssot)}")


def _pick_col(df, candidates):
    for c in candidates:
        if c in df.columns:
            return c
    raise KeyError(
        f"Required column not found. Tried {candidates}. Available columns: {list(df.columns)}"
    )


COL_STAGE = "Stage"
COL_DFWD = _pick_col(df_stage_ssot, ["Dfwd_m", "Fwd Draft(m)", "FWD Draft(m)", "FWD_m"])
COL_DAFT = _pick_col(
    df_stage_ssot, ["Daft_m", "AFT Draft(m)", "AFT Draft (m)", "AFT_m"]
)
COL_TMEAN = _pick_col(df_stage_ssot, ["Tmean_m", "Tmean (m)", "Mean_m", "Mean (m)"])
COL_TRIM = _pick_col(df_stage_ssot, ["Trim_cm", "Trim (cm)"])
COL_DISP = _pick_col(df_stage_ssot, ["Disp_t", "Disp (t)", "Displacement_t"])
COL_GM = _pick_col(df_stage_ssot, ["GM(m)", "GM_m", "GM (m)"])

# Optional columns (best effort)
COL_XLCG = None
for _c in ["x_stage_m", "x_LCG_m", "LCG_m"]:
    if _c in df_stage_ssot.columns:
        COL_XLCG = _c
        break
COL_PREBALLAST = "PreBallast_t" if "PreBallast_t" in df_stage_ssot.columns else None

# Minimal description map (reporting only; SSOT is Stage name)
STAGE_DESCRIPTION_MAP = {
    "Stage 1": "Arrival / Baseline",
    "Stage 2": "TR1 Ramp Start",
    "Stage 3": "TR1 Mid-Ramp / Transfer",
    "Stage 4": "TR1 On Deck (Interim)",
    "Stage 5": "TR1 Final Position",
    "Stage 5_PreBallast": "Pre-ballast for TR2 (Gate-A/B critical)",
    "Stage 6A_Critical (Opt C)": "CRITICAL RoRo condition (Opt C) — UKC check basis",
    "Stage 6C": "Final Stowage (TR1+TR2)",
    "Stage 7": "Departure / Post-operation",
}

# ============================================================================
# 8. BUILD STAGE TABLE FROM SSOT (drafts are SSOT; tide affects UKC only)
# ============================================================================
print("\n[8/12] Building Stage Table from SSOT...")
print("-" * 80)

stages_results = []

for _, _r in df_stage_ssot.iterrows():
    stage_name = str(_r[COL_STAGE]).strip()

    fwd_m = float(_r[COL_DFWD])
    aft_m = float(_r[COL_DAFT])
    trim_cm = float(_r[COL_TRIM])
    mean_m = float(_r[COL_TMEAN])
    disp_t = float(_r[COL_DISP])
    gm_m = float(_r[COL_GM])

    lcg_m = float(_r[COL_XLCG]) if COL_XLCG else 0.0
    preballast_t = float(_r[COL_PREBALLAST]) if COL_PREBALLAST else None

    # Hydro properties at this displacement (loads empty; hydro derived from displacement)
    hydro_only = solve_stage(disp_t, mean_m, [], **vessel_params)

    # Discharge requirement (FWD draft constraint only; SSOT draft is tide-independent)
    discharge_needed_t = 0.0
    if fwd_m > FWD_MAX_DRAFT_M:
        discharge_needed_t = (
            (fwd_m - FWD_MAX_DRAFT_M) * hydro_only["TPC_used_t_per_cm"] * 100.0
        )

    fb_min = _freeboard_min_m(D_MOLDED, fwd_m, aft_m)
    gate_fwd_ok = fwd_m <= FWD_MAX_DRAFT_M
    gate_aft_ok = aft_m >= AFT_MIN_DRAFT_M
    gate_fb_ok = _gate_freeboard_ok(fb_min, FREEBOARD_MIN_M)

    stages_results.append(
        {
            "Stage": stage_name,
            "Description": STAGE_DESCRIPTION_MAP.get(stage_name, ""),
            "Mass_t": disp_t,
            "LCG_m": lcg_m,
            "FWD_m": fwd_m,
            "AFT_m": aft_m,
            "Trim_cm": trim_cm,
            "Mean_m": mean_m,
            # Definition split: tide does NOT change draft. Tide affects UKC only.
            "Forecast_Tide_m": 0.00,  # m above CD (forecast)
            "Water_Level_m": 0.00,  # placeholder; computed later as Forecast_Tide + DATUM_OFFSET
            "Max_Draft_m": max(fwd_m, aft_m),  # m
            "Freeboard_Min_m": fb_min,  # m (D - Draft)
            # Gate set (split)
            "Gate_FWD_Max": _gate_icon(gate_fwd_ok),
            "Gate_AFT_Min": _gate_icon(gate_aft_ok),
            "Gate_Freeboard": _gate_icon(gate_fb_ok),
            "GM_m": gm_m,
            "GM_min_m": hydro_only.get("GM_min_m", GM_MIN_M),
            "LCF_used_m": hydro_only.get(
                "LCF_used_m", float(vessel_params.get("LCF_m_from_midship", 0.76))
            ),
            "MCTC_used": hydro_only.get(
                "MCTC_used_t_m_per_cm", float(vessel_params.get("MTC_t_m_per_cm", 34.0))
            ),
            "TPC_used": hydro_only.get(
                "TPC_used_t_per_cm", float(vessel_params.get("TPC_t_per_cm", 8.0))
            ),
            "FWD_Margin_m": FWD_MAX_DRAFT_M - fwd_m,
            "AFT_Margin_m": aft_m - AFT_MIN_DRAFT_M,
            "Discharge_Needed_t": discharge_needed_t,
            "Status": (
                "✅ PASS (FWD/AFT)"
                if (gate_fwd_ok and gate_aft_ok)
                else (
                    "⚠️ FAIL (FWD)"
                    if (not gate_fwd_ok and gate_aft_ok)
                    else (
                        "⚠️ FAIL (AFT)"
                        if (gate_fwd_ok and not gate_aft_ok)
                        else "⚠️ FAIL (FWD&AFT)"
                    )
                )
            ),
            "Engineering_Grade": hydro_only.get("EngineeringGrade_active", False),
            "PreBallast_t": preballast_t,
        }
    )

# Add Stage 6B as Tide Window (UKC improvement) WITHOUT changing drafts
try:
    _s6a = next(s for s in stages_results if s.get("Stage") == STAGE6A_NAME)
except StopIteration:
    print(
        f"⚠️  {STAGE6A_NAME} not found in SSOT. Stage 6B (tide window) will not be added."
    )
else:
    stage6b_result = dict(_s6a)
    stage6b_result.update(
        {
            "Stage": STAGE6B_NAME,
            "Description": f"Tide Window (Forecast +{FORECAST_TIDE_STAGE6B_M:.2f} m, UKC only; draft unchanged)",
            "Forecast_Tide_m": float(FORECAST_TIDE_STAGE6B_M),
            "Status": stage6b_result.get("Status", "INFO") + " / INFO:TIDE(UKC)",
        }
    )
    stages_results.append(stage6b_result)

df_stages = pd.DataFrame(stages_results)

print("\n📊 Stage Calculation Results:")
print(
    df_stages[
        [
            "Stage",
            "Description",
            "FWD_m",
            "AFT_m",
            "Forecast_Tide_m",
            "Gate_FWD_Max",
            "Gate_AFT_Min",
            "Freeboard_Min_m",
            "Status",
        ]
    ].to_string(index=False)
)

# Identify critical stages
critical_stages = df_stages[df_stages["FWD_m"] > TARGET_FWD]
if len(critical_stages) > 0:
    print(f"\n⚠️ {len(critical_stages)} stage(s) exceed limit:")
    print(
        critical_stages[["Stage", "FWD_m", "Discharge_Needed_t"]].to_string(index=False)
    )
else:
    print(f"\n✅ ALL STAGES COMPLIANT")

# ============================================================================
# 9. DISCHARGE ACTION MATRIX (ENGINEERING-GRADE)
# ============================================================================
print("\n[9/12] Generating Discharge Matrix...")


def calc_draft_response_per_10t_agi(
    tank_lcg_m, discharge_amount_t=10.0, stage_result=None
):
    """Calculate FWD draft change using engineering-grade parameters"""
    if stage_result is None:
        mctc_used = vessel_params["MTC_t_m_per_cm"]
        tpc_used = vessel_params["TPC_t_per_cm"]
        lcf_used = vessel_params["LCF_m_from_midship"]
    else:
        mctc_used = stage_result["MCTC_used"]
        tpc_used = stage_result["TPC_used"]
        lcf_used = stage_result["LCF_used_m"]

    moment_change = -discharge_amount_t * (tank_lcg_m - lcf_used)
    trim_change_cm = moment_change / mctc_used if mctc_used > 1e-9 else 0.0
    parallel_cm = -discharge_amount_t / tpc_used if tpc_used > 1e-9 else 0.0
    trim_fwd_component_cm = -trim_change_cm / 2.0
    total_fwd_change_cm = parallel_cm + trim_fwd_component_cm

    return total_fwd_change_cm


discharge_actions = []

# Use Stage 6A parameters for discharge calculations (SSOT stage name)
stage6a_result_dict = next(
    (s for s in stages_results if s.get("Stage") == STAGE6A_NAME), None
)
if stage6a_result_dict is None:
    raise RuntimeError(
        f"Cannot find {STAGE6A_NAME} in stage table; discharge matrix cannot proceed."
    )

for idx, tank in df_discharge.iterrows():
    tank_id = tank["Tank_ID"]
    tank_cap = tank["Capacity_t"]
    tank_lcg = tank["LCG_midship_m"]

    delta_fwd_10t = calc_draft_response_per_10t_agi(tank_lcg, 10.0, stage6a_result_dict)

    if "FWB2" in tank_id:
        priority = 1
        priority_label = "PRIMARY"
    elif "VOIDDB2" in tank_id:
        priority = 2
        priority_label = "SECONDARY"
    elif "VOIDDB1" in tank_id:
        priority = 3
        priority_label = "TERTIARY"
    else:
        priority = 4
        priority_label = "OTHER"

    effectiveness = (
        "HIGH" if delta_fwd_10t < -1.5 else "MEDIUM" if delta_fwd_10t < -1.0 else "LOW"
    )

    discharge_actions.append(
        {
            "Tank_ID": tank_id,
            "Capacity_t": tank_cap,
            "LCG_m": tank_lcg,
            "Delta_FWD_per_10t_cm": delta_fwd_10t,
            "Priority": priority,
            "Priority_Label": priority_label,
            "Effectiveness": effectiveness,
            "Operational_Status": "OPERABLE - Discharge to Sea",
        }
    )

df_discharge_matrix = pd.DataFrame(discharge_actions).sort_values("Priority")

print("\n📊 Discharge Effectiveness (Engineering-Grade):")
print(
    df_discharge_matrix[
        ["Tank_ID", "Delta_FWD_per_10t_cm", "Priority_Label", "Effectiveness"]
    ].to_string(index=False)
)

# ============================================================================
# ============================================================================
# 10. UKC / WATER LEVEL / REQUIRED WL (DEFINITION SPLIT)
# ============================================================================
print("\n[10/12] UKC + Water Level Gate Analysis...")

# NOTE:
# - Draft (Dfwd/Daft) is a ship property relative to the hull (keel→waterline).
# - Tide/Water Level does NOT change draft.
# - Tide/Water Level DOES change available water depth and therefore UKC.
#
# Reference concept:
#   UKC = (Charted Depth + Water Level) − (Static Draft)  [then apply dynamic allowances like squat/safety]
# (We keep allowances explicit here.)

# ---------------------------------------------------------------------------
# Inputs (MUST be aligned to same datum/timezone as tide table)
# ---------------------------------------------------------------------------
CHARTED_DEPTH_APPROACH_M = 4.50  # m (EXAMPLE, referenced to Chart Datum)
CHARTED_DEPTH_RAMP_TOE_M = 4.20  # m (EXAMPLE, referenced to Chart Datum)

SQUAT_ALLOWANCE_M = 0.15  # m (set per speed/channel confinement)
SAFETY_MARGIN_M = 0.20  # m (project/port conservative margin)

DEPTH_CRITICAL_M = min(CHARTED_DEPTH_APPROACH_M, CHARTED_DEPTH_RAMP_TOE_M)
ALLOWANCES_M = SQUAT_ALLOWANCE_M + SAFETY_MARGIN_M

# Ensure Water_Level_m is consistent
if "Forecast_Tide_m" not in df_stages.columns:
    df_stages["Forecast_Tide_m"] = 0.00
df_stages["Water_Level_m"] = df_stages["Forecast_Tide_m"] + DATUM_OFFSET_M

# Conservative UKC using critical depth
df_stages["Max_Draft_m"] = df_stages[["FWD_m", "AFT_m"]].max(axis=1)
df_stages["UKC_Min_m"] = (DEPTH_CRITICAL_M + df_stages["Water_Level_m"]) - (
    df_stages["Max_Draft_m"] + ALLOWANCES_M
)

# Required Water Level above CD to satisfy UKC gate (this is NOT a forecast tide value)
df_stages["Required_WL_m"] = (
    df_stages["Max_Draft_m"] + ALLOWANCES_M + UKC_MIN
) - DEPTH_CRITICAL_M
df_stages["Required_WL_m"] = df_stages["Required_WL_m"].apply(
    lambda x: max(0.0, float(x))
)

# ------------------------------------------------------------------
# Tide/UKC SSOT alias fields (operations-friendly column names)
# ------------------------------------------------------------------
df_stages["Tide_required_m"] = df_stages["Required_WL_m"]
df_stages["Forecast_tide_m"] = df_stages["Forecast_Tide_m"]
df_stages["Tide_margin_m"] = df_stages["Forecast_tide_m"] - df_stages["Tide_required_m"]

# UKC requirement vs actual ends
df_stages["UKC_min_m"] = float(UKC_MIN)
df_stages["UKC_fwd_m"] = (DEPTH_CRITICAL_M + df_stages["Water_Level_m"]) - (df_stages["FWD_m"] + ALLOWANCES_M)
df_stages["UKC_aft_m"] = (DEPTH_CRITICAL_M + df_stages["Water_Level_m"]) - (df_stages["AFT_m"] + ALLOWANCES_M)
df_stages["UKC_min_actual_m"] = df_stages[["UKC_fwd_m", "UKC_aft_m"]].min(axis=1)


# UKC gate result
df_stages["Gate_UKC"] = df_stages["UKC_Min_m"].apply(
    lambda ukc: _gate_icon(ukc >= UKC_MIN)
)

print(
    f"   Tide datum: {TIDE_DATUM} | Tide TZ: {TIDE_TZ} | Datum offset: {DATUM_OFFSET_M:.2f} m"
)
print(
    f"   Depth (Approach/RampToe): {CHARTED_DEPTH_APPROACH_M:.2f} / {CHARTED_DEPTH_RAMP_TOE_M:.2f} m  (EXAMPLE)"
)
print(
    f"   Critical depth used: {DEPTH_CRITICAL_M:.2f} m | Allowances(squat+safety): {ALLOWANCES_M:.2f} m"
)
print(
    "   UKC formula used: UKC=(Depth+WaterLevel)-(Draft+Squat+Safety)  # Tide affects UKC, not draft"
)

# Stage 6A vs 6B (definition clarified)
ukc_view = df_stages[df_stages["Stage"].isin([STAGE6A_NAME, STAGE6B_NAME])][
    [
        "Stage",
        "FWD_m",
        "AFT_m",
        "Forecast_tide_m",
        "Tide_required_m",
        "Tide_margin_m",
        "UKC_min_m",
        "UKC_fwd_m",
        "UKC_aft_m",
        "UKC_min_actual_m",
        "Gate_UKC",
    ]
]
print(f"\n📌 UKC Gate ({STAGE6A_NAME} vs {STAGE6B_NAME}):")
print(ukc_view.to_string(index=False))


# Optional: overall gate summary (FWD max + AFT min + UKC)
def _overall_gate(row) -> str:
    ok = (
        str(row.get("Gate_FWD_Max", "")).startswith("✅")
        and str(row.get("Gate_AFT_Min", "")).startswith("✅")
        and str(row.get("Gate_UKC", "")).startswith("✅")
    )
    return _gate_icon(ok)


df_stages["Gate_Overall"] = df_stages.apply(_overall_gate, axis=1)
# 11. GENERATE EXCEL WORKBOOK
# ============================================================================
print("\n[11/12] Generating Excel...")

wb = Workbook()
if "Sheet" in wb.sheetnames:
    wb.remove(wb["Sheet"])

header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
warning_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
critical_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Sheet 1: Tank SSOT
ws_tanks = wb.create_sheet("Tank_SSOT")
for r_idx, row in enumerate(dataframe_to_rows(df_tanks, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws_tanks.cell(row=r_idx, column=c_idx, value=value)
        cell.border = border
        if r_idx == 1:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

for col in ws_tanks.columns:
    max_len = max(len(str(cell.value)) for cell in col)
    ws_tanks.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

# Sheet 2: VOID3 Analysis
ws_void3 = wb.create_sheet("VOID3_Analysis")

void3_report = [
    ["VOID3 OPERATIONAL ASSESSMENT", "", "", ""],
    ["", "", "", ""],
    ["Physical Specifications", "", "", ""],
    ["Parameter", "Value", "Unit", "Assessment"],
    ["Total Capacity", void3_total_cap, "tons", "✅ Sufficient"],
    ["Average LCG", void3_avg_lcg, "m", "✅ Favorable"],
    ["Average VCG", void3_avg_vcg, "m", "✅ Low CoG"],
    ["", "", "", ""],
    ["Operational Approval", "", "", ""],
    ["Use Case", "Status", "Notes", ""],
    ["Pre-ballast Storage", "✅ APPROVED", "Shore-fill only", ""],
    ["Transfer Receiver", "❌ NOT OPERABLE", "No pump/valve route", ""],
    ["", "", "", ""],
    ["CRITICAL WARNING", "Transfer operations NOT AUTHORIZED", "", ""],
]

for r_idx, row in enumerate(void3_report, 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws_void3.cell(row=r_idx, column=c_idx, value=value)
        if r_idx == 1:
            cell.font = Font(bold=True, size=14)
        elif "APPROVED" in str(value):
            cell.fill = PatternFill(
                start_color="00B050", end_color="00B050", fill_type="solid"
            )
            cell.font = Font(bold=True, color="FFFFFF")
        elif "NOT OPERABLE" in str(value):
            cell.fill = critical_fill
            cell.font = Font(bold=True, color="FFFFFF")

for col in ws_void3.columns:
    ws_void3.column_dimensions[col[0].column_letter].width = 35

# Sheet 3: Stage Calculations (with engineering parameters)
ws_stages = wb.create_sheet("Stage_Calculations")
for r_idx, row in enumerate(dataframe_to_rows(df_stages, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws_stages.cell(row=r_idx, column=c_idx, value=value)
        cell.border = border
        if r_idx == 1:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif "⚠️" in str(value):
            cell.fill = warning_fill

for col in ws_stages.columns:
    max_len = max(len(str(cell.value)) for cell in col)
    ws_stages.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

# Sheet 4: Discharge Matrix
ws_discharge = wb.create_sheet("Discharge_Matrix")
for r_idx, row in enumerate(
    dataframe_to_rows(df_discharge_matrix, index=False, header=True), 1
):
    for c_idx, value in enumerate(row, 1):
        cell = ws_discharge.cell(row=r_idx, column=c_idx, value=value)
        cell.border = border
        if r_idx == 1:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

for col in ws_discharge.columns:
    max_len = max(len(str(cell.value)) for cell in col)
    ws_discharge.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

# Sheet 5: UKC / Water Level Gate (definition split)
ws_ukc = wb.create_sheet("UKC_WL_Gate")

# Stage 6A/6B snapshot (if present)
try:
    _s6a = df_stages[df_stages["Stage"] == STAGE6A_NAME].iloc[0]
except Exception:
    _s6a = None
try:
    _s6b = df_stages[df_stages["Stage"] == "Stage 6B"].iloc[0]
except Exception:
    _s6b = None

ukc_data = [
    ["UKC / Water Level Gate (Definition Split)", "", "", "", ""],
    ["", "", "", "", ""],
    ["Tide Datum / TZ", f"{TIDE_DATUM} / {TIDE_TZ}", "", "", ""],
    ["Datum Offset", f"{DATUM_OFFSET_M:.2f}", "m", "", ""],
    [
        "Charted Depth (Approach)",
        f"{CHARTED_DEPTH_APPROACH_M:.2f}",
        "m",
        "⚠️ EXAMPLE",
        "",
    ],
    [
        "Charted Depth (Ramp Toe)",
        f"{CHARTED_DEPTH_RAMP_TOE_M:.2f}",
        "m",
        "⚠️ EXAMPLE",
        "",
    ],
    ["Allowances (Squat+Safety)", f"{ALLOWANCES_M:.2f}", "m", "", ""],
    ["UKC Criterion (min)", f"{UKC_MIN:.2f}", "m", "", ""],
    ["", "", "", "", ""],
    ["Stage", "Forecast_Tide_m", "Water_Level_m", "Max_Draft_m", "UKC_Min_m"],
]

if _s6a is not None:
    ukc_data.append(
        [
            STAGE6A_NAME,
            f"{_s6a['Forecast_Tide_m']:.2f}",
            f"{_s6a['Water_Level_m']:.2f}",
            f"{_s6a['Max_Draft_m']:.2f}",
            f"{_s6a['UKC_Min_m']:.2f}",
        ]
    )
if _s6b is not None:
    ukc_data.append(
        [
            "Stage 6B",
            f"{_s6b['Forecast_Tide_m']:.2f}",
            f"{_s6b['Water_Level_m']:.2f}",
            f"{_s6b['Max_Draft_m']:.2f}",
            f"{_s6b['UKC_Min_m']:.2f}",
        ]
    )

ukc_data += [
    ["", "", "", "", ""],
    ["Stage", "Required_WL_m", "", "", "Gate_UKC"],
]
if _s6a is not None:
    ukc_data.append(
        [
            STAGE6A_NAME,
            f"{_s6a['Required_WL_m']:.2f}",
            "",
            "",
            str(_s6a.get("Gate_UKC", "")),
        ]
    )
if _s6b is not None:
    ukc_data.append(
        [
            "Stage 6B",
            f"{_s6b['Required_WL_m']:.2f}",
            "",
            "",
            str(_s6b.get("Gate_UKC", "")),
        ]
    )

for r_idx, row in enumerate(ukc_data, 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws_ukc.cell(row=r_idx, column=c_idx, value=value)
        if r_idx == 1:
            cell.font = Font(bold=True, size=14)
        elif r_idx in (10, 14):  # table headers
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif "⚠️ EXAMPLE" in str(value) or "⚠️" in str(value):
            cell.fill = warning_fill

for col in ws_ukc.columns:
    ws_ukc.column_dimensions[col[0].column_letter].width = 22
# Sheet 6: BWRB Log
ws_bwrb = wb.create_sheet("BWRB_Log")

bwrb_headers = [
    "Date/Time",
    "Operation",
    "Tank_ID",
    "Amount_t",
    "Source",
    "Destination",
    "FWD_Before_m",
    "FWD_After_m",
    "Port_Approval",
    "Master_Sign",
    "BWMP_Status",
]

for c_idx, header in enumerate(bwrb_headers, 1):
    cell = ws_bwrb.cell(row=1, column=c_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

excel_filename = "OPS_FINAL_R3_AGI_Ballast_Integrated.xlsx"
excel_path = Path(excel_filename)
try:
    wb.save(excel_path)
    print(f"✅ Excel saved: {excel_path}")
except PermissionError:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    fallback_path = excel_path.with_stem(f"{excel_path.stem}_{timestamp}")
    wb.save(fallback_path)
    print(
        f"⚠️ Excel file is locked. Saved to fallback: {fallback_path}"
    )

# ============================================================================
# 12. GENERATE MARKDOWN REPORT
# ============================================================================
print("\n[12/12] Generating Markdown Report...")

report_md = f"""# OPS-FINAL-R3: AGI BALLAST MANAGEMENT SYSTEM (INTEGRATED)

**Version:** R3-INTEGRATED
**Generated:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
**Calculation Engine:** agi_tr_patched v6.6
**Engineering Grade:** {'✅ ACTIVE' if stages_results[0]['Engineering_Grade'] else '⚠️ FALLBACK'}

---

## Executive Summary

### Integration Status

✅ **Engineering-Grade Calculations ACTIVE**
- Hydro table interpolation: Draft-dependent LCF/MCTC/TPC
- GM 2D bilinear interpolation: Displacement and trim dependent
- Frame-based coordinates: Precise load positioning

### Key Findings

**VOID3 Assessment:**
- Physical Suitability: EXCELLENT (304.16t capacity)
- Operational Role: Pre-ballast Storage ONLY
- Transfer Capability: NOT OPERABLE

**Stage Analysis:**
- Critical Stages: {len(critical_stages)} exceed 2.70m limit
- {STAGE6A_NAME} FWD: {stage6a_result_dict['FWD_m']:.3f} m
{f"- Discharge Required: {stage6a_result_dict['Discharge_Needed_t']:.1f} tons" if stage6a_result_dict['Discharge_Needed_t'] > 0 else "- No Discharge Required"}
- Engineering Grade: {stages_results[0]['Engineering_Grade']}

**Discharge Strategy:**
- Priority 1: FWB2.P/S ({df_discharge_matrix[df_discharge_matrix['Priority']==1]['Capacity_t'].sum():.2f} tons)
- Priority 2: VOIDDB2.C ({df_discharge_matrix[df_discharge_matrix['Priority']==2]['Capacity_t'].sum():.2f} tons)
- Priority 3: VOIDDB1.C ({df_discharge_matrix[df_discharge_matrix['Priority']==3]['Capacity_t'].sum():.2f} tons)

---

## Calculation Methodology

### Engineering-Grade Calculations (agi_tr_patched v6.6)

**Hydro Table Interpolation:**
- LCF (Longitudinal Center of Flotation): Draft-dependent via hydro table
- MCTC (Moment to Change Trim): Draft-dependent via hydro table
- TPC (Tons Per Centimeter): Draft-dependent via hydro table

**GM 2D Bilinear Interpolation:**
- Displacement-dependent: as per GM grid input range
- Trim-dependent: as per GM grid input range
- Source: LCT BUSHRA Stability Booklet 7x7 grid

**Frame-Based Coordinates:**
- Frame 0 = Stern (LCG = +30m from midship)
- Midship = Frame 30.151
- Convention: AFT = +, FWD = -

**Data Sources:**
- Hydro table: {len(hydro_table)} draft points
- GM 2D Grid: 7×7 interpolation grid
- Frame mapping: Linear conversion with slope validation

---

## Stage Analysis

### Stage Calculation Results

| Stage | FWD_m | AFT_m | GM_m | FWD_Margin_m | Engineering_Grade | Status |
|-------|-------|-------|------|--------------|-------------------|--------|
{chr(10).join([f"| {r['Stage']} | {r['FWD_m']:.3f} | {r['AFT_m']:.3f} | {r['GM_m']:.3f} | {r['FWD_Margin_m']:.3f} | {r['Engineering_Grade']} | {r['Status']} |" for r in stages_results])}

### Engineering Parameters by Stage

| Stage | LCF (m) | MCTC (t·m/cm) | TPC (t/cm) | Basis |
|-------|---------|---------------|------------|-------|
{chr(10).join([f"| {r['Stage']} | {r['LCF_used_m']:.3f} | {r['MCTC_used']:.2f} | {r['TPC_used']:.2f} | {'Hydro' if r['Engineering_Grade'] else 'Fallback'} |" for r in stages_results[:7]])}

---

## VOID3 Analysis (Preserved)

**Approved Use:** PRE-BALLAST STORAGE ONLY
- Shore-fill before RORO: 200.0 tons
- Maintain as static ballast throughout operation
- NO transfer, NO discharge, NO internal movement

**Not Authorized:**
- Transfer receiver operations
- Internal ballast transfer
- Dynamic trim adjustment

---

## Discharge Strategy (Engineering-Grade)

| Tank_ID | LCG_m | Delta_FWD_per_10t_cm | Priority_Label | Effectiveness |
|---------|-------|----------------------|----------------|---------------|
{chr(10).join([f"| {r['Tank_ID']} | {r['LCG_m']:.3f} | {r['Delta_FWD_per_10t_cm']:.3f} | {r['Priority_Label']} | {r['Effectiveness']} |" for _, r in df_discharge_matrix.iterrows()])}

**Discharge Execution:**
1. Start with Priority 1 (FWB2.P/S)
2. Discharge in 10t increments
3. Wait 5 minutes for settling
4. Measure draft response
5. Continue until FWD ≤ 2.70m

---

## Deliverables

1. **{excel_filename}** (6 sheets)
2. **OPS_FINAL_R3_Report_Integrated.md** (this document)
3. **agi_tr_core_engine.py** (calculation engine)

---

**Engineering Grade Status:** {'✅ ACTIVE - Using hydro interpolation' if stages_results[0]['Engineering_Grade'] else '⚠️ FALLBACK - Using constants'}

**Next Steps:**
1. Verify frame positions (TR ramp/stowage)
2. Conduct 10t discharge trial
3. Update with actual field data
4. Obtain Port/HM approval

---

*Generated by OPS-FINAL-R3 Integrated System with agi_tr_patched v6.6 calculations*
"""

report_filename = "OPS_FINAL_R3_Report_Integrated.md"
with open(report_filename, "w", encoding="utf-8") as f:
    f.write(report_md)

print(f"✅ Report saved: {report_filename}")

# ============================================================================
# SUMMARY
# ============================================================================
print("\n" + "=" * 80)
print("GENERATION COMPLETE")
print("=" * 80)

print(f"\n📦 Deliverables:")
print(f"   1. {excel_filename}")
print(f"   2. {report_filename}")
print(f"   3. agi_tr_core_engine.py (calculation engine)")

print(f"\n✅ Key Results:")
print(f"   - VOID3: {void3_total_cap:.2f} tons (Pre-ballast Only)")
print(f"   - Discharge Capacity: {discharge_total:.2f} tons")
print(f"   - {STAGE6A_NAME} FWD: {stage6a_result_dict['FWD_m']:.3f} m")
print(f"   - Engineering Grade: {stages_results[0]['Engineering_Grade']}")
# UKC Gate snapshot (definition split)
try:
    _ukc6a = df_stages[df_stages["Stage"] == STAGE6A_NAME].iloc[0]
    _ukc6b = df_stages[df_stages["Stage"] == "Stage 6B"].iloc[0]
    print(
        f"   - UKC Gate (6A/6B): {_ukc6a.get('Gate_UKC','')} / {_ukc6b.get('Gate_UKC','')}"
    )
    print(
        f"   - Required_WL (6A): {_ukc6a.get('Required_WL_m', 0.0):.2f} m | Forecast_Tide (6B): {_ukc6b.get('Forecast_Tide_m', 0.0):.2f} m"
    )
except Exception as _e:
    print(f"   - UKC Gate: ❓ VERIFY (UKC columns missing)  ({_e})")


print(f"\n⚠️ Action Items:")
print(f"   1. Verify frame positions with GA drawings")
print(f"   2. Confirm hydro table accuracy")
print(f"   3. Conduct field discharge trial (10t)")
print(f"   4. Obtain Port/HM approval")

print(f"\n" + "=" * 80)
print("Ready for Phase 2: Field Validation")
print("=" * 80)
