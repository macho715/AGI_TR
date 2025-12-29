#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
lct_stage_spmt_shuttle.py (AGI SSOT patched)

Key change
- Default coordinate mode is FR_M (station from AP in meters).
- The legacy "frame index * 0.60 m + AP offset" conversion is preserved only if you
  explicitly set coord_mode="FRAME_INDEX" in your config.

Recommendation
- For AGI pipeline consistency, keep coord_mode="FR_M" and provide all frames as meters from AP.

"""

from __future__ import annotations

import json
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Literal

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from agi_ssot import VesselSSOT, build_agi_default_ssot

CoordMode = Literal["FR_M", "FRAME_INDEX"]


@dataclass
class VesselCoord:
    """Wrapper around the shared AGI SSOT coordinate converter."""
    ssot: VesselSSOT = build_agi_default_ssot()

    def frame_to_station_m(self, frame_value: float) -> float:
        fr = self.ssot.to_fr_m_from_ap(frame_value)
        if fr is None:
            raise ValueError("frame_value is None/invalid")
        return round(fr, 3)

    def station_to_x_from_midship(self, fr_m_from_ap: float) -> float:
        return round(self.ssot.x_from_midship_m(fr_m_from_ap), 3)


@dataclass
class StageConfig:
    tr1_final_frame: float
    tr2_final_frame: float
    ramp_start_frame: float
    ramp_mid_frame: float
    spmt_park_frame: float

    w_tr_t: float = 217.0
    w_spmt_t: float = 54.2
    w_beams_t: float = 11.6
    w_ppu_t: float = 7.2
    assembly_total_t: float = 282.8
    dunnage_t: float = 0.0
    dunnage_frame: float = 31.0

    final_spmt_mode: str = "OFFBOARD"  # ONBOARD / OFFBOARD
    beams_mode: str = "REMOVED"        # ATTACHED / REMOVED
    ppu_mode: str = "INCLUDED_IN_SPMT" # INCLUDED_IN_SPMT / SEPARATE_ONBOARD / OFFBOARD
    ppu_included_in_assembly: bool = True

    # SSOT block (optional; defaults to AGI SSOT)
    coord_mode: CoordMode = "FR_M"
    frame_spacing_m: float = 1.0
    ap_offset_m: float = 0.0
    midship_from_ap_m: float = 30.151
    lpp_m: float = 60.302


def load_config(path: str) -> StageConfig:
    data = json.loads(Path(path).read_text(encoding="utf-8"))

    ss = data.get("ssot", {})
    pos = data.get("positions", {})
    w = data.get("weights", {})
    m = data.get("modes", {})

    return StageConfig(
        tr1_final_frame=float(pos.get("tr1_final_frame", 42.0)),
        tr2_final_frame=float(pos.get("tr2_final_frame", 40.0)),
        ramp_start_frame=float(pos.get("ramp_start_frame", pos.get("ramp_frame", 17.95))),
        ramp_mid_frame=float(pos.get("ramp_mid_frame", pos.get("ramp_frame", 17.95))),
        spmt_park_frame=float(pos.get("spmt_park_frame", 10.0)),
        dunnage_frame=float(pos.get("dunnage_frame", 31.0)),

        w_tr_t=float(w.get("w_tr_t", 217.0)),
        w_spmt_t=float(w.get("w_spmt_t", 54.2)),
        w_beams_t=float(w.get("w_beams_t", 11.6)),
        w_ppu_t=float(w.get("w_ppu_t", 7.2)),
        assembly_total_t=float(w.get("assembly_total_t", 282.8)),
        dunnage_t=float(w.get("dunnage_t", 0.0)),

        final_spmt_mode=str(m.get("final_spmt_mode", "OFFBOARD")).upper(),
        beams_mode=str(m.get("beams_mode", "REMOVED")).upper(),
        ppu_mode=str(m.get("ppu_mode", "INCLUDED_IN_SPMT")).upper(),
        ppu_included_in_assembly=bool(m.get("ppu_included_in_assembly", True)),

        coord_mode=str(ss.get("coord_mode", "FR_M")).upper(),  # default FR_M
        frame_spacing_m=float(ss.get("frame_spacing_m", 1.0)),
        ap_offset_m=float(ss.get("ap_offset_m", 0.0)),
        midship_from_ap_m=float(ss.get("midship_from_ap_m", 30.151)),
        lpp_m=float(ss.get("lpp_m", 60.302)),
    )


def build_vessel_coord(cfg: StageConfig) -> VesselCoord:
    ssot = VesselSSOT(
        lpp_m=cfg.lpp_m,
        midship_from_ap_m=cfg.midship_from_ap_m,
        coord_mode="FR_M" if cfg.coord_mode == "FR_M" else "FRAME_INDEX",
        frame_spacing_m=cfg.frame_spacing_m,
        ap_offset_m=cfg.ap_offset_m,
        fr_range_m=(0.0, cfg.lpp_m),
    )
    return VesselCoord(ssot=ssot)


def build_stages(cfg: StageConfig) -> List[Tuple[str, List[Tuple[str, float, float]]]]:
    """
    Returns list of (stage_name, [(item_name, weight_t, fr_m_from_ap), ...]).
    """
    vc = build_vessel_coord(cfg)

    # Helpers
    def fr(v: float) -> float:
        return vc.frame_to_station_m(v)

    def base_tr1_fixed() -> List[Tuple[str, float, float]]:
        items: List[Tuple[str, float, float]] = [("TR1 (fixed on stools)", cfg.w_tr_t, fr(cfg.tr1_final_frame))]
        if cfg.beams_mode == "ATTACHED":
            items.append(("Beams (TR1)", cfg.w_beams_t, fr(cfg.tr1_final_frame)))
        if cfg.dunnage_t > 0:
            items.append(("Dunnage", cfg.dunnage_t, fr(cfg.dunnage_frame)))
        return items

    stages: List[Tuple[str, List[Tuple[str, float, float]]]] = []

    stages.append(("S1 25% TR1 on ramp start", [
        ("TR1+SPMT+Beams (25%)", cfg.assembly_total_t * 0.25, fr(cfg.ramp_start_frame)),
        *([] if cfg.dunnage_t <= 0 else [("Dunnage", cfg.dunnage_t, fr(cfg.dunnage_frame))]),
    ]))

    stages.append(("S2 50% TR1 on ramp mid (HOLD)", [
        ("TR1+SPMT+Beams (50%)", cfg.assembly_total_t * 0.50, fr(cfg.ramp_mid_frame)),
        *([] if cfg.dunnage_t <= 0 else [("Dunnage", cfg.dunnage_t, fr(cfg.dunnage_frame))]),
    ]))

    stages.append(("S3 100% TR1 at final", [
        ("TR1+SPMT+Beams (100%)", cfg.assembly_total_t, fr(cfg.tr1_final_frame)),
        *([] if cfg.dunnage_t <= 0 else [("Dunnage", cfg.dunnage_t, fr(cfg.dunnage_frame))]),
    ]))

    stages.append(("S4 Pre-ballast for TR2 (SPMT off deck)", base_tr1_fixed()))

    stages.append(("S5 50% TR2 on ramp mid (HOLD)", [
        *base_tr1_fixed(),
        ("TR2+SPMT+Beams (50%)", cfg.assembly_total_t * 0.50, fr(cfg.ramp_mid_frame)),
    ]))

    final_items = base_tr1_fixed() + [("TR2 (fixed on stools)", cfg.w_tr_t, fr(cfg.tr2_final_frame))]
    if cfg.beams_mode == "ATTACHED":
        final_items.append(("Beams (TR2)", cfg.w_beams_t, fr(cfg.tr2_final_frame)))

    if cfg.final_spmt_mode == "ONBOARD":
        final_items.append(("SPMT (empty, parked)", cfg.w_spmt_t, fr(cfg.spmt_park_frame)))
        if cfg.ppu_mode == "SEPARATE_ONBOARD":
            final_items.append(("PPU (separate onboard)", cfg.w_ppu_t, fr(cfg.spmt_park_frame)))

    stages.append(("S6 Transit (final)", final_items))

    return stages


def main():
    import argparse
    ap = argparse.ArgumentParser()
    ap.add_argument("--config", required=True)
    ap.add_argument("--out_xlsx", required=True)
    args = ap.parse_args()

    cfg = load_config(args.config)
    vc = build_vessel_coord(cfg)
    stages = build_stages(cfg)

    wb = Workbook()
    ws = wb.active
    ws.title = "Stage_Loads"

    headers = ["Stage", "Item", "Weight_t", "Fr_m_from_AP", "x_m_from_Midship"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.alignment = Alignment(horizontal="center")

    for stage_name, items in stages:
        for name, wt, fr_m in items:
            x = vc.station_to_x_from_midship(fr_m)
            ws.append([stage_name, name, round(wt, 3), round(fr_m, 3), round(x, 3)])

    for i, w in enumerate([28, 38, 12, 14, 16], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(args.out_xlsx)
    print(f"[OK] Wrote: {args.out_xlsx}")


if __name__ == "__main__":
    main()
