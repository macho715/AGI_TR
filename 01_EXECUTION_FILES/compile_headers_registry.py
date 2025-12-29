#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Compile HEADERS_MASTER.xlsx to headers_registry.json
"""

import json
import sys
from pathlib import Path
from typing import Any, Dict, List
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook


def compile_from_excel(excel_path: Path, output_json_path: Path) -> None:
    """Compile Excel master to JSON registry."""
    wb = load_workbook(excel_path, read_only=True)

    # Read Fields sheet
    if "Fields" not in wb.sheetnames:
        raise ValueError("HEADERS_MASTER.xlsx must have 'Fields' sheet")
    fields_df = pd.read_excel(excel_path, sheet_name="Fields")

    # Read Deliverables sheet
    if "Deliverables" not in wb.sheetnames:
        raise ValueError("HEADERS_MASTER.xlsx must have 'Deliverables' sheet")
    deliv_df = pd.read_excel(excel_path, sheet_name="Deliverables")

    # Build fields dict
    fields: Dict[str, Any] = {}
    for _, row in fields_df.iterrows():
        key = str(row.get("Field_Key", "")).strip()
        if not key or pd.isna(row.get("Field_Key")):
            continue
        aliases_str = str(row.get("Aliases", "")).strip()
        aliases = [a.strip() for a in aliases_str.split(",") if a.strip()] if aliases_str and not pd.isna(row.get("Aliases")) else []
        fields[key] = {
            "default_header": str(row.get("Default_Header", key)),
            "aliases": aliases,
            "unit": str(row.get("Unit", "")) if pd.notna(row.get("Unit")) else "",
            "type": str(row.get("Type", "string")) if pd.notna(row.get("Type")) else "string",
            "description": str(row.get("Description", "")) if pd.notna(row.get("Description")) else "",
        }

    # Build deliverables dict
    deliverables: Dict[str, Any] = {}
    for _, row in deliv_df.iterrows():
        did = str(row.get("Deliverable_ID", "")).strip()
        if not did or pd.isna(row.get("Deliverable_ID")):
            continue

        columns_str = str(row.get("Columns", "")).strip()
        columns = [c.strip() for c in columns_str.split(",") if c.strip()] if columns_str and not pd.isna(row.get("Columns")) else []

        output_headers_str = str(row.get("Output_Headers", "")).strip()
        output_headers = {}
        if output_headers_str and not pd.isna(row.get("Output_Headers")):
            try:
                output_headers = json.loads(output_headers_str)
            except json.JSONDecodeError:
                # Fallback: parse simple key:value pairs
                for pair in output_headers_str.split(","):
                    if ":" in pair:
                        k, v = pair.split(":", 1)
                        output_headers[k.strip()] = v.strip()

        sheet_val = row.get("Sheet")
        header_row_val = row.get("Header_Row")
        
        deliverables[did] = {
            "file_pattern": str(row.get("File_Pattern", "")) if pd.notna(row.get("File_Pattern")) else "",
            "sheet": str(sheet_val) if pd.notna(sheet_val) else None,
            "header_row": int(header_row_val) if pd.notna(header_row_val) else None,
            "columns": columns,
            "output_headers": output_headers,
            "strict": bool(row.get("Strict", True)) if pd.notna(row.get("Strict")) else True,
            "version": str(row.get("Version", "2025.12.28.1")) if pd.notna(row.get("Version")) else "2025.12.28.1",
        }

    # Build registry JSON
    registry = {
        "meta": {
            "version": "2025.12.28.1",
            "normalize": "loose",
            "compiled_from": str(excel_path.name),
            "compiled_at": datetime.now().isoformat(),
        },
        "fields": fields,
        "deliverables": deliverables,
    }

    # Write JSON
    output_json_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_json_path, "w", encoding="utf-8") as f:
        json.dump(registry, f, indent=2, ensure_ascii=False)

    print(f"[OK] Compiled {len(fields)} fields, {len(deliverables)} deliverables")
    print(f"[OK] Output: {output_json_path}")


def main():
    import argparse

    parser = argparse.ArgumentParser(description="Compile HEADERS_MASTER.xlsx to JSON")
    parser.add_argument("--excel", type=str, required=True, help="Path to HEADERS_MASTER.xlsx")
    parser.add_argument("--output", type=str, default="headers_registry.json", help="Output JSON path")

    args = parser.parse_args()

    excel_path = Path(args.excel)
    output_path = Path(args.output)

    if not excel_path.exists():
        print(f"[ERROR] Excel not found: {excel_path}")
        return 1

    try:
        compile_from_excel(excel_path, output_path)
        return 0
    except Exception as e:
        print(f"[ERROR] Compilation failed: {type(e).__name__}: {e}")
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())

