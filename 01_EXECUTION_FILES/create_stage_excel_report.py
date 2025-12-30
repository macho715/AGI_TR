#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
스테이지별 상세 결과 Excel 테이블 생성
Excel Table Style로 정리된 보고서 생성
"""

import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_stage_summary_excel(final_output_dir: Path, output_file: Path):
    """스테이지별 상세 결과를 Excel 테이블 형식으로 생성"""

    wb = Workbook()

    # 기본 시트 제거
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # 데이터 로드
    qa_csv = final_output_dir / "pipeline_stage_QA.csv"
    summary_csv = final_output_dir / "solver_ballast_summary.csv"
    plan_csv = final_output_dir / "solver_ballast_stage_plan.csv"
    exec_csv = final_output_dir / "BALLAST_EXEC.csv"

    if not qa_csv.exists():
        print(f"[ERROR] {qa_csv} 파일을 찾을 수 없습니다.")
        return None

    df_qa = pd.read_csv(qa_csv, encoding='utf-8-sig')
    df_summary = pd.read_csv(summary_csv, encoding='utf-8-sig') if summary_csv.exists() else pd.DataFrame()
    df_plan = pd.read_csv(plan_csv, encoding='utf-8-sig') if plan_csv.exists() else pd.DataFrame()
    df_exec = pd.read_csv(exec_csv, encoding='utf-8-sig') if exec_csv.exists() else pd.DataFrame()

    # ============================================================
    # 시트 1: 요약 (Summary)
    # ============================================================
    ws_summary = wb.create_sheet("01_요약", 0)

    # 헤더 스타일
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 제목
    ws_summary['A1'] = "스테이지별 상세 결과 요약"
    ws_summary['A1'].font = Font(bold=True, size=14)
    ws_summary.merge_cells('A1:H1')

    # 요약 테이블
    summary_data = []
    for idx, row in df_qa.iterrows():
        stage = row['Stage']
        fwd = row.get('Draft_FWD_m', row.get('Current_FWD_m', ''))
        aft = row.get('Draft_AFT_m', row.get('Current_AFT_m', ''))
        trim = row.get('Trim_cm', '')
        gate_a = 'PASS' if row.get('GateA_AFT_MIN_2p70_PASS', False) else 'FAIL'
        gate_b = row.get('GateB_FWD_MAX_2p70_CD_PASS', 'N/A')
        if isinstance(gate_b, bool):
            gate_b = 'PASS' if gate_b else 'FAIL'
        elif pd.isna(gate_b):
            gate_b = 'N/A'
        ukc = row.get('UKC_min_actual_m', row.get('UKC_min_m', ''))

        summary_data.append({
            'Stage': stage,
            'FWD Draft (m)': fwd,
            'AFT Draft (m)': aft,
            'Trim (cm)': trim,
            'Gate-A (AFT≥2.70m)': gate_a,
            'Gate-B (FWD≤2.70m)': gate_b,
            'UKC Min (m)': ukc,
            'Critical': 'Yes' if 'Critical' in str(stage) or 'PreBallast' in str(stage) else 'No'
        })

    df_summary_table = pd.DataFrame(summary_data)

    # 테이블 작성
    headers = list(df_summary_table.columns)
    for col_idx, header in enumerate(headers, 1):
        cell = ws_summary.cell(row=3, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    for row_idx, (_, row) in enumerate(df_summary_table.iterrows(), 4):
        for col_idx, value in enumerate(row, 1):
            cell = ws_summary.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

            # Gate-A FAIL 강조
            if col_idx == 5 and value == 'FAIL':
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            # Gate-B FAIL 강조
            elif col_idx == 6 and value == 'FAIL':
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            # Critical Stage 강조
            elif col_idx == 8 and value == 'Yes':
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    # 컬럼 너비 조정
    ws_summary.column_dimensions['A'].width = 25
    for col in range(2, 9):
        ws_summary.column_dimensions[get_column_letter(col)].width = 15

    # ============================================================
    # 시트 2: 스테이지별 상세 (Stage Details)
    # ============================================================
    ws_details = wb.create_sheet("02_스테이지별_상세")

    # 주요 컬럼 선택
    detail_cols = [
        'Stage', 'Draft_FWD_m', 'Draft_AFT_m', 'Trim_cm', 'Tmean_m',
        'Input_Disp_t', 'Freeboard_FWD_m', 'Freeboard_AFT_m', 'Freeboard_Min_m',
        'UKC_fwd_m', 'UKC_aft_m', 'UKC_min_actual_m',
        'Forecast_tide_m', 'Tide_required_m', 'Tide_margin_m',
        'GateA_AFT_MIN_2p70_PASS', 'GateA_AFT_MIN_2p70_Margin_m',
        'GateB_FWD_MAX_2p70_CD_applicable', 'GateB_FWD_MAX_2p70_CD_PASS', 'GateB_FWD_MAX_2p70_CD_Margin_m'
    ]

    available_cols = [col for col in detail_cols if col in df_qa.columns]
    df_details = df_qa[available_cols].copy()

    # 컬럼명 한글화
    col_mapping = {
        'Stage': 'Stage',
        'Draft_FWD_m': 'FWD Draft (m)',
        'Draft_AFT_m': 'AFT Draft (m)',
        'Trim_cm': 'Trim (cm)',
        'Tmean_m': 'Mean Draft (m)',
        'Input_Disp_t': 'Displacement (t)',
        'Freeboard_FWD_m': 'Freeboard FWD (m)',
        'Freeboard_AFT_m': 'Freeboard AFT (m)',
        'Freeboard_Min_m': 'Freeboard Min (m)',
        'UKC_fwd_m': 'UKC FWD (m)',
        'UKC_aft_m': 'UKC AFT (m)',
        'UKC_min_actual_m': 'UKC Min (m)',
        'Forecast_tide_m': 'Forecast Tide (m)',
        'Tide_required_m': 'Tide Required (m)',
        'Tide_margin_m': 'Tide Margin (m)',
        'GateA_AFT_MIN_2p70_PASS': 'Gate-A PASS',
        'GateA_AFT_MIN_2p70_Margin_m': 'Gate-A Margin (m)',
        'GateB_FWD_MAX_2p70_CD_applicable': 'Gate-B Applicable',
        'GateB_FWD_MAX_2p70_CD_PASS': 'Gate-B PASS',
        'GateB_FWD_MAX_2p70_CD_Margin_m': 'Gate-B Margin (m)'
    }

    df_details.rename(columns=col_mapping, inplace=True)

    # 테이블 작성
    headers = list(df_details.columns)
    for col_idx, header in enumerate(headers, 1):
        cell = ws_details.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    for row_idx, (_, row) in enumerate(df_details.iterrows(), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws_details.cell(row=row_idx, column=col_idx)
            if pd.notna(value):
                cell.value = value
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

            # Gate-A FAIL 강조
            if 'Gate-A PASS' in headers and col_idx == headers.index('Gate-A PASS') + 1:
                if value == False:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            # Gate-B FAIL 강조
            elif 'Gate-B PASS' in headers and col_idx == headers.index('Gate-B PASS') + 1:
                if value == False:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # 컬럼 너비 조정
    for col_idx, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        if 'Stage' in header:
            ws_details.column_dimensions[col_letter].width = 25
        else:
            ws_details.column_dimensions[col_letter].width = 15

    # ============================================================
    # 시트 3: Gate 검증 결과 (Gate Validation)
    # ============================================================
    ws_gates = wb.create_sheet("03_Gate_검증")

    gate_cols = [
        'Stage',
        'Draft_FWD_m', 'Draft_AFT_m',
        'GateA_AFT_MIN_2p70_PASS', 'GateA_AFT_MIN_2p70_Margin_m',
        'GateB_FWD_MAX_2p70_CD_applicable', 'GateB_FWD_MAX_2p70_CD_PASS', 'GateB_FWD_MAX_2p70_CD_Margin_m',
        'Gate_FWD_Max', 'Gate_AFT_Min', 'Gate_Freeboard', 'Gate_UKC'
    ]

    available_gate_cols = [col for col in gate_cols if col in df_qa.columns]
    df_gates = df_qa[available_gate_cols].copy()

    # 컬럼명 한글화
    gate_col_mapping = {
        'Stage': 'Stage',
        'Draft_FWD_m': 'FWD Draft (m)',
        'Draft_AFT_m': 'AFT Draft (m)',
        'GateA_AFT_MIN_2p70_PASS': 'Gate-A PASS',
        'GateA_AFT_MIN_2p70_Margin_m': 'Gate-A Margin (m)',
        'GateB_FWD_MAX_2p70_CD_applicable': 'Gate-B Applicable',
        'GateB_FWD_MAX_2p70_CD_PASS': 'Gate-B PASS',
        'GateB_FWD_MAX_2p70_CD_Margin_m': 'Gate-B Margin (m)',
        'Gate_FWD_Max': 'Gate FWD Max',
        'Gate_AFT_Min': 'Gate AFT Min',
        'Gate_Freeboard': 'Gate Freeboard',
        'Gate_UKC': 'Gate UKC'
    }

    df_gates.rename(columns=gate_col_mapping, inplace=True)

    # 테이블 작성
    headers = list(df_gates.columns)
    for col_idx, header in enumerate(headers, 1):
        cell = ws_gates.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    for row_idx, (_, row) in enumerate(df_gates.iterrows(), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws_gates.cell(row=row_idx, column=col_idx)
            if pd.notna(value):
                if isinstance(value, bool):
                    cell.value = 'PASS' if value else 'FAIL'
                else:
                    cell.value = value
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

            # FAIL 강조
            if isinstance(value, str) and value == 'FAIL':
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            elif isinstance(value, str) and value == 'OK':
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    # 컬럼 너비 조정
    for col_idx, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        if 'Stage' in header:
            ws_gates.column_dimensions[col_letter].width = 25
        else:
            ws_gates.column_dimensions[col_letter].width = 15

    # ============================================================
    # 시트 4: Ballast 계획 (Ballast Plan)
    # ============================================================
    if not df_plan.empty:
        ws_plan = wb.create_sheet("04_Ballast_계획")

        # 컬럼명 한글화
        plan_col_mapping = {
            'Stage': 'Stage',
            'Tank': 'Tank',
            'Action': 'Action',
            'Delta_t': 'Delta (t)',
            'PumpTime_h': 'Time (h)'
        }

        df_plan_display = df_plan.copy()
        df_plan_display.rename(columns=plan_col_mapping, inplace=True)

        # 테이블 작성
        headers = list(df_plan_display.columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws_plan.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        for row_idx, (_, row) in enumerate(df_plan_display.iterrows(), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws_plan.cell(row=row_idx, column=col_idx)
                if pd.notna(value):
                    cell.value = value
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

        # 컬럼 너비 조정
        ws_plan.column_dimensions['A'].width = 25
        ws_plan.column_dimensions['B'].width = 15
        for col in range(3, 6):
            ws_plan.column_dimensions[get_column_letter(col)].width = 15

    # ============================================================
    # 시트 5: 실행 시퀀스 (Execution Sequence)
    # ============================================================
    if not df_exec.empty:
        ws_exec = wb.create_sheet("05_실행_시퀀스")

        # 주요 컬럼 선택
        exec_cols = [
            'Stage', 'Step', 'Tank', 'Action', 'Start_t', 'Target_t',
            'Delta_t', 'Time_h', 'Pump_ID', 'PumpRate_tph', 'Hold_Point',
            'Draft_FWD', 'Draft_AFT', 'Trim_cm', 'UKC'
        ]

        available_exec_cols = [col for col in exec_cols if col in df_exec.columns]
        df_exec_display = df_exec[available_exec_cols].copy()

        # 컬럼명 한글화
        exec_col_mapping = {
            'Stage': 'Stage',
            'Step': 'Step',
            'Tank': 'Tank',
            'Action': 'Action',
            'Start_t': 'Start (t)',
            'Target_t': 'Target (t)',
            'Delta_t': 'Delta (t)',
            'Time_h': 'Time (h)',
            'Pump_ID': 'Pump ID',
            'PumpRate_tph': 'Rate (tph)',
            'Hold_Point': 'Hold Point',
            'Draft_FWD': 'Draft FWD (m)',
            'Draft_AFT': 'Draft AFT (m)',
            'Trim_cm': 'Trim (cm)',
            'UKC': 'UKC (m)'
        }

        df_exec_display.rename(columns=exec_col_mapping, inplace=True)

        # 테이블 작성
        headers = list(df_exec_display.columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws_exec.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border

        for row_idx, (_, row) in enumerate(df_exec_display.iterrows(), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws_exec.cell(row=row_idx, column=col_idx)
                if pd.notna(value):
                    cell.value = value
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

                # Hold Point 강조
                if 'Hold Point' in headers and col_idx == headers.index('Hold Point') + 1:
                    if value == 'Y':
                        cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

        # 컬럼 너비 조정
        ws_exec.column_dimensions['A'].width = 25
        for col in range(2, len(headers) + 1):
            ws_exec.column_dimensions[get_column_letter(col)].width = 12

    # ============================================================
    # 시트 6: Critical Stage 요약
    # ============================================================
    ws_critical = wb.create_sheet("06_Critical_요약")

    critical_stages = df_qa[df_qa['Stage'].str.contains('Critical|PreBallast', case=False, na=False)]

    if not critical_stages.empty:
        critical_data = []
        for idx, row in critical_stages.iterrows():
            stage = row['Stage']
            fwd = row.get('Draft_FWD_m', '')
            aft = row.get('Draft_AFT_m', '')
            gate_a = 'PASS' if row.get('GateA_AFT_MIN_2p70_PASS', False) else 'FAIL'
            gate_a_margin = row.get('GateA_AFT_MIN_2p70_Margin_m', '')
            gate_b = 'PASS' if row.get('GateB_FWD_MAX_2p70_CD_PASS', False) else 'FAIL'
            gate_b_margin = row.get('GateB_FWD_MAX_2p70_CD_Margin_m', '')
            ukc = row.get('UKC_min_actual_m', '')

            # Solver 계획 정보
            plan_rows = df_plan[df_plan['Stage'] == stage] if not df_plan.empty else pd.DataFrame()
            plan_count = len(plan_rows)
            total_delta = plan_rows['Delta_t'].sum() if not plan_rows.empty else 0
            total_time = plan_rows['PumpTime_h'].sum() if not plan_rows.empty else 0

            critical_data.append({
                'Stage': stage,
                'FWD Draft (m)': fwd,
                'AFT Draft (m)': aft,
                'Gate-A': gate_a,
                'Gate-A Margin (m)': gate_a_margin,
                'Gate-B': gate_b,
                'Gate-B Margin (m)': gate_b_margin,
                'UKC Min (m)': ukc,
                'Ballast Operations': plan_count,
                'Total Delta (t)': total_delta,
                'Total Time (h)': total_time
            })

        df_critical = pd.DataFrame(critical_data)

        # 테이블 작성
        headers = list(df_critical.columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws_critical.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border

        for row_idx, (_, row) in enumerate(df_critical.iterrows(), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws_critical.cell(row=row_idx, column=col_idx)
                if pd.notna(value):
                    cell.value = value
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

                # FAIL 강조
                if 'Gate-A' in headers and col_idx == headers.index('Gate-A') + 1:
                    if value == 'FAIL':
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                elif 'Gate-B' in headers and col_idx == headers.index('Gate-B') + 1:
                    if value == 'FAIL':
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # 컬럼 너비 조정
        ws_critical.column_dimensions['A'].width = 25
        for col in range(2, len(headers) + 1):
            ws_critical.column_dimensions[get_column_letter(col)].width = 15

    # 파일 저장
    wb.save(output_file)
    print(f"[OK] Excel 파일 생성 완료: {output_file}")
    return output_file


if __name__ == "__main__":
    # 최신 final_output 폴더 찾기
    base_dir = Path("01_EXECUTION_FILES")
    final_dirs = sorted(base_dir.glob("final_output_*"), key=lambda x: x.stat().st_mtime, reverse=True)

    if not final_dirs:
        print("[ERROR] final_output 폴더를 찾을 수 없습니다.")
        exit(1)

    final_dir = final_dirs[0]
    print(f"[INFO] 최신 final_output 폴더: {final_dir.name}")

    # 출력 파일명
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = base_dir / f"STAGE_DETAILED_RESULTS_{timestamp}.xlsx"

    # Excel 파일 생성
    create_stage_summary_excel(final_dir, output_file)

