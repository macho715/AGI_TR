#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
AGI SPMT Shuttle — Integrated Builder (SSOT-locked)

What this tool does (AGI scope)
1) Enforces a single coordinate convention (SSOT):
   - Input "Frame" values are **Fr(m) = station from AP in meters** (coord_mode=FR_M)
   - x_from_midship(m) = midship_from_ap_m - Fr(m)
2) Generates an Excel workbook with:
   - Stage_Config
   - Cargo_SPMT_Inputs (auto-filled, including CoG_x_m = x_from_midship)
   - Stage_Loads (itemized, per stage)
   - Stage_Summary (per stage)
   - Stage_Results + Stage-wise Cargo on Deck (Bryan-format summary)
   - LOG (warnings / assumptions / audit notes)
3) Exports CSVs (stage_loads.csv, stage_summary.csv, stage_results.csv)

Notes
- This script intentionally does NOT use "structural frame spacing" for coordinate conversion.
- If your input values are truly "frame indexes", DO NOT use this tool without a verified
  frame-index mapping; use a dedicated conversion table instead.

CLI Example
  python agi_spmt_unified.py --config spmt_shuttle_config_AGI.json --out_xlsx AGI_SPMT_Shuttle.xlsx --out_dir out/

"""

from __future__ import annotations

import argparse
import sys
import csv
import json
import math
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from agi_ssot import VesselSSOT, build_agi_default_ssot


# ------------------------
# Excel styling primitives
# ------------------------
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=14)
BOLD = Font(bold=True)

THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
WRAP = Alignment(vertical="top", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="top", wrap_text=True)

DEFAULT_DECIMALS = 3


# ------------------------
# Data models
# ------------------------
@dataclass
class StageItem:
    item: str
    weight_t: float
    fr_m_from_ap: float
    x_m_from_midship: float

@dataclass
class StageSummary:
    stage: int
    stage_name: str
    note: str
    total_ondeck_t: float
    lcg_x_m: float
    lcg_fr_m: float
    items: List[StageItem]

# Bryan-like headers used by stage_shuttle_autocalc.py
BRYAN_HEADERS = [
    "Stage",
    "Cargo Total(t) [calc]",
    "LCGx(m) [calc]",
    "Cargo Total(t) [SSOT]",
    "LCGx(m) [SSOT]",
    "ΔCargo(t)",
    "ΔLCGx(m)",
    "Note",
]

CARGO_HEADERS = [
    "Item",
    "Weight_t",
    "CoG_x_m",
    "CoG_y_m",
    "CoG_z_m",
    "Final_Frame",
    "Ramp_Frame",
    "L_m",
    "W_m",
    "H_m",
    "Remarks/Source",
]


# ------------------------
# Helpers
# ------------------------
def _as_float(v: Any) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, str) and v.strip() == "":
        return None
    try:
        return float(v)
    except Exception:
        return None

def r(v: float, nd: int = DEFAULT_DECIMALS) -> float:
    return round(float(v), nd)

def ensure_sheet(wb: Workbook, name: str):
    if name in wb.sheetnames:
        return wb[name]
    return wb.create_sheet(name)

def clear_sheet(ws) -> None:
    ws.delete_rows(1, ws.max_row)

def set_col_widths(ws, widths: List[float], start_col: int = 1) -> None:
    for i, w in enumerate(widths, start=start_col):
        ws.column_dimensions[get_column_letter(i)].width = float(w)

def write_header(ws, row: int, col: int, headers: List[str]) -> None:
    for j, h in enumerate(headers):
        c = ws.cell(row=row, column=col + j, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = BORDER

def style_table(ws, top_row: int, left_col: int, nrows: int, ncols: int) -> None:
    for r0 in range(top_row, top_row + nrows):
        for c0 in range(left_col, left_col + ncols):
            c = ws.cell(row=r0, column=c0)
            c.border = BORDER
            if r0 == top_row:
                c.alignment = CENTER
            else:
                # numeric-ish columns align right
                c.alignment = WRAP

def log_append(ws, msg: str) -> None:
    r0 = ws.max_row + 1
    ws.cell(row=r0, column=1, value=msg).alignment = WRAP


# ------------------------
# Config normalization
# ------------------------
def normalize_config(cfg_raw: Dict[str, Any]) -> Tuple[VesselSSOT, Dict[str, Any], List[str]]:
    """
    Accepts the existing nested schema used by spmt_shuttle_stage_builder.py and cargo_spmt_inputs_stagewise_autofill.py.

    Returns:
      - ssot (VesselSSOT)
      - cfg (flattened dict for downstream)
      - warnings list
    """
    warnings: List[str] = []

    site = str(cfg_raw.get("site", "AGI")).strip().upper()
    if site != "AGI":
        warnings.append(f"[SITE] config.site='{site}' (expected 'AGI' for this tool). Proceeding anyway.")

    ss = cfg_raw.get("ssot", {}) if isinstance(cfg_raw.get("ssot", {}), dict) else {}
    coord_mode = str(ss.get("coord_mode", "FR_M")).strip().upper()
    if coord_mode not in ("FR_M", "FRAME_INDEX"):
        warnings.append(f"[SSOT] Unknown coord_mode='{coord_mode}'. Forcing FR_M.")
        coord_mode = "FR_M"

    # Build SSOT (AGI default if unspecified)
    ssot = build_agi_default_ssot()
    ssot = VesselSSOT(
        lpp_m=float(ss.get("lpp_m", ssot.lpp_m)),
        midship_from_ap_m=float(ss.get("midship_from_ap_m", ssot.midship_from_ap_m)),
        coord_mode="FR_M" if coord_mode == "FR_M" else "FRAME_INDEX",
        frame_spacing_m=float(ss.get("frame_spacing_m", ssot.frame_spacing_m)),
        ap_offset_m=float(ss.get("ap_offset_m", ssot.ap_offset_m)),
        fr_range_m=(
            float(ss.get("fr_min_m", ssot.fr_range_m[0])),
            float(ss.get("fr_max_m", ssot.fr_range_m[1])),
        ),
    )

    if ssot.coord_mode != "FR_M":
        warnings.append(
            "[SSOT] coord_mode is FRAME_INDEX. For AGI, FR_M is the SSOT default. "
            "Use FRAME_INDEX only if you have a verified index-to-station definition."
        )

    w = cfg_raw.get("weights", {}) if isinstance(cfg_raw.get("weights", {}), dict) else {}
    p = cfg_raw.get("positions", {}) if isinstance(cfg_raw.get("positions", {}), dict) else {}
    m = cfg_raw.get("modes", {}) if isinstance(cfg_raw.get("modes", {}), dict) else {}

    # Flatten
    cfg: Dict[str, Any] = {
        "site": site,
        "midship_frame": ssot.midship_from_ap_m,  # compatibility name
        "coord_mode": ssot.coord_mode,
        "frame_spacing_m": ssot.frame_spacing_m,
        "ap_offset_m": ssot.ap_offset_m,

        # weights (tons)
        "w_tr_t": float(w.get("w_tr_t", 217.0)),
        "w_spmt_t": float(w.get("w_spmt_t", 54.2)),
        "w_beams_t": float(w.get("w_beams_t", 11.6)),
        "w_ppu_t": float(w.get("w_ppu_t", 7.2)),
        "assembly_total_t": float(w.get("assembly_total_t", 282.8)),
        "dunnage_t": float(w.get("dunnage_t", 0.0)),
        "preballast_t": float(w.get("preballast_t", 0.0)),

        # positions (interpreted per ssot.coord_mode)
        "tr1_final_frame": float(p.get("tr1_final_frame", 42.0)),
        "tr2_final_frame": float(p.get("tr2_final_frame", 40.0)),
        "ramp_frame": float(p.get("ramp_frame", p.get("ramp_mid_frame", 17.95))),
        "ramp_start_frame": float(p.get("ramp_start_frame", p.get("ramp_frame", 17.95))),
        "ramp_mid_frame": float(p.get("ramp_mid_frame", p.get("ramp_frame", 17.95))),
        "spmt_park_frame": float(p.get("spmt_park_frame", 10.0)),
        "ppu_park_frame": float(p.get("ppu_park_frame", 46.0)),
        "dunnage_frame": float(p.get("dunnage_frame", 31.0)),
        "preballast_pos": float(p.get("preballast_pos", 0.0)),

        # modes
        "final_spmt_mode": str(m.get("final_spmt_mode", "OFFBOARD")).strip().upper(),
        "beams_mode": str(m.get("beams_mode", "REMOVED")).strip().upper(),
        "ppu_mode": str(m.get("ppu_mode", "INCLUDED_IN_SPMT")).strip().upper(),
        "ppu_included_in_assembly": bool(m.get("ppu_included_in_assembly", True)),

        # operational flags
        "include_preballast_in_stage0": bool(cfg_raw.get("include_preballast_in_stage0", False)),
        "include_preballast_in_stage5": bool(cfg_raw.get("include_preballast_in_stage5", False)),
    }

    # Quick mode validation
    if cfg["final_spmt_mode"] not in ("ONBOARD", "OFFBOARD"):
        warnings.append(f"[MODE] final_spmt_mode='{cfg['final_spmt_mode']}' invalid. Forcing OFFBOARD.")
        cfg["final_spmt_mode"] = "OFFBOARD"

    if cfg["beams_mode"] not in ("ATTACHED", "REMOVED"):
        warnings.append(f"[MODE] beams_mode='{cfg['beams_mode']}' invalid. Forcing REMOVED.")
        cfg["beams_mode"] = "REMOVED"

    if cfg["ppu_mode"] not in ("INCLUDED_IN_SPMT", "SEPARATE_ONBOARD", "OFFBOARD"):
        warnings.append(f"[MODE] ppu_mode='{cfg['ppu_mode']}' invalid. Forcing INCLUDED_IN_SPMT.")
        cfg["ppu_mode"] = "INCLUDED_IN_SPMT"

    return ssot, cfg, warnings


def validate_positions(ssot: VesselSSOT, cfg: Dict[str, Any], *, strict: bool) -> List[str]:
    warnings: List[str] = []
    # Interpret frame inputs according to ssot
    pos_keys = [
        "tr1_final_frame", "tr2_final_frame",
        "ramp_frame", "ramp_start_frame", "ramp_mid_frame",
        "spmt_park_frame", "ppu_park_frame", "dunnage_frame",
        "preballast_pos",
    ]
    for k in pos_keys:
        fr = ssot.to_fr_m_from_ap(cfg.get(k))
        try:
            ssot.validate_fr_m(fr, name=k, strict=strict)
        except Exception as e:
            if strict:
                raise
            warnings.append(f"[COORD] {e}")
    return warnings


# ------------------------
# Stage builder (0..8)
# ------------------------
def build_stage_0_to_8(ssot: VesselSSOT, cfg: Dict[str, Any]) -> List[StageSummary]:
    midship = float(ssot.midship_from_ap_m)

    def fr_to_x(fr_m: float) -> float:
        return ssot.x_from_midship_m(fr_m)

    def make_item(name: str, w_t: float, fr_in: float) -> StageItem:
        fr_m = ssot.to_fr_m_from_ap(fr_in)
        if fr_m is None:
            raise ValueError(f"{name}: missing position (frame).")
        x = fr_to_x(fr_m)
        return StageItem(name, float(w_t), float(fr_m), float(x))

    w_tr = float(cfg["w_tr_t"])
    w_spmt = float(cfg["w_spmt_t"])
    w_beams = float(cfg["w_beams_t"])
    w_ppu = float(cfg["w_ppu_t"])
    w_asmb = float(cfg["assembly_total_t"])
    w_dun = float(cfg["dunnage_t"])
    w_pre = float(cfg.get("preballast_t", 0.0))

    tr1_fr = float(cfg["tr1_final_frame"])
    tr2_fr = float(cfg["tr2_final_frame"])
    ramp_fr = float(cfg["ramp_frame"])
    ramp_start_fr = float(cfg.get("ramp_start_frame", ramp_fr))
    ramp_mid_fr = float(cfg.get("ramp_mid_frame", ramp_fr))
    spmt_park_fr = float(cfg["spmt_park_frame"])
    ppu_park_fr = float(cfg["ppu_park_frame"])
    dunnage_fr = float(cfg["dunnage_frame"])
    preballast_pos = float(cfg.get("preballast_pos", 0.0))

    final_spmt_mode = str(cfg["final_spmt_mode"]).upper()
    beams_mode = str(cfg["beams_mode"]).upper()
    ppu_mode = str(cfg["ppu_mode"]).upper()
    ppu_in_assembly = bool(cfg["ppu_included_in_assembly"])

    include_pre0 = bool(cfg.get("include_preballast_in_stage0", False))
    include_pre5 = bool(cfg.get("include_preballast_in_stage5", False))

    def items_tr_fixed(tr1_loaded: bool, tr2_loaded: bool) -> List[StageItem]:
        items: List[StageItem] = []
        if tr1_loaded:
            items.append(make_item("TR1 (fixed on stools)", w_tr, tr1_fr))
            if beams_mode == "ATTACHED":
                items.append(make_item("Beams (TR1)", w_beams, tr1_fr))
        if tr2_loaded:
            items.append(make_item("TR2 (fixed on stools)", w_tr, tr2_fr))
            if beams_mode == "ATTACHED":
                items.append(make_item("Beams (TR2)", w_beams, tr2_fr))
        if w_dun > 0:
            items.append(make_item("Dunnage/Load spreading", w_dun, dunnage_fr))
        return items

    def items_stage0() -> List[StageItem]:
        items: List[StageItem] = []
        if include_pre0 and w_pre > 0:
            items.append(make_item("Pre-ballast equivalent (config)", w_pre, preballast_pos))
        return items

    def items_stage5_base() -> List[StageItem]:
        items = items_tr_fixed(tr1_loaded=True, tr2_loaded=False)
        if include_pre5 and w_pre > 0:
            items.append(make_item("Pre-ballast equivalent (config)", w_pre, preballast_pos))
        return items

    def items_stage8_final() -> List[StageItem]:
        items = items_tr_fixed(tr1_loaded=True, tr2_loaded=True)

        # Final SPMT logic
        if final_spmt_mode == "ONBOARD":
            items.append(make_item("SPMT (empty, parked)", w_spmt, spmt_park_fr))
            if ppu_mode == "SEPARATE_ONBOARD":
                items.append(make_item("PPU (separate onboard)", w_ppu, ppu_park_fr))
        # If OFFBOARD: do nothing

        return items

    # Build stages (names match stage_shuttle_autocalc.py for pipeline compatibility)
    stage_defs: List[Tuple[int, str, str, List[StageItem]]] = []

    stage_defs.append((0, "Stage 0 Pre-ballast ready", "", items_stage0()))

    stage_defs.append((1, "Stage 1 25% TR1 on deck @ ramp start", "", [
        make_item("TR1+SPMT+Beams (assembly, 25%)", w_asmb * 0.25, ramp_start_fr),
        *([] if w_dun <= 0 else [make_item("Dunnage/Load spreading", w_dun, dunnage_fr)]),
    ]))

    stage_defs.append((2, "Stage 2 50% TR1 on deck @ ramp mid (HOLD)", "HOLD", [
        make_item("TR1+SPMT+Beams (assembly, 50%)", w_asmb * 0.50, ramp_mid_fr),
        *([] if w_dun <= 0 else [make_item("Dunnage/Load spreading", w_dun, dunnage_fr)]),
    ]))

    stage_defs.append((3, "Stage 3 75% TR1 on deck @ ramp mid", "", [
        make_item("TR1+SPMT+Beams (assembly, 75%)", w_asmb * 0.75, ramp_mid_fr),
        *([] if w_dun <= 0 else [make_item("Dunnage/Load spreading", w_dun, dunnage_fr)]),
    ]))

    stage_defs.append((4, "Stage 4 TR1 fully on deck @ final (pre-stow)", "", [
        make_item("TR1+SPMT+Beams (assembly, 100%)", w_asmb, tr1_fr),
        *([] if w_dun <= 0 else [make_item("Dunnage/Load spreading", w_dun, dunnage_fr)]),
    ]))

    stage_defs.append((5, "Stage 5 Pre-ballast TR2 ready (SPMT off deck)", "", items_stage5_base()))

    stage_defs.append((6, "Stage 6 50% TR2 on deck @ ramp mid (HOLD)", "HOLD", [
        *items_stage5_base(),
        make_item("TR2+SPMT+Beams (assembly, 50%)", w_asmb * 0.50, ramp_mid_fr),
    ]))

    stage_defs.append((7, "Stage 7 TR2 fully on deck @ final (pre-stow)", "", [
        *items_stage5_base(),
        make_item("TR2+SPMT+Beams (assembly, 100%)", w_asmb, tr2_fr),
    ]))

    # Stage 8 transit / final deck load
    stage_defs.append((8, "Stage 8 Transit (final stow)", "", items_stage8_final()))

    # Sanity / warning notes for PPU
    ppu_note = ""
    if ppu_mode == "SEPARATE_ONBOARD" and ppu_in_assembly:
        ppu_note = "WARNING: ppu_mode=SEPARATE_ONBOARD but ppu_included_in_assembly=True (double count risk)."

    # Compute summaries
    out: List[StageSummary] = []
    for stage_idx, stage_name, note, items in stage_defs:
        if ppu_note and stage_idx == 8:
            note = (note + " | " if note else "") + ppu_note

        tw = sum(i.weight_t for i in items)
        if tw > 0:
            lcg_x = sum(i.weight_t * i.x_m_from_midship for i in items) / tw
        else:
            lcg_x = 0.0
        lcg_fr = ssot.fr_from_x_from_midship_m(lcg_x)  # inverse of x=midship-fr
        out.append(StageSummary(
            stage=stage_idx,
            stage_name=stage_name,
            note=note,
            total_ondeck_t=r(tw, 2),
            lcg_x_m=r(lcg_x, 3),
            lcg_fr_m=r(lcg_fr, 3),
            items=items,
        ))
    return out


# ------------------------
# Outputs (CSV + Excel)
# ------------------------
def export_csv(out_dir: Path, stages: List[StageSummary]) -> Dict[str, Path]:
    out_dir.mkdir(parents=True, exist_ok=True)

    # stage_loads.csv
    stage_loads_path = out_dir / "stage_loads.csv"
    with stage_loads_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Stage", "Stage_Name", "Item", "Weight_t", "Fr_m_from_AP", "x_m_from_Midship"])
        for s in stages:
            for it in s.items:
                w.writerow([s.stage, s.stage_name, it.item, r(it.weight_t, 3), r(it.fr_m_from_ap, 3), r(it.x_m_from_midship, 3)])

    # stage_summary.csv
    stage_summary_path = out_dir / "stage_summary.csv"
    with stage_summary_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Stage", "Stage_Name", "Total_onDeck_t", "LCG_x_from_Midship_m", "LCG_Fr_from_AP_m", "Note"])
        for s in stages:
            w.writerow([s.stage, s.stage_name, r(s.total_ondeck_t, 2), r(s.lcg_x_m, 3), r(s.lcg_fr_m, 3), s.note])

    # stage_results.csv (Bryan-style)
    stage_results_path = out_dir / "stage_results.csv"
    with stage_results_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(BRYAN_HEADERS)
        for s in stages:
            # No external SSOT comparison here; we output calc == SSOT for internal consistency
            w.writerow([s.stage, r(s.total_ondeck_t, 2), r(s.lcg_x_m, 3), r(s.total_ondeck_t, 2), r(s.lcg_x_m, 3), 0.0, 0.0, s.note])

    return {
        "stage_loads_csv": stage_loads_path,
        "stage_summary_csv": stage_summary_path,
        "stage_results_csv": stage_results_path,
    }


def write_stage_config(ws, ssot: VesselSSOT, cfg: Dict[str, Any], warnings: List[str]) -> None:
    clear_sheet(ws)
    ws["A1"] = "Stage_Config (AGI SPMT Shuttle — SSOT Locked)"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")

    ws["A2"] = "SSOT: Fr(m)=station from AP,  x_from_midship = midship_from_ap - Fr"
    ws.merge_cells("A2:D2")

    # Key/Value table
    headers = ["Key", "Value", "Unit/Type", "Comment"]
    write_header(ws, 4, 1, headers)
    row = 5

    def kv(k: str, v: Any, unit: str = "", comment: str = ""):
        nonlocal row
        ws.cell(row=row, column=1, value=k)
        ws.cell(row=row, column=2, value=str(v))
        ws.cell(row=row, column=3, value=unit)
        ws.cell(row=row, column=4, value=comment)
        for c in range(1, 5):
            ws.cell(row=row, column=c).border = BORDER
            ws.cell(row=row, column=c).alignment = WRAP
        row += 1

    for k, v in ssot.explain().items():
        kv(f"ssot.{k}", v, "", "")

    kv("site", cfg.get("site"), "", "AGI expected")
    for k in [
        "w_tr_t","w_spmt_t","w_beams_t","w_ppu_t","assembly_total_t","dunnage_t","preballast_t",
        "tr1_final_frame","tr2_final_frame","ramp_frame","ramp_start_frame","ramp_mid_frame",
        "spmt_park_frame","ppu_park_frame","dunnage_frame","preballast_pos",
        "final_spmt_mode","beams_mode","ppu_mode","ppu_included_in_assembly",
        "include_preballast_in_stage0","include_preballast_in_stage5",
    ]:
        kv(k, cfg.get(k), "", "")

    # Warnings
    row += 1
    ws.cell(row=row, column=1, value="Warnings / Notes").font = BOLD
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 1
    if not warnings:
        warnings = ["(none)"]
    for w in warnings:
        ws.cell(row=row, column=1, value=w).alignment = WRAP
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 1

    set_col_widths(ws, [26, 22, 14, 50], start_col=1)


def write_cargo_inputs(ws, ssot: VesselSSOT, cfg: Dict[str, Any]) -> None:
    clear_sheet(ws)
    ws["A1"] = "Cargo / SPMT Inputs (AGI) — AutoFill (CoG_x_m = x_from_midship)"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:K1")

    ws["A2"] = f"SSOT: x_from_midship(m) = {ssot.midship_from_ap_m:.3f} - Fr(m from AP). (Not structural frame spacing.)"
    ws.merge_cells("A2:K2")

    write_header(ws, 4, 1, CARGO_HEADERS)
    set_col_widths(ws, [32, 12, 12, 12, 12, 12, 12, 10, 10, 10, 44], start_col=1)

    def row_write(r0: int, values: Dict[str, Any]):
        for j, h in enumerate(CARGO_HEADERS, start=1):
            v = values.get(h, "")
            c = ws.cell(row=r0, column=j, value=v)
            c.border = BORDER
            if isinstance(v, (int, float)):
                c.number_format = "0.00"
                c.alignment = RIGHT
            else:
                c.alignment = WRAP

    def make_row(item: str, weight_t: float, final_frame: Optional[float], ramp_frame: Optional[float], remarks: str) -> Dict[str, Any]:
        row: Dict[str, Any] = {"Item": item, "Weight_t": r(weight_t, 2), "Final_Frame": final_frame, "Ramp_Frame": ramp_frame, "Remarks/Source": remarks}
        if final_frame is not None:
            fr_m = ssot.to_fr_m_from_ap(final_frame)
            if fr_m is not None:
                row["CoG_x_m"] = r(ssot.x_from_midship_m(fr_m), 3)
        return row

    r0 = 5
    # Assembly rows
    row_write(r0, make_row("TR1+SPMT+Beams (Assembly Total)", cfg["assembly_total_t"], cfg["tr1_final_frame"], cfg["ramp_frame"],
                           "Used for ramp/transit stage modeling; CoG_x computed from Final_Frame via SSOT."))
    r0 += 1
    row_write(r0, make_row("TR2+SPMT+Beams (Assembly Total)", cfg["assembly_total_t"], cfg["tr2_final_frame"], cfg["ramp_frame"],
                           "Used for ramp/transit stage modeling; CoG_x computed from Final_Frame via SSOT."))
    r0 += 1

    # Optional separate items (final stage)
    row_write(r0, make_row("SPMT (empty, if ONBOARD at Stage 8)", cfg["w_spmt_t"], cfg["spmt_park_frame"], None,
                           "Included only when final_spmt_mode=ONBOARD."))
    r0 += 1
    row_write(r0, make_row("PPU (if separate onboard)", cfg["w_ppu_t"], cfg["ppu_park_frame"], None,
                           "Included only when ppu_mode=SEPARATE_ONBOARD. Verify double-count vs assembly_total_t."))
    r0 += 1
    row_write(r0, make_row("Dunnage / Load Spreading", cfg["dunnage_t"], cfg["dunnage_frame"], None,
                           "If used, ensure weight/position are measured or agreed."))
    r0 += 1

    # Placeholder row
    row_write(r0, {"Item": "Other / Consumables (optional)", "Remarks/Source": ""})


def write_stage_loads(ws, stages: List[StageSummary]) -> None:
    clear_sheet(ws)
    ws["A1"] = "Stage_Loads (itemized)"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")

    headers = ["Stage", "Stage_Name", "Item", "Weight_t", "Fr_m_from_AP", "x_m_from_Midship"]
    write_header(ws, 3, 1, headers)
    set_col_widths(ws, [8, 42, 40, 12, 14, 16], start_col=1)

    r0 = 4
    for s in stages:
        for it in s.items:
            ws.cell(row=r0, column=1, value=s.stage).border = BORDER
            ws.cell(row=r0, column=2, value=s.stage_name).border = BORDER
            ws.cell(row=r0, column=3, value=it.item).border = BORDER
            ws.cell(row=r0, column=4, value=r(it.weight_t, 3)).border = BORDER
            ws.cell(row=r0, column=5, value=r(it.fr_m_from_ap, 3)).border = BORDER
            ws.cell(row=r0, column=6, value=r(it.x_m_from_midship, 3)).border = BORDER
            for c in range(1, 7):
                ws.cell(row=r0, column=c).alignment = WRAP if c in (2,3) else RIGHT
            r0 += 1


def write_stage_summary(ws, stages: List[StageSummary]) -> None:
    clear_sheet(ws)
    ws["A1"] = "Stage_Summary"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")

    headers = ["Stage", "Stage_Name", "Total_onDeck_t", "LCG_x_from_Midship_m", "LCG_Fr_from_AP_m", "Note"]
    write_header(ws, 3, 1, headers)
    set_col_widths(ws, [8, 42, 14, 18, 16, 44], start_col=1)

    r0 = 4
    for s in stages:
        ws.cell(row=r0, column=1, value=s.stage).border = BORDER
        ws.cell(row=r0, column=2, value=s.stage_name).border = BORDER
        ws.cell(row=r0, column=3, value=r(s.total_ondeck_t, 2)).border = BORDER
        ws.cell(row=r0, column=4, value=r(s.lcg_x_m, 3)).border = BORDER
        ws.cell(row=r0, column=5, value=r(s.lcg_fr_m, 3)).border = BORDER
        ws.cell(row=r0, column=6, value=s.note).border = BORDER
        for c in range(1, 7):
            ws.cell(row=r0, column=c).alignment = WRAP if c in (2,6) else RIGHT
        r0 += 1


def write_bryan_sheets(ws_results, ws_stagewise, stages: List[StageSummary]) -> None:
    for ws in (ws_results, ws_stagewise):
        clear_sheet(ws)
        ws["A1"] = "Stage results (Bryan-format)"
        ws["A1"].font = TITLE_FONT
        ws.merge_cells("A1:H1")

        write_header(ws, 3, 1, BRYAN_HEADERS)
        set_col_widths(ws, [8, 18, 14, 18, 14, 10, 10, 44], start_col=1)

        r0 = 4
        for s in stages:
            row = [s.stage, r(s.total_ondeck_t, 2), r(s.lcg_x_m, 3), r(s.total_ondeck_t, 2), r(s.lcg_x_m, 3), 0.0, 0.0, s.note]
            for j, v in enumerate(row, start=1):
                c = ws.cell(row=r0, column=j, value=v)
                c.border = BORDER
                c.alignment = RIGHT if j not in (1,8) else WRAP
            r0 += 1


def build_workbook(
    ssot: VesselSSOT,
    cfg: Dict[str, Any],
    warnings: List[str],
    stages: List[StageSummary],
    out_xlsx: Path,
    template_xlsx: Optional[Path] = None,
) -> Path:
    if template_xlsx:
        wb = load_workbook(template_xlsx)
    else:
        wb = Workbook()

    # Ensure sheets
    ws_cfg = ensure_sheet(wb, "Stage_Config")
    ws_cargo = ensure_sheet(wb, "Cargo_SPMT_Inputs")
    ws_loads = ensure_sheet(wb, "Stage_Loads")
    ws_sum = ensure_sheet(wb, "Stage_Summary")
    ws_res = ensure_sheet(wb, "Stage_Results")
    ws_stagewise = ensure_sheet(wb, "Stage-wise Cargo on Deck")
    ws_log = ensure_sheet(wb, "LOG")

    # Write content
    write_stage_config(ws_cfg, ssot, cfg, warnings)
    write_cargo_inputs(ws_cargo, ssot, cfg)
    write_stage_loads(ws_loads, stages)
    write_stage_summary(ws_sum, stages)
    write_bryan_sheets(ws_res, ws_stagewise, stages)

    # LOG
    clear_sheet(ws_log)
    ws_log["A1"] = "LOG"
    ws_log["A1"].font = TITLE_FONT
    ws_log.merge_cells("A1:D1")
    ws_log["A2"] = "Notes / warnings emitted during build"
    ws_log.merge_cells("A2:D2")

    if not warnings:
        warnings = ["(none)"]
    for w in warnings:
        log_append(ws_log, w)

    # Save
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_xlsx)
    return out_xlsx


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--config", required=True, help="Config JSON (AGI SSOT schema).")
    ap.add_argument("--out_xlsx", required=True, help="Output XLSX path.")
    ap.add_argument("--out_dir", default="", help="Output directory for CSV exports (optional).")
    ap.add_argument("--template_xlsx", default="", help="Existing XLSX template to update (optional).")
    ap.add_argument("--strict_coords", action="store_true", help="Fail if any position is outside expected range.")
    ap.add_argument("--audit", action="store_true", help="Run coordinate consistency audit after build.")
    args = ap.parse_args()

    cfg_raw = json.loads(Path(args.config).read_text(encoding="utf-8"))
    ssot, cfg, warnings = normalize_config(cfg_raw)

    # Enforce FR_M for AGI unless explicitly provided as FRAME_INDEX (warned above)
    coord_warnings = validate_positions(ssot, cfg, strict=bool(args.strict_coords))
    warnings.extend(coord_warnings)

    stages = build_stage_0_to_8(ssot, cfg)

    out_xlsx = Path(args.out_xlsx)
    tpl = Path(args.template_xlsx) if args.template_xlsx else None
    build_workbook(ssot, cfg, warnings, stages, out_xlsx, template_xlsx=tpl)

    if args.out_dir:
        export_csv(Path(args.out_dir), stages)

    print(f"[OK] Wrote XLSX: {out_xlsx}")
    if args.out_dir:
        print(f"[OK] Wrote CSVs to: {Path(args.out_dir)}")

    if args.audit:
        audit_script = Path(__file__).resolve().with_name("audit_coordinate_consistency.py")
        report_path = out_xlsx.parent / "coord_audit.md"
        if audit_script.exists():
            import subprocess

            argv = [
                sys.executable,
                str(audit_script),
                "--root",
                str(Path(__file__).resolve().parent),
                "--report",
                str(report_path),
            ]
            result = subprocess.run(argv, capture_output=True, text=True)
            if result.returncode != 0:
                print(f"[WARN] Audit failed (rc={result.returncode}).")
                if result.stdout:
                    print(result.stdout.strip())
                if result.stderr:
                    print(result.stderr.strip())
            else:
                print(f"[OK] Audit report: {report_path}")
        else:
            print(f"[WARN] Audit script missing: {audit_script}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
